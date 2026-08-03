#!/usr/bin/env python3
"""
build_competitor_analysis.py
Source: data_auction_insights.json (from Google Ads Auction Insights screenshots)
Per city per campaign: who is outbidding us, do they have GMB, review gap, decision.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(DIR, "data_auction_insights.json")) as f:
    AI = json.load(f)
with open(os.path.join(DIR, "data_competition.json")) as f:
    COMP = json.load(f)
with open(os.path.join(DIR, "data_rival_enrichment.json")) as f:
    ENRICH = json.load(f)

# ── Domain knowledge: (entity_type, gmb_status, their_reviews, note) ──────
# gmb_status: "Confirmed" / "Likely" / "Web-only" / "Partial" / "N/A"

DOMAIN_META = {
    "drkamarajhospital.com":  ("Clinic - men's health", "Confirmed", 923,
        "Primary SH paid competitor Chennai. 923 GMB reviews vs our 584. Direct threat on bids + map-pack."),
    "arunmuthuvel.com":       ("Clinic - sexologist", "Confirmed", None,
        "Dr Arun Muthuvel Chennai. Confirmed GMB. IS=11%, posAbove=43-53%. Outbid marginally."),
    "homeocare.in":           ("Clinic chain - homeopathy", "Confirmed", 590,
        "National homeopathy chain. Confirmed GMB, 590 reviews (Pune Hadapsar 0.5km, HYD Borabanda 0.3km). Different pathology — let go on reviews."),
    "topsexologistinpune.com":("Clinic - sexologist", "Confirmed", None,
        "Top Sexologist Pune. Confirmed GMB. IS=35% — second-biggest SH paid competitor Pune. posAbove=73%. Priority outbid."),
    "drerandes.com":          ("Clinic - sexologist", "Confirmed", 284,
        "Dr Erande Pune. Confirmed GMB, 284 SERP appearances. posAbove=44%, outranks us 58% when both show."),
    "milann.co.in":           ("Clinic - fertility", "Confirmed", None,
        "Milann Fertility Centre Bangalore. Confirmed GMB (large chain). Fertility category — let go on reviews, outbid as needed."),
    "themenscareclinic.com":  ("Clinic - men's health (Medi Life)", "Confirmed", None,
        "Medi Life Bangalore (themenscareclinic.com). Confirmed GMB. Direct SH competitor. posAbove=77% — outranks us most of the time. Priority review + bid target."),
    "nuhospitals.com":        ("Hospital - urology", "Confirmed", None,
        "NU Hospitals Bangalore. Confirmed GMB (large hospital). Urology/andrology adjacent. posAbove=74%. Outbid, let go on reviews."),
    "oasisindia.in":          ("Clinic - fertility (Oasis IVF)", "Confirmed", None,
        "Oasis Fertility / Oasis IVF. Confirmed GMB multi-city. Fertility adjacent. posAbove=77-86%. Outbid, let go on reviews."),
    "ovumfertility.in":       ("Clinic - fertility", "Confirmed", None,
        "Ovum Fertility Bangalore. Confirmed GMB. Adjacent category — let go on reviews."),
    "drsafehands.com":        ("Clinic - STI/HIV national", "Confirmed", 140,
        "DrSafeHands — primary STI competitor in all cities. IS=53-64%, posAbove=88-97%. 140 GMB reviews wins on S2/S3 content (confidential, HIV/PEP). Biggest single threat."),
    "orangehealth.in":        ("Diagnostic Lab", "Confirmed", 19000,
        "Orange Health Labs. 19k reviews. Lab category — let go on reviews. Outbid only (STI IS 32-38%)."),
    "vijayadiagnostic.com":   ("Diagnostic Lab", "Confirmed", 33300,
        "Vijaya Diagnostics. 33k reviews. Lab category — let go on reviews entirely."),
    "suburbandiagnostics.com":("Diagnostic Lab", "Confirmed", None,
        "Suburban Diagnostics Mumbai. Lab category — let go on reviews."),
    "metropolisindia.com":    ("Diagnostic Lab", "Confirmed", None,
        "Metropolis Labs. Lab category — let go on reviews."),
    "mensclinicgroup.com":    ("Clinic - men's health / STI", "Confirmed", 468,
        "Men's Clinic Group. Confirmed GMB (multi-city chain). IS=37% Hyderabad SH + STI. posAbove=65%. 468 reviews near Kukatpally 2km. Priority outbid + review target Hyderabad."),
    "practo.com":             ("Aggregator / Platform", "Web-only", None,
        "Practo is a listing platform. The domain has no single GMB - clinics listed on it have their own. Let go - their ads don't compete for map-pack calls. No review competition."),
    "apollohospitals.com":    ("Hospital chain", "Confirmed", None,
        "Apollo Hospitals. Has GMB but different category (gynaecology, general surgery). posAbove high but different intent. Outbid if IS impact large, let go on reviews."),
    "manipalhospitals.com":   ("Hospital chain", "Confirmed", 3334,
        "Manipal Hospitals. Confirmed GMB, 3334 reviews (Manipal Clinics Brookefield, 0.2km from our clinic). Different category. Let go on reviews. Outbid if needed."),
    "hexahealth.com":         ("Platform - surgical aggregator", "Web-only", None,
        "HexaHealth is a procedure-aggregation platform. No clinic GMB. Let go."),
    "even.in":                ("Health App / Platform", "Web-only", None,
        "Even Health - digital health subscription. No physical GMB. Let go entirely."),
    "pristyncare.com":        ("Platform - surgical aggregator", "Web-only", None,
        "Pristyn Care - procedure aggregation. No single GMB. Let go."),
    "1mg.com":                ("Platform - pharmacy/health", "Web-only", None,
        "Tata 1mg pharmacy and health app. No clinic GMB. Let go entirely."),
    "pharmeasy.in":           ("Platform - pharmacy", "Web-only", None,
        "PharmEasy online pharmacy. No clinic GMB. Let go."),
    "freehivtest.org.in":     ("Awareness Site", "Web-only", None,
        "Free HIV test awareness website. No commercial intent. Let go."),
    "letsfighthpv.com":       ("Awareness Site", "Web-only", None,
        "HPV awareness campaign site. No clinic GMB. Let go."),
    "google.com":             ("Google Health Panel", "N/A", None,
        "Google's own health knowledge cards in search. Cannot compete with these - they are Google's own results."),
    "cadabamshospitals.com":  ("Hospital - mental health", "Confirmed", None,
        "Cadabams Hospitals. Has GMB Bangalore. MH category. IS=42%, posAbove=77%. Direct MH competitor."),
    "amahahealth.com":        ("Platform - mental health", "Partial", None,
        "Amaha Health - digital + some physical clinics. Has some GMB listings. IS=36%, posAbove=65%."),
    "maargamindcare.in":      ("Clinic - mental health", "Confirmed", None,
        "Maarga Mind Care Bangalore. Confirmed GMB. IS=27%, posAbove=38%. Compete on bids + reviews."),
    "ayoo.care":              ("Platform", "Web-only", None, "Digital mental health app. No GMB."),
    "mindtalk.in":            ("Platform", "Web-only", None, "Online counselling platform. No GMB."),
    "mibo.care":              ("Platform", "Web-only", None, "Digital health app. No GMB."),
    "anc.clinic":             ("Clinic - mental health", "Confirmed", None,
        "ANC Clinic Bangalore. Confirmed GMB. IS=12%, posAbove=48%. Compete on bids + reviews."),
    "oppam.me":               ("Platform", "Web-only", None, "Mental wellness app. No GMB."),
    "rockethealth.app":       ("Platform", "Web-only", None, "Digital health platform. No GMB."),
    "sukoonhealth.info":      ("Platform", "Web-only", None, "Mental health platform. No GMB."),
    "mpowerminds.com":        ("Clinic chain - mental health", "Confirmed", None,
        "MPower Minds. Confirmed GMB across cities."),
    # SH Mumbai / Navi Mumbai
    "drskjainclinic.com":           ("Clinic - sexologist (Burlington)", "Confirmed", None,
        "Dr SK Jain Burlington Clinic Mumbai. Confirmed GMB. IS=38-40% Mumbai + Navi Mumbai. posAbove=42-58%. Very well-known HIV/STI clinic. Priority outbid."),
    "ashakiranclinic.com":          ("Clinic - sexologist", "Confirmed", None,
        "Asha Kiran Clinic Mumbai. Confirmed GMB. IS=26-32% Mumbai + Navi Mumbai. posAbove=55-57%. Outbid + check review gap."),
    "drpriyankkothari.in":          ("Clinic - urologist", "Confirmed", None,
        "Dr Priyank Kothari Mumbai. Confirmed GMB. Low IS (<10%) but posAbove=40%. Outbid."),
    "kayakalpinternational.net.in": ("Clinic - ayurveda/sexual", "Confirmed", 520,
        "Kayakalp International Mumbai. Confirmed GMB, 520 reviews. near Dadar 0.8km. posAbove=94% — almost always above us. Low IS (<10%) so budget impact small. Monitor."),
    "gautamayurveda.co.in":         ("Clinic - ayurveda", "Confirmed", 130,
        "Gautam Ayurveda (Dr. Swapnil Gautam) Mumbai/Vasai. Confirmed GMB, 130 reviews. posAbove=79%, absTop=55%. Low IS (<10%). Monitor."),
    # SH Hyderabad
    "apollofertility.com":          ("Clinic - fertility", "Confirmed", None,
        "Apollo Fertility Hyderabad. Confirmed GMB. IS<10%, fertility adjacent. Let go on reviews."),
    # STI Chennai
    "metromaleclinic.com":          ("Clinic - men's health/STI", "Confirmed", 593,
        "Metromale Clinic & Fertility Center Chennai. Confirmed GMB, 593 reviews. IS=16% Chennai STI, posAbove=41%. Direct STI competitor. Outbid + close review gap."),
    # Labs
    "lalpathlabs.com":              ("Diagnostic Lab", "Confirmed", None,
        "Dr Lal PathLabs. National diagnostic chain. IS=13% Bangalore STI. Lab — let go on reviews. Outbid if IS>20%."),
    # Awareness sites
    "drhivaidsonline.com":          ("Awareness Site / Info", "Web-only", None,
        "DrHIVAIDSOnline — information site. No GMB. Let go."),
    # Navi Mumbai MH
    "manpravah.com":                ("Clinic - mental health", "Confirmed", None,
        "Manpravah MH clinic Navi Mumbai. Confirmed GMB. IS=27%, posAbove=45%. Compete on bids + reviews."),
    "trijog.com":                   ("Platform - mental health", "Partial", None,
        "Trijog. Has some physical presence + GMB. IS=12%, posAbove=84%, absTop=58% — very dominant when shown. Outbid."),
    "abhasa.org":                   ("Clinic - rehabilitation/MH", "Confirmed", None,
        "Abhasa Rehabilitation & Mental Health. Confirmed GMB. IS<10% but posAbove=87%, absTop=64% — extremely dominant when shown. Appears Pune MH too. Priority outbid."),
    "jagrutiwellness.com":          ("Clinic - rehabilitation", "Confirmed", None,
        "Jagruti Wellness / Jagruti Rehabilitation. Confirmed GMB. IS<10%, posAbove=58%. Outbid."),
    "butterflylearnings.com":       ("Platform - mental health", "Web-only", None,
        "Butterfly Learnings. Online mental health learning. No clinic GMB. Let go."),
    "zenuphealth.com":              ("Platform - mental health", "Web-only", None,
        "Zenup Health. Digital mental health app. No clinic GMB. Let go."),
    "betterhelp.com":               ("Platform - online therapy", "Web-only", None,
        "BetterHelp. International online therapy platform. No India clinic GMB. Let go."),
    # Pune MH
    "nityanandrehab.com":           ("Clinic - rehabilitation/MH", "Confirmed", None,
        "Nityanand Rehabilitation Centre Pune. Confirmed GMB. IS=17%, posAbove=59%. Outbid + compete on reviews."),
    "mindsightclinic.com":          ("Clinic - mental health", "Confirmed", None,
        "Mindsight Clinic Pune. Confirmed GMB. IS<10% but posAbove=82%, absTop=50% — very dominant when shown. Outbid."),
    "lissun.app":                   ("Platform - mental health", "Web-only", None,
        "Lissun. Mental health app + some physical presence. No single clinic GMB. Let go on reviews."),
}

# Allo reviews per city
ALLO_REVIEWS = {}
for cat in ("SH", "STI", "MH"):
    for city, cd in COMP[cat]["cities"].items():
        if city not in ALLO_REVIEWS:
            ALLO_REVIEWS[city] = cd.get("our_reviews", 0)

# ── Decision logic ─────────────────────────────────────────────────────────

def decide(domain, their_is, pos_above, allo_rev, their_rev):
    meta = DOMAIN_META.get(domain, ("Unknown", "Unknown", None, ""))
    typ, gmb, _, _ = meta

    if gmb in ("Web-only", "N/A"):
        return "LET GO", "Web/platform only — no GMB clinic profile. Cannot compete for map-pack calls/directions.", "F1F5F9", "6B7280"

    if "Lab" in typ or "Diagnostic" in typ:
        if their_is >= 20:
            return "OUTBID ONLY", "Lab taking IS — outbid. Let go on reviews (different category, 10k+ reviews irrelevant).", "FEF9C3", "A16207"
        return "MONITOR", "Diagnostic lab, low IS. Different category. Monitor quarterly.", "F1F5F9", "6B7280"

    if gmb in ("Confirmed", "Likely", "Partial"):
        gap_note = ""
        if their_rev and allo_rev:
            gap = their_rev - allo_rev
            gap_note = f" Review gap: {gap:+,}." if gap != 0 else " Reviews: we are ahead."

        if their_is >= 30 and pos_above >= 60:
            return (f"COMPETE: Outbid + Review Target{gap_note}",
                    f"Direct GMB clinic competitor. IS={their_is:.0f}% + above us {pos_above:.0f}% of time.",
                    "FECACA", "C42B2B")
        if their_is >= 15 or pos_above >= 50:
            return (f"COMPETE: Outbid + GMB Watch{gap_note}",
                    f"GMB competitor. IS={their_is:.0f}%, above us {pos_above:.0f}% of time. Track review velocity.",
                    "FED7AA", "C95A14")
        return ("COMPETE: Raise bids",
                f"GMB clinic, IS<15% but posAbove={pos_above:.0f}% — outranks us when it shows.",
                "FEF9C3", "A16207")

    return "MONITOR", "Insufficient data.", "F1F5F9", "6B7280"

def parse_pct(s):
    if not s or s == "< 10%": return 5.0
    return float(str(s).replace("%", "").strip())

# ── Build rows ─────────────────────────────────────────────────────────────

ROWS = []
CAT_ORDER  = {"SH": 0, "MH": 1, "STI": 2}
CITY_ORDER = {"Bangalore": 0, "Chennai": 1, "Pune": 2, "Hyderabad": 3, "Mumbai": 4}

for city, cats in AI["byCity"].items():
    for cat, data in cats.items():
        allo_rev    = ALLO_REVIEWS.get(city, 0)
        allo_is_str = data.get("you", "—")
        for comp in data.get("competitors", []):
            domain    = comp["domain"]
            their_is  = parse_pct(comp.get("is", "0"))
            overlap   = parse_pct(comp.get("overlap") or "0")
            pos_above = parse_pct(comp.get("posAbove") or "0")
            abs_top   = parse_pct(comp.get("absTop") or "0")
            outrank   = parse_pct(comp.get("outrank") or "0")
            meta = DOMAIN_META.get(domain, ("Unknown", "Unknown", None, ""))
            typ, gmb, their_rev, detail = meta
            # Override review count from enrichment if richer
            enrich = ENRICH.get(domain, {})
            enrich_rev = enrich.get("reviews")
            if enrich_rev and (their_rev is None or enrich_rev > their_rev):
                their_rev = enrich_rev
            gmb_url = enrich.get("maps_url", "")
            # Nearest clinics string
            nc = enrich.get("nearest_clinics", [])
            nc_filtered = [c for c in nc if c.get("km") is not None and c.get("city") == city]
            if not nc_filtered:
                nc_filtered = [c for c in nc if c.get("km") is not None]
            if nc_filtered:
                parts = [f"{c['loc'] or c['city']} ({c['km']:.1f}km · ours:{(c.get('our_reviews') or 0):,})" for c in nc_filtered[:3] if c.get("loc") or c.get("city")]
                near_str = " · ".join(parts)
            else:
                near_str = "—"
            dec, why, dec_bg, dec_fc = decide(domain, their_is, pos_above, allo_rev, their_rev)
            ROWS.append({
                "cat": cat, "city": city, "domain": domain,
                "typ": typ, "gmb": gmb,
                "their_is": their_is, "their_is_str": comp.get("is", "< 10%"),
                "overlap": overlap, "overlap_str": comp.get("overlap") or "—",
                "pos_above": pos_above, "pos_above_str": comp.get("posAbove") or "—",
                "abs_top": abs_top,     "abs_top_str":   comp.get("absTop") or "—",
                "outrank": outrank,     "outrank_str":   comp.get("outrank") or "—",
                "their_rev": their_rev, "allo_rev": allo_rev, "allo_is_str": allo_is_str,
                "detail": detail, "gmb_url": gmb_url, "near_str": near_str,
                "decision": dec, "why": why, "dec_bg": dec_bg, "dec_fc": dec_fc,
            })

ROWS.sort(key=lambda x: (CAT_ORDER.get(x["cat"], 9), CITY_ORDER.get(x["city"], 9), -x["their_is"]))

# ── Workbook ───────────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Auction Insights Analysis"
ws.sheet_view.showGridLines = False

COLS = [
    ("Cat",              5),
    ("City",            12),
    ("Competitor Domain\n(click = GMB link)", 26),
    ("Entity Type",     20),
    ("GMB?",             7),
    ("Their\nIS%",       8),
    ("Overlap\nRate%",   8),
    ("We Outrank\nThem%",8),
    ("They Above\nUs%",  9),
    ("Abs Top\nPage%",   7),
    ("Nearest Allo\nClinic",  24),
    ("Their\nReviews",   8),
    ("Allo\nReviews",    7),
    ("Review\nGap",      8),
    ("Our\nIS%",         6),
    ("Decision",        26),
    ("Why + Detail",    48),
]
NC = len(COLS)
for i, (_, w) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def cs(ws, row, col, val, bg="FFFFFF", fc="111111", bold=False, size=9,
       align="left", wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.fill      = PatternFill("solid", fgColor=bg)

def merge(ws, row, c1, c2, val, **kw):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    cs(ws, row, c1, val, **kw)

def rh(ws, row, h): ws.row_dimensions[row].height = h

def rag(v, red_t, yellow_t, invert=False):
    if invert:
        if v >= red_t:    return "BBF7D0", "15803D"
        if v >= yellow_t: return "FDE68A", "A16207"
        return "FCA5A5", "C42B2B"
    else:
        if v >= red_t:    return "FCA5A5", "C42B2B"
        if v >= yellow_t: return "FDE68A", "A16207"
        return "BBF7D0", "15803D"

SEC_BG = {"SH": "1E3A5F", "STI": "7C2D12", "MH": "3B0764"}
ROW_BG = {"SH": "EFF6FF", "STI": "FFF7ED", "MH": "FAF5FF"}
DIV_BG = {"SH": "BFDBFE", "STI": "FED7AA", "MH": "EDE9FE"}

# ── Title rows ─────────────────────────────────────────────────────────────

r = 1
merge(ws, r, 1, NC,
      "Google Ads Auction Insights — Who Is Outbidding Us + Do They Have GMB?",
      bg="0F172A", fc="FFFFFF", bold=True, size=13, align="center")
rh(ws, r, 24); r += 1

merge(ws, r, 1, NC,
      "Source: Google Ads Auction Insights (per-campaign screenshots, 1-26 Jul 2026)  "
      "|  SH: Bangalore · Chennai · Hyderabad · Mumbai · Navi Mumbai · Pune  "
      "|  STI: Bangalore · Chennai · Hyderabad · Mumbai · Pune  "
      "|  MH: Bangalore · Navi Mumbai · Pune",
      bg="1E293B", fc="94A3B8", size=8, italic=True, align="center")
rh(ws, r, 13); r += 1

merge(ws, r, 1, NC,
      "DECISION LOGIC:  GMB clinic + high IS → COMPETE (outbid + review target)  "
      "|  GMB clinic + low IS → COMPETE (outbid + watch)  "
      "|  Web/platform only → LET GO (no GMB = can't convert via calls/directions)  "
      "|  Diagnostic lab → LET GO on reviews, outbid only if IS >20%",
      bg="EFF6FF", fc="1E3A5F", bold=True, size=8, align="center")
rh(ws, r, 16); r += 1

merge(ws, r, 1, NC,
      "Domain cell = clickable GMB link   "
      "|   Nearest Allo Clinic = SERP data (distance · our reviews at that clinic)   "
      "|   Their IS%: 🔴 >30% · 🟡 15-30% · 🟢 <15%   "
      "|   Overlap%: 🔴 >70% · 🟡 40-70% · 🟢 <40% (how often we both appear in same auction)   "
      "|   They Above Us%: 🔴 >60% · 🟡 40-60% · 🟢 <40%   "
      "|   Abs Top%: 🔴 >50% · 🟡 25-50% · 🟢 <25%   "
      "|   Review Gap: 🔴 they lead >200 · 🟡 1-200 · 🟢 we lead",
      bg="F8FAFC", fc="374151", size=8, italic=True, align="center")
rh(ws, r, 13); r += 1

# ── Persistent frozen header row (always visible when scrolling) ────────────
FREEZE_ROW = r
for col, (h, _) in enumerate(COLS, 1):
    cs(ws, r, col, h, bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center", wrap=True)
rh(ws, r, 30); r += 1
ws.freeze_panes = f"A{r}"   # freeze everything above the first data row

def write_col_headers(ws, row, bg="1E293B"):
    for col, (h, _) in enumerate(COLS, 1):
        cs(ws, row, col, h, bg=bg, fc="FFFFFF", bold=True, size=8, align="center", wrap=True)
    rh(ws, row, 30)

# ── Main rows ──────────────────────────────────────────────────────────────

prev_cat = prev_city = None

for rd in ROWS:
    cat  = rd["cat"]
    city = rd["city"]

    if cat != prev_cat:
        merge(ws, r, 1, NC, "", bg="F8FAFC"); rh(ws, r, 8); r += 1
        sec_label = {"SH": "SH — Sexual Health", "STI": "STI — Sexually Transmitted Infections", "MH": "MH — Mental Health"}[cat]
        merge(ws, r, 1, NC, f"{sec_label}  |  Source: Google Ads Auction Insights",
              bg=SEC_BG[cat], fc="FFFFFF", bold=True, size=10, align="left")
        rh(ws, r, 18); r += 1
        write_col_headers(ws, r, bg=SEC_BG[cat]); r += 1
        prev_cat = cat; prev_city = None

    if city != prev_city:
        merge(ws, r, 1, NC,
              f"  {city.upper()}   |   Our IS: {rd['allo_is_str']}   |   Allo reviews this city: {rd['allo_rev']:,}",
              bg=DIV_BG[cat], fc=SEC_BG[cat], bold=True, size=9, align="left")
        rh(ws, r, 15); r += 1
        prev_city = city

    row_bg = ROW_BG[cat]

    # Colours
    their_is_bg, their_is_fc = rag(rd["their_is"], 30, 15)
    pos_above_bg, pos_above_fc = rag(rd["pos_above"], 60, 40)
    we_above    = 100 - rd["pos_above"]
    we_above_bg, we_above_fc = rag(we_above, 60, 40, invert=True)
    abs_top_bg, abs_top_fc = rag(rd["abs_top"], 50, 25)

    gmb_map = {
        "Confirmed": ("BBF7D0", "15803D"), "Likely": ("FDE68A", "A16207"),
        "Web-only":  ("FCA5A5", "C42B2B"), "Partial": ("FDE68A", "A16207"),
        "N/A":       ("F1F5F9", "6B7280"),
    }
    gmb_bg, gmb_fc = gmb_map.get(rd["gmb"], ("F1F5F9", "6B7280"))

    # Review gap
    if rd["their_rev"] is not None:
        gap = rd["their_rev"] - rd["allo_rev"]
        gap_str = f"+{gap:,}" if gap > 0 else str(gap)
        gap_bg, gap_fc = rag(gap, 200, 1)
        their_rev_str = f"{rd['their_rev']:,}"
    else:
        gap_str = "—"; gap_bg, gap_fc = row_bg, "6B7280"; their_rev_str = "—"

    overlap_bg, overlap_fc = rag(rd["overlap"], 70, 40)

    cells = [
        (cat,                    row_bg,      "center", 9,  True,  SEC_BG[cat],       False),
        (city,                   row_bg,      "left",   9,  False, "111111",           False),
        (rd["domain"],           row_bg,      "left",   8,  True,  "1D4ED8",           False),  # hyperlink col
        (rd["typ"],              row_bg,      "left",   8,  False, "374151",           True),
        (rd["gmb"],              gmb_bg,      "center", 8,  True,  gmb_fc,             False),
        (rd["their_is_str"],     their_is_bg, "center", 10, True,  their_is_fc,        False),
        (rd["overlap_str"],      overlap_bg,  "center", 9,  False, overlap_fc,         False),
        (f"{we_above:.0f}%",     we_above_bg, "center", 9,  False, we_above_fc,        False),
        (rd["pos_above_str"],    pos_above_bg,"center", 10, True,  pos_above_fc,       False),
        (rd["abs_top_str"],      abs_top_bg,  "center", 9,  True,  abs_top_fc,         False),
        (rd["near_str"],         "F0FDF4",    "left",   8,  False, "166534",           True),
        (their_rev_str,          row_bg,      "center", 9,  False, "374151",           False),
        (f"{rd['allo_rev']:,}",  row_bg,      "center", 9,  False, "374151",           False),
        (gap_str,                gap_bg,      "center", 9,  True,  gap_fc,             False),
        (rd["allo_is_str"],      row_bg,      "center", 8,  False, "374151",           False),
        (rd["decision"],         rd["dec_bg"],"left",   8,  True,  rd["dec_fc"],       True),
        (rd["detail"],           row_bg,      "left",   8,  False, "1E3A5F",           True),
    ]

    rh(ws, r, 52)
    for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(cells, 1):
        cs(ws, r, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)
    # Add hyperlink on domain cell (col 3) if we have a GMB URL
    if rd.get("gmb_url"):
        cell = ws.cell(row=r, column=3)
        cell.hyperlink = rd["gmb_url"]
        cell.font = Font(name="Arial", bold=True, size=8, color="1D4ED8", underline="single")
    r += 1

# ── Key Findings ───────────────────────────────────────────────────────────

merge(ws, r, 1, NC, "", bg="F8FAFC"); rh(ws, r, 8); r += 1
merge(ws, r, 1, NC, "KEY FINDINGS — Priority Actions", bg="0F172A", fc="FFFFFF", bold=True, size=10, align="left")
rh(ws, r, 18); r += 1

findings = [
    ("FECACA", "C42B2B",
     "STI: DrSafeHands is the #1 paid + GMB competitor in BOTH Hyderabad and Mumbai",
     "HYD STI: DrSafeHands IS=53% vs our 55% — neck and neck. They show ABOVE us 88% of the time. "
     "MUM STI: DrSafeHands IS=60% vs our 40% — beating us on IS. Above us 97% of the time. "
     "They have 140 GMB reviews but write S2/S3 reviews (confidential, HIV/PEP) which Google surfaces for those exact queries. "
     "Action: (1) Raise bids +25-30% on Mumbai + Hyderabad STI. (2) Add 5+ S2/S3 reviews per STI clinic per month — see Tab 5 of review workbook for exact language."),

    ("FED7AA", "C95A14",
     "SH Chennai: Dr Kamaraj Hospital — confirmed paid bidder, GMB confirmed, 923 reviews vs our 584",
     "drkamarajhospital.com IS=45% vs our 55%. Shows above us 70% of the time. Review gap = 339. "
     "Action: Raise SH Chennai bids +20%. Target +15-25 SH reviews/month for Chennai specifically. At +25/month, close the gap in 14 months."),

    ("FEF9C3", "A16207",
     "SH Pune: topsexologistinpune.com — IS=35%, posAbove=74%, likely has GMB",
     "Single clinic website in Pune with 35% IS. When they show, they're above us 74% of the time. "
     "Dr Erande's (drerandes.com) also present — outranks us 58% of the time. "
     "Action: Raise Pune SH bids +15-20%. Manually search topsexologistinpune.com on Google Maps to check their review count."),

    ("F0FDF4", "15803D",
     "SH Bangalore: fragmented — fix ad quality before raising bids",
     "No single dominant competitor. Practo (web-only, 29% IS) — let go. "
     "themenscareclinic.com (17% IS, posAbove=78%) — direct SH competitor with likely GMB — check their reviews. "
     "Milann, NU Hospitals, Oasis — different categories (fertility/urology), let go on reviews. "
     "Bangalore IS rank loss is partly self-inflicted: ad relevance=26% below-avg drags QS. Fix RSAs first, then raise bids."),

    ("EFF6FF", "1E3A5F",
     "Let go: Practo, Apollo, 1mg, HexaHealth, Pristyn, PharmEasy, Even, awareness sites",
     "These are web/platform players — Practo alone takes 25-38% IS in every SH city. "
     "They cannot convert via map-pack calls or directions because they have no single GMB clinic profile. "
     "Their traffic goes to listing pages where our clinics are sometimes shown too. "
     "Do NOT try to outbid them — the IS they take is not converting to clinic-intent calls. Focus budget on clinic competitors with real GMB."),
]

for hdr_bg, hdr_fc, header, body in findings:
    merge(ws, r, 1, NC, header, bg=hdr_bg, fc=hdr_fc, bold=True, size=9, align="left")
    rh(ws, r, 16); r += 1
    merge(ws, r, 1, NC, body, bg="F8FAFC", fc="111111", size=8, wrap=True)
    rh(ws, r, 54); r += 1
    merge(ws, r, 1, NC, "", bg="F1F5F9"); rh(ws, r, 5); r += 1

# ── Second sheet: GMB Review Detail ────────────────────────────────────────

ws2 = wb.create_sheet("GMB Review Detail")
ws2.sheet_view.showGridLines = False

DCOLS = [
    ("Cat",                 6),
    ("City",               12),
    ("Competitor\nDomain", 20),
    ("Their GMB Profile\nName (click = link)",  28),
    ("Their\nReviews",      8),
    ("Our Allo\nClinic",   18),
    ("Our\nReviews",        8),
    ("Gap\n(+= they lead)", 9),
    ("Dist\nkm",            7),
]
DNC = len(DCOLS)
for i, (_, w) in enumerate(DCOLS, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

def clean_name(n):
    return n.replace("My Ad Centre", "").replace("  ", " ").strip() if n else ""

# Build auction presence map: domain → set of cities in our auction
auction_cities = {}
for city, cats in AI["byCity"].items():
    for cat, data in cats.items():
        for comp in data.get("competitors", []):
            d = comp["domain"]
            auction_cities.setdefault(d, set()).add(city)

# Build detail rows: one per (competitor × our clinic), filtered to auction cities
DETAIL = []
for domain, enr in sorted(ENRICH.items()):
    meta = DOMAIN_META.get(domain, ("Unknown", "Unknown", None, ""))
    typ, gmb, _, detail_note = meta
    if gmb not in ("Confirmed", "Partial"):
        continue
    valid_cities = auction_cities.get(domain, set())
    for nc in enr.get("nearest_clinics", []):
        if nc["city"] not in valid_cities:
            continue
        DETAIL.append({
            "domain": domain,
            "typ": typ,
            "gmb": gmb,
            "cat": nc.get("cat", ""),
            "city": nc["city"],
            "their_name": clean_name(nc.get("their_name", "")),
            "their_reviews": nc.get("their_reviews") or 0,
            "their_maps": nc.get("their_maps", ""),
            "our_loc": nc.get("loc", "") or "",
            "our_reviews": nc.get("our_reviews") or 0,
            "km": nc.get("km", 0) or 0,
        })
    # Flag domains confirmed GMB but not in SERP enrichment (no per-clinic data)
    if not enr.get("nearest_clinics") and valid_cities:
        DETAIL.append({
            "domain": domain, "typ": typ, "gmb": gmb, "cat": "", "city": "—",
            "their_name": "(No SERP location data — check GMB manually)",
            "their_reviews": 0, "their_maps": "", "our_loc": "—", "our_reviews": 0, "km": 0,
        })

DETAIL.sort(key=lambda x: (x["domain"], x["city"], x["their_maps"], x["km"]))

# ── Title rows for sheet 2 ──────────────────────────────────────────────────
r2 = 1
merge(ws2, r2, 1, DNC,
      "GMB Review Detail — Every Competitor GMB Profile × Every Nearby Allo Clinic",
      bg="0F172A", fc="FFFFFF", bold=True, size=13, align="center")
rh(ws2, r2, 24); r2 += 1

merge(ws2, r2, 1, DNC,
      "Filtered to GMB-confirmed competitors appearing in our Google Ads auction   "
      "|   Each row = one of our clinics near that competitor's GMB profile   "
      "|   Gap = their reviews − ours (positive = they lead)   "
      "|   Click the GMB Profile Name cell to open their GMB listing",
      bg="1E293B", fc="94A3B8", size=8, italic=True, align="center")
rh(ws2, r2, 13); r2 += 1

# Frozen header
for col, (h, _) in enumerate(DCOLS, 1):
    cs(ws2, r2, col, h, bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center", wrap=True)
rh(ws2, r2, 30); r2 += 1
ws2.freeze_panes = f"A{r2}"

prev_domain = None
for rd in DETAIL:
    domain = rd["domain"]
    if domain != prev_domain:
        merge(ws2, r2, 1, DNC,
              f"  {domain}   |   {rd['typ']}   |   GMB: {rd['gmb']}",
              bg="1E3A5F", fc="FFFFFF", bold=True, size=9, align="left")
        rh(ws2, r2, 16); r2 += 1
        prev_domain = domain

    their_rev = rd["their_reviews"]
    our_rev = rd["our_reviews"]
    gap = their_rev - our_rev
    if gap > 200:
        gap_bg, gap_fc = "FCA5A5", "C42B2B"
    elif gap > 0:
        gap_bg, gap_fc = "FDE68A", "A16207"
    elif gap < 0:
        gap_bg, gap_fc = "BBF7D0", "15803D"
    else:
        gap_bg, gap_fc = "F1F5F9", "6B7280"
    gap_str = f"+{gap:,}" if gap > 0 else (f"{gap:,}" if gap < 0 else "0")

    cat_bg = {"SH": "EFF6FF", "STI": "FFF7ED", "MH": "FAF5FF"}.get(rd["cat"], "F8FAFC")
    cat_fc = {"SH": "1E3A5F", "STI": "7C2D12", "MH": "3B0764"}.get(rd["cat"], "6B7280")

    row_data = [
        (rd["cat"],                             cat_bg,   "center", 8, True,  cat_fc,   False),
        (rd["city"],                            "F8FAFC", "left",   8, False, "111111", False),
        (domain,                                "F8FAFC", "left",   7, False, "374151", False),
        (rd["their_name"],                      "FAFAFA", "left",   8, True,  "1D4ED8", True),
        (f"{their_rev:,}" if their_rev else "—","F8FAFC", "center", 9, False, "374151", False),
        (rd["our_loc"] or "—",                 "F0FDF4", "left",   8, False, "166534", False),
        (f"{our_rev:,}" if our_rev else "—",   "F0FDF4", "center", 9, False, "166534", False),
        (gap_str,                               gap_bg,   "center", 9, True,  gap_fc,   False),
        (f"{rd['km']:.1f}" if rd["km"] else "—","F8FAFC","center", 8, False, "6B7280", False),
    ]

    rh(ws2, r2, 16)
    for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(row_data, 1):
        cs(ws2, r2, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)

    # Make "Their GMB Profile Name" (col 4) a clickable link
    if rd.get("their_maps"):
        cell = ws2.cell(row=r2, column=4)
        cell.hyperlink = rd["their_maps"]
        cell.font = Font(name="Arial", bold=True, size=8, color="1D4ED8", underline="single")

    r2 += 1

# freeze_panes already set above after persistent header row
OUT = os.path.join(DIR, "allo_competitor_intel.xlsx")
wb.save(OUT)
print(f"Saved: {OUT}")
n_compete = sum(1 for x in ROWS if "COMPETE" in x["decision"])
n_letgo   = sum(1 for x in ROWS if "LET GO"  in x["decision"])
print(f"  {len(ROWS)} competitors | COMPETE={n_compete} | LET GO={n_letgo}")
print("  Top threats by their IS%:")
for x in sorted(ROWS, key=lambda x: -x["their_is"])[:6]:
    print(f"    [{x['cat']} {x['city']}] {x['domain']:<40} IS={x['their_is_str']:<8} posAbove={x['pos_above_str']}")
