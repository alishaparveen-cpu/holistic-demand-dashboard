#!/usr/bin/env python3
"""
Competitor Analysis refresh script.
Run: python3 refresh_competitor_analysis.py
Pulls the latest week's SERP scan and writes all_cities_serp_latest.xlsx to ~/Downloads.
"""

import subprocess, json, sys, time
from datetime import date
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
REDSHIFT = dict(profile='redshift-data', region='ap-south-1',
                cluster='warehouse', db='allo_prod')
TIER1 = {'Chennai','Hyderabad','Bangalore','Pune','Mumbai','Navi Mumbai'}
TIER1_ORDER = ['Bangalore','Chennai','Mumbai','Navi Mumbai','Hyderabad','Pune']
OUT = f"{__import__('os').path.expanduser('~')}/Downloads/all_cities_serp_latest.xlsx"

# ── Redshift helper ───────────────────────────────────────────────────────────
def run_sql(sql):
    r = subprocess.run([
        'aws','redshift-data','execute-statement',
        '--profile', REDSHIFT['profile'], '--region', REDSHIFT['region'],
        '--cluster-identifier', REDSHIFT['cluster'], '--database', REDSHIFT['db'],
        '--sql', sql
    ], capture_output=True, text=True)
    stmt_id = json.loads(r.stdout)['Id']
    while True:
        s = subprocess.run(['aws','redshift-data','describe-statement',
            '--profile', REDSHIFT['profile'], '--region', REDSHIFT['region'],
            '--id', stmt_id, '--query','Status','--output','text'],
            capture_output=True, text=True).stdout.strip()
        if s == 'FINISHED': break
        if s in ('FAILED','ABORTED'):
            err = subprocess.run(['aws','redshift-data','describe-statement',
                '--profile', REDSHIFT['profile'], '--region', REDSHIFT['region'],
                '--id', stmt_id, '--query','Error','--output','text'],
                capture_output=True, text=True).stdout.strip()
            sys.exit(f"Query {s}: {err}")
        time.sleep(1)
    records = []
    next_token = None
    while True:
        cmd = ['aws','redshift-data','get-statement-result',
               '--profile', REDSHIFT['profile'], '--region', REDSHIFT['region'],
               '--id', stmt_id, '--output','json']
        if next_token: cmd += ['--next-token', next_token]
        d = json.loads(subprocess.run(cmd, capture_output=True, text=True).stdout)
        records.extend(d.get('Records', []))
        next_token = d.get('NextToken')
        if not next_token: break
    return records

def sv(c): return c.get('stringValue', c.get('longValue', c.get('doubleValue', None)))

# ── Step 1: find latest scan date ─────────────────────────────────────────────
print("Finding latest scan date...")
rows = run_sql("SELECT MAX(DATE(search_timestamp)) as latest FROM allo_analytics.serp_analyses")
latest_dt = rows[0][0]['stringValue']
print(f"  Latest scan: {latest_dt}")

# ── Step 2: pull all data for that date ───────────────────────────────────────
print(f"Pulling all cities for {latest_dt}...")
records = run_sql(f"""
SELECT city, nearest_clinic, latitude, longitude, allo_rank, parsed_serp, keyword
FROM allo_analytics.serp_analyses
WHERE DATE(search_timestamp) = '{latest_dt}'
ORDER BY city, keyword
""")
print(f"  Fetched {len(records):,} rows")

# ── Step 3: process ───────────────────────────────────────────────────────────
def is_allo(n): return 'allo' in str(n).lower() if n else False
def parse_super(v):
    if v is None: return {}
    if isinstance(v, dict): return v
    try: return json.loads(v)
    except: return {}
def safe_int(v, d=0):
    try: return int(v)
    except: return d

def extract_pack(ps):
    mp = ps.get('mapPack', [])
    return sorted([{
        'position': e.get('position',0), 'name': e.get('name',''),
        'reviews': e.get('reviewsCount',0) or 0, 'rating': e.get('rating',0) or 0,
        'dist_km': e.get('distance',0) or 0, 'is_allo': is_allo(e.get('name','')),
    } for e in (mp if isinstance(mp,list) else [])], key=lambda x: x['position'])

def extract_ads(ps):
    ads = ps.get('ads', {})
    if isinstance(ads, str):
        try: ads = json.loads(ads)
        except: ads = {}
    fold1 = ads.get('fold_1', []) if isinstance(ads, dict) else []
    riv_ad, riv_nm, we_ad = 'No','','No'
    for ad in (fold1 or []):
        domain = ad.get('domain','') or ad.get('url','')
        name = ad.get('title','') or ad.get('name','')
        if 'allohealth' in domain.lower(): we_ad = 'Yes'
        else:
            riv_ad = 'Yes'
            if not riv_nm: riv_nm = name
    return riv_ad, riv_nm, we_ad

HDR = [
    'tier','city','our_clinic_locality','lat','lon',
    'our_rank','our_dist_km','our_reviews','our_rating',
    'rival_name','rival_reviews','rival_rating','rival_dist_km',
    'allo_sibling_count','best_allo_rank','is_brand_win',
    'p1_name','p1_reviews','p1_rating','p1_dist_km',
    'p2_name','p2_reviews','p2_rating','p2_dist_km',
    'p3_name','p3_reviews','p3_rating','p3_dist_km',
    'rival_ad_present','rival_ad_name','we_ad_present',
    'keyword'
]

print("Processing rows...")
all_rows = []
for rec in records:
    vals = [sv(c) for c in rec]
    city_raw, nc_raw, lat, lon, our_rank_raw, ps_raw, keyword = vals
    city = 'Navi Mumbai' if 'navi mumbai' in str(keyword).lower() else city_raw
    tier = 'Tier 1' if city in TIER1 else 'Tier 2'
    lat = float(lat) if lat is not None else 0
    lon = float(lon) if lon is not None else 0
    our_rank_str = str(int(our_rank_raw)) if our_rank_raw not in (None,'') else 'Absent'
    our_rank_int = safe_int(our_rank_str, 0)
    ps = parse_super(ps_raw)
    nc = parse_super(nc_raw)
    clinic_locality = nc.get('locality','') if nc else ''
    our_reviews = nc.get('reviewsCount',0) or 0 if nc else 0
    our_rating = nc.get('reviewsAvgRating',0) or 0 if nc else 0
    pack = extract_pack(ps)
    non_allo = [p for p in pack if not p['is_allo']]
    allo_entries = [p for p in pack if p['is_allo']]
    riv_ad, riv_nm, we_ad = extract_ads(ps)
    sib_count = len(allo_entries) if allo_entries else 1
    best_rank = min((p['position'] for p in allo_entries), default=our_rank_int)
    is_bw = 'Yes' if (best_rank > 0 and best_rank < our_rank_int) else ('Yes' if our_rank_int == 1 else 'No')
    our_dist = next((p['dist_km'] for p in allo_entries if p['position']==our_rank_int), 0)
    def pp(i): return non_allo[i] if i < len(non_allo) else {'name':'','reviews':0,'rating':0,'dist_km':0}
    p1,p2,p3 = pp(0),pp(1),pp(2)
    all_rows.append([
        tier, city, clinic_locality, lat, lon,
        our_rank_str, our_dist, our_reviews, our_rating,
        p1['name'], p1['reviews'], p1['rating'], p1['dist_km'],
        sib_count, best_rank if best_rank else 'Absent', is_bw,
        p1['name'], p1['reviews'], p1['rating'], p1['dist_km'],
        p2['name'], p2['reviews'], p2['rating'], p2['dist_km'],
        p3['name'], p3['reviews'], p3['rating'], p3['dist_km'],
        riv_ad, riv_nm, we_ad,
        keyword
    ])

# ── Step 4: build Excel ───────────────────────────────────────────────────────
print("Building Excel...")
H1  = PatternFill('solid', fgColor='1F3864'); HF = Font(bold=True, color='FFFFFF', size=10)
T1F = PatternFill('solid', fgColor='1a3a5c'); T2F = PatternFill('solid', fgColor='2d4a1e')
BW  = PatternFill('solid', fgColor='C6EFCE')
SIB = PatternFill('solid', fgColor='FFF2CC')
OOR = PatternFill('solid', fgColor='F2F2F2')
ADV = PatternFill('solid', fgColor='FCE4D6')
T1R = PatternFill('solid', fgColor='dce8f5')
T2R = PatternFill('solid', fgColor='e8f5dc')

CW = {'tier':8,'city':13,'our_clinic_locality':20,'keyword':28,'is_brand_win':11,
      'allo_sibling_count':12,'best_allo_rank':12,'rival_name':22,
      'our_rank':8,'our_dist_km':10,'our_reviews':10,'our_rating':9,
      'rival_ad_present':12,'rival_ad_name':20,'we_ad_present':10,
      'p1_name':22,'p2_name':22,'p3_name':22}

ri  = HDR.index('our_rank')
si  = HDR.index('allo_sibling_count')
bi  = HDR.index('is_brand_win')
ti  = HDR.index('tier')
ci  = HDR.index('city')
ki  = HDR.index('keyword')
rai = HDR.index('rival_ad_present')
wai = HDR.index('we_ad_present')

def apply_hdr(ws, fill=None):
    ws.append(HDR)
    for cell in ws[1]:
        cell.fill = fill or H1; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for col_idx, col_name in enumerate(HDR, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = CW.get(col_name, 9)
    ws.freeze_panes = 'F2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HDR))}1"

def colour_rows(ws, rows_data, start=2):
    for i, r in enumerate(rows_data, start):
        if str(r[ri]) not in ('1','2','3'): ws.cell(row=i, column=ri+1).fill = OOR
        if r[bi]=='Yes' and safe_int(r[si],1)>1: ws.cell(row=i, column=bi+1).fill = BW
        if safe_int(r[si],1)>1: ws.cell(row=i, column=si+1).fill = SIB
        if r[rai]=='Yes': ws.cell(row=i, column=rai+1).fill = ADV

wb = openpyxl.Workbook()

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws = wb.active; ws.title = '📊 Summary'

# Week header row
ws.append([f"Scan date: {latest_dt}  |  Pulled: {date.today().isoformat()}  |  Rows: {len(all_rows):,}  |  To refresh: python3 refresh_competitor_analysis.py"])
ws['A1'].font = Font(bold=True, size=11, color='1F3864')
ws.merge_cells('A1:O1')

SUM_HDR = ['Tier','City','Cells','Keywords','Rank 1','Rank 2','Rank 3','Absent',
           'Brand Win','Multi-Sib','Rival Ad','We Ad','Rank1 %','Absent %','Brand Win %']
ws.append(SUM_HDR)
for cell in ws[2]:
    cell.fill = H1; cell.font = HF
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

stats = defaultdict(lambda:{'total':0,'r1':0,'r2':0,'r3':0,'ab':0,'bw':0,'ms':0,'ra':0,'wa':0,'kws':set()})
for r in all_rows:
    s = stats[(r[ti],r[ci])]
    s['total']+=1
    rank=str(r[ri])
    if rank=='1': s['r1']+=1
    elif rank=='2': s['r2']+=1
    elif rank=='3': s['r3']+=1
    else: s['ab']+=1
    if r[bi]=='Yes': s['bw']+=1
    if safe_int(r[si],1)>1: s['ms']+=1
    if r[rai]=='Yes': s['ra']+=1
    if r[wai]=='Yes': s['wa']+=1
    s['kws'].add(r[ki])

TIER2_CITIES = sorted(set(r[ci] for r in all_rows if r[ti]=='Tier 2'))
row_num = 3
for tier, cities in [('Tier 1', TIER1_ORDER), ('Tier 2', TIER2_CITIES)]:
    for city in cities:
        s = stats.get((tier, city))
        if not s: continue
        t = s['total']
        ws.append([tier, city, t, len(s['kws']),
                   s['r1'],s['r2'],s['r3'],s['ab'],
                   s['bw'],s['ms'],s['ra'],s['wa'],
                   f"{s['r1']/t*100:.0f}%", f"{s['ab']/t*100:.0f}%", f"{s['bw']/t*100:.0f}%"])
        fill = T1R if tier=='Tier 1' else T2R
        for cell in ws[row_num]: cell.fill = fill
        row_num += 1

for col in range(1, len(SUM_HDR)+1):
    ws.column_dimensions[get_column_letter(col)].width = 12
ws.column_dimensions['B'].width = 14
ws.freeze_panes = 'A3'

# ── Tier 1 / Tier 2 sheets ────────────────────────────────────────────────────
ws2 = wb.create_sheet('🏙️ Tier 1'); apply_hdr(ws2, T1F)
t1_rows = [r for r in all_rows if r[ti]=='Tier 1']
for r in t1_rows: ws2.append(r)
colour_rows(ws2, t1_rows)

ws3 = wb.create_sheet('🌆 Tier 2'); apply_hdr(ws3, T2F)
t2_rows = [r for r in all_rows if r[ti]=='Tier 2']
for r in t2_rows: ws3.append(r)
colour_rows(ws3, t2_rows)

# ── Per-city Tier 1 sheets ────────────────────────────────────────────────────
city_fills = {'Bangalore':'123a63','Chennai':'3a1d1a','Mumbai':'0d3b34',
              'Navi Mumbai':'1a5c40','Hyderabad':'1e1b4b','Pune':'2d1b4e'}
for city in TIER1_ORDER:
    city_rows = [r for r in t1_rows if r[ci]==city]
    if not city_rows: continue
    ws_c = wb.create_sheet(f'  {city[:10]}')
    apply_hdr(ws_c, PatternFill('solid', fgColor=city_fills.get(city,'1F3864')))
    for r in city_rows: ws_c.append(r)
    colour_rows(ws_c, city_rows)

# ── Rank 1 / Absent / Shuffle ─────────────────────────────────────────────────
for title, fgColor, filter_fn in [
    ('🥇 Rank 1',  '375623', lambda r: str(r[ri])=='1'),
    ('❌ Absent',  '7B0000', lambda r: str(r[ri]) not in ('1','2','3')),
    ('👥 Shuffle', '7B5A00', lambda r: safe_int(r[si],1)>1),
]:
    ws_x = wb.create_sheet(title)
    apply_hdr(ws_x, PatternFill('solid', fgColor=fgColor))
    filtered = [r for r in all_rows if filter_fn(r)]
    for r in filtered: ws_x.append(r)
    colour_rows(ws_x, filtered)

wb.save(OUT)
print(f"\n✓ Saved: {OUT}")
print(f"  Scan: {latest_dt} | Rows: {len(all_rows):,} | T1: {len(t1_rows):,} | T2: {len(t2_rows):,}")
print(f"  Cities: {len(stats)}")
