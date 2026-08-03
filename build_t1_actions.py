#!/usr/bin/env python3
"""
build_t1_actions.py
Final T1 actionables: paid + organic, per city × campaign × category.
Sources:
  - data_ga_campaigns.json   (bid/budget/util/IS/QS, Jun-2026 pull)
  - data_auction_insights.json (competitor IS, posAbove, Jul-2026 screenshots)
  - data_rival_enrichment.json (competitor GMB proximity + reviews per clinic)
  - DOMAIN_META from build_competitor_analysis.py
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))

# ── HAND-SYNTHESISED ACTION ROWS ───────────────────────────────────────────
# Columns: city, cat, campaign, action_type, priority,
#          current_state, diagnosis, action, quantum, prereq

ACTIONS = [

# ════════════════════════════════════════════════════════
# BANGALORE
# ════════════════════════════════════════════════════════

# ── BLR SH ──
{"city":"Bangalore","cat":"SH","campaign":"SH Exact-Local",
 "type":"Quality — Ad Relevance","pri":"P1",
 "state":"util=73% | IS=51% | AR=26%↓ | QS=6.1",
 "diagnosis":"26% of keywords below-avg ad relevance dragging QS. Bids are structurally handicapped until RSAs match keyword intent.",
 "action":"Rewrite RSAs: add clinic-area headline ('ED Clinic in Indiranagar'), symptom headlines ('Weak Erection Treatment'), match keyword groups to dedicated ad groups.",
 "quantum":"Target AR <10%↓ → QS 7+","prereq":"—"},

{"city":"Bangalore","cat":"SH","campaign":"SH Exact-Local",
 "type":"Location Bid +adj","pri":"P1",
 "state":"Medi Life (Medi Life BLR): IS=17%, posAbove=77% | near Indiranagar 4.3km, Sahakara Nagar 6.1km",
 "diagnosis":"Medi Life outranks us 77% when both show. Confirmed GMB. They're drawing SH patients near our two biggest Bangalore clinics.",
 "action":"Add +25% location bid adjustment for Indiranagar & Sahakara Nagar pin zones in T1_Bangalore_SH_Exact_Local. Keep rest at 0%.",
 "quantum":"+25% bid adj for 2 clinics only","prereq":"RSA fix first (P1 above)"},

{"city":"Bangalore","cat":"SH","campaign":"SH Exact",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=31% | LP=41%↓ | IS=54%",
 "diagnosis":"Landing page below-avg on 41% of keywords. util=31% means the campaign barely runs — LP quality tank is suppressing delivery.",
 "action":"Map SH Exact to a faster, intent-specific LP (not the generic booking page). Test a direct 'Book SH Consult' CTA above-fold with clinic-specific schema.",
 "quantum":"LP% target: <20%↓","prereq":"—"},

{"city":"Bangalore","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Budget","pri":"P2",
 "state":"util=101% | BL=9% | IS=11%",
 "diagnosis":"Budget ceiling hit. BL=9% = losing 9% of eligible impressions to budget cap.",
 "action":"Increase daily budget ₹240 → ₹360 (+50%).",
 "quantum":"₹240 → ₹360/day","prereq":"—"},

{"city":"Bangalore","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Bid","pri":"P2",
 "state":"RL=80% | IS=11%",
 "diagnosis":"80% of rank loss is on rank (not budget). After budget increase, raise bid to improve rank.",
 "action":"After budget raise settles (1 wk), increase phrase bid by +20%.",
 "quantum":"+20% after budget raise","prereq":"Budget raise first"},

# ── BLR STI ──
{"city":"Bangalore","cat":"STI","campaign":"STI Exact-Local",
 "type":"Bid","pri":"P1",
 "state":"IS=42% vs DrSafeHands IS=64% | RL=57% | posAbove=90%",
 "diagnosis":"DrSafeHands IS=64% vs our 42% — we're losing 22pp of impression share to them on rank. They're above us 90% of the time. Budget (84% util) is secondary constraint.",
 "action":"Raise target CPA / manual bid by +25% on T1_Bangalore_STD_Exact_Local. DrSafeHands IS should drop below 55% as we recapture rank.",
 "quantum":"+25% bid | target: IS 55%+","prereq":"—"},

{"city":"Bangalore","cat":"STI","campaign":"STI Exact-Local",
 "type":"Budget","pri":"P1",
 "state":"util=84% | BL=1%",
 "diagnosis":"84% util — approaching ceiling. Once bid raises pull more volume, budget will cap. Raise now.",
 "action":"Increase daily budget ₹863 → ₹1,100 (+27%). Re-evaluate after 2 weeks.",
 "quantum":"₹863 → ₹1,100/day","prereq":"Raise bid simultaneously"},

{"city":"Bangalore","cat":"STI","campaign":"STI Exact-Local",
 "type":"Location Bid +adj","pri":"P2",
 "state":"DrSafeHands near KR Puram 1.2km (178 rev) & HSR Layout 1.4km (154 rev) | our reviews: KR Puram 384, HSR 683",
 "diagnosis":"DrSafeHands has physical presence 1-1.4km from two clinics. Even though we lead on reviews, they outrank us in the STI auction near these clinics.",
 "action":"Add +20% location bid adj for KR Puram and HSR Layout zones in STI campaign.",
 "quantum":"+20% bid adj for 2 clinic zones","prereq":"Overall bid raise first"},

{"city":"Bangalore","cat":"STI","campaign":"STI Exact-Local",
 "type":"GMB — Reviews","pri":"P2",
 "state":"DrSafeHands near Rajajinagar 8.0km (rev=140 vs our 26, gap +114) | Mahalakshmi Layout 8.5km (rev=141 vs our 26, gap +115)",
 "diagnosis":"Two BLR STI clinics where DrSafeHands leads on reviews: Rajajinagar (gap +114) and Mahalakshmi Layout (gap +115). New/low-review clinics are vulnerable.",
 "action":"Prioritise S2/S3 review generation at Rajajinagar and Mahalakshmi Layout STI clinics: target +15 reviews/month each until we pass 150.",
 "quantum":"+15 reviews/mo × 2 clinics","prereq":"—"},

# ── BLR MH ──
{"city":"Bangalore","cat":"MH","campaign":"MH Exact-Local",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"QS=3.8 | LP=62%↓ | util=97%",
 "diagnosis":"QS 3.8 is critically low — driven by LP=62% below-avg. Every rupee spent is penalised. Raising budget without fixing LP = burning money at a worse rate.",
 "action":"STOP raising budget until LP fixed. Build a dedicated MH LP: therapy/counselling angle, symptom headlines (anxiety, depression), local clinic CTA. Speed < 2s on mobile.",
 "quantum":"LP% target: <25%↓ | QS target: 5+","prereq":"LP fix is P0 — do this week"},

{"city":"Bangalore","cat":"MH","campaign":"MH Exact-Local",
 "type":"Budget","pri":"P2",
 "state":"util=97% | IS=31%",
 "diagnosis":"Budget ceiling hit, but quality gate blocks this. After LP fix, IS should improve organically. Then raise budget.",
 "action":"After LP fix and QS improves to 5+, raise daily budget ₹1,380 → ₹1,800 (+30%).",
 "quantum":"₹1,380 → ₹1,800 (post LP fix)","prereq":"LP fix must ship first"},

{"city":"Bangalore","cat":"MH","campaign":"MH Exact-Local",
 "type":"Location Bid +adj","pri":"P2",
 "state":"Cadabams near Indiranagar 1.0km (IS=42%, posAbove=77%) | Amaha near Bilekahalli 2.3km",
 "diagnosis":"Cadabams IS=42%, posAbove=77% — they're dominating MH near our Indiranagar clinic. Direct GMB competitor.",
 "action":"After QS 5+, add +20% location bid adj for Indiranagar zone. ANC Clinic near Nallagandla 1.4km — add +15% for Nallagandla zone.",
 "quantum":"+20% Indiranagar | +15% Nallagandla","prereq":"MH LP fix first"},

{"city":"Bangalore","cat":"MH","campaign":"MH Phrase-Local",
 "type":"Quality — Ad Relevance","pri":"P2",
 "state":"util=28% | AR=1%↓ | IS=10%",
 "diagnosis":"AR=1% below avg = almost every ad is irrelevant. Campaign barely runs (28% util) because QS is suppressed. Fix RSAs before touching budget or bid.",
 "action":"Pause Phrase-Local. Port best-converting MH phrase keywords into MH Exact-Local after adding dedicated ad groups. Or rebuild phrase campaign with MH-specific RSAs.",
 "quantum":"Pause or rebuild — don't spend on this now","prereq":"—"},

# ════════════════════════════════════════════════════════
# CHENNAI
# ════════════════════════════════════════════════════════

# ── CHN SH ──
{"city":"Chennai","cat":"SH","campaign":"SH Exact",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=6% | LP=94%↓ | IS=38%",
 "diagnosis":"LP=94% below-avg: virtually every keyword gets penalised. util=6% means the campaign is functionally off — it wins almost no auctions because QS is floor-level.",
 "action":"PAUSE SH Exact Chennai until LP is fixed. Fix LP: Chennai-specific SH page, local clinic schema, fast mobile load. Then un-pause and monitor.",
 "quantum":"Pause → fix LP → re-enable","prereq":"LP fix is P0"},

{"city":"Chennai","cat":"SH","campaign":"SH Exact-Local",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=60% | LP=53%↓ | IS=58%",
 "diagnosis":"LP=53% below-avg on Exact-Local too. IS=58% is decent but LP penalty caps QS. Fix LP for full impact of current bids.",
 "action":"Deploy fixed Chennai SH LP (same fix as Exact above). Focus on 'Clinic in [area]' + symptom content for Nungambakkam / Velachery pages.",
 "quantum":"LP% target: <25%↓","prereq":"Same LP fix as Exact"},

{"city":"Chennai","cat":"SH","campaign":"SH Exact-Local",
 "type":"Location Bid +adj","pri":"P1",
 "state":"Dr Kamaraj near Nungambakkam 2.2km (924 rev vs our 326, gap +598) | near Mogappair 5.1km (gap +882) | near Manapakkam 6.6km (gap +921)",
 "diagnosis":"Dr Kamaraj IS=45%, posAbove=70% and has 924 GMB reviews. Nungambakkam, Mogappair and Manapakkam are battleground clinics — they're both beating us on ads AND on reviews.",
 "action":"Add +25% location bid adj for Nungambakkam + Mogappair zones. Manapakkam is new (+3 reviews) — hold until reviews built up, no point driving traffic to an empty profile.",
 "quantum":"+25% Nungambakkam | +25% Mogappair | hold Manapakkam","prereq":"LP fix first"},

{"city":"Chennai","cat":"SH","campaign":"SH Exact-Local",
 "type":"GMB — Reviews","pri":"P1",
 "state":"Dr Kamaraj 924 reviews | Nungambakkam: our 326 (gap 598) | Mogappair: our 42 (gap 882) | Manapakkam: our 3 (gap 921)",
 "diagnosis":"Three Chennai SH clinics with massive review gaps to Dr Kamaraj. This is the biggest organic review emergency in T1. Manapakkam is almost invisible (3 reviews).",
 "action":"Chennai SH review targets per month: Nungambakkam +30, Mogappair +25, Manapakkam +20. Use S1 (PE/ED) testimonials. At this pace, parity at Nungambakkam in ~20 months — so set 12-month milestone of 700.",
 "quantum":"+30 Nungambakkam | +25 Mogappair | +20 Manapakkam (per month)","prereq":"—"},

{"city":"Chennai","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Bid","pri":"P2",
 "state":"RL=80% | IS=15% | BL=6% | util=73%",
 "diagnosis":"RL=80% = losing 80% of eligible impressions to low rank. Bid too low vs competition.",
 "action":"Raise phrase bid +20%. Monitor IS over 2 weeks.",
 "quantum":"+20% phrase bid","prereq":"—"},

{"city":"Chennai","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Budget","pri":"P2",
 "state":"util=73% → BL=6% showing budget pressure",
 "diagnosis":"After bid raise pulls more auctions, budget will cap. Pre-empt: ₹200 → ₹260.",
 "action":"Raise budget ₹200 → ₹260 alongside bid raise.",
 "quantum":"₹200 → ₹260/day","prereq":"Bid raise simultaneously"},

# ── CHN STI ──
{"city":"Chennai","cat":"STI","campaign":"STI Exact-Local",
 "type":"Bid","pri":"P1",
 "state":"IS=44% vs DrSafeHands IS=58% | RL=55% | posAbove=88%",
 "diagnosis":"DrSafeHands IS=58% vs our 44% — above us 88% of the time. RL=55% = rank limited. Budget (92% util) secondary.",
 "action":"Raise bid +25-30% on STI Exact-Local Chennai.",
 "quantum":"+28% bid | target IS 57%+","prereq":"—"},

{"city":"Chennai","cat":"STI","campaign":"STI Exact-Local",
 "type":"Budget","pri":"P1",
 "state":"util=92% | BL=1%",
 "diagnosis":"92% util near ceiling. Bid raise will push spend higher — budget will cap fast.",
 "action":"Raise budget ₹400 → ₹600 (+50%).",
 "quantum":"₹400 → ₹600/day","prereq":"Raise bid simultaneously"},

{"city":"Chennai","cat":"STI","campaign":"STI Exact-Local",
 "type":"Location Bid +adj","pri":"P2",
 "state":"DrSafeHands near Velachery 1.4km (37 rev vs our 597) | near Manapakkam 6.6km | MetroMale near Manapakkam 8.5km (593 rev vs our 3)",
 "diagnosis":"DrSafeHands near Velachery — we dominate on reviews (597 vs 37). Double down here with location bid. MetroMale near Manapakkam has 593 reviews vs our 3 — don't bid up there yet.",
 "action":"Add +20% location bid adj for Velachery zone. Hold Manapakkam and Nungambakkam until reviews build.",
 "quantum":"+20% Velachery | hold others","prereq":"Overall bid raise first"},

{"city":"Chennai","cat":"STI","campaign":"STI Exact-Local",
 "type":"GMB — Reviews","pri":"P2",
 "state":"MetroMale near Manapakkam: their 593 vs our 3 (gap 590) | DrSafeHands near Nungambakkam: 37 vs our 326 (we lead)",
 "diagnosis":"Manapakkam STI clinic has 3 reviews — MetroMale leads by 590. Nungambakkam we're fine. Velachery we lead.",
 "action":"STI review push at Manapakkam: target +20 reviews/month. In parallel, prioritise SH reviews there too (Dr Kamaraj gap).",
 "quantum":"+20 reviews/mo at Manapakkam STI","prereq":"—"},

# ════════════════════════════════════════════════════════
# HYDERABAD
# ════════════════════════════════════════════════════════

# ── HYD SH ──
{"city":"Hyderabad","cat":"SH","campaign":"SH Exact-Local",
 "type":"Budget","pri":"P1",
 "state":"util=88% | IS=59% | LP=50%↓",
 "diagnosis":"88% util — approaching ceiling with LP penalty still active. MensClinicsGroup IS=37%, posAbove=65% in Hyderabad SH. Need budget headroom.",
 "action":"Raise budget ₹2,875 → ₹3,500 (+22%).",
 "quantum":"₹2,875 → ₹3,500/day","prereq":"—"},

{"city":"Hyderabad","cat":"SH","campaign":"SH Exact-Local",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"LP=50%↓ | IS=59%",
 "diagnosis":"LP=50% below avg is capping QS and wasting bids. Pattern consistent with all SH cities.",
 "action":"Deploy city-specific SH LP for Hyderabad with Kondapur / Ameerpet / Madhapur clinic CTAs. Fast load + local schema.",
 "quantum":"LP% target: <25%↓","prereq":"—"},

{"city":"Hyderabad","cat":"SH","campaign":"SH Exact-Local",
 "type":"Bid","pri":"P2",
 "state":"MensClinicsGroup IS=37%, posAbove=65% | RL=41%",
 "diagnosis":"After LP fix, QS will rise and same bid will generate more rank. Modest bid raise on top to overtake MensClinicsGroup.",
 "action":"After LP fix settles (2 wks), raise bid +10%.",
 "quantum":"+10% post LP fix","prereq":"LP fix first"},

{"city":"Hyderabad","cat":"SH","campaign":"SH Exact",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=14% | LP=90%↓",
 "diagnosis":"Functionally dead. LP=90% below avg. Same pattern as Chennai Exact. Pause until LP fixed.",
 "action":"Pause SH Exact Hyderabad. Merge budget into Exact-Local once LP is fixed and LP% <30% on Exact-Local.",
 "quantum":"Pause → re-enable post LP fix","prereq":"LP fix"},

{"city":"Hyderabad","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Budget","pri":"P2",
 "state":"util=94% | BL=4% | RL=87%",
 "diagnosis":"Ceiling hit + worst rank-loss (87%) of all phrase campaigns. Need both budget and bid uplift.",
 "action":"Raise budget ₹180 → ₹280 (+56%) + bid +20%.",
 "quantum":"₹180 → ₹280/day + bid +20%","prereq":"—"},

# ── HYD STI ──
{"city":"Hyderabad","cat":"STI","campaign":"STI Exact-Local",
 "type":"Bid","pri":"P1",
 "state":"IS=38% vs DrSafeHands IS=53% | RL=60% | posAbove=88% | util=60%",
 "diagnosis":"Biggest STI rank gap (15pp behind DrSafeHands). RL=60% = rank not budget. 60% util means budget won't cap immediately after bid raise.",
 "action":"Raise bid +30% on STI Exact-Local Hyderabad. Target IS 53%+.",
 "quantum":"+30% bid | target IS 53%+","prereq":"—"},

{"city":"Hyderabad","cat":"STI","campaign":"STI Exact-Local",
 "type":"Budget","pri":"P2",
 "state":"util=60% | bud=₹1,000",
 "diagnosis":"60% util gives runway but bid raise will close this gap fast. Pre-empt ceiling in 2-3 weeks.",
 "action":"Raise budget ₹1,000 → ₹1,300 (+30%) in week 2 after bid raise.",
 "quantum":"₹1,000 → ₹1,300/day (week 2)","prereq":"Bid raise week 1"},

{"city":"Hyderabad","cat":"STI","campaign":"STI Exact-Local",
 "type":"Location Bid +adj","pri":"P2",
 "state":"DrSafeHands near Borabanda 1.4km (72 rev vs our 18, gap +54) | Madhapur 2.3km (72 vs 0) | Kondapur 4.2km (72 vs 994, we lead) | Ameerpet 6.0km (72 vs 951)",
 "diagnosis":"Borabanda and Madhapur: DrSafeHands is closer + we have very low reviews there. Kondapur + Ameerpet: we dominate on reviews, bid-up is less critical.",
 "action":"Add +25% location bid adj for Borabanda + Madhapur zones. No adjustment for Kondapur / Ameerpet (we dominate).",
 "quantum":"+25% Borabanda | +25% Madhapur | hold others","prereq":"Overall bid raise first"},

{"city":"Hyderabad","cat":"STI","campaign":"STI Exact-Local",
 "type":"GMB — Reviews","pri":"P2",
 "state":"Borabanda: DrSafeHands 72 vs our 18 (gap +54) | Madhapur: 72 vs 0 | Vanasthalipuram 24.0km: 72 vs 11",
 "diagnosis":"Borabanda + Madhapur are weakest Hyderabad STI clinics on reviews. Vanasthalipuram is far and low-review — lower priority.",
 "action":"STI review push at Borabanda: +10/month. Madhapur: +10/month. Deprioritise Kondapur / Ameerpet (already 994 and 951 reviews).",
 "quantum":"+10 reviews/mo Borabanda | +10 Madhapur","prereq":"—"},

# ════════════════════════════════════════════════════════
# MUMBAI
# ════════════════════════════════════════════════════════

# ── MUM SH ──
{"city":"Mumbai","cat":"SH","campaign":"SH Exact-Local",
 "type":"Budget","pri":"P1",
 "state":"util=90% | IS=62% | LP=42%↓",
 "diagnosis":"90% util near ceiling. IS=62% is the highest of all SH cities — Mumbai SH is competitive. Budget increase captures more of this demand.",
 "action":"Raise budget ₹2,087 → ₹2,600 (+25%).",
 "quantum":"₹2,087 → ₹2,600/day","prereq":"—"},

{"city":"Mumbai","cat":"SH","campaign":"SH Exact-Local",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"LP=42%↓ | IS=62%",
 "diagnosis":"LP penalty active despite good IS. Fix LP to improve QS and reduce effective CPC.",
 "action":"Deploy Mumbai SH LP: Andheri / Ghatkopar / Kemps Corner / Dadar clinic pages with local schema.",
 "quantum":"LP% target: <20%↓","prereq":"—"},

{"city":"Mumbai","cat":"SH","campaign":"SH Exact-Local",
 "type":"Location Bid +adj","pri":"P1",
 "state":"Dr SK Jain Burlington: IS=38%, near Dadar 1.2km | Dr Priyank Kothari near Kemps Corner 0.3km | Kayakalp posAbove=94% | Gautam Ayurveda posAbove=79%",
 "diagnosis":"Dadar and Kemps Corner are Mumbai's highest-competition SH locations. Dr SK Jain (IS=38%) + Priyank Kothari (IS<10% but posAbove=40%) both active near these clinics.",
 "action":"Add +25% location bid adj for Dadar + Kemps Corner zones. Others (Andheri, Ghatkopar, Thane): hold at 0%.",
 "quantum":"+25% Dadar | +25% Kemps Corner","prereq":"LP fix first"},

{"city":"Mumbai","cat":"SH","campaign":"SH Exact",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=18% | LP=81%↓",
 "diagnosis":"LP=81% below avg. Functionally dead, same pattern as Chennai/Hyderabad Exact campaigns.",
 "action":"Pause SH Exact Mumbai. Fix shared LP. Re-enable when LP% <30%.",
 "quantum":"Pause → re-enable post LP fix","prereq":"LP fix"},

{"city":"Mumbai","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Budget","pri":"P1",
 "state":"util=99% | BL=11% | IS=15%",
 "diagnosis":"Ceiling hit. BL=11% = losing 11% impressions to budget. Worst budget loss of all phrase campaigns.",
 "action":"Raise budget ₹120 → ₹220 (+83%). This is undersized — Mumbai phrase is QS=9.0 (best in class), allocate more.",
 "quantum":"₹120 → ₹220/day","prereq":"—"},

{"city":"Mumbai","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Bid","pri":"P2",
 "state":"RL=74% | QS=9.0",
 "diagnosis":"QS=9.0 = excellent quality. RL=74% is almost entirely on rank being too low. With QS 9 the bid can work very efficiently.",
 "action":"After budget increase settles, raise bid +15%. Every rupee here buys better rank at high efficiency.",
 "quantum":"+15% (post budget raise)","prereq":"Budget raise first"},

# ── MUM STI ──
{"city":"Mumbai","cat":"STI","campaign":"STI Exact-Local",
 "type":"Bid","pri":"P1",
 "state":"IS=29% vs DrSafeHands IS=60% | RL=69% | posAbove=97% | util=72%",
 "diagnosis":"WORST STI position: DrSafeHands IS=60% vs our 29% — 31pp gap. posAbove=97% = they're above us virtually every impression. RL=69% = rank-limited.",
 "action":"Raise bid +35% immediately on STI Exact-Local Mumbai. This is the single highest-priority bid action across all T1 cities.",
 "quantum":"+35% bid | target IS 45%+","prereq":"—"},

{"city":"Mumbai","cat":"STI","campaign":"STI Exact-Local",
 "type":"Budget","pri":"P1",
 "state":"util=72% | bud=₹397",
 "diagnosis":"₹397 budget is tiny for Mumbai STI. After +35% bid raise, budget will cap within days.",
 "action":"Raise budget ₹397 → ₹700 (+76%) simultaneously with bid raise.",
 "quantum":"₹397 → ₹700/day","prereq":"Raise bid simultaneously"},

{"city":"Mumbai","cat":"STI","campaign":"STI Exact-Local",
 "type":"Location Bid +adj","pri":"P2",
 "state":"DrSafeHands near Ghatkopar 2.7km (100 rev vs our 496, we lead) | near Andheri East 4.0km (100 vs 712, we lead) | near Kemps Corner 17km (low priority)",
 "diagnosis":"We lead on reviews near Ghatkopar + Andheri East. After overall bid raise takes us to IS 45%+, add location boost at these clinics to dominate locally.",
 "action":"After IS >45%: add +20% location adj for Ghatkopar + Andheri East. Leverage our review advantage.",
 "quantum":"+20% Ghatkopar | +20% Andheri East (after IS >45%)","prereq":"Overall bid + budget raise first"},

# ════════════════════════════════════════════════════════
# PUNE
# ════════════════════════════════════════════════════════

# ── PUNE SH ──
{"city":"Pune","cat":"SH","campaign":"SH Exact-Local",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=77% | IS=66% | LP=39%↓",
 "diagnosis":"IS=66% is best SH IS across all T1 cities. LP=39% below avg is the remaining drag — fixing it will further improve QS and reduce CPC.",
 "action":"Deploy Pune SH LP: Wakad / Katraj / Kharadi / Baner clinic-specific pages with fast load + local schema.",
 "quantum":"LP% target: <20%↓","prereq":"—"},

{"city":"Pune","cat":"SH","campaign":"SH Exact-Local",
 "type":"Location Bid +adj","pri":"P2",
 "state":"topsexologistinpune.com IS=35%, posAbove=74% | DrSafeHands near Wakad 1.3km | Dr Erande near Wakad 3.3km + Baner 3.8km",
 "diagnosis":"Wakad is the primary SH battleground in Pune: topsexologistinpune.com + DrSafeHands + Dr Erande all active nearby. IS=35% for topsexologist is significant.",
 "action":"Add +25% location bid adj for Wakad zone. Katraj and Kharadi: hold (less competition nearby).",
 "quantum":"+25% Wakad | hold others","prereq":"—"},

{"city":"Pune","cat":"SH","campaign":"SH Exact",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=10% | LP=79%↓",
 "diagnosis":"Same pattern: LP=79% kills the campaign. util=10% = functionally paused by QS penalties.",
 "action":"Pause SH Exact Pune. Fix shared SH LP. Re-enable when LP% <30%.",
 "quantum":"Pause → re-enable post LP fix","prereq":"LP fix"},

{"city":"Pune","cat":"SH","campaign":"SH Phrase-Local",
 "type":"Budget","pri":"P2",
 "state":"util=97% | BL=3% | RL=74%",
 "diagnosis":"Near ceiling, and QS=4.6 is worst of all phrase campaigns. Budget increase without quality fix has diminishing returns.",
 "action":"Raise budget ₹198 → ₹260 (+31%). Also fix RSAs — QS=4.6 is a problem.",
 "quantum":"₹198 → ₹260/day + RSA fix","prereq":"—"},

# ── PUNE STI ──
{"city":"Pune","cat":"STI","campaign":"STI Exact-Local",
 "type":"Bid","pri":"P1",
 "state":"IS=48% vs DrSafeHands IS=59% | RL=51% | posAbove=91%",
 "diagnosis":"11pp gap. RL=51% = rank limited. DrSafeHands above us 91% of the time. Bid raise needed.",
 "action":"Raise bid +20% on STI Exact-Local Pune.",
 "quantum":"+20% bid | target IS 57%+","prereq":"—"},

{"city":"Pune","cat":"STI","campaign":"STI Exact-Local",
 "type":"Budget","pri":"P2",
 "state":"util=64% | bud=₹350",
 "diagnosis":"₹350 is undersized for Pune STI. After bid raise, budget will cap.",
 "action":"Raise budget ₹350 → ₹500 (+43%).",
 "quantum":"₹350 → ₹500/day","prereq":"Bid raise week 1"},

{"city":"Pune","cat":"STI","campaign":"STI Exact-Local",
 "type":"Location Bid +adj","pri":"P2",
 "state":"DrSafeHands near Wakad 1.3km (69 rev vs our 306, we lead) | near Baner 3.8km (69 vs 83, close) | near Kothrud 6km (62 vs 351, we lead)",
 "diagnosis":"We lead on reviews everywhere in Pune STI. Wakad is closest battleground. Bid up near Wakad to leverage our review advantage.",
 "action":"Add +20% location bid adj for Wakad. Baner: +10% (close fight with DrSafeHands). Kothrud: hold (we dominate on reviews).",
 "quantum":"+20% Wakad | +10% Baner | hold Kothrud","prereq":"Overall bid raise first"},

# ── PUNE MH ──
{"city":"Pune","cat":"MH","campaign":"MH Exact-Local",
 "type":"Quality — Landing Page","pri":"P1",
 "state":"util=88% | LP=46%↓ | QS=5.0 | IS=33%",
 "diagnosis":"LP=46% below avg suppressing QS. IS=33% vs Abhasa (posAbove=87%, absTop=64%) and Mindsight (posAbove=82%) — both dominating above-the-fold positions.",
 "action":"Fix Pune MH LP: anxiety/therapy angle, Katraj/Kharadi/Hadapsar clinic pages. Fast mobile load.",
 "quantum":"LP% target: <25%↓","prereq":"LP fix first"},

{"city":"Pune","cat":"MH","campaign":"MH Exact-Local",
 "type":"Budget","pri":"P2",
 "state":"util=88% | IS=33%",
 "diagnosis":"Budget ceiling approaching. After LP fix and QS improves, raise budget.",
 "action":"After LP fix: raise budget ₹400 → ₹550 (+38%).",
 "quantum":"₹400 → ₹550 (post LP fix)","prereq":"LP fix first"},

{"city":"Pune","cat":"MH","campaign":"MH Exact-Local",
 "type":"Location Bid +adj","pri":"P3",
 "state":"Abhasa near Katraj 0.1km (IS<10% but posAbove=87%) | Nityanand Rehab near Katraj 0.1km (IS=17%)",
 "diagnosis":"Katraj is an MH battleground. Two competitors with confirmed GMB within 0.1km. Dominate locally.",
 "action":"After QS 5.5+: add +20% location bid adj for Katraj zone.",
 "quantum":"+20% Katraj (post QS improvement)","prereq":"LP fix → QS improvement"},

# ════════════════════════════════════════════════════════
# CROSS-CITY / STRUCTURAL
# ════════════════════════════════════════════════════════
{"city":"ALL T1","cat":"SH","campaign":"SH Exact + SH Exact-Local",
 "type":"Quality — Landing Page (SYSTEMIC)","pri":"P1",
 "state":"SH Exact LP%: BLR=41↓ CHN=94↓ HYD=90↓ MUM=81↓ PUN=79↓ | Exact-Local LP%: CHN=53↓ HYD=50↓",
 "diagnosis":"SYSTEMIC PATTERN: All SH Exact campaigns across every T1 city have LP=41-94% below average. This is a shared LP problem — same bad LP served in all cities. Single fix unblocks ₹6k+ daily budget across these paused campaigns.",
 "action":"Audit shared SH landing page. Likely issues: slow load >3s mobile, no above-fold CTA, weak symptom relevance. Build one fast, intent-matched SH LP and A/B across cities.",
 "quantum":"Single LP fix unblocks 4+ paused Exact campaigns","prereq":"P0 — do this week"},

{"city":"ALL T1","cat":"STI","campaign":"STI Exact-Local (all cities)",
 "type":"GMB — Reviews (DrSafeHands priority)","pri":"P1",
 "state":"DrSafeHands IS=53-64% across all 5 STI cities | 140 GMB reviews nationally | S2/S3 content (HIV/PEP) drives their map-pack dominance",
 "diagnosis":"DrSafeHands wins STI map-pack on S2/S3 review content (confidential testing, HIV PEP). Their 140 GMB reviews beat us in cities where our STI clinics have <200 reviews.",
 "action":"Add 3-5 S2/S3 review templates per month at every STI clinic with <200 reviews: HIV testing, PEP consult, STD panel language (no patient names). Priority clinics: Rajajinagar (26), Mahalakshmi Layout (26), Borabanda (18), Madhapur (0), Manapakkam (3).",
 "quantum":"+3-5 S2/S3 reviews/mo at low-review STI clinics","prereq":"—"},

]

# ─────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "T1 Actionables"
ws.sheet_view.showGridLines = False

COLS = [
    ("Pri",         6),
    ("City",       14),
    ("Cat",         6),
    ("Campaign",   18),
    ("Action Type",22),
    ("Current State\n(latest data)",   34),
    ("Diagnosis",  46),
    ("Specific Action",46),
    ("Change Quantum",  22),
    ("Prerequisite /\nDo-order",       22),
]
NC = len(COLS)
for i, (_, w) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def cs(ws, row, col, val, bg="FFFFFF", fc="111111", bold=False, size=9,
       align="left", wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.fill      = PatternFill("solid", fgColor=bg)

def merge(ws, row, c1, c2, val, **kw):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    cs(ws, row, c1, val, **kw)

def rh(ws, row, h): ws.row_dimensions[row].height = h

# ── Header ──
r = 1
merge(ws, r, 1, NC,
      "T1 City Actionables — Google Ads (Paid + GMB) | Synthesised from Auction Insights · Campaign Performance · Competitor GMB Data",
      bg="0F172A", fc="FFFFFF", bold=True, size=13, align="center")
rh(ws, r, 24); r += 1

merge(ws, r, 1, NC,
      "Data: Auction Insights Jul 2026  |  Campaign metrics Jun 2026  |  Competitor GMB: SERP crawl data"
      "   |   Exec this sequence: P1 Quality (LP/AR) → P1 Bid → P1 Budget → P2 actions   |   Never raise budget on campaigns with LP% >50% — LP fix first",
      bg="1E293B", fc="94A3B8", size=8, italic=True, align="center")
rh(ws, r, 13); r += 1

# Legend row
merge(ws, r, 1, 2, "P1 = Do this week", bg="FCA5A5", fc="7F1D1D", bold=True, size=8, align="center")
merge(ws, r, 3, 4, "P2 = Do this month", bg="FDE68A", fc="78350F", bold=True, size=8, align="center")
merge(ws, r, 5, 6, "P3 = After P1+P2 land", bg="D9F99D", fc="3F6212", bold=True, size=8, align="center")
merge(ws, r, 7, 8, "Quality → Bid → Budget order must be respected for LP-broken campaigns", bg="F0F9FF", fc="075985", bold=True, size=8, align="center")
merge(ws, r, 9, NC, "", bg="F8FAFC"); rh(ws, r, 14); r += 1

# Column headers
for col, (h, _) in enumerate(COLS, 1):
    cs(ws, r, col, h, bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center", wrap=True)
rh(ws, r, 30); r += 1
ws.freeze_panes = f"A{r}"

# ── Data rows ──
PRI_COLORS = {
    "P1": ("FCA5A5", "7F1D1D"),
    "P2": ("FDE68A", "78350F"),
    "P3": ("D9F99D", "3F6212"),
}
TYPE_COLORS = {
    "Budget":       ("DBEAFE", "1E40AF"),
    "Bid":          ("DCFCE7", "14532D"),
    "Quality":      ("F3E8FF", "581C87"),
    "Location Bid": ("E0F2FE", "0C4A6E"),
    "GMB":          ("FFF1F2", "881337"),
}

prev_city = None
CITY_ORDER = ["ALL T1","Bangalore","Chennai","Hyderabad","Mumbai","Pune"]
actions_sorted = sorted(ACTIONS, key=lambda x: (
    CITY_ORDER.index(x["city"]) if x["city"] in CITY_ORDER else 99,
    x["cat"], ["P1","P2","P3"].index(x["pri"]),
))

for row in actions_sorted:
    city = row["city"]
    if city != prev_city:
        merge(ws, r, 1, NC,
              f"  {'── ' if city != 'ALL T1' else '══ CROSS-CITY ══  '}{city.upper()}",
              bg=("0F172A" if city == "ALL T1" else "1E3A5F"),
              fc="FFFFFF", bold=True, size=10 if city == "ALL T1" else 9, align="left")
        rh(ws, r, 18); r += 1
        prev_city = city

    pri_bg, pri_fc = PRI_COLORS.get(row["pri"], ("F8FAFC","374151"))
    # Determine type colour by prefix matching
    type_bg, type_fc = "F8FAFC", "374151"
    for key, (tb, tf) in TYPE_COLORS.items():
        if key.lower() in row["type"].lower():
            type_bg, type_fc = tb, tf
            break

    row_bg = {
        "P1": "FFF5F5",
        "P2": "FFFBEB",
        "P3": "F7FEE7",
    }.get(row["pri"], "F8FAFC")

    cells = [
        (row["pri"],      pri_bg,  "center", 9,  True,  pri_fc,  False),
        (row["city"],     row_bg,  "left",   8,  False, "111111", False),
        (row["cat"],      row_bg,  "center", 8,  True,  "374151", False),
        (row["campaign"], row_bg,  "left",   8,  False, "374151", True),
        (row["type"],     type_bg, "left",   8,  True,  type_fc,  True),
        (row["state"],    "F0F9FF","left",   7,  False, "075985", True),
        (row["diagnosis"],row_bg,  "left",   8,  False, "1E293B", True),
        (row["action"],   "F0FDF4","left",   8,  True,  "14532D", True),
        (row["quantum"],  "FEF3C7","center", 8,  True,  "78350F", True),
        (row["prereq"],   "F5F3FF","left",   8,  False, "5B21B6", True),
    ]
    rh(ws, r, 72)
    for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(cells, 1):
        cs(ws, r, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)
    r += 1

# ── Summary count ──────────────────────────────────────────────────────────
merge(ws, r, 1, NC, "", bg="F8FAFC"); rh(ws, r, 8); r += 1
p1 = sum(1 for x in ACTIONS if x["pri"]=="P1")
p2 = sum(1 for x in ACTIONS if x["pri"]=="P2")
p3 = sum(1 for x in ACTIONS if x["pri"]=="P3")
merge(ws, r, 1, NC,
      f"  TOTAL: {len(ACTIONS)} actionables  |  P1 (this week): {p1}  |  P2 (this month): {p2}  |  P3 (after P1+P2): {p3}",
      bg="0F172A", fc="94A3B8", bold=True, size=9, align="left")
rh(ws, r, 16); r += 1

OUT = os.path.join(DIR, "allo_t1_actions.xlsx")
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  {len(ACTIONS)} actions | P1={p1} | P2={p2} | P3={p3}")
print()
by_type = {}
for a in ACTIONS:
    t = a["type"].split("—")[0].strip().split("(")[0].strip()
    by_type[t] = by_type.get(t, 0) + 1
for t, c in sorted(by_type.items(), key=lambda x: -x[1]):
    print(f"  {t}: {c}")
