#!/usr/bin/env python3
"""Detailed competition analysis per clinic (SH), for ALL clinics — from SERP map-pack +
web-verified pathy. Three tabs: final analysis · city mix · detail.

Inputs: data_serp_competitors.tsv · data_serp_pathy.tsv (+ _v2 override) · data_clinic_ages.tsv ·
        data_campaign_compose.json
Output: ~/Downloads/competition_analysis.xlsx
"""
import os, csv, re, datetime, json, statistics as _st
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TODAY = datetime.date(2026, 7, 24)
TOP_N = 5
CAT_LABEL = {'SH': 'Sexologist'}

COMPOSE = json.load(open(os.path.join(ROOT, 'data_campaign_compose.json')))
CWK = COMPOSE['_meta']['weeks']; CNW = len(CWK)
def _wa(rows, f, p): return sum((x.get(f, 0) or 0) for x in rows if p(x)) / CNW
def sh_funnel(city):
    c = COMPOSE.get(city)
    if not c: return None
    ap = lambda x: x['cat'] == 'SH' and x['mt'] == 'Exact-Local'
    fp = lambda x: x['cat'] == 'SH' and (x['ch'], x['med']) != ('Google', 'Web')
    A = lambda f: _wa(c['acq'], f, ap); F = lambda f: _wa(c['fun'], f, fp)
    im, el, ck, lc = A('impr'), A('elig'), A('click'), A('locclick'); ld, bk = F('lead'), F('bk'); sp = A('sp')
    rt = lambda n, d: n / d if d else 0
    return dict(spend=sp, leads=ld, IS=rt(im, el), mkt=el, locpct=rt(lc, ck),
                loc2ld=rt(ld, lc), ld2bk=rt(bk, ld), cpl=rt(sp, ld))

# ---------- pathy taxonomy ----------
def load_pathy():
    out = {}
    for fn in ('data_serp_pathy.tsv', 'data_serp_pathy_v2.tsv'):   # v2 (re-verify) overrides
        p = os.path.join(ROOT, fn)
        if not os.path.exists(p): continue
        for row in csv.DictReader(open(p), delimiter='\t'): out[row['place_id']] = row
    return out
PATHY = load_pathy()
PATHY_ORDER = ['Allopathic', 'Ayurvedic', 'Unani', 'Homeopathic', 'Non-medical', 'Thin', 'Mixed']
ALTMED = ('Ayurvedic', 'Unani', 'Homeopathic')
def norm_pathy(p): return p if p in PATHY_ORDER else 'Thin'   # Unknown/blank -> Thin (weak)
PATHY_FILL = {'Allopathic': 'EEF0F3', 'Ayurvedic': 'E7DCF5', 'Unani': 'CFE9E6', 'Homeopathic': 'FBD9E8',
              'Non-medical': 'FCEFD4', 'Thin': 'F1F1F1', 'Mixed': 'FDE9D0'}

def load_ages():
    ages = {}; p = os.path.join(ROOT, 'data_clinic_ages.tsv')
    if os.path.exists(p):
        for row in csv.reader(open(p), delimiter='\t'):
            if len(row) >= 3 and row[2]:
                try: ages[(row[0], row[1])] = round((TODAY - datetime.date.fromisoformat(row[2])).days / 30.4, 1)
                except ValueError: pass
    return ages

def num(x, d=0.0):
    try: return float(x)
    except (TypeError, ValueError): return d

AYUR = re.compile(r'ayur|homoeo|homeo|unani|siddha|naturopath', re.I)
STUFF = re.compile(r'\b(best|top|no\.?\s?1|#1|no1)\b|sexolog.*(clinic|centre|center|hospital).*sexolog', re.I)
def keyword_stuffed(name):
    if STUFF.search(name): return True
    if '|' in name and re.search(r'sexolog|sexual|men', name, re.I): return True
    return name.count(' ') >= 7 and bool(re.search(r'sexolog|sexual health|men.?s health', name, re.I))
def rank_str(v):
    r = round(v); return f'#{r}' if r >= 1 else 'Rarely'

# ---------- styling ----------
ACC = '2C6CAE'; WHITE = 'FFFFFF'; INK = '1A2230'
fillc = lambda h: PatternFill('solid', fgColor=h)
thin = Side(style='thin', color='E4E8EF')
RED = Side(style='medium', color='C0392B'); GRN = Side(style='medium', color='1E8449'); PUR = Side(style='medium', color='7D5BA6')
BOX = lambda s: Border(left=s, right=s, top=s, bottom=s)

def main():
    ages = load_ages()
    rows = list(csv.DictReader(open(os.path.join(ROOT, 'data_serp_competitors.tsv')), delimiter='\t'))
    clinics = defaultdict(list)   # (city, loc) -> [comp]
    for r in rows:
        if r.get('cat', 'SH') != 'SH': continue
        key = (r['city'], r['locality'])
        pv = PATHY.get(r['place_id'], {})
        c = dict(name=r['comp_name'], rating=num(r['rating']), reviews=int(num(r['reviews'])),
                 appearances=int(num(r['appearances'])), avg_pos=num(r['avg_pos']),
                 km=num(r['clinic_km']) if r['clinic_km'] else None,
                 ever_sponsored=str(r['ever_sponsored']).lower() == 'true',
                 our_reviews=int(num(r['our_reviews'])), our_rank=num(r['our_avg_rank']), place_id=r['place_id'],
                 pathy=norm_pathy(pv.get('pathy')), qual=pv.get('qualification') or '—', vreason=pv.get('reason') or '')
        c['stuffed'] = keyword_stuffed(c['name'])
        clinics[key].append(c)
    for lst in clinics.values():
        mx = max((c['appearances'] for c in lst), default=1) or 1
        for c in lst: c['cov'] = c['appearances'] / mx

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    write_final_sheet(wb.create_sheet('SH · final analysis'), clinics)
    write_mix_sheet(wb.create_sheet('SH · city mix'), clinics)
    write_detail_sheet(wb.create_sheet('SH · detail'), clinics, ages)
    out = os.path.join(os.path.expanduser('~'), 'Downloads', 'competition_analysis.xlsx')
    wb.save(out)
    print(f'wrote {out} · {len(clinics)} clinics')

# ============ per-city competitor summary (top-5 deduped) ============
def city_top(clinics, city):
    seen = {}
    for (cty, loc), lst in clinics.items():
        if cty != city: continue
        for c in sorted(lst, key=lambda c: (-c['appearances'], c['avg_pos']))[:TOP_N]:
            pid = c['place_id']
            if pid not in seen or c['appearances'] > seen[pid]['appearances']: seen[pid] = c
    return sorted(seen.values(), key=lambda c: -c['appearances'])

def city_verdict(comps, our):
    n = len(comps) or 1
    cnt = {p: 0 for p in PATHY_ORDER}
    for c in comps: cnt[c['pathy']] += 1
    allo = cnt['Allopathic']; altpct = (cnt['Ayurvedic'] + cnt['Unani'] + cnt['Homeopathic']) / n
    toprev = max((c['reviews'] for c in comps), default=0)
    dom_alt = max(ALTMED, key=lambda p: cnt[p]) if altpct else None
    if allo / n >= 0.6: field = f'Allopathic-dominant ({allo}/{n})'
    elif altpct >= 0.5: field = f'{dom_alt}-led alt-medicine field ({altpct:.0%})'
    elif altpct >= 0.3: field = f'mixed, notable {dom_alt} presence'
    else: field = 'fragmented field'
    if our >= toprev: moat = f'we lead on reviews ({our} vs {toprev})'
    elif toprev > our * 1.5: moat = f'rivals out-review us ({toprev} vs {our}) — review gap'
    else: moat = f'review-competitive ({our} vs {toprev})'
    return f'{field}; {moat}'

def clinic_verdict(comps, our):
    top = comps[0]; tr = top['reviews']; tp = top['pathy']
    alt = sum(1 for c in comps if c['pathy'] in ALTMED)
    far = all((c['km'] or 9) > 4 for c in comps[:3])
    if our >= max(tr, 1): head = f'✅ we lead ({our} vs top {tr})'
    elif tr > our * 1.3: head = f'⚠ {tp} rival leads reviews ({tr} vs {our})'
    else: head = f'competitive (top {tr} vs {our})'
    tail = []
    if far and our < tr: tail.append('but top rivals >4km away')
    if alt >= 3: tail.append(f'{alt}/{len(comps)} alt-medicine')
    return head + (' · ' + ' · '.join(tail) if tail else '')

# ============ TAB 1: final analysis ============
def write_final_sheet(ws, clinics):
    hdr = ['City', 'Spend', 'Leads', 'IS', 'Market', 'Loc %', 'Loc→Lead %', 'Lead→Book %', 'CPL',
           'Our reviews', 'Top rival (rev · pathy)', 'Ads', 'IS — reason', 'Loc→Lead — reason', 'VERDICT']
    for j, h in enumerate(hdr):
        cc = ws.cell(1, 1 + j, h); cc.font = Font(bold=True, color=WHITE, size=10)
        cc.fill = fillc(ACC); cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    cities = [c for c in COMPOSE if c != '_meta' and sh_funnel(c) and sh_funnel(c)['spend'] > 800]
    cities.sort(key=lambda c: -sh_funnel(c)['spend'])
    medcpl = _st.median([sh_funnel(c)['cpl'] for c in cities if sh_funnel(c)['leads'] >= 3] or [200])
    rr = 2
    for city in cities:
        f = sh_funnel(city); comps = city_top(clinics, city)
        our = max((c['our_reviews'] for c in comps), default=0) if comps else 0
        topr = comps[0]['reviews'] if comps else 0; topp = comps[0]['pathy'] if comps else '—'
        topn = comps[0]['name'][:20] if comps else ''
        ads = sum(1 for c in comps if c['ever_sponsored'])
        moat = topr > our * 1.3
        is_reason = f'{ads} rivals run ads — real auction pressure' if ads >= 2 else 'our bid/QS too low (rivals rarely bid; IS lost to rank)'
        if our >= max(topr, 1): l2l = f'we lead reviews ({our} vs {topr}) — holds'
        elif moat: l2l = f'review gap: {topp} rival {topr} vs our {our}' + (f' + {topp} positioning' if topp in ALTMED else '')
        else: l2l = f'thin moat (our {our})'
        if f['cpl'] > 400 and f['leads'] < 5: verdict, vf = 'CUT — no moat, no volume', 'F6C9C9'
        elif f['ld2bk'] < 0.6 and f['leads'] >= 4: verdict, vf = 'FIX BOOKING (ops) + reviews', 'FCEFD4'
        elif our < 100 and moat and f['loc2ld'] < 0.18: verdict, vf = 'REVIEWS FIRST — don’t scale bids', 'FCEFD4'
        elif f['IS'] >= 0.68: verdict, vf = 'HOLD — saturated', 'EEF0F3'
        elif f['cpl'] <= medcpl * 1.1 and f['leads'] >= 8 and f['IS'] < 0.65: verdict, vf = 'SCALE BIDS ↑ (+ keep reviews)', 'C9E7CE'
        else: verdict, vf = 'MAINTAIN + build reviews', 'EEF0F3'
        vals = [city, round(f['spend']), round(f['leads']), f['IS'], round(f['mkt']), f['locpct'], f['loc2ld'],
                f['ld2bk'], round(f['cpl']), round(our), f'{topn} ({topr}·{topp})', ads, is_reason, l2l, verdict]
        fmts = [None, '₹#,##0', '#,##0', '0%', '#,##0', '0%', '0%', '0%', '₹#,##0', '#,##0', None, '0', None, None, None]
        for j, v in enumerate(vals):
            cell = ws.cell(rr, 1 + j, v); cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal='left' if j in (0, 10, 12, 13, 14) else 'center', vertical='center', wrap_text=(j in (12, 13, 14)))
            if fmts[j]: cell.number_format = fmts[j]
            if j == 0: cell.font = Font(bold=True, color=INK)
            if j == 3 and f['IS'] < 0.5: cell.fill = fillc('F6C9C9')
            if j == 6 and f['loc2ld'] < 0.15: cell.fill = fillc('F6C9C9')
            elif j == 6 and f['loc2ld'] >= 0.25: cell.fill = fillc('C9E7CE')
            if j == 7 and f['ld2bk'] < 0.6: cell.fill = fillc('F6C9C9')
            if j == 8 and f['cpl'] > medcpl * 1.8: cell.fill = fillc('F6C9C9')
            if j == 10 and moat: cell.fill = fillc('F6C9C9')
            if j == 14: cell.fill = fillc(vf); cell.font = Font(bold=True, color=INK)
        rr += 1
    for j, w in enumerate([13, 8, 7, 6, 9, 7, 9, 10, 8, 9, 24, 6, 30, 34, 30]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.freeze_panes = 'B2'; ws.auto_filter.ref = f'A1:{get_column_letter(len(hdr))}{rr - 1}'
    ws.cell(rr + 1, 1, 'IS is lost to OUR bid/QS (rivals rarely advertise) · Loc→Lead is lost to the REVIEW MOAT + alt-medicine positioning · verdict = the fix.').font = Font(italic=True, color='8A93A3', size=9)

# ============ TAB 2: city mix — summary with named rivals + impact ============
def rival_str(c):
    km = f'{c["km"]:.0f}km' if c['km'] is not None else '?km'
    return f'{c["name"][:20]} ({c["pathy"]}·{c["reviews"]}rev·{km})'

def city_summary(city, comps, clinics):
    """Fuller narrative: top threat + our exposed clinics."""
    n = len(comps)
    cnt = {p: 0 for p in PATHY_ORDER}
    for c in comps: cnt[c['pathy']] += 1
    altpct = (cnt['Ayurvedic'] + cnt['Unani'] + cnt['Homeopathic']) / n if n else 0
    dom_alt = max(ALTMED, key=lambda p: cnt[p]) if altpct else None
    top = comps[0]; our = max((c['our_reviews'] for c in comps), default=0)
    # our clinics in this city, most-exposed first (thinnest reviews / not #1)
    myclinics = []
    for (cty, loc), lst in clinics.items():
        if cty == city:
            c0 = lst[0]; myclinics.append((loc, c0['our_reviews'] or 0, c0['our_rank']))
    myclinics.sort(key=lambda x: x[1])
    exposed = [loc for loc, rev, rk in myclinics if rev < top['reviews'] * 0.7 or round(rk) > 1][:3]
    field = (f'{dom_alt}-led ({altpct:.0%} alt-medicine)' if altpct >= 0.5
             else f'Allopathic-heavy' if cnt['Allopathic'] / n >= 0.6 else 'mixed field')
    lead = f'we out-review them ({our} vs {top["reviews"]})' if our >= top['reviews'] else f'they out-review us ({top["reviews"]} vs {our})'
    exp = f' → most exposed: {", ".join(exposed)}' if exposed else ' → we hold across clinics'
    return f'{field}. Top threat {top["name"][:22]} ({top["pathy"]}, {top["reviews"]} reviews); {lead}{exp}.'

def write_mix_sheet(ws, clinics):
    hdr = ['City', '# Comp', 'Allopathic', 'Non-Allo', '· Ayurvedic', '· Unani', '· Homeopathic', '· Other',
           'Dominant', 'Our reviews', 'Top 3 rivals (pathy · reviews · dist)', 'Our clinics most exposed', 'Summary insight']
    for j, h in enumerate(hdr):
        cc = ws.cell(1, 1 + j, h); cc.font = Font(bold=True, color=WHITE, size=10)
        cc.fill = fillc(ACC); cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    cities = sorted({c for (c, l) in clinics}, key=lambda c: -((sh_funnel(c) or {}).get('spend', 0)))
    natcnt = {p: 0 for p in PATHY_ORDER}; natn = 0
    rr = 2
    for city in cities:
        comps = city_top(clinics, city)
        if not comps: continue
        cnt = {p: 0 for p in PATHY_ORDER}
        for c in comps: cnt[c['pathy']] += 1; natcnt[c['pathy']] += 1
        n = len(comps); natn += n
        allo = cnt['Allopathic']; nonallo = n - allo; other = cnt['Non-medical'] + cnt['Thin'] + cnt['Mixed']
        allo_p = allo / n if n else 0
        lead_alt = max(ALTMED, key=lambda p: cnt[p]) if any(cnt[p] for p in ALTMED) else 'Non-medical'
        top_pathy = comps[0]['pathy']
        if allo_p >= 0.55: dom, domfill = 'Allopathic', 'EEF0F3'
        elif allo_p <= 0.45:
            sub = top_pathy if top_pathy in ALTMED else lead_alt   # face it with the actual top rival's pathy
            dom, domfill = f'Non-Allo · {sub}', PATHY_FILL.get(sub, 'E7DCF5')
        else: dom, domfill = f'Split (Allo/{lead_alt})', 'DCE6F4'
        pct = lambda x: (x / n) if n else 0
        our = max((c['our_reviews'] for c in comps), default=0)
        top3 = ' ; '.join(rival_str(c) for c in comps[:3])
        # exposed clinics
        myc = sorted([(loc, clinics[(city, loc)][0]['our_reviews'] or 0, clinics[(city, loc)][0]['our_rank'])
                      for (cty, loc) in clinics if cty == city], key=lambda x: x[1])
        toprev = comps[0]['reviews']
        exp_all = [loc for loc, rev, rk in myc if rev < toprev * 0.7 or round(rk) > 1]
        exposed = (', '.join(exp_all[:4]) + (f' +{len(exp_all) - 4} more' if len(exp_all) > 4 else '')) or '— we hold'
        cells = [city, n, pct(allo), pct(nonallo), pct(cnt['Ayurvedic']), pct(cnt['Unani']), pct(cnt['Homeopathic']), pct(other),
                 dom, round(our) or None, top3, exposed, city_summary(city, comps, clinics)]
        for j, v in enumerate(cells):
            cell = ws.cell(rr, 1 + j, v); cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal='left' if j in (0, 10, 11, 12) else 'center', vertical='center', wrap_text=(j in (10, 11, 12)))
            if 2 <= j <= 7: cell.number_format = '0%'
            if j == 0: cell.font = Font(bold=True, color=INK)
            if j == 2 and pct(allo) >= 0.6: cell.fill = fillc('EEF0F3')
            if j == 3 and pct(nonallo) >= 0.6: cell.fill = fillc('E7DCF5')
            if j == 8: cell.fill = fillc(domfill); cell.font = Font(bold=True)
        ws.row_dimensions[rr].height = 40
        rr += 1
    # NATIONAL summary row
    natn = natn or 1
    top_nat = 'Allopathic 49% of rivals; alt-medicine (Ayurvedic+Unani+Homeopathic) ~37% — biggest non-allo blocs: Ayurvedic then Unani. Review moat is the deciding factor city-to-city.'
    cells = ['NATIONAL', natn, natcnt['Allopathic'] / natn, 1 - natcnt['Allopathic'] / natn,
             natcnt['Ayurvedic'] / natn, natcnt['Unani'] / natn, natcnt['Homeopathic'] / natn,
             (natcnt['Non-medical'] + natcnt['Thin'] + natcnt['Mixed']) / natn, 'Even (~50/50)', None, '', '', top_nat]
    rr += 1
    for j, v in enumerate(cells):
        cell = ws.cell(rr, 1 + j, v); cell.font = Font(bold=True, color=INK)
        cell.alignment = Alignment(horizontal='left' if j in (0, 12) else 'center', vertical='center', wrap_text=(j == 12))
        if 2 <= j <= 7: cell.number_format = '0%'
        cell.fill = fillc('DCE6F4')
    ws.row_dimensions[rr].height = 40
    for j, w in enumerate([13, 6, 10, 9, 10, 7, 11, 7, 11, 9, 40, 22, 66]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.freeze_panes = 'B2'; ws.auto_filter.ref = f'A1:{get_column_letter(len(hdr))}{rr - 2}'
    ws.cell(rr + 1, 1, 'Per-city competition summary: mix % (Allopathic vs Non-Allopathic), the named top-3 rivals with pathy·reviews·distance, '
            'our clinics most exposed (thin reviews or not ranking #1), and a one-line insight. NATIONAL row = overall.').font = Font(italic=True, color='8A93A3', size=9)

# ============ TAB 3: detail — flat, filterable by City, border-highlighted, with insights ============
def write_detail_sheet(ws, clinics, ages):
    hdr = ['City', 'Clinic', 'Clinic verdict', 'Age', 'Our rank', 'Our rev', 'Competitor', 'Rank',
           'Rating', 'Reviews', 'Ratio', 'Dist', 'Qualification', 'Pathy', 'Reason']
    for j, h in enumerate(hdr):
        cc = ws.cell(1, 1 + j, h); cc.font = Font(bold=True, color=WHITE, size=10)
        cc.fill = fillc(ACC); cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    bycity = defaultdict(list)
    for (city, loc) in clinics: bycity[city].append(loc)
    ordered = sorted(bycity, key=lambda c: (-((sh_funnel(c) or {}).get('spend', 0)), c))
    topb = Side(style='thin', color='9AA6B5')
    rr = 2
    for city in ordered:
        for loc in sorted(bycity[city], key=lambda l: -(clinics[(city, l)][0]['our_reviews'] or 0)):
            comps = sorted(clinics[(city, loc)], key=lambda c: (-c['appearances'], c['avg_pos']))[:TOP_N]
            if not comps: continue
            our = comps[0]['our_reviews']; cv = clinic_verdict(comps, our)
            age = ages.get((city, loc)); orank = rank_str(comps[0]['our_rank'])
            for i, c in enumerate(comps):
                ratio = round(our / c['reviews'], 2) if c['reviews'] else None
                km = c['km']; pathy = c['pathy']
                vals = [city, loc, cv if i == 0 else '', f'{age} mo' if age is not None else '—',
                        orank if i == 0 else '', our if i == 0 else '', c['name'], rank_str(c['avg_pos']),
                        c['rating'] or None, c['reviews'] or None, ratio, f'{km:.1f} km' if km is not None else '—',
                        c['qual'], pathy, c['vreason']]
                for j, v in enumerate(vals):
                    cell = ws.cell(rr, 1 + j, v)
                    cell.alignment = Alignment(horizontal='left' if j in (1, 2, 6, 12, 14) else 'center', vertical='center', wrap_text=(j in (2, 14)))
                    cell.border = Border(bottom=thin, top=(topb if i == 0 else None))
                    if j in (0, 1) and i == 0: cell.font = Font(bold=True, color=INK)
                    if j == 2: cell.font = Font(bold=True, color=INK, size=9)
                    # highlight only the THREAT-driving cells with borders (not blanket fills)
                    if j == 9 and c['reviews'] and c['reviews'] >= 500: cell.border = BOX(RED)          # big review base
                    if j == 10 and ratio is not None and ratio < 0.5: cell.border = BOX(RED)           # they out-review us 2×+
                    elif j == 10 and ratio is not None and ratio >= 2: cell.border = BOX(GRN)          # we out-review 2×+
                    if j == 11 and km is not None and km <= 3: cell.border = BOX(RED)                  # very close
                    if j == 13:                                                                        # pathy: light fill + alt-med box
                        cell.fill = fillc(PATHY_FILL.get(pathy, 'FFFFFF'))
                        if pathy in ALTMED: cell.border = BOX(PUR)
                rr += 1
    for j, w in enumerate([12, 15, 40, 7, 8, 7, 30, 7, 7, 9, 7, 9, 22, 12, 58]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.freeze_panes = 'C2'; ws.auto_filter.ref = f'A1:{get_column_letter(len(hdr))}{rr - 1}'
    ws.cell(rr + 1, 1, 'Filter by City (col A) to deep-dive. Clinic verdict = the outcome for that clinic. '
            'Borders flag the threat driver: red = rival out-reviews us 2×+ / ≥500 reviews / ≤3km · green = we out-review 2×+ · purple = Ayurvedic/Unani/Homeopathic. '
            'Pathy web-verified from the doctor’s degree.').font = Font(italic=True, color='8A93A3', size=9)

if __name__ == '__main__':
    main()
