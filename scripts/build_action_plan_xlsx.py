#!/usr/bin/env python3
"""Competition × Google-Ads ACTION PLAN workbook.
Combines the paid funnel (IS / Loc→Lead / Lead→Book / CPL...) with the local-map-pack competition
(review gap, top rival, who beats us) to produce a scale / fix-profile / fix-booking / cut call per
city and per clinic — and flags contradictions (e.g. Google says 'scale' but the review gap is so big
that paid spend just leaks to a better-reviewed rival).

Sheets:  ① Action Highlights  ·  ② SH cities ③ STI cities ④ MH cities  ·  ⑤ SH clinics ⑥ STI clinics ⑦ MH clinics
"""
import os, json, statistics as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CUBE = json.load(open(os.path.join(ROOT, 'data_competition.json')))
CATLBL = {'SH': 'Sexual Health', 'STI': 'STI', 'MH': 'Mental Health'}

# benchmarks
IS_T, LOC2LD_T, LOC2LD_HARD, LD2BK_T = 0.70, 0.30, 0.20, 0.65

WHITE = Font(color='FFFFFF', bold=True)
BOLD = Font(bold=True)
HDR = PatternFill('solid', fgColor='2C6CAE')
SUB = PatternFill('solid', fgColor='EEF2F8')
RED = PatternFill('solid', fgColor='F8C9C4')
AMB = PatternFill('solid', fgColor='FBE7C2')
GRN = PatternFill('solid', fgColor='CDE9D3')
GREY = PatternFill('solid', fgColor='E4E8EF')
BLU = PatternFill('solid', fgColor='D6E4F5')
thin = Side(style='thin', color='D5DBE5')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

ACTFILL = {'SCALE': GRN, 'SCALE + FIX PROFILE': AMB, 'FIX PROFILE': AMB, 'FIX BOOKING': PatternFill('solid', fgColor='F6D9B8'),
           'CUT': RED, 'HOLD — efficient': GREY, 'REVIEW-ONLY (no paid spend)': BLU, 'SCALE ⚠ close reviews first': AMB}


def city_stat(cat, city):
    C = CUBE[cat]
    cc = C['cities'][city]
    keys = cc['clinics']
    revs = [C['clinics'][k]['our_reviews'] for k in keys]
    med = int(st.median(revs)) if revs else 0
    beaten = sum(1 for k in keys if any(t.startswith('beaten:') for t in C['clinics'][k]['tags']))
    tr = cc['top_rivals'][0] if cc['top_rivals'] else {}
    f = cc.get('funnel') or {}
    return dict(city=city, n=len(keys), med=med, beaten=beaten, winrate=cc.get('winrate', 0),
                rival=tr.get('name', ''), rtype=tr.get('pathy', ''), rrev=tr.get('reviews', 0), f=f)


def decide(s):
    """Return (action, why, insight)."""
    f = s['f']
    gapx = (s['rrev'] / s['med']) if s['med'] else (99 if s['rrev'] else 1)   # rival reviews ÷ our median
    big_gap = gapx >= 3
    if not f or not f.get('spend'):
        why = f"No paid spend. {s['beaten']}/{s['n']} clinics out-reviewed (top rival {s['rival']} {s['rrev']} vs our median {s['med']})."
        ins = "Organic/GMB market — build reviews & GMB before/instead of paid." if s['beaten'] else "Holding on reviews organically."
        return 'REVIEW-ONLY (no paid spend)', why, ins
    IS, l2l, l2b = f.get('IS', 0), f.get('loc2ld', 0), f.get('ld2bk', 0)
    room = IS < IS_T
    profile = l2l < LOC2LD_HARD
    profile_soft = l2l < LOC2LD_T
    booking = l2b < LD2BK_T
    why = f"IS {IS:.0%} · Loc→Lead {l2l:.0%} · Lead→Book {l2b:.0%} · CPL ₹{f.get('cpl',0)}"
    ins = ''
    if booking and not profile_soft:
        act = 'FIX BOOKING'
        ins = f"Leads arrive but stall at booking ({l2b:.0%} vs {LD2BK_T:.0%} bench) — ops: calling, availability, follow-up."
    elif profile and room:
        act = 'SCALE + FIX PROFILE'
        ins = f"Room to scale (IS {IS:.0%}) but map clicks don't convert ({l2l:.0%}); rival {s['rtype']} {s['rrev']} vs our {s['med']} reviews wins the listing. Scale + build reviews."
    elif profile or profile_soft:
        act = 'FIX PROFILE'
        ins = f"Map clicks leak — listing loses to {s['rtype']} rival ({s['rrev']} vs {s['med']} reviews). Build reviews / GMB, don't just raise budget."
    elif room and big_gap:
        act = 'SCALE ⚠ close reviews first'
        ins = f"⚠ Google shows room (IS {IS:.0%}) BUT rival out-reviews us {gapx:.0f}× ({s['rrev']} vs {s['med']}). Scaling paid spend leaks to the better-reviewed rival — fix reviews alongside."
    elif room:
        act = 'SCALE'
        ins = f"Healthy funnel + room (IS {IS:.0%}). Raise budget/bid to capture more of the {f.get('mkt',0):,}/wk market."
    else:
        act = 'HOLD — efficient'
        ins = f"Capturing most of the market (IS {IS:.0%}) with a healthy funnel. Hold; defend reviews."
    # contradiction overlay even when action isn't scale
    if 'SCALE' in act and big_gap and 'close reviews' not in act and 'PROFILE' not in act:
        ins += f"  ⚠ Note: rival out-reviews us {gapx:.0f}× — pair the scale with a review push."
    return act, why, ins


def lever_move(s):
    """The specific KNOB + budget DIRECTION. Returns (lever, move, rationale).
    move ∈ {INVEST ↑, HOLD, FIX PROFILE, FIX PROFILE FIRST, FIX BOOKING, CUT ↓, ORGANIC}."""
    f = s['f']
    gapx = (s['rrev'] / s['med']) if s['med'] else (99 if s['rrev'] else 1); big = gapx >= 3
    if not f or not f.get('spend'):
        return ('Reviews / GMB', 'ORGANIC', f"No paid spend; {s['beaten']}/{s['n']} out-reviewed → grow reviews/GMB first")
    IS, l2l, l2b = f.get('IS', 0), f.get('loc2ld', 0), f.get('ld2bk', 0)
    cpd = (f['spend'] / f['done']) if f.get('done') else None
    room, profile, profsoft, booking = IS < IS_T, l2l < LOC2LD_HARD, l2l < LOC2LD_T, l2b < LD2BK_T
    expensive = cpd is not None and cpd > 600
    comparable = gapx <= 2.0     # review gap coverable → a bid/profile push can win the click now
    heavy = gapx >= 5.0          # rival out-reviews us so heavily that paid just leaks to them
    # 1) cut: expensive + tiny + leaky → pull budget
    if expensive and f.get('leads', 0) < 10 and (profsoft or booking):
        return ('Cut', 'CUT ↓', f"CPD ₹{cpd:.0f}, only {f['leads']:.0f} leads/wk — inefficient; pull budget & reallocate")
    # 2) booking is the drop → ops, hold spend
    if booking and not profsoft:
        return ('Booking ops', 'FIX BOOKING', f"Lead→Book {l2b:.0%} vs {LD2BK_T:.0%} — ops (calling/availability), not budget")
    # 3) heavy competitor + clicks leak → paid wasted; it's a reviews play, don't bid to win
    if heavy and profsoft:
        return ('Reviews (not bid)', 'FIX PROFILE FIRST', f"Rival out-reviews us {gapx:.0f}× — too heavy to win on ads; build reviews, don't push bid here")
    # 4) COMPARABLE + room → the winnable case: a bid/profile push can win while reviews catch up
    if room and comparable:
        return ('Bid / Budget ↑', 'INVEST ↑', f"Gap coverable ({gapx:.1f}×) + IS {IS:.0%} room — bid/profile push can win the click while reviews catch up")
    # 5) room to buy but a real gap → invest carefully + build reviews
    if room:
        tag = 'INVEST ↑ (careful)' if (expensive or heavy) else 'INVEST ↑'
        note = ' + build reviews' if (profsoft or gapx > 2) else ''
        lev = 'Budget / Bid' + (' + Reviews' if note else '')
        return (lev, tag, f"IS {IS:.0%} room, gap {gapx:.1f}×" + (f", CPD ₹{cpd:.0f}" if expensive else "") + f" — raise budget/bid{note}")
    # 6) no IS room — growth lever is profile, or hold
    if profsoft:
        return ('Reviews / GMB rank', 'FIX PROFILE', f"IS {IS:.0%} (little room) & Loc→Lead {l2l:.0%} — grow via reviews/GMB, not budget")
    return ('Hold / defend', 'HOLD', f"IS {IS:.0%}, funnel healthy — hold spend, defend reviews")


ORANGE = PatternFill('solid', fgColor='F6D9B8')
MOVEFILL = {'INVEST ↑': GRN, 'INVEST ↑ (careful)': PatternFill('solid', fgColor='DDEBC9'), 'HOLD': GREY,
            'FIX PROFILE': AMB, 'FIX PROFILE FIRST': AMB, 'FIX BOOKING': ORANGE, 'CUT ↓': RED, 'ORGANIC': BLU}
MOVEORDER = ['INVEST ↑', 'INVEST ↑ (careful)', 'FIX PROFILE FIRST', 'FIX PROFILE', 'FIX BOOKING', 'CUT ↓', 'HOLD', 'ORGANIC']


def hdr_row(ws, row, cols, widths):
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row, i, c); cell.fill = HDR; cell.font = WHITE; cell.alignment = CEN; cell.border = BORD
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def flag(cell, val, bad, warn, invert=False):
    """Colour a metric cell red/amber against a benchmark (invert=True → lower is worse)."""
    if val is None: return
    below = val < bad
    warnb = val < warn
    if invert:  # e.g. cost, gap — higher is worse; here handled by caller instead
        return
    if below: cell.fill = RED
    elif warnb: cell.fill = AMB


def build():
    wb = openpyxl.Workbook()
    # ── Sheet ① Highlights ───────────────────────────────────────────────
    ws = wb.active; ws.title = '① Highlights + why'
    ws.merge_cells('A1:I1')
    t = ws.cell(1, 1, 'HIGHLIGHTS — the binding metric per city×category, the reason, the lever, and the budget move')
    t.font = Font(bold=True, size=13, color='1A2230'); t.alignment = LFT
    ws.row_dimensions[1].height = 22
    hdr_row(ws, 2, ['Category', 'City', 'Clinics', 'MOVE (budget)', 'LEVER', 'Why — the binding metric', 'Insight / contradiction', 'Top rival (reviews)', 'Our med rev'],
            [13, 15, 8, 17, 16, 30, 52, 25, 10])
    r = 3
    rows_all = []
    for cat in CUBE['_meta']['cats']:
        for city in CUBE[cat]['cities']:
            s = city_stat(cat, city)
            act, why, ins = decide(s); lever, move, rat = lever_move(s)
            rows_all.append((cat, s, act, why, ins, lever, move, rat))
    hi = [h for h in rows_all if (h[1]['f'].get('spend') or h[1]['beaten'])]
    hi.sort(key=lambda h: (MOVEORDER.index(h[6]) if h[6] in MOVEORDER else 9, -h[1]['n']))
    for cat, s, act, why, ins, lever, move, rat in hi:
        ws.cell(r, 1, CATLBL[cat]).border = BORD
        ws.cell(r, 2, s['city']).border = BORD
        ws.cell(r, 3, s['n']).border = BORD; ws.cell(r, 3).alignment = CEN
        m = ws.cell(r, 4, move); m.fill = MOVEFILL.get(move, GREY); m.font = BOLD; m.alignment = CEN; m.border = BORD
        ws.cell(r, 5, lever).border = BORD; ws.cell(r, 5).alignment = CEN
        ws.cell(r, 6, why).border = BORD; ws.cell(r, 6).alignment = LFT
        ws.cell(r, 7, ins).border = BORD; ws.cell(r, 7).alignment = LFT
        ws.cell(r, 8, f"{s['rival']} ({s['rrev']})").border = BORD; ws.cell(r, 8).alignment = LFT
        ws.cell(r, 9, s['med']).border = BORD; ws.cell(r, 9).alignment = CEN
        ws.row_dimensions[r].height = 40
        r += 1
    ws.freeze_panes = 'A3'

    # ── Sheet ② Budget reallocation — pull from / invest in ──────────────
    ws = wb.create_sheet('② Budget reallocation')
    ws.merge_cells('A1:K1')
    t = ws.cell(1, 1, 'BUDGET REALLOCATION — pull ₹ from CUT/inefficient, feed INVEST; FIX before scaling (grouped by move)')
    t.font = Font(bold=True, size=13); t.alignment = LFT; ws.row_dimensions[1].height = 22
    bcols = ['MOVE', 'Category', 'City', 'Spend/wk', 'IS', 'Loc→Lead', 'Lead→Book', 'CPD', 'Rival rev vs ours', 'LEVER', 'Rationale']
    hdr_row(ws, 2, bcols, [17, 13, 15, 10, 7, 9, 9, 9, 16, 16, 56])
    r = 3
    paid = [(cat, s, lever, move, rat) for cat, s, act, why, ins, lever, move, rat in rows_all if s['f'].get('spend')]
    paid.sort(key=lambda x: (MOVEORDER.index(x[3]) if x[3] in MOVEORDER else 9, -(x[1]['f'].get('spend') or 0)))
    cur_move = None
    for cat, s, lever, move, rat in paid:
        if move != cur_move:   # section band
            cur_move = move
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
            b = ws.cell(r, 1, f"{move}  ·  {'PULL BUDGET' if move=='CUT ↓' else 'FEED THIS' if move.startswith('INVEST') else 'FIX BEFORE SCALING' if 'FIX' in move else 'HOLD / defend'}")
            b.fill = MOVEFILL.get(move, GREY); b.font = BOLD; b.alignment = LFT; r += 1
        f = s['f']; cpd = (f['spend'] / f['done']) if f.get('done') else None
        gapx = (s['rrev'] / s['med']) if s['med'] else 0
        vals = [move, CATLBL[cat], s['city'], f.get('spend'), f.get('IS'), f.get('loc2ld'), f.get('ld2bk'),
                round(cpd) if cpd else '', f"{s['rrev']} vs {s['med']} ({gapx:.0f}×)" if s['med'] else str(s['rrev']), lever, rat]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v); c.border = BORD; c.alignment = LFT if i in (1, 3, 9, 10, 11) else CEN
        for col in (5, 6, 7):
            if ws.cell(r, col).value is not None: ws.cell(r, col).number_format = '0%'
        for col in (4, 8):
            if ws.cell(r, col).value not in (None, ''): ws.cell(r, col).number_format = '#,##0'
        if f.get('loc2ld') is not None: ws.cell(r, 6).fill = RED if f['loc2ld'] < LOC2LD_HARD else (AMB if f['loc2ld'] < LOC2LD_T else GRN)
        if f.get('ld2bk') is not None and f['ld2bk'] < LD2BK_T: ws.cell(r, 7).fill = RED
        if f.get('IS') is not None and f['IS'] < IS_T: ws.cell(r, 5).fill = AMB
        r += 1
    ws.freeze_panes = 'A3'

    # ── Sheets ③④⑤ city-level per category ──────────────────────────────
    ccols = ['City', '#Clinics', 'MOVE', 'LEVER', 'Market/wk', 'IS', 'Spend/wk', 'CPL', 'Loc %', 'Loc→Lead', 'Lead→Book', 'Book→Done',
             'Our med rev', 'Top rival', 'Rival type', 'Rival rev', 'Gap ×', '#Beaten', 'Insight / contradiction']
    cwid = [15, 8, 17, 16, 10, 7, 10, 8, 7, 9, 9, 9, 10, 22, 14, 9, 7, 8, 55]
    for cat, sh in zip(CUBE['_meta']['cats'], ['③ SH — cities', '④ STI — cities', '⑤ MH — cities']):
        ws = wb.create_sheet(sh)
        ws.merge_cells('A1:S1')
        t = ws.cell(1, 1, f'{CATLBL[cat]} — city action plan (Google Ads funnel × competition)'); t.font = Font(bold=True, size=12); t.alignment = LFT
        hdr_row(ws, 2, ccols, cwid)
        rows = sorted((city_stat(cat, c) for c in CUBE[cat]['cities']), key=lambda s: -(s['f'].get('spend') or 0))
        r = 3
        for s in rows:
            f = s['f']; act, why, ins = decide(s); lever, move, rat = lever_move(s)
            gapx = (s['rrev'] / s['med']) if s['med'] else (s['rrev'] and 99 or 0)
            vals = [s['city'], s['n'], move, lever, f.get('mkt'), f.get('IS'), f.get('spend'), f.get('cpl'), f.get('locpct'),
                    f.get('loc2ld'), f.get('ld2bk'), f.get('bk2dn'), s['med'], s['rival'], s['rtype'], s['rrev'],
                    round(gapx, 1) if gapx else '', s['beaten'], ins]
            for i, v in enumerate(vals, 1):
                cell = ws.cell(r, i, v); cell.border = BORD
                cell.alignment = LFT if i in (1, 4, 14, 15, 19) else CEN
            # percentage formats (IS 6, Loc% 9, Loc→Lead 10, Lead→Book 11, Book→Done 12)
            for col in (6, 9, 10, 11, 12):
                if ws.cell(r, col).value is not None: ws.cell(r, col).number_format = '0%'
            for col in (5, 7, 8):
                if ws.cell(r, col).value is not None: ws.cell(r, col).number_format = '#,##0'
            # benchmark colours
            if f.get('IS') is not None and f['IS'] < IS_T: ws.cell(r, 6).fill = AMB if f['IS'] >= 0.5 else RED
            if f.get('loc2ld') is not None: ws.cell(r, 10).fill = RED if f['loc2ld'] < LOC2LD_HARD else (AMB if f['loc2ld'] < LOC2LD_T else GRN)
            if f.get('ld2bk') is not None and f['ld2bk'] < LD2BK_T: ws.cell(r, 11).fill = RED
            if gapx and gapx >= 3: ws.cell(r, 17).fill = RED
            elif gapx and gapx >= 1.5: ws.cell(r, 17).fill = AMB
            m = ws.cell(r, 3); m.fill = MOVEFILL.get(move, GREY); m.font = BOLD
            ws.row_dimensions[r].height = 40
            r += 1
        ws.freeze_panes = 'A3'

    # ── Sheets ⑥⑦⑧ clinic-level per category ────────────────────────────
    kcols = ['City', 'Clinic', 'Our rev', 'Top rival', 'Rival type', 'Rival rev', 'Gap (rival−ours)', 'Ratio',
             'Dist km', 'GMB views/wk', 'Verdict', 'Clinic action']
    kwid = [14, 20, 9, 24, 15, 9, 14, 8, 8, 12, 34, 30]
    for cat, sh in zip(CUBE['_meta']['cats'], ['⑥ SH — clinics', '⑦ STI — clinics', '⑧ MH — clinics']):
        ws = wb.create_sheet(sh)
        ws.merge_cells('A1:L1')
        t = ws.cell(1, 1, f'{CATLBL[cat]} — clinic action plan (per Allo clinic vs its #1 local rival)'); t.font = Font(bold=True, size=12); t.alignment = LFT
        hdr_row(ws, 2, kcols, kwid)
        cl = CUBE[cat]['clinics']
        rows = sorted(cl.values(), key=lambda v: (v['city'], v['loc']))
        r = 3
        for v in rows:
            tr = v['competitors'][0]
            gap = v['our_reviews'] - tr['reviews']
            ratio = (v['our_reviews'] / tr['reviews']) if tr['reviews'] else None
            gmb = v.get('gmb') or {}
            act = ('Hold + keep reviews' if v['vkind'] == 'win' else
                   'Fix GMB rank/bid (we have reviews)' if v['vkind'] == 'outrank' else
                   f"Build reviews (need +{-gap} to match)")
            vals = [v['city'], v['loc'], v['our_reviews'], tr['name'], tr['pathy'], tr['reviews'], gap,
                    round(ratio, 2) if ratio is not None else '', tr.get('km') if tr.get('km') is not None else '',
                    gmb.get('searches', ''), v['verdict'], act]
            for i, x in enumerate(vals, 1):
                cell = ws.cell(r, i, x); cell.border = BORD
                cell.alignment = LFT if i in (1, 2, 4, 5, 11, 12) else CEN
            ws.cell(r, 7).fill = GRN if gap >= 0 else (RED if gap < -300 else AMB)
            if ratio is not None: ws.cell(r, 8).fill = GRN if ratio >= 1 else (AMB if ratio >= 0.5 else RED)
            if gmb.get('searches'): ws.cell(r, 10).number_format = '#,##0'
            r += 1
        ws.freeze_panes = 'A3'

    out = os.path.join(ROOT, 'Competition_Action_Plan.xlsx')
    wb.save(out)
    print('wrote', out, '·', len(hi), 'highlight rows ·', len(CUBE['_meta']['cats']), 'categories')


if __name__ == '__main__':
    build()
