#!/usr/bin/env python3
"""Build mh-practo.html — a REFERENCE view of Practo General-Psychiatry locality pageviews.

This data is attached for reference only. It is NOT used in the demand score (Practo pageviews reflect
Practo's own traffic footprint, and only ~1/3 of clinics map cleanly to a Practo zone — so the Scoring
Explorer keeps demand at city-level search volume). This page just lets you see the locality-level Practo
numbers per city, with the zones that correspond to an Allo clinic highlighted.
"""
import openpyxl, json, os, re
_here = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(_here) if os.path.basename(_here) == 'scripts' else _here
def P(f): return os.path.join(ROOT, f)

def main():
    wb = openpyxl.load_workbook(P('practo_psychiatry.xlsx'), data_only=True)
    ws = wb[wb.sheetnames[0]]
    from collections import defaultdict
    zones = defaultdict(list); citypv = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]: continue
        c = str(r[0]).strip().lower(); pv = r[4] if isinstance(r[4], (int, float)) else 0
        if r[2] == 'Zone inventory': zones[c].append((str(r[3]).strip(), pv))
        else: citypv[c] = pv

    mh = json.load(open(P('data_competition.json')))['MH']['clinics']
    practo = json.load(open(P('data_practo_demand.json')))
    matched_zone = {(k.split('|')[0].lower(), m['zone'].lower()) for k, m in practo['meta'].items()}
    cmap = {'delhi ncr': 'delhi'}
    allo_cities = sorted({k.split('|')[0] for k in mh})
    nmatch = len(practo['byclinic'])

    rows = []
    for ac in allo_cities:
        pc = cmap.get(ac.lower(), ac.lower())
        zs = sorted(zones.get(pc, []), key=lambda x: -x[1])
        if not zs and pc not in citypv: continue
        cpv = citypv.get(pc)
        head = (f'<tr class="ch"><td>{ac}</td><td class="r">{fmt(cpv)}</td>'
                f'<td class="r">{len(zs)} zones</td><td></td></tr>')
        rows.append(head)
        for zn, pv in zs:
            hit = (pc, zn.lower()) in matched_zone
            mk = '<span class="tag">Allo clinic</span>' if hit else ''
            rows.append(f'<tr class="{ "hit" if hit else "" }"><td class="z">{zn}</td>'
                        f'<td class="r">{fmt(pv)}</td><td></td><td>{mk}</td></tr>')

    html = TMPL.replace('{{ROWS}}', '\n'.join(rows)) \
               .replace('{{NMATCH}}', str(nmatch)).replace('{{NTOT}}', str(len(mh)))
    open(P('mh-practo.html'), 'w').write(html)
    print(f'wrote mh-practo.html · {nmatch}/{len(mh)} clinics have a matching Practo zone')

def fmt(v):
    return '—' if v is None else f'{int(v):,}'

TMPL = """<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>Practo locality pageviews — reference</title>
<style>
:root{--ink:#1b2733;--mut:#748092;--line:#e6e9ec;--blue:#2c6cae;--hit:#eef5ff}
*{box-sizing:border-box}body{font:14px/1.5 -apple-system,Segoe UI,Roboto,sans-serif;color:var(--ink);background:#f4f6f8;margin:0;padding:26px}
.wrap{max-width:860px;margin:0 auto}h1{font-size:20px;margin:0 0 4px}
.note{background:#fff;border:1px solid var(--line);border-left:3px solid var(--blue);border-radius:6px;padding:12px 15px;margin:14px 0 18px;color:#40515f;font-size:12.5px}
.note b{color:var(--ink)}
a{color:var(--blue);font-weight:700;text-decoration:none}
table{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--line);border-radius:8px;overflow:hidden}
th,td{text-align:left;padding:6px 12px;border-bottom:1px solid var(--line);font-size:12.5px}
th{background:#fbfcfd;color:var(--mut);font-weight:700;text-transform:uppercase;font-size:10.5px;letter-spacing:.04em}
.r{text-align:right;font-variant-numeric:tabular-nums}
tr.ch td{background:#eef1f4;font-weight:800;font-size:13px;border-top:2px solid #dfe4e8}
tr.hit{background:var(--hit)}td.z{padding-left:22px;color:#40515f}
.tag{background:var(--blue);color:#fff;font-size:9.5px;font-weight:700;padding:1px 6px;border-radius:10px;letter-spacing:.03em}
</style></head><body><div class="wrap">
<a href="mh-launch.html">← Back to Scoring Explorer</a>
<h1>Practo — General Psychiatry locality pageviews</h1>
<div class="note"><b>Reference only — not used in the demand score.</b> These are Practo's own zone-level
pageviews for "General Psychiatry". They measure Practo's traffic footprint (not total market demand), and
only <b>{{NMATCH}} of {{NTOT}}</b> clinics map cleanly to a Practo zone — so the Scoring Explorer keeps
Demand at city-level search volume. Rows shaded blue with an <span class="tag">Allo clinic</span> tag are the
localities where we have a clinic; the rest are shown for context. City rows show Practo's city-level total.</div>
<table><thead><tr><th>City / zone</th><th class="r">Practo pageviews</th><th class="r">zones</th><th></th></tr></thead>
<tbody>
{{ROWS}}
</tbody></table>
<p style="color:var(--mut);font-size:11.5px;margin-top:14px">Source: Practo General-Psychiatry inventory (zone + city). Attached alongside the Google keyword search-query data.</p>
</div></body></html>"""

if __name__ == '__main__':
    main()
