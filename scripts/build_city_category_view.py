#!/usr/bin/env python3
"""Tab-2 style "city × category" view generated from data_campaign_compose.json.

Replicates Alisha's 'final analysis (top metrics)' sheet: metrics as ROWS, cities as
COLUMNS, category blocks (SH / STI / MH) stacked vertically, each cell a 6-week average.

Two sheets:
  'Excl Google-Web (5-6)' — funnel keeps 5 of 6 channel×mediums (drops Google×Web)
  'Incl Web (all 6)'      — funnel keeps all 6 channel×mediums

Filters (both sheets):
  ACQUISITION (Spend/IS/Impr/Clicks/LocClicks…) → category, match-type = Exact-Local
  FUNNEL      (Leads/Book%/Done%)               → category, channel×medium per sheet

Level metrics = mean over the 6 weeks; ratio metrics = ratio of the averaged components
(matches the compose page's blended Avg-6wk — verified to tie Tab 2's acquisition side exactly).

Usage:  python3 scripts/build_city_category_view.py            # all cities, SH/STI/MH, both sheets
Output: ~/Downloads/city_category_view.xlsx
"""
import os, json, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CUBE = json.load(open(os.path.join(ROOT, 'data_campaign_compose.json')))
WEEKS = CUBE['_meta']['weeks']; NW = len(WEEKS)

MATCH_TYPE = 'Exact-Local'
CATS = ['SH', 'STI', 'MH']
EXCLUDE_CHMED = {('Google', 'Web')}              # what "exclude web" drops

def wk_avg(rows, field, pred): return sum((r.get(field, 0) or 0) for r in rows if pred(r)) / NW
def r(n, d): return n / d if d else 0

def city_col(city, cat, exclude_web):
    c = CUBE.get(city)
    if not c: return None
    acqp = lambda x: x['cat'] == cat and x['mt'] == MATCH_TYPE
    funp = lambda x: x['cat'] == cat and (not exclude_web or (x['ch'], x['med']) not in EXCLUDE_CHMED)
    A = lambda f: wk_avg(c['acq'], f, acqp); F = lambda f: wk_avg(c['fun'], f, funp)
    spend, impr, elig, click, locclick = A('sp'), A('impr'), A('elig'), A('click'), A('locclick')
    lead, bk, dn = F('lead'), F('bk'), F('dn')
    # rank-lost impressions + cost-weighted Quality Score (acquisition, cat×Exact-Local)
    aw = [x for x in c['acq'] if acqp(x)]; spw = sum((x.get('sp', 0) or 0) for x in aw)
    rank_lost = sum((x.get('rlis', 0) or 0) for x in aw) / NW
    qs = (sum((x.get('qs', 0) or 0) * (x.get('sp', 0) or 0) for x in aw) / spw) if spw else 0
    return {
        'Spend': spend, 'IS (Imp. Share)': r(impr, elig), 'Impressions': impr,
        'Market size estd': elig, 'Clicks': click, 'Loc Clicks': locclick,
        'Loc %': r(locclick, click), 'Cost per click': r(spend, click),
        'Cost per loc click': r(spend, locclick), 'Leads': lead,
        'Click → Lead %': r(lead, click), 'Loc-Click → Lead %': r(lead, locclick),
        'Lead → Book %': r(bk, lead), 'Book → Done %': r(dn, bk),
        'CPL': r(spend, lead), 'CPB': r(spend, bk), 'CPD': r(spend, dn),
        'Rank-lost impr': rank_lost, 'Quality Score': qs,
    }

MONEY = '₹#,##0'; PCT = '0.0%'; NUM = '#,##0'
ROWS = [
    ('Spend', MONEY), ('IS (Imp. Share)', PCT), ('Impressions', NUM), ('Market size estd', NUM),
    ('Clicks', NUM), ('Loc Clicks', NUM), ('Loc %', PCT), ('Cost per click', MONEY),
    ('Cost per loc click', MONEY), ('Leads', NUM), ('Click → Lead %', PCT), ('Loc-Click → Lead %', PCT),
    ('Lead → Book %', PCT), ('Book → Done %', PCT), ('CPL', MONEY), ('CPB', MONEY), ('CPD', MONEY),
]
ACC = '2C6CAE'; GRN = '1F6F5C'; INK = '1A2230'; WHITE = 'FFFFFF'; ZEB = 'F4F7FB'; CATBG = 'DCE6F4'
fill = lambda h: PatternFill('solid', fgColor=h)
thin = Side(style='thin', color='E4E8EF')

# heat-map: metric -> direction. 'good' = higher is greener; 'cost' = lower is greener (cheaper).
COLOR_METRICS = {
    'IS (Imp. Share)': 'good', 'Market size estd': 'good', 'Loc Clicks': 'good', 'Loc %': 'good',
    'Cost per click': 'cost', 'Cost per loc click': 'cost', 'Loc-Click → Lead %': 'good',
    'Lead → Book %': 'good', 'CPL': 'cost',
}
# Key-metrics heat set — colour ONLY these six (Alisha's focus metrics)
KEY_HEAT = {
    'IS (Imp. Share)': 'good', 'Market size estd': 'good', 'Loc %': 'good',
    'Loc-Click → Lead %': 'good', 'Cost per loc click': 'cost', 'Lead → Book %': 'good',
}
# Red "focus here" border only on genuinely weak PERFORMANCE cells (benchmark-based, selective like Alisha's
# manual borders). Market size is context (not a problem to fix) → never bordered.
FOCUS_RULES = {
    'IS (Imp. Share)': lambda v: v < 0.45,          # severe rank loss (target 70%)
    'Loc %': lambda v: v < 0.45,                     # weak local pull (target 60%)
    'Loc-Click → Lead %': lambda v: v < 0.15,        # profile/credibility leak (target 30%)
    'Cost per loc click': lambda v: v > 55,          # expensive local click
    'Lead → Book %': lambda v: v < 0.60,             # IST/booking drop
}
# Green "winner" border — only cells that clearly BEAT the benchmark (selective, so it doesn't clutter).
GREEN_RULES = {
    'IS (Imp. Share)': lambda v: v >= 0.70,          # hit the 70% target
    'Loc %': lambda v: v >= 0.62,                     # beats the 60% target
    'Loc-Click → Lead %': lambda v: v >= 0.30,        # hit the 30% target
    'Cost per loc click': lambda v: v <= 22,          # very cheap local click
    'Lead → Book %': lambda v: v >= 0.85,             # excellent booking rate
}
# All-metrics sheets colour+highlight a wider set: + Loc Clicks + CPL, and border IS / Market size / CPL too.
ALL_HEAT = {**KEY_HEAT, 'Loc Clicks': 'good', 'CPL': 'cost'}
ALL_FOCUS = {**FOCUS_RULES, 'CPL': lambda v: v > 250, 'Loc Clicks': lambda v: v < 40}          # red = weak/expensive
ALL_GREEN = {**GREEN_RULES, 'CPL': lambda v: v <= 130, 'Loc Clicks': lambda v: v >= 300,
             'Market size estd': lambda v: v >= 8000}                                            # green = cheap / big-market opportunity
_STOPS = [(248, 200, 200), (255, 244, 200), (200, 230, 201)]  # light red -> light yellow -> light green (pastel, dark text stays readable)
def grad_hex(t):
    """t in [0,1] -> hex across red→yellow→green."""
    t = max(0.0, min(1.0, t))
    if t < 0.5: a, b, u = _STOPS[0], _STOPS[1], t / 0.5
    else: a, b, u = _STOPS[1], _STOPS[2], (t - 0.5) / 0.5
    return ''.join(f'{round(a[i] + (b[i] - a[i]) * u):02X}' for i in range(3))

def city_order():
    """Cities ranked by total Exact-Local spend across SH/STI/MH (biggest first)."""
    tot = {}
    for city in CUBE:
        if city == '_meta': continue
        tot[city] = sum((x.get('sp', 0) or 0) for x in CUBE[city].get('acq', []) if x['mt'] == MATCH_TYPE)
    return [c for c in sorted(tot, key=lambda c: -tot[c]) if tot[c] > 0]

# Reference annotations for Sheet 1 (Alisha's Tab-2 "leavers / meaning / benchmark" columns)
ANN_LEAVERS = {
    'IS (Imp. Share)': 'KW, Geography | Scope',
    'Impressions': '+ Competition, Bid / Budget',
    'Market size estd': 'basis our KWs, Geog | Scope',
    'Loc Clicks': '- Ranking (paid/free)\n- Competition (paid/free)\n- Relevance of Asset (Thumbnail)',
    'Loc-Click → Lead %': '- Relevance of Asset (Profile)\n- Credibility of asset',
    'Lead → Book %': 'IST efficiency - SL, WA flow, Language etc.',
    'Book → Done %': 'PRM efficiency',
}
ANN_MEANS = {
    'Loc Clicks': '- Brand, Clinic Name,\n- Headline / Copy,\n- Hygiene,\n- Rating/Reviews',
}
ANN_BENCH = {
    'IS (Imp. Share)': {'*': '70% for the right KWs', 'MH': '70%* for the right KWs'},
    'Loc %': {'*': 'tending to 60%'},
    'Loc-Click → Lead %': {'SH': '30%', 'STI': '50-60%', 'MH': ''},
}

def write_sheet(ws, cities, exclude_web, action=False, annotate=False,
                metric_rows=None, colfn=None, heatmap=None, mark_focus=False,
                focus_rules=None, green_rules=None):
    metric_rows = metric_rows or ROWS; colfn = colfn or city_col; heatmap = heatmap if heatmap is not None else COLOR_METRICS
    focus_rules = focus_rules or FOCUS_RULES; green_rules = green_rules or GREEN_RULES
    c0 = 5 if annotate else 2                 # first city column (leave B/C/D for annotations)
    ncol = c0 - 1 + len(cities)
    focus_cells = []; green_cells = []        # worst cells → red border; standout cells → green border
    ws.cell(1, 1, f'Metric · 6-wk avg ({WEEKS[0]} → {WEEKS[-1]})').font = Font(bold=True, color=WHITE, size=10)
    ws.cell(1, 1).fill = fill(ACC)
    if annotate:
        for cc, txt in ((2, 'What are the leavers'), (3, 'What it means'), (4, 'Benchmark / Target')):
            h = ws.cell(1, cc, txt); h.font = Font(bold=True, color=WHITE, size=9)
            h.fill = fill(GRN); h.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for j, city in enumerate(cities):
        h = ws.cell(1, c0 + j, city); h.font = Font(bold=True, color=WHITE); h.fill = fill(ACC)
        h.alignment = Alignment(horizontal='center')
    rr = 2
    for cat in CATS:
        cols = {city: colfn(city, cat, exclude_web) for city in cities}
        band = ws.cell(rr, 1, f'▼ {cat}'); band.font = Font(bold=True, color=ACC)
        for j in range(ncol): ws.cell(rr, 1 + j).fill = fill(CATBG)
        for j, city in enumerate(cities):   # repeat city names on the band row so each block is self-labelled
            cc = ws.cell(rr, c0 + j, city); cc.font = Font(bold=True, color=ACC, size=9)
            cc.alignment = Alignment(horizontal='center')
        ws.row_dimensions[rr].height = 15
        rr += 1
        for i, (metric, fmt) in enumerate(metric_rows):
            lab = ws.cell(rr, 1, metric); lab.font = Font(bold=True, color=INK, size=10.5)
            lab.alignment = Alignment(vertical='center')
            nlines = 1
            if annotate:
                lv = ANN_LEAVERS.get(metric, ''); mn = ANN_MEANS.get(metric, '')
                bm = ANN_BENCH.get(metric, {}).get(cat) or ANN_BENCH.get(metric, {}).get('*', '')
                nlines = max(nlines, lv.count('\n') + 1, mn.count('\n') + 1)
                for cc, txt in ((2, lv), (3, mn), (4, bm)):
                    a = ws.cell(rr, cc, txt or None); a.font = Font(color='4A5262', size=9)
                    a.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            heat = heatmap.get(metric)
            vals = [cols[c][metric] for c in cities if cols[c] and (cols[c][metric] or 0) > 0] if heat else []
            lo, hi = (min(vals), max(vals)) if vals else (0, 0)
            for j, city in enumerate(cities):
                data = cols[city]; v = data[metric] if data else None
                cell = ws.cell(rr, c0 + j, round(v, 4) if data else None)
                cell.number_format = fmt; cell.font = Font(size=10.5)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if heat and v and v > 0 and hi > lo:
                    t = (v - lo) / (hi - lo)
                    if heat == 'cost': t = 1 - t          # cheaper = greener
                    cell.fill = fill(grad_hex(t))
                    if mark_focus:
                        fr = focus_rules.get(metric); gr = green_rules.get(metric)
                        if fr and fr(v): focus_cells.append((rr, c0 + j))
                        elif gr and gr(v): green_cells.append((rr, c0 + j))
                elif i % 2 == 1:
                    cell.fill = fill(ZEB)
            if i % 2 == 1:
                for cc in range(1, c0): ws.cell(rr, cc).fill = fill(ZEB)
            ws.row_dimensions[rr].height = max(17, nlines * 13.5)
            rr += 1
        if action:
            from openpyxl.comments import Comment
            cpls = [cols[c]['CPL'] for c in cities if cols[c] and cols[c]['Leads'] >= 3]
            medcpl = _st.median(cpls) if cpls else 0
            al = ws.cell(rr, 1, '► ACTION'); al.font = Font(bold=True, color=WHITE); al.fill = fill(GRN)
            for j, city in enumerate(cities):
                m = cols[city]
                if not m: continue
                move, why = decide(cat, m, medcpl); code, bg, fc = short_move(move)
                cell = ws.cell(rr, c0 + j, code); cell.fill = fill(bg)
                cell.font = Font(bold=True, color=fc, size=9); cell.alignment = Alignment(horizontal='center')
                cm = Comment(f'{city} · {cat}\n{move}\n\n{why}', 'analysis'); cm.width = 260; cm.height = 110
                cell.comment = cm
            rr += 1
        rr += 1  # blank row between category blocks
    for row in ws.iter_rows(min_row=1, max_row=rr, max_col=ncol):
        for c in row: c.border = Border(bottom=thin)
    # red "focus here" border on the worst cells of the coloured metrics (opportunity / where to fix)
    red = Side(style='medium', color='C0392B')
    for (r_, c_) in focus_cells:
        ws.cell(r_, c_).border = Border(left=red, right=red, top=red, bottom=red)
    green = Side(style='medium', color='1E8449')
    for (r_, c_) in green_cells:
        ws.cell(r_, c_).border = Border(left=green, right=green, top=green, bottom=green)
    ws.column_dimensions['A'].width = 21
    if annotate:
        ws.column_dimensions['B'].width = 30; ws.column_dimensions['C'].width = 24; ws.column_dimensions['D'].width = 17
    for j in range(len(cities)): ws.column_dimensions[get_column_letter(c0 + j)].width = 12.5
    ws.freeze_panes = f'{get_column_letter(c0)}2'
    tag = 'EXCLUDES Google×Web (5 of 6 channel×mediums)' if exclude_web else 'INCLUDES all 6 channel×mediums (Google×Web added)'
    note = ws.cell(rr + 1, 1, f'Acquisition: match-type=Exact-Local · Funnel: {tag}. Coloured metrics: IS · Market size · Loc% · Loc-Click→Lead% · Cost/loc-click · Lead→Book%. '
           f'Red border = weakest cells (fix here) · Green border = beats benchmark (why we win). Level metrics = 6-wk mean; ratios = ratio of averaged components.')
    note.font = Font(italic=True, color='8A93A3', size=9)

# ---------- Action Plan sheet ----------
import statistics as _st
ACT_METRICS = [m for m, _ in ROWS] + ['Rank-lost impr', 'Quality Score']
ACT_FMT = {m: f for m, f in ROWS}; ACT_FMT['Rank-lost impr'] = NUM; ACT_FMT['Quality Score'] = '0.0'
MOVE_FILL = {  # substring -> (fill, font color)
    'Scale': ('C6EFCE', '0A5A28'), 'Cut': ('FFC7CE', '9C0006'),
    'Fix': ('FFEB9C', '7A5A00'), 'Hold': ('E7EAEE', '5A6472'),
}
L2L_TARGET = {'SH': 0.20, 'STI': 0.30, 'MH': 0.15}

def decide(cat, m, medcpl):
    sp = m['Spend']; IS = m['IS (Imp. Share)']; mkt = m['Market size estd']; ld = m['Leads']
    cpl = m['CPL']; locp = m['Loc %']; l2l = m['Loc-Click → Lead %']; l2b = m['Lead → Book %']
    qs = m['Quality Score']; rl = m['Rank-lost impr']
    if sp < 300: return ('Organic only — no paid', f'{ld:.0f} leads, negligible Exact-Local spend (₹{sp:,.0f}/wk)')
    minld = 8 if cat == 'SH' else 5
    scale = cpl <= medcpl * 1.1 and ld >= minld and IS < 0.65 and mkt > 1200
    weak_prof = locp < 0.45 or l2l < L2L_TARGET[cat]
    weak_ops = l2b < 0.65 and ld >= 4
    sat = IS >= 0.68
    waste = (cpl > medcpl * 1.8 and ld < 5) or (sp >= 800 and ld < 3)
    if cat == 'MH' and not (scale and qs >= 6):
        if waste: return ('Cut / pause', f'₹{sp:,.0f}/wk to {ld:.0f} leads, CPL ₹{cpl:,.0f}')
        return ('Fix QS + booking (do NOT bid up)', f'QS {qs:.1f}, Book {l2b*100:.0f}%, Loc→Ld {l2l*100:.0f}% — funnel broken')
    if waste: return ('Cut / pause', f'₹{sp:,.0f}/wk to {ld:.0f} leads, CPL ₹{cpl:,.0f}')
    if scale and weak_prof: return ('Scale bids + fix profile', f'CPL ₹{cpl:,.0f} at/below median, {rl:,.0f} rank-lost impr/wk; but Loc→Ld only {l2l*100:.0f}%')
    if scale: return ('Scale bids up', f'CPL ₹{cpl:,.0f} at/below median, IS {IS*100:.0f}%, {rl:,.0f} rank-lost impr/wk recoverable')
    if sat: return ('Hold (near ceiling)', f'IS {IS*100:.0f}% saturated, small market ({mkt:,.0f})')
    if weak_ops: return ('Fix booking flow', f'Lead→Book only {l2b*100:.0f}% — {ld:.0f} leads exist but drop')
    if weak_prof: return ('Fix GMB profile', f'Loc→Ld {l2l*100:.0f}% (target 30%), Loc% {locp*100:.0f}%')
    return ('Maintain', f'CPL ₹{cpl:,.0f}, IS {IS*100:.0f}% — healthy, limited headroom')

def short_move(move):
    """Full move string -> (short code, bg hex, font hex) for the vertical Action row."""
    if 'Scale bids + fix' in move: return ('SCALE↑ +prof', 'C6EFCE', '0A5A28')
    if 'Scale bids up' in move:     return ('SCALE ↑', 'C6EFCE', '0A5A28')
    if 'Cut' in move:               return ('CUT', 'FFC7CE', '9C0006')
    if 'QS' in move:                return ('FIX QS/funnel', 'FFEB9C', '7A5A00')
    if 'booking flow' in move:      return ('FIX booking', 'FFEB9C', '7A5A00')
    if 'GMB profile' in move:       return ('FIX profile', 'FFEB9C', '7A5A00')
    if 'Hold' in move:              return ('HOLD', 'E7EAEE', '5A6472')
    if 'Organic' in move:           return ('—', 'FFFFFF', 'A0A7B2')
    return ('keep', 'FFFFFF', '5A6472')

# ---------- Tab 3: complete funnel metric set, one row per city×category, sortable ----------
def city_col_full(city, cat, exclude_web):
    """Every funnel metric available (acquisition + funnel + economics), 6-wk averaged."""
    c = CUBE.get(city)
    if not c: return None
    acqp = lambda x: x['cat'] == cat and x['mt'] == MATCH_TYPE
    funp = lambda x: x['cat'] == cat and (not exclude_web or (x['ch'], x['med']) not in EXCLUDE_CHMED)
    A = lambda f: wk_avg(c['acq'], f, acqp); F = lambda f: wk_avg(c['fun'], f, funp)
    budget, impr, elig, locimpr = A('budget'), A('impr'), A('elig'), A('locimpr')
    click, locclick, spend = A('click'), A('locclick'), A('sp')
    rlis, blis = A('rlis'), A('blis')
    aw = [x for x in c['acq'] if acqp(x)]; spw = sum((x.get('sp', 0) or 0) for x in aw)
    bid = (sum((x.get('bid', 0) or 0) * (x.get('sp', 0) or 0) for x in aw) / spw) if spw else 0
    qs = (sum((x.get('qs', 0) or 0) * (x.get('sp', 0) or 0) for x in aw) / spw) if spw else 0
    lead, bk, dn, rev = F('lead'), F('bk'), F('dn'), F('rev')
    return {
        'Budget (₹/day)': budget, 'Bid (avg CPC)': bid, 'Spend': spend,
        'Budget Util %': r(spend, budget * 7), 'Impressions': impr, 'IS (Imp. Share)': r(impr, elig),
        'Market size estd': elig, 'Rank-lost impr': rlis, 'Budget-lost impr': blis,
        'Location Impressions': locimpr, 'Loc Imp %': r(locimpr, impr), 'Clicks': click,
        'CTR': r(click, impr), 'Loc Clicks': locclick, 'Loc %': r(locclick, click),
        'Cost per click': r(spend, click), 'Cost per loc click': r(spend, locclick), 'Quality Score': qs,
        'Leads': lead, 'Click → Lead %': r(lead, click), 'Loc-Click → Lead %': r(lead, locclick),
        'Booked': bk, 'Click → Book %': r(bk, click), 'Loc-Click → Book %': r(bk, locclick),
        'Lead → Book %': r(bk, lead), 'Done': dn, 'Book → Done %': r(dn, bk), 'Lead → Done %': r(dn, lead),
        'Revenue (est)': rev, 'ROAS': r(rev, spend), 'CPL': r(spend, lead), 'CPB': r(spend, bk),
        'CPD': r(spend, dn), 'RPC (rev÷done)': r(rev, dn),
    }
XX = '0.00"×"'
FULL_COLS = [  # (label, numfmt)
    ('Budget (₹/day)', MONEY), ('Bid (avg CPC)', MONEY), ('Spend', MONEY), ('Budget Util %', PCT),
    ('Impressions', NUM), ('IS (Imp. Share)', PCT), ('Market size estd', NUM), ('Rank-lost impr', NUM),
    ('Budget-lost impr', NUM), ('Location Impressions', NUM), ('Loc Imp %', PCT), ('Clicks', NUM),
    ('CTR', PCT), ('Loc Clicks', NUM), ('Loc %', PCT), ('Cost per click', MONEY),
    ('Cost per loc click', MONEY), ('Quality Score', '0.0'), ('Leads', NUM), ('Click → Lead %', PCT),
    ('Loc-Click → Lead %', PCT), ('Booked', NUM), ('Click → Book %', PCT), ('Loc-Click → Book %', PCT),
    ('Lead → Book %', PCT), ('Done', NUM), ('Book → Done %', PCT), ('Lead → Done %', PCT),
    ('Revenue (est)', MONEY), ('ROAS', XX), ('CPL', MONEY), ('CPB', MONEY), ('CPD', MONEY), ('RPC (rev÷done)', MONEY),
]
FULL_HEAT = dict(COLOR_METRICS); FULL_HEAT.update({'Quality Score': 'good', 'CTR': 'good', 'ROAS': 'good',
    'Loc Imp %': 'good', 'Revenue (est)': 'good', 'Click → Lead %': 'good', 'Book → Done %': 'good', 'CPB': 'cost', 'CPD': 'cost'})

def build_full_metrics_sheet(ws, cities):
    """Horizontal, SORTABLE: one row per city×category, every funnel metric as its own column.
    AutoFilter on the header → click any metric's ▾ to Sort Largest→Smallest or filter by category/action."""
    from openpyxl.comments import Comment
    cols = ['City', 'Category'] + [c for c, _ in FULL_COLS] + ['Action']
    ncol = len(cols)
    for j, h in enumerate(cols):
        c = ws.cell(1, 1 + j, h); c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = fill(ACC); c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 34
    # per-category heat ranges + median CPL
    data = {(cat, city): city_col_full(city, cat, True) for cat in CATS for city in cities}
    core = {(cat, city): city_col(city, cat, True) for cat in CATS for city in cities}
    ranges = {}; medcpl = {}
    for cat in CATS:
        act = [city for city in cities if data[(cat, city)] and (data[(cat, city)]['Spend'] > 0 or data[(cat, city)]['Leads'] > 0)]
        cpls = [data[(cat, c)]['CPL'] for c in act if data[(cat, c)]['Leads'] >= 3]
        medcpl[cat] = _st.median(cpls) if cpls else 0
        for metric in FULL_HEAT:
            vals = [data[(cat, c)][metric] for c in act if (data[(cat, c)][metric] or 0) > 0]
            ranges[(cat, metric)] = (min(vals), max(vals)) if vals else (0, 0)
    rr = 2
    for cat in CATS:
        for city in cities:
            m = data[(cat, city)]
            if not m or (m['Spend'] == 0 and m['Leads'] == 0): continue
            ws.cell(rr, 1, city).font = Font(bold=True, color=INK)
            cc = ws.cell(rr, 2, cat); cc.font = Font(bold=True, color=ACC); cc.alignment = Alignment(horizontal='center')
            for k, (metric, fmt) in enumerate(FULL_COLS):
                v = m[metric]; cell = ws.cell(rr, 3 + k, round(v, 4))
                cell.number_format = fmt; cell.alignment = Alignment(horizontal='right')
                heat = FULL_HEAT.get(metric)
                if heat and v and v > 0:
                    lo, hi = ranges[(cat, metric)]
                    if hi > lo:
                        t = (v - lo) / (hi - lo)
                        if heat == 'cost': t = 1 - t
                        cell.fill = fill(grad_hex(t))
            move, why = decide(cat, core[(cat, city)], medcpl[cat]); code, bg, fc = short_move(move)
            ac = ws.cell(rr, ncol, code); ac.fill = fill(bg); ac.font = Font(bold=True, color=fc, size=9)
            ac.alignment = Alignment(horizontal='center')
            ac.comment = Comment(f'{city} · {cat}\n{move}\n\n{why}', 'analysis'); ac.comment.width = 260; ac.comment.height = 110
            rr += 1
    for row in ws.iter_rows(min_row=1, max_row=rr - 1, max_col=ncol):
        for c in row: c.border = Border(bottom=thin)
    ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 8
    for k in range(len(FULL_COLS)): ws.column_dimensions[get_column_letter(3 + k)].width = 9.5
    ws.column_dimensions[get_column_letter(ncol)].width = 14
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{get_column_letter(ncol)}{rr - 1}'   # sort/filter by ANY column

def build():
    cities = city_order()
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = 'Key metrics · Excl web'
    write_sheet(ws1, cities, exclude_web=True, action=True, annotate=True, heatmap=KEY_HEAT, mark_focus=True)
    ws2 = wb.create_sheet('Key metrics · Incl web')
    write_sheet(ws2, cities, exclude_web=False, action=True, heatmap=KEY_HEAT, mark_focus=True)
    ws3 = wb.create_sheet('All metrics · Excl web')
    write_sheet(ws3, cities, exclude_web=True, action=True, metric_rows=FULL_COLS, colfn=city_col_full,
                heatmap=ALL_HEAT, mark_focus=True, focus_rules=ALL_FOCUS, green_rules=ALL_GREEN)
    ws4 = wb.create_sheet('All metrics · Incl web')
    write_sheet(ws4, cities, exclude_web=False, action=True, metric_rows=FULL_COLS, colfn=city_col_full,
                heatmap=ALL_HEAT, mark_focus=True, focus_rules=ALL_FOCUS, green_rules=ALL_GREEN)
    out = os.path.join(os.path.expanduser('~'), 'Downloads', 'city_category_view.xlsx')
    wb.save(out); return out, cities

if __name__ == '__main__':
    out, cities = build()
    print(f'wrote {out} · {len(cities)} cities · cats {"/".join(CATS)} · {NW} weeks · 4 vertical sheets '
          '(key/all × excl/incl web)')
