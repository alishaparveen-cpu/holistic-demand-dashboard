#!/usr/bin/env python3
"""Build data_practo_demand.json — locality (zone) level General-Psychiatry demand from Practo.

Input : practo_psychiatry.xlsx  (cols: City, Keyword, Inventory type, Zone, Pageviews)
          - 'Zone inventory' rows  = locality-level pageviews (the granular signal)
          - 'Cityinventory' rows   = one city-level pageview total per city
Output: data_practo_demand.json  { byclinic:{key:zone_pv}, meta:{key:{zone,zone_pv,city_pv}}, citypv:{city:pv} }

Matched to the clinic localities in data_competition.json['MH'] by exact/normalised name + a small
curated alias list. Practo pageviews measure Practo's OWN traffic footprint (e.g. Thane reads a spurious
55), so the explorer uses this only as a bounded, sqrt-damped ±40% within-city tilt on the city search
volume — never as a standalone demand number. See build_mh_explorer.py -> practo_tilt().
"""
import openpyxl, json, re, os

# repo root = parent of scripts/ (matches build_competition_cube.py); falls back to script dir
_here = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(_here) if os.path.basename(_here) == 'scripts' else _here
def P(f): return os.path.join(ROOT, f)

CMAP  = {'delhi ncr': 'delhi'}                       # Allo city label -> Practo city label
ALIAS = {'chinchwad': 'pimpri chinchwad', 'electronic city': 'electronics city',
         'new palasia': 'palasia', 'mvp colony': 'mvp', 'dilsukhnagar': 'dilsukh nagar'}
def norm(s): return re.sub(r'[^a-z0-9]+', '', str(s).lower())

def main():
    wb = openpyxl.load_workbook(P('practo_psychiatry.xlsx'), data_only=True)
    ws = wb[wb.sheetnames[0]]
    zbycity, citypv = {}, {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]: continue
        city = str(r[0]).strip().lower()
        pv = r[4] if isinstance(r[4], (int, float)) else 0
        if r[2] == 'Zone inventory':
            zbycity.setdefault(city, {})[str(r[3]).strip().lower()] = pv
        else:
            citypv[city] = pv

    mh = json.load(open(P('data_competition.json')))['MH']['clinics']
    byclinic, meta = {}, {}
    for k in mh:
        ac, al = k.split('|'); pc = CMAP.get(ac.lower(), ac.lower())
        z = zbycity.get(pc, {}); aln = al.lower().strip(); m = None
        if aln in z:                       m = aln
        elif ALIAS.get(aln) in z:          m = ALIAS[aln]
        else:
            na = norm(al)
            for zn in z:
                if norm(zn) == na: m = zn; break
        if m:
            byclinic[k] = z[m]
            meta[k] = {'zone': m, 'zone_pv': z[m], 'city_pv': citypv.get(pc)}

    out = {'byclinic': byclinic, 'meta': meta, 'citypv': citypv}
    json.dump(out, open(P('data_practo_demand.json'), 'w'), indent=0)
    print(f'wrote data_practo_demand.json · {len(byclinic)}/{len(mh)} clinics matched to a Practo zone')

if __name__ == '__main__':
    main()
