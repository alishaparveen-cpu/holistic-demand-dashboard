#!/usr/bin/env python3
"""Regenerate city_wise_final_funnel.xlsx (the "Campaign Compose" workbook) from
data_campaign_compose.json — now with ALL 6 weeks (incl Jul 13-19) + every city + Online.

Interactive sheets keep live formulas:
  Campaign Compose  — pick City (B1) + tick CATEGORY/MATCH-TYPE (→SPEND) and CHANNEL/MEDIUM/CATEGORY
                      (→LEADS) toggles; SUMIFS recompute via CmpData!Sel / SrcData!Sel helper cols.
  CmpData / SrcData — backing tables; Sel column is the tick-match flag the SUMIFS filter on.
Computed value sheets: City WoW (INDEX-driven), City Funnel (latest wk), City Averages, Total Funnel,
and the six category×match-type acquisition slices.
Run: python3 scripts/build_compose_xlsx.py   → ~/Downloads/city_wise_final_funnel.xlsx
"""
import os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CUBE=json.load(open(os.path.join(ROOT,'data_campaign_compose.json')))
WEEKS=CUBE['_meta']['weeks']                 # 6 weeks, oldest→newest
NW=len(WEEKS)
CITIES=sorted(c for c in CUBE if c!='_meta')
def dispcat(c): return 'STD' if c=='STI' else c   # campaign side shows STD; funnel keeps STI

# ---------- styling ----------
ACC='2C6CAE'; GRN='1F6F5C'; YEL='FFF2A8'; BLU='DCE6F4'; EFF='F3E7D0'; GRNF='DDF0E8'
WHITE='FFFFFF'; INK='1A2230'; MUT='8A93A3'
def fill(hex): return PatternFill('solid', fgColor=hex)
hdrF=Font(bold=True,color=WHITE,size=11); hdrFill=fill(ACC); grnFill=fill(GRN)
secFill=fill(BLU); secFont=Font(bold=True,color=ACC)
funSecFill=fill(GRNF); funSecFont=Font(bold=True,color=GRN)
boldF=Font(bold=True,color=INK); mutedF=Font(italic=True,color=MUT,size=9)
thin=Side(style='thin',color='E4E8EF'); bord=Border(bottom=thin)
rightA=Alignment(horizontal='right'); leftA=Alignment(horizontal='left')

# ---------- metric aggregation helpers (for value sheets) ----------
def acq_agg(rows):
    a=dict(budget=0,spend=0,impr=0,elig=0,locimpr=0,click=0,locclick=0,bidspend=0,bidwt=0)
    for r in rows:
        a['budget']+=r.get('budget',0) or 0; a['spend']+=r.get('sp',0) or 0
        a['impr']+=r.get('impr',0) or 0; a['elig']+=r.get('elig',0) or 0
        a['locimpr']+=r.get('locimpr',0) or 0; a['click']+=r.get('click',0) or 0
        a['locclick']+=r.get('locclick',0) or 0
        if r.get('bid'): a['bidspend']+=r['bid']*(r.get('sp',0) or 0); a['bidwt']+=r.get('sp',0) or 0
    return a
def fun_agg(rows):
    f=dict(leads=0,booked=0,done=0,rev=0,spend=0)
    for r in rows:
        f['leads']+=r.get('lead',0) or 0; f['booked']+=r.get('bk',0) or 0
        f['done']+=r.get('dn',0) or 0; f['rev']+=r.get('rev',0) or 0; f['spend']+=r.get('sp',0) or 0
    return f
def rat(n,d): return n/d if d else 0
# metric spec for VALUE sheets: (label, numfmt, fn(a,f))
MONEY='₹#,##0'; PCT='0.0%'; INT='#,##0'; X='0.00"×"'
ACQ_METRICS=[
 ('Budget (₹/day)',MONEY,lambda a,f:a['budget']),
 ('Bid (avg CPC ceiling)',MONEY,lambda a,f:rat(a['bidspend'],a['bidwt'])),
 ('Spend',MONEY,lambda a,f:a['spend']),
 ('Budget Util %',PCT,lambda a,f:rat(a['spend'],a['budget']*7)),
 ('Impressions',INT,lambda a,f:a['impr']),
 ('IS (Imp. Share)',PCT,lambda a,f:rat(a['impr'],a['elig'])),
 ('Market size estd',INT,lambda a,f:a['elig']),
 ('Location Impressions',INT,lambda a,f:a['locimpr']),
 ('Loc Imp %',PCT,lambda a,f:rat(a['locimpr'],a['impr'])),
 ('Clicks',INT,lambda a,f:a['click']),
 ('CTR',PCT,lambda a,f:rat(a['click'],a['impr'])),
 ('Loc Clicks',INT,lambda a,f:a['locclick']),
 ('Loc %',PCT,lambda a,f:rat(a['locclick'],a['click'])),
 ('Cost per click',MONEY,lambda a,f:rat(a['spend'],a['click'])),
 ('Cost per loc click',MONEY,lambda a,f:rat(a['spend'],a['locclick'])),
]
FUN_METRICS=[
 ('Leads',INT,lambda a,f:f['leads']),
 ('Click → Lead %',PCT,lambda a,f:rat(f['leads'],a['click'])),
 ('Loc-Click → Lead %',PCT,lambda a,f:rat(f['leads'],a['locclick'])),
 ('Booked',INT,lambda a,f:f['booked']),
 ('Click → Book %',PCT,lambda a,f:rat(f['booked'],a['click'])),
 ('Loc-Click → Book %',PCT,lambda a,f:rat(f['booked'],a['locclick'])),
 ('Lead → Book %',PCT,lambda a,f:rat(f['booked'],f['leads'])),
 ('Done',INT,lambda a,f:f['done']),
 ('Book → Done %',PCT,lambda a,f:rat(f['done'],f['booked'])),
 ('Lead → Done %',PCT,lambda a,f:rat(f['done'],f['leads'])),
 ('Revenue (est · done×cat-RPC)',MONEY,lambda a,f:f['rev']),
 ('Spend (campaigns)',MONEY,lambda a,f:a['spend']),
 ('ROAS',X,lambda a,f:rat(f['rev'],a['spend'])),
 ('CPL',MONEY,lambda a,f:rat(a['spend'],f['leads'])),
 ('CPB',MONEY,lambda a,f:rat(a['spend'],f['booked'])),
 ('CPD',MONEY,lambda a,f:rat(a['spend'],f['done'])),
 ('RPC (rev ÷ done)',MONEY,lambda a,f:rat(f['rev'],f['done'])),
]
ALL_METRICS=ACQ_METRICS+FUN_METRICS

def city_week_vals(city, wk, acq_filter=None, fun_filter=None):
    node=CUBE.get(city,{})
    arows=[r for r in node.get('acq',[]) if r['wk']==wk and (acq_filter is None or acq_filter(r))]
    frows=[r for r in node.get('fun',[]) if r['wk']==wk and (fun_filter is None or fun_filter(r))]
    a=acq_agg(arows); f=fun_agg(frows)
    return a,f

wb=openpyxl.Workbook()

# ==================== 1) CmpData ====================
cmp=wb.active; cmp.title='CmpData'
CMP_HDR=['City','Cat','MatchType','Grp','Wk','Spend','Impr','Elig','LocImpr','Click','LocClick',
         'Web','Rev','Book','Done','Sel','Budget','BidSpend','BidWt']
cmp.append(CMP_HDR)
cmp_rows=0
for city in CITIES:
    for r in CUBE[city].get('acq',[]):
        cat=dispcat(r['cat']); mt=r['mt']; sp=r.get('sp',0) or 0; bid=r.get('bid')
        cmp.append([city,cat,mt,f"{cat} · {mt}",r['wk'],sp,r.get('impr',0) or 0,
                    round(r.get('elig',0) or 0,1),r.get('locimpr',0) or 0,r.get('click',0) or 0,
                    r.get('locclick',0) or 0,0,0,0,0,None,r.get('budget',0) or 0,
                    (bid*sp if bid else 0),(sp if bid else 0)])
        cmp_rows+=1
# Sel formula (tick-match) — references Campaign Compose toggles (labels col J, values col K)
for i in range(2,cmp_rows+2):
    cmp.cell(i,16).value=(f"=IF(AND(INDEX('Campaign Compose'!$K$3:$K$5,MATCH(B{i},'Campaign Compose'!$J$3:$J$5,0)),"
                          f"INDEX('Campaign Compose'!$K$8:$K$10,MATCH(C{i},'Campaign Compose'!$J$8:$J$10,0))),1,0)")

# ==================== 2) SrcData ====================
src=wb.create_sheet('SrcData')
SRC_HDR=['City','Channel','Medium','Category','Wk','Leads','Booked','Done','Spend','Rev','Sel']
src.append(SRC_HDR)
src_rows=0
for city in CITIES:
    for r in CUBE[city].get('fun',[]):
        src.append([city,r['ch'],r['med'],r['cat'],r['wk'],r.get('lead',0) or 0,r.get('bk',0) or 0,
                    r.get('dn',0) or 0,r.get('sp',0) or 0,r.get('rev',0) or 0,None])
        src_rows+=1
for i in range(2,src_rows+2):
    src.cell(i,11).value=(f"=IF(AND(INDEX('Campaign Compose'!$K$13:$K$14,MATCH(B{i},'Campaign Compose'!$J$13:$J$14,0)),"
                          f"INDEX('Campaign Compose'!$K$17:$K$19,MATCH(C{i},'Campaign Compose'!$J$17:$J$19,0)),"
                          f"INDEX('Campaign Compose'!$K$22:$K$26,MATCH(D{i},'Campaign Compose'!$J$22:$J$26,0))),1,0)")

# ==================== 3) Campaign Compose (interactive) ====================
cc=wb.create_sheet('Campaign Compose',0)
WKCOLS=[get_column_letter(2+i) for i in range(NW)]   # B..G
AVGC=get_column_letter(2+NW)                          # H
GRAINC=get_column_letter(3+NW)                        # I
LABELC=get_column_letter(4+NW); LABELCN=4+NW          # J
VALC=get_column_letter(5+NW);   VALCN=5+NW            # K
def SIF(sheet,col,city='$B$1',selcol=None,selval=1,wk=None,extra=None):
    parts=[f"{sheet}!${col}:${col}",f"{sheet}!$A:$A",city]
    if selcol: parts+=[f"{sheet}!${selcol}:${selcol}",str(selval)]
    if extra:  parts+=[f"{sheet}!${extra[0]}:${extra[0]}",f'"{extra[1]}"']
    if wk:     parts+=[f"{sheet}!$E:$E",wk]
    return "SUMIFS("+",".join(parts)+")"
# formula generators per metric row: fn(wk_ref or None) -> excel formula string
def cS(col,wk): return SIF('CmpData',col,selcol='P',wk=wk)     # campaign, tick-filtered
def sS(col,wk): return SIF('SrcData',col,selcol='K',wk=wk)     # funnel, tick-filtered
def ratio(n,d): return f"=IFERROR(({n})/({d}),0)"
def money_int(x): return f"={x}"
# row spec: (label, fmt, formula_fn(wk), grain, rowclass)  rowclass in {'','b','sec','funsec','hi','eff','key'}
def rows_for(wk):
    return [
    ('▼ PAID ACQUISITION — Google Ads (campaign pickers)',None,None,'','sec'),
    ('Budget (₹/day)',MONEY,f"={cS('Q',wk)}",'↑ CAMPAIGN pickers · Σ daily budgets','b'),
    ('Bid (avg CPC ceiling)',MONEY,ratio(cS('R',wk),cS('S',wk)),'spend-wtd avg ceiling',''),
    ('Spend',MONEY,f"={cS('F',wk)}",'↑ CAMPAIGN pickers (Cat × Match-type)','b hi'),
    ('Budget Util %',PCT,ratio(cS('F',wk),f"{cS('Q',wk)}*7"),'spend ÷ (budget×7)','eff'),
    ('Impressions',INT,f"={cS('G',wk)}",'Campaign · Google Ads',''),
    ('IS (Imp. Share)',PCT,ratio(cS('G',wk),cS('H',wk)),'won ÷ eligible impr','key'),
    ('Market size estd',INT,f"={cS('H',wk)}",'eligible impr = Impr ÷ IS','key'),
    ('Location Impressions',INT,f"={cS('I',wk)}",'',''),
    ('Loc Imp %',PCT,ratio(cS('I',wk),cS('G',wk)),'loc impr ÷ impr',''),
    ('Clicks',INT,f"={cS('J',wk)}",'',''),
    ('CTR',PCT,ratio(cS('J',wk),cS('G',wk)),'clicks ÷ impr',''),
    ('Loc Clicks',INT,f"={cS('K',wk)}",'','key'),
    ('Loc %',PCT,ratio(cS('K',wk),cS('J',wk)),'loc clicks ÷ clicks',''),
    ('Cost per click',MONEY,ratio(cS('F',wk),cS('J',wk)),'spend ÷ clicks','eff key'),
    ('Cost per loc click',MONEY,ratio(cS('F',wk),cS('K',wk)),'spend ÷ loc clicks','eff key'),
    ('▼ FULL FUNNEL — LEADS by funnel pickers · spend by campaign pickers',None,None,'','funsec'),
    ('Leads',INT,f"={sS('F',wk)}",'City · LEADS (Channel×Medium×Cat)','b'),
    ('Click → Lead %',PCT,ratio(sS('F',wk),cS('J',wk)),'funnel leads ÷ campaign clicks',''),
    ('Loc-Click → Lead %',PCT,ratio(sS('F',wk),cS('K',wk)),'funnel leads ÷ campaign loc-clicks','key'),
    ('Booked',INT,f"={sS('G',wk)}",'City · funnel pickers','b key'),
    ('Click → Book %',PCT,ratio(sS('G',wk),cS('J',wk)),'funnel booked ÷ campaign clicks',''),
    ('Loc-Click → Book %',PCT,ratio(sS('G',wk),cS('K',wk)),'funnel booked ÷ campaign loc-clicks',''),
    ('Lead → Book %',PCT,ratio(sS('G',wk),sS('F',wk)),'booked ÷ leads','key'),
    ('Done',INT,f"={sS('H',wk)}",'City · funnel pickers','b'),
    ('Book → Done %',PCT,ratio(sS('H',wk),sS('G',wk)),'done ÷ booked','key'),
    ('Lead → Done %',PCT,ratio(sS('H',wk),sS('F',wk)),'done ÷ leads',''),
    ('Revenue (est · done×cat-RPC)',MONEY,f"={sS('J',wk)}",'funnel done × per-cat RPC','b'),
    ('Spend (campaigns)',MONEY,f"={cS('F',wk)}",'↑ CAMPAIGN pickers','b'),
    ('ROAS',X,ratio(sS('J',wk),cS('F',wk)),'funnel rev ÷ campaign spend','b hi'),
    ('CPL',MONEY,ratio(cS('F',wk),sS('F',wk)),'campaign spend ÷ funnel leads','eff key'),
    ('CPB',MONEY,ratio(cS('F',wk),sS('G',wk)),'campaign spend ÷ funnel booked','eff key'),
    ('CPD',MONEY,ratio(cS('F',wk),sS('H',wk)),'campaign spend ÷ funnel done','b eff key'),
    ('RPC (rev ÷ done)',MONEY,ratio(sS('J',wk),sS('H',wk)),'per-cat avg; varies by mix','key'),
    ]
CLSFILL={'hi':fill(YEL),'eff':fill(EFF),'key':fill('EEF3FB'),'sec':secFill,'funsec':funSecFill}
# header
cc['A1']='City ▸'; cc['B1']=CITIES[0] if 'Bangalore' not in CITIES else 'Bangalore'
cc['A1'].font=boldF; cc['B1'].font=Font(bold=True,color=ACC,size=12)
dv=DataValidation(type='list',formula1='"'+",".join(CITIES)+'"',allow_blank=False); cc.add_data_validation(dv); dv.add(cc['B1'])
cc.cell(3,1,'Metric').font=hdrF; cc.cell(3,1).fill=hdrFill
for i,w in enumerate(WEEKS):
    c=cc.cell(3,2+i,w); c.font=hdrF; c.fill=hdrFill; c.alignment=rightA
c=cc.cell(3,2+NW,'Avg / Σ'); c.font=hdrF; c.fill=grnFill; c.alignment=rightA
c=cc.cell(3,3+NW,'Level / grain'); c.font=hdrF; c.fill=hdrFill
# body
specs=rows_for(None)  # template (labels/fmt/grain/class only)
R=4
for (label,fmt,_fn,grain,cls) in specs:
    cc.cell(R,1,label)
    if cls in ('sec','funsec'):
        cell=cc.cell(R,1); cell.font=(funSecFont if cls=='funsec' else secFont)
        for cnum in range(1,4+NW): cc.cell(R,cnum).fill=CLSFILL[cls]
        R+=1; continue
    # per-week formulas
    for i,wcol in enumerate(WKCOLS):
        wkref=f"{wcol}$3"
        fml=rows_for(wkref)  # regenerate with this week ref
        # find same label
        f_for=next(x for x in fml if x[0]==label)[2]
        cell=cc.cell(R,2+i,f_for); cell.number_format=fmt; cell.alignment=rightA
    # Avg / Σ column: no week filter
    favg=next(x for x in rows_for(None) if x[0]==label)[2]
    cell=cc.cell(R,2+NW,favg); cell.number_format=fmt; cell.alignment=rightA; cell.fill=fill('EAF3EF')
    # grain
    g=cc.cell(R,3+NW,grain); g.font=mutedF
    # label styling / row-tint
    lblcell=cc.cell(R,1)
    if 'b' in cls.split(): lblcell.font=boldF
    tint=next((k for k in ('hi','eff','key') if k in cls.split()),None)
    if tint:
        for cnum in list(range(1,2+NW)):
            if cc.cell(R,cnum).fill.fgColor.rgb in (None,'00000000'): cc.cell(R,cnum).fill=CLSFILL[tint]
    R+=1

# ---- toggle side-panel (labels col J, values col K) ----
def toggle(row,label,val,header=False):
    c=cc.cell(row,LABELCN,label)
    if header: c.font=Font(bold=True,color=GRN,size=10)
    else:
        cc.cell(row,VALCN,val)
        cc.cell(row,VALCN).alignment=Alignment(horizontal='center')
cc.cell(2,LABELCN,'CAMPAIGN · CATEGORY → drives SPEND').font=Font(bold=True,color=ACC,size=10)
toggle(3,'SH',True); toggle(4,'STD',True); toggle(5,'MH',True)
toggle(7,'CAMPAIGN · MATCH TYPE → drives SPEND',None,header=True)
toggle(8,'Exact-Local',True); toggle(9,'Exact',True); toggle(10,'Phrase-Local',True)
toggle(12,'FUNNEL · CHANNEL → drives LEADS',None,header=True)
toggle(13,'GMB',True); toggle(14,'Google',True)
toggle(16,'FUNNEL · MEDIUM → drives LEADS',None,header=True)
toggle(17,'Call',True); toggle(18,'Web',True); toggle(19,'WhatsApp',True)
toggle(21,'FUNNEL · CATEGORY → drives LEADS',None,header=True)
toggle(22,'SH',True); toggle(23,'STI',True); toggle(24,'MH',True); toggle(25,'Other',True); toggle(26,'Uncategorized',True)
tv=DataValidation(type='list',formula1='"TRUE,FALSE"',allow_blank=False); cc.add_data_validation(tv)
for row in [3,4,5,8,9,10,13,14,17,18,19,22,23,24,25,26]: tv.add(cc.cell(row,VALCN))
cc.cell(28,LABELCN,'Tick TRUE/FALSE to compose. Categories: STI=STD. Waterfall: campaign-utm → call-audit → done-dx.').font=mutedF

# ---- splits (medium / category / channel leads; match-type spend) ----
def split_block(startR, title, rows):   # rows: list of (label, formula_fn(wkref)->str, avg_fn)
    cc.cell(startR,1,title).font=Font(bold=True,color=GRN,size=10)
    r=startR+1
    for (label,ff) in rows:
        cc.cell(r,1,'  '+label)
        for i,wcol in enumerate(WKCOLS):
            cell=cc.cell(r,2+i,ff(f"{wcol}$3")); cell.number_format=INT; cell.alignment=rightA
        cell=cc.cell(r,2+NW,ff(None)); cell.number_format=INT; cell.alignment=rightA; cell.fill=fill('EAF3EF')
        r+=1
    return r+1
def srcSif(col,val,wk):
    p=[f"SrcData!${'F'}:$F","SrcData!$A:$A","$B$1",f"SrcData!${col}:${col}",f'"{val}"']
    if wk: p+=["SrcData!$E:$E",wk]
    return "=SUMIFS("+",".join(p)+")"
def cmpSif(val,wk):
    p=["CmpData!$F:$F","CmpData!$A:$A","$B$1","CmpData!$C:$C",f'"{val}"']
    if wk: p+=["CmpData!$E:$E",wk]
    return "=SUMIFS("+",".join(p)+")"
r=split_block(30,'── Leads by medium ──',[(m,(lambda v,mm=m:(lambda wk:srcSif('C',mm,wk)))(m)) for m in ['Call','Web','WhatsApp']])
r=split_block(r,'── Leads by category ──',[(c,(lambda cc2:(lambda wk:srcSif('D',cc2,wk)))(c)) for c in ['SH','STI','MH','Other','Uncategorized']])
r=split_block(r,'── Leads by channel ──',[(c,(lambda cc2:(lambda wk:srcSif('B',cc2,wk)))(c)) for c in ['GMB','Google']])
r=split_block(r,'── Spend by match-type (campaign) ──',[(m,(lambda mm:(lambda wk:cmpSif(mm,wk)))(m)) for m in ['Exact-Local','Exact','Phrase-Local']])

# column widths
cc.column_dimensions['A'].width=30
for i in range(NW+1): cc.column_dimensions[get_column_letter(2+i)].width=12
cc.column_dimensions[GRAINC].width=30; cc.column_dimensions[LABELC].width=34; cc.column_dimensions[VALC].width=8
cc.freeze_panes='B4'

# ==================== 4) DataWoW (unfiltered per-city metric matrix) ====================
dw=wb.create_sheet('DataWoW')
dw.append(['City','Metric']+WEEKS+['Avg'])
for city in CITIES:
    for (label,fmt,fn) in ALL_METRICS:
        vals=[]
        for w in WEEKS:
            a,f=city_week_vals(city,w); vals.append(round(fn(a,f),4))
        avg=round(sum(vals)/NW,4)
        dw.append([city,label]+vals+[avg])

# ==================== 5) City WoW (INDEX-driven, own city selector) ====================
cw=wb.create_sheet('City WoW')
cw['A1']='City ▸'; cw['B1']=cc['B1'].value; cw['B1'].font=Font(bold=True,color=ACC,size=12)
dv2=DataValidation(type='list',formula1='"'+",".join(CITIES)+'"'); cw.add_data_validation(dv2); dv2.add(cw['B1'])
cw['A1'].font=boldF
hdr=['Metric']+WEEKS+['Avg']
for j,h in enumerate(hdr):
    c=cw.cell(2,1+j,h); c.font=hdrF; c.fill=(grnFill if h=='Avg' else hdrFill); c.alignment=(rightA if j else leftA)
nmet=len(ALL_METRICS)
for i,(label,fmt,fn) in enumerate(ALL_METRICS):
    r=3+i; cw.cell(r,1,label)
    for j in range(NW+1):
        cell=cw.cell(r,2+j); cell.number_format=fmt; cell.alignment=rightA
# make them array formulas (INDEX/MATCH on City AND Metric)
from openpyxl.worksheet.formula import ArrayFormula
for i in range(nmet):
    r=3+i
    for j in range(NW+1):
        col=get_column_letter(3+j)
        addr=cw.cell(r,2+j).coordinate
        cw[addr]=ArrayFormula(addr,f"=IFERROR(INDEX(DataWoW!${col}:${col},MATCH(1,(DataWoW!$A:$A=$B$1)*(DataWoW!$B:$B=$A{r}),0)),0)")
cw.column_dimensions['A'].width=28
for j in range(NW+1): cw.column_dimensions[get_column_letter(2+j)].width=12
cw.freeze_panes='B3'

# ==================== 6) City Funnel (latest week) & City Averages ====================
LATEST=WEEKS[-1]
def city_matrix(sheetname, header, valuefn):
    ws=wb.create_sheet(sheetname)
    ws.cell(1,1,header).font=hdrF; ws.cell(1,1).fill=hdrFill
    for j,city in enumerate(CITIES):
        c=ws.cell(1,2+j,city); c.font=hdrF; c.fill=hdrFill; c.alignment=rightA
    for i,(label,fmt,fn) in enumerate(ALL_METRICS):
        r=2+i; ws.cell(r,1,label)
        for j,city in enumerate(CITIES):
            v=valuefn(city,fn); cell=ws.cell(r,2+j,round(v,4)); cell.number_format=fmt; cell.alignment=rightA
    ws.column_dimensions['A'].width=26
    for j in range(len(CITIES)): ws.column_dimensions[get_column_letter(2+j)].width=12
    ws.freeze_panes='B2'
def latest_val(city,fn):
    a,f=city_week_vals(city,LATEST); return fn(a,f)
def avg_val(city,fn):
    tot=0
    for w in WEEKS:
        a,f=city_week_vals(city,w); tot+=fn(a,f)
    return tot/NW
city_matrix('City Funnel (latest)', f'Metric · {LATEST}', latest_val)
city_matrix('City Averages', 'Metric · 6-wk avg', avg_val)

# ==================== 7) Total Funnel (all cities, latest week) ====================
tf=wb.create_sheet('Total Funnel')
tf.cell(1,1,'Metric').font=hdrF; tf.cell(1,1).fill=hdrFill
tf.cell(1,2,f'ALL CITIES · {LATEST}').font=hdrF; tf.cell(1,2).fill=hdrFill; tf.cell(1,2).alignment=rightA
allA=acq_agg([r for city in CITIES for r in CUBE[city].get('acq',[]) if r['wk']==LATEST])
allF=fun_agg([r for city in CITIES for r in CUBE[city].get('fun',[]) if r['wk']==LATEST])
for i,(label,fmt,fn) in enumerate(ALL_METRICS):
    tf.cell(2+i,1,label); cell=tf.cell(2+i,2,round(fn(allA,allF),4)); cell.number_format=fmt; cell.alignment=rightA
tf.column_dimensions['A'].width=26; tf.column_dimensions['B'].width=18

# ==================== 8) category × match-type acquisition slices ====================
for cat_disp,cat_cube in [('SH','SH'),('STD','STI'),('MH','MH')]:
    for mt in ['Exact-Local','Exact']:
        ws=wb.create_sheet(f'{cat_disp} {mt}')
        ws.cell(1,1,f'{cat_disp} · {mt} only — per city, week-on-week (campaign feed). Avg: volumes=mean, costs/ratios=blended.').font=mutedF
        r=3
        af=lambda x,cc2=cat_cube,mm=mt: (x['cat']==cc2 and x['mt']==mm)
        for city in CITIES:
            # skip cities with no rows for this slice
            has=any(af(x) and x['wk'] in WEEKS for x in CUBE[city].get('acq',[]))
            if not has: continue
            ws.cell(r,1,f'{city} — {cat_disp} {mt}').font=boldF; r+=1
            hdr=['Metric']+WEEKS+['Avg']
            for j,h in enumerate(hdr):
                c=ws.cell(r,1+j,h); c.font=hdrF; c.fill=(grnFill if h=='Avg' else hdrFill); c.alignment=(rightA if j else leftA)
            r+=1
            for (label,fmt,fn) in ACQ_METRICS:
                ws.cell(r,1,label); vals=[]
                for w in WEEKS:
                    a,_=city_week_vals(city,w,acq_filter=af); v=fn(a,{}); vals.append(v)
                    cell=ws.cell(r,1+WEEKS.index(w)+1,round(v,4)); cell.number_format=fmt; cell.alignment=rightA
                avg=sum(vals)/NW; cell=ws.cell(r,2+NW,round(avg,4)); cell.number_format=fmt; cell.alignment=rightA
                r+=1
            r+=1
        ws.column_dimensions['A'].width=24
        for j in range(NW+1): ws.column_dimensions[get_column_letter(2+j)].width=12

# order sheets: Campaign Compose first, then summaries, then data
order=['Campaign Compose','Total Funnel','City Funnel (latest)','City Averages','City WoW',
       'SH Exact-Local','SH Exact','STD Exact-Local','STD Exact','MH Exact-Local','MH Exact',
       'DataWoW','CmpData','SrcData']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

OUT=os.path.expanduser('~/Downloads/city_wise_final_funnel.xlsx')
wb.save(OUT)
print(f"saved {OUT}")
print(f"  weeks: {WEEKS}")
print(f"  cities: {len(CITIES)} (incl {'Online' if 'Online' in CITIES else 'no Online'})")
print(f"  CmpData rows: {cmp_rows} · SrcData rows: {src_rows}")
