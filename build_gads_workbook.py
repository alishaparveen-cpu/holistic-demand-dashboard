#!/usr/bin/env python3
"""
build_gads_workbook.py — T1 Exact Local campaign table
Scope: SH (7) then STI (6) — each row = one campaign
Columns: metrics (RAG coloured) + Problem + Why + Fix Proposed
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(DIR, "data_ga_campaigns.json")) as f:
    GA = json.load(f)

latest_week = GA.get('_meta', {}).get('latest_week', '2026-06-08')

# ── Pull + filter ─────────────────────────────────────────────────────────

def vn(c, key):
    v = c.get(key)
    if isinstance(v, list): return v[0] if v else 0
    return v or 0

def city_from(name):
    parts = name.split('_')
    if parts[0] == 'T1' and 'Local' in name:
        return ' '.join(parts[1:-3])
    return ''

def cat_from(name):
    for p in name.split('_'):
        if p == 'SH': return 'SH'
        if p in ('STD', 'STI'): return 'STI'
    return ''

RAW = []
for c in GA.get('campaigns', []):
    n = c.get('n', '')
    if not n.startswith('T1'): continue
    if 'Exact' not in n or 'Local' not in n: continue
    cat = cat_from(n)
    if not cat: continue
    RAW.append({
        'name':   n,
        'cat':    cat,
        'city':   city_from(n),
        'sp':     vn(c, 'sp'),
        'bud':    c.get('bud', 0),
        'is_':    vn(c, 'is'),
        'bl':     vn(c, 'bl'),
        'rl':     vn(c, 'rl'),
        'util':   vn(c, 'util'),
        'impr':   vn(c, 'impr'),
        'loc':    vn(c, 'loc'),
        'locpct': vn(c, 'locpct'),
        'cpc':    vn(c, 'cpc'),
        'cplc':   vn(c, 'cplc'),
        'ar':     vn(c, 'ar'),
        'lp':     vn(c, 'lp'),
    })

SH  = sorted([c for c in RAW if c['cat'] == 'SH'],  key=lambda x: -x['sp'])
STI = sorted([c for c in RAW if c['cat'] == 'STI'], key=lambda x: -x['sp'])

# ── Per-campaign Why / Fix (hardcoded from analysis) ──────────────────────

PAID_COMP = {
    'Hyderabad':   'Mens Clinic Kothapet (103 SERP apps)',
    'Bangalore':   'Medi Life Health Care (132 apps)',
    'Pune':        'Ashakiran Clinic (89 apps)',
    'Mumbai':      "Dr Erande's Clinic (284 apps) + Dr SK Jain Burlington (106 apps)",
    'Chennai':     'Dr Kamaraj Hospital (300 apps)',
    'Navi Mumbai': 'Unison Medicare (243 apps)',
    'Thane':       'Competitor bidding on Thane exact queries',
}

# Keys match city+cat for lookup
WHY_FIX = {
    ('Hyderabad',   'SH'):  (
        "Losing 41% IS to rank. Mens Clinic Kothapet (103 SERP apps) outbids us in peak hours. "
        "Util=88% near-cap but bl=0 — budget isn't the bottleneck.",
        "+20% bid increase. Monitor IS at day 5. If no lift → competitor also raised → prioritise QS fix.",
    ),
    ('Bangalore',   'SH'):  (
        "49% IS lost to rank — highest in SH group. Ad Relevance=26% below-avg drags QS, so each bid "
        "point costs us more than competitors. Medi Life (132 apps) competes here.",
        "Step 1: Rewrite RSAs — exact keyword in Headline 1 (e.g. 'ED Treatment Bangalore'). "
        "Step 2: After AR improves, raise bids +15%. Fix QS first — bidding on poor QS burns budget.",
    ),
    ('Pune',        'SH'):  (
        "Losing 34% IS to rank. Ashakiran Clinic (89 apps) is the paid competitor. "
        "IS=66% is already the strongest in SH — gains will be incremental.",
        "+15% bid increase. Lowest urgency in SH group — IS at 66% already above average.",
    ),
    ('Mumbai',      'SH'):  (
        "Losing 37% IS to rank. Dr Erande's (284 apps) + Dr SK Jain Burlington (106 apps) are aggressive "
        "Mumbai bidders. Loc%=43% below 50% target — non-clinic-intent keywords diluting clicks. LP=42% below-avg.",
        "+15% bid increase. Negative-match informational queries (pull Search Terms report). "
        "Fix landing page H1 to match primary keyword to address LP score.",
    ),
    ('Chennai',     'SH'):  (
        "Losing 42% IS to rank. Dr Kamaraj Hospital (300 apps) is dominant — highest competitor pressure "
        "in SH group. LP=53% below-avg: landing page not matching keyword intent.",
        "Fix landing page first — H1 must mirror primary keyword, booking CTA above fold. "
        "Then raise bids +20%. Chennai has highest competitor pressure; QS fix reduces cost per rank point.",
    ),
    ('Navi Mumbai', 'SH'):  (
        "Losing 40% IS to rank. Unison Medicare (243 apps) competes in Navi Mumbai. "
        "Budget ~₹448/day is small relative to opportunity.",
        "+20% bid increase. If util rises above 85% after bid raise, increase budget ₹448→₹620/day.",
    ),
    ('Thane',       'SH'):  (
        "Losing 31% IS to rank — lowest rank loss in entire group, suggesting moderate competitor pressure. "
        "Loc%=45% borderline; watch for informational queries.",
        "+10% bid increase (conservative — IS already 69%). Monitor Loc% — if it drops below 40%, "
        "prune informational keywords from Search Terms report.",
    ),
    ('Bangalore',   'STI'): (
        "57% IS lost to rank — second-highest rank loss overall. Budget util=84%, occasionally capped. "
        "₹723/day is low for Bangalore's STI auction. DrSafeHands (447 apps) + Medi Life compete.",
        "Increase budget ₹723→₹1,000/day (+38%). Raise bids +20%. "
        "Both together should lift IS from 42% to ~60% within 2 weeks.",
    ),
    ('Hyderabad',   'STI'): (
        "IS=38% — second-lowest overall. Losing 60% to rank. Loc%=43% below target. "
        "DrSafeHands (1,043 SERP apps) dominates Hyderabad STI SERP and likely competes in auction.",
        "+25% bid increase. Negative-match informational STI queries (Loc%=43% signals info-seeker traffic). "
        "Consider budget increase ₹605→₹900/day.",
    ),
    ('Chennai',     'STI'): (
        "Util=92% — hitting budget cap most days. Loc%=31.8% is critical: only 1 in 3 clicks shows "
        "clinic intent. Running out of budget in peak hours means missing premium traffic. "
        "Dr Kamaraj (300 apps) runs ads all day while we stop mid-day.",
        "Increase budget ₹400→₹600/day (urgent). Aggressive informational keyword negatives — "
        "Loc%=31.8% means heavy info-seeker traffic. Then raise bids +20%.",
    ),
    ('Mumbai',      'STI'): (
        "IS=29% — LOWEST of all 13 campaigns. Losing 69% IS to rank. "
        "Dr Erande's (284 apps) + Dr SK Jain Burlington (106 apps) dominate Mumbai STI auction. "
        "Budget ₹284/day is critically under-resourced for Mumbai.",
        "Increase budget ₹284→₹500/day — severely under-budgeted for Mumbai. "
        "Raise bids +30%. Expected IS improvement: 29% → ~50%.",
    ),
    ('Pune',        'STI'): (
        "Losing 51% IS to rank. Ashakiran Clinic (89 apps) competes in Pune STI. "
        "Budget ₹225/day is among the smallest in the group. Loc%=44% borderline.",
        "+20% bid increase. Increase budget ₹225→₹350/day. "
        "Keyword audit if Loc% drops below 40%.",
    ),
    ('Navi Mumbai', 'STI'): (
        "bl=7% — directly losing IS to budget cap. CPC=₹55 is 3× the Mumbai STI campaign (₹15). "
        "Budget ₹140/day runs out quickly at ₹55 CPC. Either over-bidding or very narrow keyword set.",
        "Investigate high CPC first — check individual keyword bids in the auction. "
        "Then increase budget ₹140→₹250/day. IS→Budget% should fall to <2%.",
    ),
}

def problem_label(camp):
    issues = []
    if camp['bl'] > 2 and camp['util'] > 80: issues.append('Budget Capped')
    if camp['rl'] > 25:                       issues.append('Rank Loss')
    if camp['ar'] > 20:                       issues.append('Ad Relevance ↓')
    if camp['lp'] > 50:                       issues.append('Landing Page ↓')
    if camp['locpct'] < 40:                   issues.append('Low Loc%')
    elif camp['locpct'] < 45:                 issues.append('Loc% Borderline')
    return ' · '.join(issues) if issues else 'Monitor'

def problem_bg(camp):
    if camp['bl'] > 2 and camp['util'] > 80: return 'FED7AA'  # orange
    if camp['rl'] > 55:                       return 'FECACA'  # red
    if camp['ar'] > 20 or camp['lp'] > 50:   return 'EDE9FE'  # purple
    if camp['locpct'] < 40:                   return 'D1FAE5'  # teal
    return 'F1F5F9'

# ── Workbook ──────────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "T1 Exact Local"
ws.sheet_view.showGridLines = False

# Column definitions: (header, width)
COLS = [
    ("Cat",         5),
    ("City",        13),
    ("Campaign",    28),
    ("Spend\n₹/wk", 9),
    ("Budget\n₹/day",9),
    ("IS%",         7),
    ("IS Lost\nRank%", 8),
    ("IS Lost\nBudget%", 7),
    ("Util%",       7),
    ("Impr\n/wk",   8),
    ("Loc Clk\n/wk",8),
    ("Loc%",        7),
    ("CPC ₹",       7),
    ("CPLC ₹",      7),
    ("Problem",    22),
    ("Why",        42),
    ("Fix Proposed",42),
]
NC = len(COLS)

for i, (_, w) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Style helpers ─────────────────────────────────────────────────────────

def cs(ws, row, col, val, bg="FFFFFF", fc="111111", bold=False, size=9,
       align="left", wrap=False, italic=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def merge(ws, row, c1, c2, val, **kw):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    cs(ws, row, c1, val, **kw)

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

# RAG colour helpers
def is_col(v):   # IS%
    if v < 50: return 'FCA5A5','C42B2B'
    if v < 65: return 'FDE68A','A16207'
    return 'BBF7D0','15803D'

def rl_col(v):   # IS lost to rank
    if v > 50: return 'FCA5A5','C42B2B'
    if v > 30: return 'FDE68A','A16207'
    return 'BBF7D0','15803D'

def bl_col(v):   # IS lost to budget
    if v > 5:  return 'FCA5A5','C42B2B'
    if v > 2:  return 'FDE68A','A16207'
    return 'D1FAE5','15803D'

def util_col(v): # budget utilisation
    if v > 90: return 'FCA5A5','C42B2B'
    if v > 78: return 'FDE68A','A16207'
    return 'D1FAE5','15803D'

def loc_col(v):  # loc%
    if v < 40: return 'FCA5A5','C42B2B'
    if v < 50: return 'FDE68A','A16207'
    return 'BBF7D0','15803D'

# ── Title & header rows ───────────────────────────────────────────────────

r = 1
merge(ws, r, 1, NC,
      f"Google Ads — T1 Exact Local Campaigns  ·  SH + STI  ·  Week of {latest_week}",
      bg="0F172A", fc="FFFFFF", bold=True, size=13, align="center")
rh(ws, r, 24); r += 1

merge(ws, r, 1, NC,
      "Metrics coloured: 🟢 good  ·  🟡 watch  ·  🔴 act now  |  "
      "Loc-Click→Lead% and Lead→Book% require CRM join — not in Google Ads data  |  "
      "Spend ₹/wk = trailing 7-day actual",
      bg="1E293B", fc="94A3B8", size=8, italic=True, align="center")
rh(ws, r, 13); r += 1

def write_col_headers(ws, row, bg="1E293B"):
    for col, (h, _) in enumerate(COLS, 1):
        cs(ws, row, col, h, bg=bg, fc="FFFFFF", bold=True, size=8, align="center", wrap=True)
    rh(ws, row, 28)

# ── Write one campaign row ────────────────────────────────────────────────

def write_row(ws, row, camp, row_bg):
    city, cat = camp['city'], camp['cat']
    why, fix = WHY_FIX.get((city, cat), ("—", "—"))
    prob = problem_label(camp)
    pbg  = problem_bg(camp)

    is_b,  is_f  = is_col(camp['is_'])
    rl_b,  rl_f  = rl_col(camp['rl'])
    bl_b,  bl_f  = bl_col(camp['bl'])
    ut_b,  ut_f  = util_col(camp['util'])
    lc_b,  lc_f  = loc_col(camp['locpct'])

    cells = [
        (cat,                      row_bg,  "center", 9,  True,  "1A3A5C", False),
        (city,                     row_bg,  "left",   9,  False, "111111", False),
        (camp['name'],             row_bg,  "left",   8,  False, "374151", False),
        (f"₹{camp['sp']:,}",       row_bg,  "center", 9,  True,  "1A3A5C", False),
        (f"₹{camp['bud']:,}",      row_bg,  "center", 8,  False, "374151", False),
        (f"{camp['is_']}%",        is_b,    "center", 10, True,  is_f,    False),
        (f"{camp['rl']}%",         rl_b,    "center", 10, True,  rl_f,    False),
        (f"{camp['bl']}%",         bl_b,    "center", 9,  True,  bl_f,    False),
        (f"{camp['util']}%",       ut_b,    "center", 9,  True,  ut_f,    False),
        (f"{camp['impr']:,}",      row_bg,  "center", 9,  False, "374151", False),
        (f"{camp['loc']:,}",       row_bg,  "center", 9,  False, "374151", False),
        (f"{camp['locpct']:.0f}%", lc_b,    "center", 10, True,  lc_f,    False),
        (f"₹{camp['cpc']:.0f}",    row_bg,  "center", 9,  False, "374151", False),
        (f"₹{camp['cplc']:.0f}",   row_bg,  "center", 9,  False, "374151", False),
        (prob,                     pbg,     "left",   8,  True,  "111111", True),
        (why,                      row_bg,  "left",   8,  False, "1E3A5F", True),
        (fix,                      row_bg,  "left",   8,  False, "064E3B", True),
    ]
    for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(cells, 1):
        cs(ws, row, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)
    rh(ws, row, 52)

# ── SH section ───────────────────────────────────────────────────────────

merge(ws, r, 1, NC, "SH  —  Sexual Health  |  7 Exact Local T1 Campaigns",
      bg="1E3A5F", fc="FFFFFF", bold=True, size=10, align="left")
rh(ws, r, 18); r += 1

write_col_headers(ws, r); r += 1

for camp in SH:
    write_row(ws, r, camp, "EFF6FF")
    r += 1

# SH totals
sh_spend = sum(c['sp'] for c in SH)
sh_impr  = sum(c['impr'] for c in SH)
sh_loc   = sum(c['loc'] for c in SH)
sh_is    = sum(c['is_']*c['sp'] for c in SH) / sh_spend if sh_spend else 0
sh_rl    = sum(c['rl']*c['sp'] for c in SH) / sh_spend if sh_spend else 0
sh_lcp   = sum(c['locpct']*c['loc'] for c in SH) / sh_loc if sh_loc else 0

totals = [
    ("SH TOTAL","center",9,True,"FFFFFF"),("","","","",""),("","","","",""),
    (f"₹{sh_spend:,}","center",9,True,"FFFFFF"),("","","","",""),
    (f"{sh_is:.0f}%","center",9,True,"FFFFFF"),
    (f"{sh_rl:.0f}%","center",9,True,"FFFFFF"),
    ("","","","",""),("","","","",""),
    (f"{sh_impr:,}","center",9,False,"FFFFFF"),
    (f"{sh_loc:,}","center",9,False,"FFFFFF"),
    (f"{sh_lcp:.0f}%","center",9,False,"FFFFFF"),
    ("","","","",""),("","","","",""),
    ("","","","",""),("","","","",""),("","","","",""),
]
rh(ws, r, 15)
for col, t in enumerate(totals, 1):
    if t[0] == "": cs(ws, r, col, "", bg="2563EB", fc="FFFFFF", size=8)
    else:
        val, aln, sz, bd, fc = t
        cs(ws, r, col, val, bg="2563EB", fc=fc, bold=bd, size=sz, align=aln)
r += 1

# blank gap
merge(ws, r, 1, NC, "", bg="F8FAFC"); rh(ws, r, 8); r += 1

# ── STI section ───────────────────────────────────────────────────────────

merge(ws, r, 1, NC, "STI  —  Sexually Transmitted Infections  |  6 Exact Local T1 Campaigns",
      bg="7C2D12", fc="FFFFFF", bold=True, size=10, align="left")
rh(ws, r, 18); r += 1

write_col_headers(ws, r, bg="9A3412"); r += 1

for camp in STI:
    write_row(ws, r, camp, "FFF7ED")
    r += 1

# STI totals
sti_spend = sum(c['sp'] for c in STI)
sti_impr  = sum(c['impr'] for c in STI)
sti_loc   = sum(c['loc'] for c in STI)
sti_is    = sum(c['is_']*c['sp'] for c in STI) / sti_spend if sti_spend else 0
sti_rl    = sum(c['rl']*c['sp'] for c in STI) / sti_spend if sti_spend else 0
sti_lcp   = sum(c['locpct']*c['loc'] for c in STI) / sti_loc if sti_loc else 0

totals_sti = [
    ("STI TOTAL","center",9,True,"FFFFFF"),("","","","",""),("","","","",""),
    (f"₹{sti_spend:,}","center",9,True,"FFFFFF"),("","","","",""),
    (f"{sti_is:.0f}%","center",9,True,"FFFFFF"),
    (f"{sti_rl:.0f}%","center",9,True,"FFFFFF"),
    ("","","","",""),("","","","",""),
    (f"{sti_impr:,}","center",9,False,"FFFFFF"),
    (f"{sti_loc:,}","center",9,False,"FFFFFF"),
    (f"{sti_lcp:.0f}%","center",9,False,"FFFFFF"),
    ("","","","",""),("","","","",""),
    ("","","","",""),("","","","",""),("","","","",""),
]
rh(ws, r, 15)
for col, t in enumerate(totals_sti, 1):
    if t[0] == "": cs(ws, r, col, "", bg="C2410C", fc="FFFFFF", size=8)
    else:
        val, aln, sz, bd, fc = t
        cs(ws, r, col, val, bg="C2410C", fc=fc, bold=bd, size=sz, align=aln)
r += 1

# blank gap
merge(ws, r, 1, NC, "", bg="F8FAFC"); rh(ws, r, 8); r += 1

# ── Metric legend ─────────────────────────────────────────────────────────

merge(ws, r, 1, NC, "METRIC THRESHOLDS", bg="0F172A", fc="FFFFFF", bold=True, size=9, align="left")
rh(ws, r, 16); r += 1

legend_rows = [
    ("IS%",             "≥65% 🟢",  "55-65% 🟡",  "<55% 🔴",  "Target: >70% SH, >55% STI for T1 Exact Local"),
    ("IS Lost to Rank", "<25% 🟢",  "25-50% 🟡",  ">50% 🔴",  "Primary driver of IS gap — fix via bid increase + QS"),
    ("IS Lost to Bud",  "<2% 🟢",   "2-5% 🟡",    ">5% 🔴",   "Budget running out mid-day — fix via budget increase"),
    ("Budget Util%",    "<78% 🟢",  "78-90% 🟡",  ">90% 🔴",  "Util>90% = capping; Util<70% = bid/rank issue (budget not constraint)"),
    ("Loc%",            ">50% 🟢",  "40-50% 🟡",  "<40% 🔴",  "Loc Click = call or directions. Target 50-65% for clinic Exact Local"),
    ("Loc-Click→Lead%", "5-15%",    "benchmark",  "CRM join", "NOT in Google Ads. Need booking system data. Multiply by Loc Clicks for lead estimate"),
    ("Lead→Book%",      "40-60%",   "benchmark",  "CRM join", "NOT in Google Ads. Need screening call outcome data"),
    ("CPL (true)",      "—",        "CPLC ÷ LC→L%","CRM join","Approximate only. Google Ads CPA excluded (conversion event definition unverified)"),
]

for metric, g, y, red, note in legend_rows:
    row_bg = "F8FAFC"
    cs(ws, r, 1, metric,    bg="E2E8F0", fc="0F172A", bold=True, size=8, align="left")
    cs(ws, r, 2, g,         bg="BBF7D0", fc="15803D", bold=False, size=8, align="center")
    cs(ws, r, 3, y,         bg="FDE68A", fc="A16207", bold=False, size=8, align="center")
    cs(ws, r, 4, red,       bg="FCA5A5", fc="C42B2B", bold=False, size=8, align="center")
    ws.merge_cells(f"E{r}:{get_column_letter(NC)}{r}")
    cs(ws, r, 5, note,      bg=row_bg,   fc="374151", bold=False, size=8, italic=True)
    rh(ws, r, 14); r += 1

# freeze and filter
ws.freeze_panes = "A4"

# ── save ──────────────────────────────────────────────────────────────────

OUT = os.path.join(DIR, "allo_gads_fixes.xlsx")
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  {len(SH)} SH campaigns + {len(STI)} STI campaigns = {len(SH)+len(STI)} rows")
