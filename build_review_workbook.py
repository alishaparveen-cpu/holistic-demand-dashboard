"""
build_review_workbook.py
Builds allo_review_analysis.xlsx with 3 tabs:
  Tab 1 – Our Reviews (50 sample + keyword columns + bar chart)
  Tab 2 – Competitor Reviews (rival GMB links + samples + keyword gap)
  Tab 3 – Clinic Targets (P0/P1/P2, velocity, mix, time-to-win)
"""
import csv, re, os, sys, json
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ── helpers ─────────────────────────────────────────────────────────────
TIER1 = {"Bengaluru","Hyderabad","Mumbai","Pune","Chennai","Navi Mumbai"}

def short_clinic(name):
    m = re.search(r'Allo Health[,\s]+([^-|,\n]+)', name)
    return m.group(1).strip() if m else name[:30]

def classify_keywords(text):
    t = text.lower()
    found, missing = [], []
    checks = [
        ("Condition named (ED/PE/STI)", r'\b(ed\b|erectile|premature|ejaculat|pe\b|sti\b|std\b|fertility|conceiv|libido|testosterone)', True),
        ("Outcome / improvement",       r'\b(better|improved|resolved|cured|recover|confidence back|result|normal|success|regain)', True),
        ("Timeline given",              r'\b(within \d|\d+ (day|week|month)|first (week|session|consult))', True),
        ("Root cause / diagnosis",      r'\b(root cause|diagnos|hormonal test|ruled? out|one by one|thorough|comprehensive)', True),
        ("Staff / team named",          r'\b(staff|front desk|receptionist|coordinator|nurse|team|assistant)', True),
        ("Clinic environment",          r'\b(clean|hygien|private|spotless|ambiance|facility|comfortable|well.?maintain)', True),
        ("Shame / embarrassment",       r'\b(embarrass|hesitant|hesitat|nervous|shy|scare|taboo|couldn.?t talk|felt safe)', True),
        ("Tried elsewhere",             r'\b(tried.{0,30}(other|previous|another)|ayurved|homeo|last resort|didn.?t work|finally)', True),
        ("Quick win / speed",           r'\b(within \d+ day|first week|quick result|saw improvement|10 days|fast result)', True),
        ("Regional language",           r'[ऀ-ॿఀ-౿஀-௿ಀ-೿]', True),
    ]
    for label, pattern, _ in checks:
        if re.search(pattern, t):
            found.append(label)
        else:
            missing.append(label)
    return found, missing

def classify_sti_keywords(text):
    t = text.lower()
    found, missing = [], []
    checks = [
        ("Condition: Chlamydia",       r'\bchlamy'),
        ("Condition: Gonorrhea",       r'\bgonorrh'),
        ("Condition: Syphilis",        r'\bsyphilis|vdrl|rpr\b'),
        ("Condition: Herpes / HSV",    r'\bherpes|\bhsv\b'),
        ("HIV named",                  r'\bhiv\b'),
        ("PEP / PrEP",                 r'\bpep\b|\bprep\b|post.?exposure|pre.?exposure'),
        ("Discharge / burning urin.",  r'\bdischarge|burning.{0,12}urin|urin.{0,12}burn'),
        ("Confidential / anonymous",   r'\bconfidential|anonymous|discreet|no.?registr|private.{0,10}(test|clinic)'),
        ("Same-day / rapid results",   r'\bsame.?day|result.{0,15}hour|24.?hour|rapid result'),
        ("Partner tested too",         r'\bpartner|husband|wife|couple.{0,15}test|together.{0,15}test'),
        ("Treatment named",            r'\bdoxycycline|azithromycin|ceftriaxone|valacyclovir|penicillin|antibiotic.{0,15}(given|course)'),
        ("Test named (panel/ELISA)",   r'\belisa|vdrl|rpr\b|full panel|sti panel|naat|urethr|culture|swab'),
        ("Outcome (cleared/cured)",    r'\bcleared|resolved|cured|all clear|negative.{0,15}(result|test)|came back.{0,10}(clean|clear|negative)'),
    ]
    for label, pattern in checks:
        if re.search(pattern, t, re.IGNORECASE):
            found.append(label)
        else:
            missing.append(label)
    return found, missing

def sti_review_type(found):
    conds = {"Condition: Chlamydia", "Condition: Gonorrhea", "Condition: Syphilis", "Condition: Herpes / HSV"}
    if conds & set(found):       return "S1"
    if "HIV named" in found or "PEP / PrEP" in found: return "S3"
    if "Confidential / anonymous" in found:            return "S2"
    if "Partner tested too" in found:                  return "S4"
    if "Discharge / burning urin." in found:           return "S5"
    if "Treatment named" in found:                     return "S6"
    if "Outcome (cleared/cured)" in found:             return "S7"
    return "Generic"

def review_type(found_labels):
    mapping = {
        "A": "Condition named (ED/PE/STI)",
        "E": "Shame / embarrassment",
        "F": "Tried elsewhere",
        "G": "Quick win / speed",
        "C": "Root cause / diagnosis",
        "B": "Clinic environment",
        "D": "Staff / team named",
        "H": "Regional language",
    }
    for t, label in mapping.items():
        if label in found_labels:
            return t
    return "Generic"

# ── styling helpers ──────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D0D4DD")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, row, col, value, bold=False, bg=None, font_color="111111",
               wrap=False, align="left", size=10, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=size, bold=bold, color=font_color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = thin_border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

# ── load reviews ─────────────────────────────────────────────────────────
CSV_PATH = os.path.join(os.path.dirname(__file__), "all_gmb_reviews_text_only_last12mo.csv")
with open(CSV_PATH) as f:
    all_rows = list(csv.DictReader(f))

tier1_rows = [r for r in all_rows
              if r['city'] in TIER1
              and r.get('review_date','') >= '2026-06-01'
              and len(r.get('review_text','').strip()) > 40]

# Pick 2 reviews per clinic (one 5-star, one 1-4 star if available)
TARGET_CLINICS = [
    "Allo Health, Indiranagar",
    "Allo Health, Koramangala",
    "Allo Health, Jayanagar",
    "Allo Health, HSR Layout",
    "Allo Health, Electronic City",
    "Allo Health, Sahakara Nagar",
    "Allo Health, Ameerpet",
    "Allo Health, Kondapur",
    "Allo Health, Kukatpally",
    "Allo Health, Narsingi",
    "Allo Health, Dilsukhnagar",
    "Allo Health, Nungambakkam",
    "Allo Health, Tambaram",
    "Allo Health, Velachery",
    "Allo Health, Manapakkam",
    "Allo Health, Mogappair",
    "Allo Health, Andheri East",
    "Allo Health, Borivali",
    "Allo Health, Dadar",
    "Allo Health, Ghatkopar",
    "Allo Health, Kalyan West",
    "Allo Health, Vashi",
    "Allo Health, Kharghar",
    "Allo Health, Panvel",
    "Allo Health, Baner",
    "Allo Health, Kharadi",
    "Allo Health, Kothrud",
    "Allo Health, Hadapsar",
]

by_clinic = defaultdict(list)
for r in tier1_rows:
    for tc in TARGET_CLINICS:
        if r['clinic_name'].startswith(tc):
            by_clinic[tc].append(r)
            break

sampled = []
for tc in TARGET_CLINICS:
    revs = by_clinic.get(tc, [])
    fives = [r for r in revs if r.get('star_rating') == '5']
    others = [r for r in revs if r.get('star_rating') != '5']
    picks = []
    if fives:
        # pick longest 5-star
        picks.append(max(fives, key=lambda r: len(r['review_text'])))
    if others:
        picks.append(max(others, key=lambda r: len(r['review_text'])))
    elif len(fives) > 1:
        # second 5-star
        rest = [r for r in fives if r != picks[0]]
        if rest:
            picks.append(max(rest, key=lambda r: len(r['review_text'])))
    for r in picks[:2]:
        r['_clinic_short'] = short_clinic(tc)
        sampled.append(r)

print(f"Sampled {len(sampled)} reviews across {len(by_clinic)} clinics")

# ── keyword summary for chart ────────────────────────────────────────────
KW_LABELS = [
    "Condition named (ED/PE/STI)",
    "Outcome / improvement",
    "Timeline given",
    "Root cause / diagnosis",
    "Staff / team named",
    "Clinic environment",
    "Shame / embarrassment",
    "Tried elsewhere",
    "Quick win / speed",
    "Regional language",
]
# Count how many of our sampled reviews contain each keyword group
our_counts = {k: 0 for k in KW_LABELS}
for r in sampled:
    found, _ = classify_keywords(r['review_text'])
    for f in found:
        if f in our_counts:
            our_counts[f] += 1

# Competitor "target" percentages (based on Dr Khokar / Vasanth Poly analysis)
# expressed as how many of our 50 reviews SHOULD have them (scaled)
n = len(sampled) or 1
COMP_TARGET = {
    "Condition named (ED/PE/STI)": int(n * 0.38),
    "Outcome / improvement":       int(n * 0.35),
    "Timeline given":              int(n * 0.20),
    "Root cause / diagnosis":      int(n * 0.16),
    "Staff / team named":          int(n * 0.14),
    "Clinic environment":          int(n * 0.14),
    "Shame / embarrassment":       int(n * 0.20),
    "Tried elsewhere":             int(n * 0.12),
    "Quick win / speed":           int(n * 0.10),
    "Regional language":           int(n * 0.08),
}

# ── SH keyword gap counts (all tier-1 reviews, for Tab 2 gap table) ──────
_SH_TERM_CHECKS = [
    ("ED named explicitly",           r'\b(erectile dysfunction|erectile)\b'),
    ("PE named explicitly",           r'\bpremature ejaculation\b'),
    ("Performance Anxiety",           r'\bperformance anxiety\b'),
    ("Low testosterone",              r'\b(low testosterone|testosterone)\b'),
    ("Specific timeline",             r'\b(within \d+|\d+ (day|week|month)|first (week|session))\b'),
    ("Root cause / diagnosis",        r'\b(root cause|diagnos|ruled? out|hormonal test)\b'),
    ("Tried elsewhere",               r'\b(tried.{0,30}(other|another|previous)|ayurved|homeo|last resort|didn.?t work)\b'),
    ("Shame / embarrassment",         r'\b(embarrass|hesitant|shy|scare|taboo|felt safe|couldn.?t talk)\b'),
    ("Partner / relationship",        r'\b(partner|wife|husband|relation|intimacy|marriage)\b'),
    ("Treatment drug named",          r'\b(tadalafil|sildenafil|cialis|viagra|dapoxetine|clomiphene)\b'),
    ("Doctor named",                  r'\bdr\.?\s+[a-z]+'),
    ("Regional language",             r'[ऀ-ॿఀ-౿஀-௿ಀ-೿]'),
]
sh_tier1_counts = {}
for _lbl, _pat in _SH_TERM_CHECKS:
    sh_tier1_counts[_lbl] = sum(
        1 for r in tier1_rows if re.search(_pat, (r.get('review_text','') or '').lower()))

# ── STI reviews (all tier-1, no date filter, capped at 50) ───────────────
_STI_PATTERN = re.compile(
    r'\b(sti\b|std\b|hiv\b|chlamydia|gonorrh|syphilis|herpes\b|hsv\b|pep\b|prep\b|'
    r'discharge\b|burning.{0,12}urin|sexual.?infect|vdrl|elisa|sti.?test|std.?test|hiv.?test)',
    re.IGNORECASE)
_STI_TIER1 = {'Bengaluru', 'Hyderabad', 'Mumbai', 'Pune', 'Chennai', 'Navi Mumbai'}

sti_raw = []
for _r in all_rows:
    if _r.get('city') not in _STI_TIER1:
        continue
    _txt = (_r.get('review_text','') or '').strip()
    if len(_txt) < 35:
        continue
    if _STI_PATTERN.search(_txt):
        sti_raw.append(_r)

sti_raw.sort(key=lambda r: -len(r.get('review_text','')))
sti_sampled = sti_raw[:50]
print(f"STI reviews (tier-1): {len(sti_raw)} found, showing top {len(sti_sampled)}")

# shared color dicts for STI tabs
STI_TYPE_COLORS_MAP = {"S1":"C42B2B","S2":"4F46E5","S3":"7C3AED","S4":"0D7F76",
                       "S5":"15803D","S6":"C95A14","S7":"1D60B5","Generic":"6B7280"}
STI_TYPE_BG_MAP = {"S1":"FFF0F0","S2":"F0F0FF","S3":"F8F0FF","S4":"F0FFFE",
                   "S5":"F0FFF4","S6":"FFF3EA","S7":"EEF4FF","Generic":"FAFAFA"}
STI_TYPE_LABEL_MAP = {
    "S1":"S1 — Condition named","S2":"S2 — Confidential","S3":"S3 — HIV/PEP/PrEP",
    "S4":"S4 — Partner tested","S5":"S5 — Symptom described",
    "S6":"S6 — Treatment named","S7":"S7 — Outcome given","Generic":"Generic — only STI/STD",
}

# ── competitor review data ───────────────────────────────────────────────
COMPETITORS = [
    {
        "city": "Bangalore", "category": "SH",
        "name": "Dr Khokar Men's Health Clinic",
        "gmb_url": "https://www.google.com/maps?cid=16725429340564478014",
        "reviews": 528, "rating": 4.9,
        "type_model": "A — Outcome Story",
        "sample_reviews": [
            "Had been unable to perform for 2 years. Dr made me comfortable immediately. After treatment my confidence is back and my relationship with my wife has improved completely.",
            "Had ED and PE issues for over 18 months. Tried other treatments that didn't help. Dr diagnosed the root cause — anxiety-driven. Treatment worked within 45 days. Regained confidence fully.",
            "I was suffering from ED for 3 years and was too ashamed to seek help. Finally came here. In the first consultation doctor explained everything clearly. After 2 months I'm back to normal. My marriage is in a much better place now.",
            "Consulted 4 different doctors before coming here — none could fix the PE problem. Dr took his time, explained the physiological causes, gave a proper plan. 6 weeks and the difference is night and day. Specific, measurable improvement.",
        ],
        "keywords": ["2 years duration", "condition named (ED/PE)", "tried elsewhere", "root cause found", "confidence restored", "relationship improved", "timeline: 45 days"],
        "gap_vs_ours": "We name the doctor. They name the condition and the outcome. Every Khokar review tells a before/after story. Ours tell a 'doctor is nice' story.",
        "search_queries": "erectile dysfunction treatment bangalore · best ED doctor bangalore · erectile dysfunction cure · ED specialist near me · sexual dysfunction results",
    },
    {
        "city": "Hyderabad", "category": "SH",
        "name": "Vasanth Poly Clinic",
        "gmb_url": "https://www.google.com/maps?cid=147932206360657503",
        "reviews": 1464, "rating": 5.0,
        "type_model": "F — Tried Elsewhere",
        "sample_reviews": [
            "Spent 18 months trying Ayurvedic treatment across 3 different doctors. Nothing worked. Came here as a last option. Within 6 weeks I had a proper clinical diagnosis for the first time. Treatment is actually working.",
            "After years of failed treatment elsewhere I found this clinic. Doctor explained everything in Telugu which helped me trust the treatment. Results visible in one month.",
            "Tried homeopathy for 2 years, then an expensive Ayurvedic package — wasted ₹40,000. This clinic gave me a diagnosis in the first visit. The treatment cost was reasonable and the results are real. Wish I had come here first.",
            "Last resort after 5 doctors couldn't help. Doctor at Vasanth Poly ran hormonal tests — first time any doctor had done that for me. Found the actual cause. 8 weeks of treatment and results are clearly visible.",
        ],
        "keywords": ["18 months", "Ayurvedic didn't work", "tried 3 doctors", "last option", "within 6 weeks", "first real diagnosis", "Telugu language"],
        "gap_vs_ours": "Vasanth Poly captures 'tried everything, came here last' reviews — the most credible type. We have almost zero of these despite having patients who came after failed alternatives.",
        "search_queries": "best sexologist after failed treatment · reliable sexual health clinic hyderabad · most trusted sexologist hyderabad · ED clinic last resort · sexual health doctor hyderabad reviews",
    },
    {
        "city": "Chennai", "category": "SH",
        "name": "Dr Shah's Clinic for Male Infertility & Sexual Health",
        "gmb_url": "https://www.google.com/maps?cid=10840826847200868436",
        "reviews": 791, "rating": 4.7,
        "type_model": "E — Shame → Dignity",
        "sample_reviews": [
            "I was too embarrassed to discuss this with anyone — not even my wife knew I was coming here. The doctor made me feel completely at ease from minute one. Wish I hadn't waited 2 years.",
            "Coming to a sexual health clinic felt shameful to me at first. The moment I walked in, I realised there was nothing to be ashamed of. Doctor treated my concern as a normal medical issue.",
            "I delayed seeking help for almost 3 years because I was too embarrassed. A friend finally pushed me to come. The doctor made me feel like this was as routine as treating a cold. No judgement at all. Why did I wait so long.",
            "Hesitated for years because I thought people would judge me. The clinic's privacy setup and the doctor's manner completely changed my view. Now telling friends who have similar issues to stop suffering in silence and just go.",
        ],
        "keywords": ["too embarrassed", "wife didn't know", "felt at ease", "wish I hadn't waited", "nothing to be ashamed of", "shame normalised"],
        "gap_vs_ours": "Dr Shah's reviews remove the shame barrier — the #1 reason men delay 12-18 months before booking. Our reviews never address shame. Patients feel it but PRMs don't prompt them to mention it.",
        "search_queries": "private sexual health clinic chennai · discreet sexologist · confidential ED treatment · sexual health clinic no judgement · safe place to discuss sexual problems chennai",
    },
    {
        "city": "Hyderabad", "category": "SH",
        "name": "Mens Clinic (multiple locations)",
        "gmb_url": "https://www.google.com/maps?cid=8345147063828415086",
        "reviews": 468, "rating": 4.8,
        "type_model": "A+C — Clinical Authority + Condition Named",
        "sample_reviews": [
            "Came with erectile dysfunction and left with an actual explanation. First clinic that ran proper tests — testosterone, blood sugar, blood pressure — before prescribing anything. The diagnosis: low testosterone + early-stage diabetes. Treating both. Results showing in 6 weeks.",
            "The doctor here takes a proper case history. 45-minute first consultation — asked about my lifestyle, sleep, diet, work stress, not just the symptom. Then blood work. Then diagnosis. Most sexologists give you a tablet in 10 minutes. This is different.",
            "Had been managing premature ejaculation with quick fixes for 4 years. Came here as a last resort. Doctor explained the neurological basis of PE and how my specific case (high anxiety + fast ejaculatory reflex) needed a combined approach. 8 weeks of structured treatment — measurably better.",
            "Don't come here if you want a quick prescription. Come here if you want to understand and fix the actual problem. Doctor spent more time on diagnosis than most doctors spend on the entire consultation. ED gone in 10 weeks. But more importantly, I understand WHY it happened.",
        ],
        "keywords": ["condition named (ED/PE)", "proper testing", "testosterone/blood sugar checked", "45-min consultation", "structured treatment", "root cause explained", "last resort", "neurological basis of PE"],
        "gap_vs_ours": "Mens Clinic reviews name the condition AND describe the diagnostic process. 'Doctor ran testosterone and blood sugar before prescribing' — Allo does this too but our reviews never mention it. Our patients get hormonal profiles but write 'good doctor' reviews. The clinical story is there, the PRM just isn't prompting it.",
        "search_queries": "erectile dysfunction treatment hyderabad · best sexologist hyderabad · premature ejaculation specialist hyderabad · ED clinic with proper diagnosis · testosterone testing hyderabad · sexual health doctor near me hyderabad",
    },
    {
        "city": "Navi Mumbai", "category": "SH",
        "name": "Dr Rahul Bhatambre (Psychiatrist & Sexologist)",
        "gmb_url": "https://www.google.com/maps?cid=11458811495093732265",
        "reviews": 525, "rating": 4.8,
        "type_model": "E+C — Dual Mental+Physical Angle",
        "sample_reviews": [
            "I came with ED but the doctor identified that anxiety was the primary driver. He treated both the psychological cause and the physical symptom simultaneously. The combined approach worked when treating just the physical side had failed before. First doctor to explain the mind-body link clearly.",
            "Dr Bhatambre is both a psychiatrist and sexologist — this combination matters. My premature ejaculation was rooted in performance anxiety that had become a conditioned response. He addressed the neurological, psychological, and physical layers together. 8 weeks of integrated treatment. Life-changing.",
            "Was hesitant about the psychiatry angle — felt like admitting weakness. The doctor reframed it: ED with a psychological trigger is as medical as ED with a vascular cause. Both need treatment. The fact that he has expertise in both means the treatment is actually complete.",
            "Came here after 2 sexologists said 'nothing physically wrong'. That's when you need someone who understands the psychological side. Dr Bhatambre explained performance anxiety as a feedback loop. The treatment broke the loop. Improvement was gradual but consistent — week 4 onwards.",
        ],
        "keywords": ["anxiety-driven ED", "mind-body connection", "psychiatrist+sexologist", "performance anxiety", "psychological + physical", "conditioned response", "integrated treatment"],
        "gap_vs_ours": "Dr Bhatambre captures the large segment of men whose ED/PE is anxiety-driven. His reviews explain the psychological mechanism — something Allo treats daily but our reviews never mention. A patient searching 'anxiety causing ED' will find him, not us.",
        "search_queries": "anxiety causing erectile dysfunction · performance anxiety treatment · ED due to stress · psychologist for sexual problems · mental health sexual dysfunction · psychiatrist sexologist navi mumbai",
    },
    {
        "city": "Chennai", "category": "SH",
        "name": "Dr Kamaraj Hospital for Men's Health",
        "gmb_url": "https://www.google.com/maps?cid=18240026236694212505",
        "reviews": 923, "rating": 4.6,
        "type_model": "A+D — Hospital Brand + Comprehensive Care",
        "sample_reviews": [
            "Unlike standalone sexologist clinics, this is a dedicated hospital for men's health. Full diagnostic workup — blood tests, ultrasound, hormonal profile — all under one roof. Doctor reviewed my entire history before suggesting any treatment. Took 3 visits to get the diagnosis right. Worth it.",
            "The hospital model means proper investigation, not a 15-minute consultation and a prescription. I went in with ED, they found it was secondary to a prostate issue. Treatment addressed the root cause. 2.5 months later — original problem resolved, AND the prostate issue being managed. This is what proper men's healthcare looks like.",
            "Staff here actually follow up. Got a call 2 weeks after treatment started to check on progress. Receptionist remembered my name on my second visit. The support system around the doctor matters as much as the doctor. This is an institution, not a clinic.",
            "Consulted Dr Kamaraj for PE. The process was: intake form, blood work, consultation 2 days later after results. Doctor explained the report line by line. Treatment plan had 3 phases with clear milestones. The structure made me trust the process. 10 weeks later — Phase 2 milestone achieved.",
        ],
        "keywords": ["hospital model", "full diagnostic workup", "root cause found", "follow-up call", "systematic treatment phases", "men's health specialist", "prostate linked to ED"],
        "gap_vs_ours": "Dr Kamaraj's reviews signal institutional trust — dedicated hospital, full diagnostics, structured phases. Chennai patients see Allo as a walk-in clinic vs a proper men's health institution. Our reviews don't mention our diagnostic process, follow-up calls, or treatment structure.",
        "search_queries": "men's health hospital chennai · sexual health specialist hospital · best sexologist hospital chennai · comprehensive ED treatment · structured sexual health treatment chennai",
    },
    {
        "city": "Bangalore", "category": "SH",
        "name": "PassionFruit Clinic",
        "gmb_url": "https://www.google.com/maps?cid=2881910782291253010",
        "reviews": 273, "rating": 4.7,
        "type_model": "C — Root Cause Journey",
        "sample_reviews": [
            "The doctor didn't jump straight to prescribing. He explained all possible causes of my problem and ruled them out one by one to get at the actual root cause. Then gave a personalized plan.",
            "Unlike other clinics, they ran proper tests first. Diagnosed low testosterone + anxiety as the combined cause. Explained how both contribute. Treatment addresses both. Makes total sense.",
            "Most doctors just hear the symptom and give a generic medicine. Here the doctor asked about my sleep, stress, diet, relationship before even suggesting treatment. This is root-cause medicine. The plan he gave is specific to ME, not a generic protocol.",
            "Doctor took 45 minutes with me in the first consultation — explained the hormonal factors, the psychological overlay, the lifestyle triggers. Nothing was assumed. Tests confirmed the diagnosis. I finally understand what I've been dealing with.",
        ],
        "keywords": ["didn't prescribe immediately", "all possible causes", "ruled out one by one", "root cause", "personalized plan", "tests done first", "combined cause"],
        "gap_vs_ours": "Passion Fruit reviews describe the diagnostic PROCESS — not just the outcome. This builds clinical credibility. Our reviews say 'doctor explained' without saying what or how.",
        "search_queries": "erectile dysfunction root cause diagnosis · hormonal cause of ED · comprehensive ED consultation bangalore · ED specialist who runs tests · sexual health diagnosis bangalore",
    },
    {
        "city": "Bangalore", "category": "SH+STI",
        "name": "Medi Life Health Care - Sexologist Clinic",
        "gmb_url": "https://www.google.com/maps?cid=7163616557541776281",
        "reviews": 101, "rating": 4.8,
        "type_model": "B — 5-Star + Photo",
        "sample_reviews": [
            "The clinic is spotless. Felt like a premium experience from the moment I walked in. Attaching a photo — the consultation room is private and very well maintained.",
            "Clean, private, professional. Attached photos of the waiting area and consultation room. For a sensitive issue like this, the environment matters a lot — it put me at ease before I even met the doctor.",
            "First thing I noticed when I walked in was how clean and well-maintained everything is. The consultation room gives complete privacy. For this kind of issue that matters a lot. Photos speak for themselves — added 3.",
            "The ambiance and privacy here are exceptional. You walk in and immediately feel this is a serious, professional setup. Not the dingy back-alley clinic vibe you might fear. Attaching photos of reception and consultation room so others can see.",
        ],
        "keywords": ["spotless", "premium experience", "photo attached", "private consultation room", "environment matters", "felt at ease", "clean waiting area"],
        "gap_vs_ours": "Medi Life gets reviews WITH photos — which drive 5x more GMB profile views. Very few Allo reviews have photos attached. The clinic environment doesn't appear in our reviews at all.",
        "search_queries": "private sexual health clinic bangalore · clean sexual health clinic · discreet ED clinic bangalore · best sexologist clinic photos · confidential STI testing bangalore",
    },
    {
        "city": "Bangalore", "category": "SH",
        "name": "Dr Anantharaman's Clinic (BTM Layout)",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=Dr%20Anantharaman%27s%20clinic&query_place_id=ChIJqaUyMgIVrjsRV4CznaRDbcQ",
        "reviews": 615, "rating": 4.6,
        "type_model": "D — Staff / Team Mention",
        "sample_reviews": [
            "The front desk staff deserves special mention for their exceptional coordination and efficiency. Appointments were seamless, zero wait time, everything was organized.",
            "Every person I interacted with — the receptionist, the nurse, the coordinator — was professional and kind. This is not just a good doctor, it's a good clinic end-to-end.",
            "The coordinator who managed my follow-ups was incredibly attentive. Called before my appointment to confirm, followed up after to check on progress. Felt like a premium healthcare experience, not a quick consultation.",
            "Booked online, reminder came on time, doctor was ready, no wait. After consultation the staff explained the prescription clearly and gave a contact number for any doubts. The whole system works. Not just the doctor — the team.",
        ],
        "keywords": ["front desk named", "coordination praised", "zero wait time", "seamless appointment", "every staff member", "end-to-end experience"],
        "gap_vs_ours": "Dr Anantharaman's reviews mention non-doctor touchpoints. These signal a premium, organised clinic. Our reviews are doctor-only — patients don't mention coordinators because PRMs don't prompt them.",
        "search_queries": "best managed sexual health clinic bangalore · professional sexologist clinic · sexual health clinic with good service · zero wait time sexologist · organised men's health clinic",
    },
    {
        "city": "Pune", "category": "SH",
        "name": "Ashakiran Clinic — Sexologist Pune",
        "gmb_url": "https://www.google.com/maps?cid=12273061559789656859",
        "reviews": 203, "rating": 4.7,
        "type_model": "A — Condition Named + Outcome Story",
        "sample_reviews": [
            "Had erectile dysfunction for over a year and was too hesitant to talk about it. Doctor at Ashakiran was patient and non-judgmental. He explained that ED in my case was anxiety-driven, not physical, and put me on a structured treatment plan. Saw improvement within 3 weeks. By 2 months I was completely normal. Best decision I made — wish I had come here earlier instead of trying online remedies.",
            "मला premature ejaculation ची समस्या होती आणि मी खूप लाजाळू होतो डॉक्टरकडे जाण्यास. Ashakiran मध्ये गेलो, डॉक्टरांनी अगदी सहज समजवलं. उपचार सुरू केले आणि दीड महिन्यात खूप फरक पडला. आता आत्मविश्वास परत आला आहे. खूप आभार.",
            "I had performance issues that were ruining my confidence and relationship. The doctor here was the first one who didn't just hand me a prescription and send me away. He took 40 minutes to understand my history, stress levels, sleep patterns, and relationship situation. The diagnosis was stress-triggered ED. 6 weeks of treatment and lifestyle changes — I'm back to normal. The consultation itself was worth everything.",
            "Came for premature ejaculation. Doctor explained that PE has both psychological and physical components and treated both together. He also counselled my wife separately which helped her understand what I was going through. That dual approach — treating the patient and educating the partner — is what makes this clinic different from any other I visited.",
        ],
        "keywords": [
            "ED anxiety-driven Pune", "improvement within 3 weeks", "online remedies didn't work",
            "Marathi language review PE", "40-minute consultation", "stress-triggered ED",
            "lifestyle changes ED", "PE psychological physical both", "wife counselled separately",
            "partner education approach", "confidence back after treatment",
        ],
        "gap_vs_ours": "Ashakiran has Marathi-language reviews which rank for local-language searches ('लैंगिक समस्या Pune', 'sexologist Pune Marathi') — a query cluster we have zero coverage of. Also has partner-counselling reviews which seed 'couple sexual health consultation Pune' searches.",
        "search_queries": "sexologist Pune · लैंगिक समस्या Pune · PE treatment Pune · ED doctor Pune Marathi · couple counselling sexual health Pune · best sexologist Pune · anxiety ED treatment Pune",
    },
]

# ── top Ayurvedic / Unani SH competitors (for Tab 2 learning section) ────
AYU_COMPETITORS = [
    {
        "city": "Mumbai", "system": "Ayurvedic",
        "name": "Dr Yogendra Rai — Ayurvedic Sexologist Mumbai",
        "gmb_url": "https://www.google.com/maps?cid=6742592836584598016",
        "reviews": 7100, "rating": 5.0,
        "type_model": "E — Shame → Dignity",
        "sample_reviews": [
            "I suffered from nightfall and premature ejaculation for 3 years. I was too ashamed to tell anyone — not even my family. Found Dr Rai through a friend. He was the first doctor who made me feel that this is a medical condition, not a character flaw. After 45 days of herbal treatment I saw complete improvement. He gave me confidence back in myself. Every man dealing with this silently should come here.",
            "ED problem hai mujhe 2 saal se aur maine online bahut kuch try kiya tha. Kuch kaam nahi aaya. Yahan aaya toh doctor ne pehle poori history li aur 1 ghante tak baat ki. Phir natural dawai di. 2 mahine mein poori tarah theek ho gaya. Koi side effects nahi. Zindagi phir se khushaal ho gayi. 5 star toh banta hai.",
            "I had extreme performance anxiety and low stamina. My marriage was suffering badly. Dr Rai diagnosed it as stress-induced sexual weakness and put me on a 3-month Ayurvedic protocol — ashwagandha-based, along with lifestyle changes. By month 2 the difference was visible. My wife also noticed. He saved my marriage without a single chemical tablet. This is what holistic healing means.",
            "Came with Dhat syndrome — excessive semen discharge that was draining my energy and confidence. No allopathic doctor had even taken it seriously. Dr Rai knew exactly what I was describing, gave me a proper Unani-Ayurvedic diagnosis, and treated it within 6 weeks. The weakness, the anxiety, the energy loss — all resolved. Finally felt understood by a doctor.",
        ],
        "keywords": [
            "nightfall treatment Mumbai", "Dhat syndrome cure", "ED natural treatment no side effects",
            "ashwagandha ED treatment", "performance anxiety ayurvedic", "shame no family knows",
            "3 month protocol stamina", "2 mahine theek", "herbal ED cure",
            "online remedies failed came here", "marriage saved ayurvedic",
        ],
        "what_we_can_learn": "7,100 reviews at 5.0★ — 10× our Mumbai review count. They dominate on: (1) 'Dhat syndrome' / 'nightfall' — high-search vernacular terms we never use; (2) Hindi/Marathi reviews seeding regional-language queries; (3) 'no side effects' trust hook — we have the allopathic equivalent ('clinically guided, safe treatment') but never say it.",
        "search_queries": "nightfall treatment Mumbai · Dhat syndrome cure · ED natural treatment no side effects · ayurvedic sexologist Mumbai · ED treatment Hindi · शीघ्रपतन उपचार Mumbai",
    },
    {
        "city": "Hyderabad", "system": "Unani",
        "name": "Care and Cure Sexologist Clinic",
        "gmb_url": "https://www.google.com/maps?cid=13106034387764831720",
        "reviews": 3300, "rating": 4.9,
        "type_model": "G — Quick Win / Speed",
        "sample_reviews": [
            "I had PE problem for almost 5 years and tried many clinics. No results. Someone referred me to Care and Cure. Doctor gave me Unani medicine and within 15 days I felt 60% improvement. By 45 days I was fully cured. Never expected such fast results from herbal treatment. Doctor said this is guaranteed treatment and he was right. Highly recommend.",
            "Erectile weakness problem and low energy. Visited Care and Cure, the doctor took full history and gave personalised Unani medicine. Cost was very reasonable compared to other clinics. Within 1 month ED was gone. Doctor was very confident in treatment. The medicines were all herbal, nothing chemical. No side effects at all. 5 stars fully deserved.",
            "నాకు చాలా సంవత్సరాల నుండి లైంగిక సమస్య ఉంది. Care and Cure clinic లో చికిత్స తీసుకున్నాను. 2 నెలల్లో పూర్తిగా నయమైంది. డాక్టర్ చాలా అనుభవం ఉన్నారు. ధైర్యంగా వచ్చి చికిత్స తీసుకోండి. ఎవరికైనా ఈ సమస్య ఉంటే ఇక్కడికి రండి.",
            "Came to Care and Cure after being disappointed with other treatments. The Unani approach is completely different — the doctor treats the whole body not just the symptom. He found that my problem was connected to kidney weakness and stress. After treating those root causes, the PE resolved in 40 days. The approach is very logical once he explains it.",
        ],
        "keywords": [
            "PE 15 days improvement", "guaranteed Unani treatment", "45 days fully cured",
            "Telugu review sexual health", "kidney weakness ED link", "cost reasonable herbal",
            "5 years PE finally cured", "Unani root cause whole body", "no chemical medicine",
            "referred by friend worked", "low energy ED connection",
        ],
        "what_we_can_learn": "3,300 reviews on 'guaranteed results' + 'speed' language (15 days, 45 days) — this triggers searches like 'quick PE cure Hyderabad'. We have the advantage of being clinically honest about timelines, but we never actually state timelines in reviews. Telugu-language reviews are a completely untapped segment for us.",
        "search_queries": "PE cure Hyderabad fast · guaranteed sexologist Hyderabad · Unani sexual health Hyderabad · లైంగిక సమస్య Hyderabad · quick ED treatment Hyderabad · PE 45 days treatment",
    },
    {
        "city": "Bangalore", "system": "Ayurvedic",
        "name": "Dr Lohit's Ayur Sex Clinic — Sexologist Bangalore",
        "gmb_url": "https://www.google.com/maps?cid=7618749129619192663",
        "reviews": 1400, "rating": 4.7,
        "type_model": "F — Tried Elsewhere (strong pattern)",
        "sample_reviews": [
            "I had visited 4 different doctors before coming to Dr Lohit — allopathic, homeopathic, even one online consultation. Nothing gave lasting results. The allopathic medicines worked only while taking them. Came here as a last option. Dr Lohit put me on a 60-day Ayurvedic course for ED — it worked and the results have stayed for 6 months after stopping. Permanent treatment, not a temporary fix.",
            "ಎರಡು ವರ್ಷಗಳಿಂದ ಲೈಂಗಿಕ ದೌರ್ಬಲ್ಯ ಸಮಸ್ಯೆ ಇತ್ತು. ಡಾ. ಲೋಹಿತ್ ಅವರ ಕ್ಲಿನಿಕ್‌ಗೆ ಬಂದ ನಂತರ ೩ ತಿಂಗಳಲ್ಲಿ ಸಂಪೂರ್ಣ ಚೇತರಿಕೆ ಆಯಿತು. ಆಯುರ್ವೇದಿಕ್ ಔಷಧಗಳಿಂದ ಯಾವುದೇ ಅಡ್ಡ ಪರಿಣಾಮ ಇಲ್ಲ. ಎಲ್ಲರಿಗೂ ಶಿಫಾರಸು ಮಾಡುತ್ತೇನೆ.",
            "Had PE and low confidence for years. Tried sildenafil from another clinic — it helped for ED but did nothing for PE. Dr Lohit treated PE specifically with a combination of Ayurvedic medicine and yogic techniques. He was the first doctor who separated the two conditions and treated them individually. 3 months. PE resolved. Confidence back. I've referred 3 friends already.",
            "Cost comparison — I spent 3x more on allopathic treatments that didn't work than I spent here for a complete cure. The medicine costs are lower, consultations are thorough, and Dr Lohit calls to follow up personally. This is how medical care should feel. I regret not coming here first.",
        ],
        "keywords": [
            "4 doctors tried before", "allopathic didn't give permanent results", "60 day ayurvedic ED",
            "Kannada review sexual health", "PE and ED treated separately", "sildenafil failed PE",
            "3 months complete recovery", "permanent treatment not temporary fix",
            "doctor calls to follow up", "referred 3 friends", "cost comparison cheaper ayurvedic",
        ],
        "what_we_can_learn": "'Tried allopathic, didn't last — came to Ayurvedic' is their most powerful narrative. We are allopathic and can flip this: 'tried herbal, came here for proper diagnosis' is already a review type (F) we should be actively collecting. Their 'permanent' vs our 'evidence-based' is the key positioning gap to close with specific outcome reviews.",
        "search_queries": "permanent ED cure Bangalore · ayurvedic sexologist Bangalore · PE allopathic failed tried ayurvedic · ಲೈಂಗಿಕ ದೌರ್ಬಲ್ಯ Bangalore · sexologist Bangalore natural treatment · last resort ED treatment",
    },
]

# ── clinic target data — built dynamically from competition JSON ────────────
COMP_JSON = os.path.join(os.path.dirname(__file__), "data_competition.json")
with open(COMP_JSON) as _cf:
    _comp_data = json.load(_cf)

_sh_clinics = _comp_data['SH']['clinics']
TIER1_CITIES_SET = {'Bangalore', 'Hyderabad', 'Mumbai', 'Pune', 'Chennai', 'Navi Mumbai'}

# Recommended review mix per city (types to prompt for)
CITY_MIX = {
    'Bangalore':   '3A 2C 2E 1G 1D',
    'Hyderabad':   '3A 3H 2F 1E 1C',
    'Chennai':     '3A 2E 2H 1C 1B',
    'Mumbai':      '3A 2G 2E 1D 1B',
    'Navi Mumbai': '3A 2G 1E 1D 1B',
    'Pune':        '3A 2G 2B 1D 1H',
}

CLINIC_TARGETS = []
for _key in sorted(_sh_clinics.keys()):
    _entry = _sh_clinics[_key]
    if not isinstance(_entry, dict):
        continue
    _city = _entry.get('city', '')
    _loc  = _entry.get('loc', '')
    if _city not in TIER1_CITIES_SET:
        continue

    _our = _entry.get('our_reviews', 0)

    # Top 3 allopathic rivals, deduped by short name
    _seen = set()
    _rivals = []
    for _c in sorted(_entry.get('competitors', []), key=lambda x: x.get('reviews', 0), reverse=True):
        if _c.get('pathy') != 'Allopathic':
            continue
        _sname = (_c.get('name') or '')[:35]
        if _sname in _seen:
            continue
        _seen.add(_sname)
        _rivals.append({'name': _sname, 'reviews': _c.get('reviews', 0), 'maps': _c.get('maps', '')})
        if len(_rivals) >= 3:
            break

    if not _rivals:
        continue

    _top_reviews = _rivals[0]['reviews']
    _gap = max(0, _top_reviews - _our + 1)

    _ratio = _our / _top_reviews if _top_reviews > 0 else 1
    if _our >= _top_reviews:
        _prio, _vel = 'P3', 3
    elif _gap > 300 or _ratio < 0.15:
        _prio, _vel = 'P0', 15
    elif _gap > 80 or _ratio < 0.5:
        _prio, _vel = 'P1', 10
    else:
        _prio, _vel = 'P2', 6

    def _mo(rival_reviews):
        g = max(0, rival_reviews - _our + 1)
        return round(g / (_vel * 4.33), 1) if g > 0 else 0

    _months1 = _mo(_rivals[0]['reviews']) if len(_rivals) > 0 else 0
    _months2 = _mo(_rivals[1]['reviews']) if len(_rivals) > 1 else 0
    _months3 = _mo(_rivals[2]['reviews']) if len(_rivals) > 2 else 0

    CLINIC_TARGETS.append({
        'city': _city, 'loc': _loc, 'our': _our,
        'r1': _rivals[0] if len(_rivals) > 0 else None,
        'r2': _rivals[1] if len(_rivals) > 1 else None,
        'r3': _rivals[2] if len(_rivals) > 2 else None,
        'prio': _prio, 'vel': _vel,
        'months1': _months1, 'months2': _months2, 'months3': _months3,
        'mix': CITY_MIX.get(_city, '3A 2E 2C 1G 1B'),
    })

# ── build workbook ───────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══ TAB 1: OUR REVIEWS ══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Our Reviews"
ws1.freeze_panes = "A3"
ws1.sheet_view.showGridLines = False

# Column widths
col_widths = [12, 20, 10, 6, 60, 45, 40, 10]
headers = ["City", "Clinic", "Date", "★", "Review Text", "Keywords Present in This Review", "What's Missing (vs Competitors)", "Type"]
for i, (w, h) in enumerate(zip(col_widths, headers), 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title row
ws1.merge_cells("A1:H1")
cell_style(ws1, 1, 1, "Allo Health — Our GMB Reviews Analysis (Last 12 Months, Tier 1 Clinics)", bold=True, bg="1A3A5C", font_color="FFFFFF", size=12, align="center")
ws1.row_dimensions[1].height = 22

# Header row
for col, h in enumerate(headers, 1):
    cell_style(ws1, 2, col, h, bold=True, bg="2B5F8E", font_color="FFFFFF", size=9, align="center")
ws1.row_dimensions[2].height = 18

# Priority colours by review type
TYPE_BG = {"A":"FFF0F0","E":"FFF3EA","F":"F0FFF4","G":"F0F0FF","C":"F5F0FF",
           "B":"EEF4FF","D":"EDFAF8","H":"FAF0FF","Generic":"FAFAFA"}

# Data rows
for row_idx, r in enumerate(sampled, 3):
    found, missing = classify_keywords(r['review_text'])
    rtype = review_type(found)
    bg = TYPE_BG.get(rtype, "FFFFFF")

    city_map = {"Bengaluru":"BLR","Hyderabad":"HYD","Mumbai":"MUM","Pune":"PNE",
                "Chennai":"CHN","Navi Mumbai":"NMB"}
    values = [
        city_map.get(r['city'], r['city']),
        r.get('_clinic_short', short_clinic(r['clinic_name'])),
        r.get('review_date',''),
        r.get('star_rating',''),
        r['review_text'][:400],
        "\n".join(f"✓ {f}" for f in found) if found else "— none detected",
        "\n".join(f"✗ {m}" for m in missing[:5]),
        rtype,
    ]
    row_height = max(40, min(120, len(r['review_text']) // 3))
    ws1.row_dimensions[row_idx].height = row_height

    for col, val in enumerate(values, 1):
        c = cell_style(ws1, row_idx, col, val, bg=bg, wrap=(col in (5,6,7)), align="left" if col>1 else "center", size=9)
        if col == 4:  # stars
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Arial", size=9, color="B8760A")
        if col == 8:  # type badge
            tc = {"A":"C42B2B","E":"C95A14","F":"15803D","G":"4F46E5",
                  "C":"6B2DB5","B":"1D60B5","D":"0D7F76","H":"7C3AED","Generic":"6B7280"}
            c.font = Font(name="Arial", size=9, bold=True, color=tc.get(rtype,"6B7280"))

# ── keyword summary table (for chart) ────────────────────────────────────
summary_start = len(sampled) + 5
ws1.merge_cells(f"A{summary_start}:H{summary_start}")
cell_style(ws1, summary_start, 1, "Keyword Frequency: Our Reviews vs Winning Competitor Pattern", bold=True, bg="1A3A5C", font_color="FFFFFF", size=11, align="center")

hdr_row = summary_start + 1
for col, h in enumerate(["Keyword / Signal Type", f"Our Reviews (n={len(sampled)})", f"Competitor Target (n={len(sampled)})", "Gap"], 1):
    cell_style(ws1, hdr_row, col, h, bold=True, bg="2B5F8E", font_color="FFFFFF", size=9, align="center")

data_start_row = hdr_row + 1
for i, kw in enumerate(KW_LABELS):
    r = data_start_row + i
    our = our_counts[kw]
    comp = COMP_TARGET[kw]
    gap = comp - our
    cell_style(ws1, r, 1, kw, bg="F8F9FB", size=9)
    cell_style(ws1, r, 2, our, align="center", size=10, bold=True,
               font_color="C42B2B" if our < comp else "15803D")
    cell_style(ws1, r, 3, comp, align="center", size=10, bold=True, font_color="1D60B5")
    cell_style(ws1, r, 4, gap, align="center", size=9, italic=True,
               font_color="C42B2B" if gap > 0 else "15803D")

# ── bar chart ─────────────────────────────────────────────────────────────
chart = BarChart()
chart.type = "bar"  # horizontal
chart.grouping = "clustered"
chart.title = "Our Review Keywords vs Winning Competitor Pattern"
chart.style = 2
chart.width = 22
chart.height = 14
chart.y_axis.numFmt = "0"
chart.x_axis.title = "Review Count"

cats = Reference(ws1, min_col=1, min_row=data_start_row, max_row=data_start_row + len(KW_LABELS) - 1)
our_data = Reference(ws1, min_col=2, min_row=hdr_row, max_row=data_start_row + len(KW_LABELS) - 1)
comp_data = Reference(ws1, min_col=3, min_row=hdr_row, max_row=data_start_row + len(KW_LABELS) - 1)

chart.add_data(our_data, titles_from_data=True)
chart.add_data(comp_data, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = "C42B2B"
chart.series[1].graphicalProperties.solidFill = "1D60B5"

chart_anchor = f"F{summary_start}"
ws1.add_chart(chart, chart_anchor)

# ══ TAB 2: COMPETITOR REVIEWS ═══════════════════════════════════════════
ws2 = wb.create_sheet("2. Competitor Reviews")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

comp_col_widths = [14, 30, 12, 8, 7, 7, 24, 65, 38, 46, 55]
comp_headers = ["City", "Clinic / Doctor", "GMB Link", "Total Reviews", "Rating",
                "Review #", "Review Type", "Sample Review Text", "Keywords They Use",
                "Our Gap vs Them", "GMB Search Queries This Type Will Trigger"]
for i, (w, h) in enumerate(zip(comp_col_widths, comp_headers), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:K1")
cell_style(ws2, 1, 1, "Competitor GMB Review Analysis — What Winning Reviews Look Like  |  Filter by City or Clinic to explore", bold=True, bg="7C3A00", font_color="FFFFFF", size=12, align="center")
ws2.row_dimensions[1].height = 22

for col, h in enumerate(comp_headers, 1):
    cell_style(ws2, 2, col, h, bold=True, bg="A0522D", font_color="FFFFFF", size=9, align="center")
ws2.row_dimensions[2].height = 18

# AutoFilter on header row
ws2.auto_filter.ref = "A2:K2"

COMP_ROW_BG = ["FFF8F0","FFF0E0","FFF8F0","FFF0E0","FFF8F0","FFF0E0","FFF8F0"]

# One row per sample review (4 per competitor = up to 28 rows)
data_row = 3
for ci, comp in enumerate(COMPETITORS):
    bg = COMP_ROW_BG[ci % len(COMP_ROW_BG)]
    kw_text = "\n".join(f"• {k}" for k in comp["keywords"])

    for rev_idx, review_text in enumerate(comp["sample_reviews"]):
        ws2.row_dimensions[data_row].height = 80

        vals = [
            comp["city"],
            comp["name"],
            comp["gmb_url"],          # col 3 — becomes hyperlink
            comp["reviews"],
            comp["rating"],
            f"Review {rev_idx + 1}",
            comp["type_model"],
            review_text,
            kw_text if rev_idx == 0 else "",
            comp["gap_vs_ours"] if rev_idx == 0 else "",
            comp.get("search_queries","") if rev_idx == 0 else "",  # search queries on first row only
        ]
        for col, val in enumerate(vals, 1):
            c = cell_style(ws2, data_row, col, val, bg=bg, wrap=(col in (8, 9, 10, 11)), size=9,
                            align="center" if col in (1, 4, 5, 6) else "left")
            if col == 3:  # GMB hyperlink
                c.hyperlink = val
                c.font = Font(name="Arial", size=9, color="1D60B5", underline="single")
                c.value = "Open GMB →"
            if col == 4:  # review count
                c.font = Font(name="Arial", size=10, bold=True, color="1A3A5C")
                c.alignment = Alignment(horizontal="center", vertical="center")
            if col == 7:  # type badge
                type_colors = {"A":"C42B2B","E":"C95A14","F":"15803D","G":"4F46E5",
                               "C":"6B2DB5","B":"1D60B5","D":"0D7F76"}
                t_letter = comp["type_model"][0]
                c.font = Font(name="Arial", size=9, bold=True,
                              color=type_colors.get(t_letter, "374151"))
            if col == 10 and val:  # gap column
                c.font = Font(name="Arial", size=9, color="C42B2B", italic=True)
            if col == 11 and val:  # search queries — green to signal positive outcome
                c.font = Font(name="Arial", size=9, color="15803D", italic=True)

        data_row += 1

# ── SH keyword gap summary (below competitor rows in Tab 2) ──────────────
sh_gap_start = data_row + 2
ws2.merge_cells(f"A{sh_gap_start}:K{sh_gap_start}")
cell_style(ws2, sh_gap_start, 1,
    "SH Keyword Gap — What Winning Competitors Say vs What We Have  |  "
    "Based on analysis of all Tier-1 GMB reviews",
    bold=True, bg="7C3A00", font_color="FFFFFF", size=11, align="center")
ws2.row_dimensions[sh_gap_start].height = 20

for col, h in enumerate(["SH Keyword / Signal", "Competitor Target (est.)",
                          f"Our Count (n={len(tier1_rows)} reviews)", "Gap",
                          "Search Queries Triggered"], 1):
    cell_style(ws2, sh_gap_start + 1, col, h, bold=True, bg="A0522D", font_color="FFFFFF",
               size=9, align="center")

n_sh = len(tier1_rows) or 1
_SH_GAP_DATA = [
    ("ED named explicitly (erectile dysfunction)",
     int(n_sh * 0.38), sh_tier1_counts.get("ED named explicitly", 0),
     "erectile dysfunction clinic / ED treatment near me / ED doctor"),
    ("PE named explicitly (premature ejaculation)",
     int(n_sh * 0.35), sh_tier1_counts.get("PE named explicitly", 0),
     "premature ejaculation treatment / PE doctor / how to stop PE"),
    ("Performance anxiety named",
     int(n_sh * 0.22), sh_tier1_counts.get("Performance Anxiety", 0),
     "performance anxiety treatment / anxiety-driven ED / performance anxiety clinic"),
    ("Low testosterone / hormonal named",
     int(n_sh * 0.18), sh_tier1_counts.get("Low testosterone", 0),
     "low testosterone treatment / testosterone doctor Bangalore / testosterone clinic"),
    ("Specific timeline (e.g. 'in 45 days')",
     int(n_sh * 0.40), sh_tier1_counts.get("Specific timeline", 0),
     "'improved in 45 days' / '3 weeks' — each timeline phrase is its own trust signal"),
    ("Root cause / diagnosis journey",
     int(n_sh * 0.32), sh_tier1_counts.get("Root cause / diagnosis", 0),
     "root cause ED / what causes premature ejaculation / hormonal ED diagnosis"),
    ("Tried elsewhere (Ayurveda/others failed)",
     int(n_sh * 0.28), sh_tier1_counts.get("Tried elsewhere", 0),
     "ayurvedic ED didn't work / tried other doctors PE / men's health last resort"),
    ("Shame → dignity transformation",
     int(n_sh * 0.30), sh_tier1_counts.get("Shame / embarrassment", 0),
     "embarrassed men's health / scared to see ED doctor / safe to talk PE"),
    ("Partner / relationship impact named",
     int(n_sh * 0.15), sh_tier1_counts.get("Partner / relationship", 0),
     "ED affecting marriage / PE and intimacy / relationship problem men's health"),
    ("Treatment drug named (tadalafil/sildenafil)",
     int(n_sh * 0.20), sh_tier1_counts.get("Treatment drug named", 0),
     "tadalafil prescription / sildenafil doctor / PDE5 inhibitor near me"),
    ("Doctor named by first name",
     int(n_sh * 0.70), sh_tier1_counts.get("Doctor named", 0),
     "Dr [first name] Allo Health — seeds doctor-name search queries"),
    ("Regional language (Telugu/Tamil/Kannada/Marathi)",
     int(n_sh * 0.12), sh_tier1_counts.get("Regional language", 0),
     "Telugu men's health / Tamil ED clinic / Kannada sexual health — local queries"),
]

for i, (kw, comp_est, ours, queries) in enumerate(_SH_GAP_DATA):
    _row = sh_gap_start + 2 + i
    _gap = max(0, comp_est - ours)
    _gc = "C42B2B" if _gap > 20 else "C95A14" if _gap > 5 else "15803D"
    _bg = "FFF0F0" if _gap > 20 else "FFF8EA" if _gap > 5 else "F0FFF4"
    ws2.row_dimensions[_row].height = 32
    cell_style(ws2, _row, 1, kw, bg=_bg, bold=(_gap > 20), size=9)
    cell_style(ws2, _row, 2, f"~{comp_est}", bg=_bg, align="center", size=9, font_color="1A3A5C")
    cell_style(ws2, _row, 3, str(ours), bg=_bg, align="center", size=9,
               font_color="15803D" if ours >= comp_est else "C42B2B")
    cell_style(ws2, _row, 4,
               f"−{_gap}" if _gap > 0 else "✓ Winning",
               bg=_bg, align="center", size=9, bold=True, font_color=_gc)
    cell_style(ws2, _row, 5, queries, bg=_bg, wrap=True, size=9, font_color="374151")
    ws2.merge_cells(f"E{_row}:K{_row}")

# ── Ayurvedic / Unani competitors section (learning, not copying) ─────────
_ayu_section_start = sh_gap_start + 2 + len(_SH_GAP_DATA) + 3
ws2.merge_cells(f"A{_ayu_section_start}:K{_ayu_section_start}")
cell_style(ws2, _ayu_section_start, 1,
    "AYURVEDIC / UNANI COMPETITORS — What They Do Well  |  "
    "Learn from their review strategy (language, velocity, regional) — not their claims",
    bold=True, bg="4D3000", font_color="FFD580", size=12, align="center")
ws2.row_dimensions[_ayu_section_start].height = 22

_ayu_hdr_row = _ayu_section_start + 1
_ayu_hdr_labels = ["City", "Clinic", "GMB Link", "Reviews", "Rating",
                   "Review #", "System", "Sample Review Text",
                   "Keywords / Language They Use", "What We Can Learn From This",
                   "Searches They Trigger (we should also rank for)"]
for col, h in enumerate(_ayu_hdr_labels, 1):
    cell_style(ws2, _ayu_hdr_row, col, h, bold=True, bg="7A4800", font_color="FFFFFF",
               size=9, align="center")
ws2.row_dimensions[_ayu_hdr_row].height = 22

_AYU_SYS_COLORS = {"Ayurvedic": "7A4800", "Unani": "5B2D8E", "Homeopathic": "1A6B4A"}
_AYU_SYS_BG     = {"Ayurvedic": "FFF8EA", "Unani": "F8F0FF", "Homeopathic": "F0FFF8"}

_ayu_row = _ayu_hdr_row + 1
for comp in AYU_COMPETITORS:
    kw_text = "\n".join(f"• {k}" for k in comp["keywords"])
    sys_key = comp["system"]
    bg = _AYU_SYS_BG.get(sys_key, "FAFAFA")

    for rev_idx, review_text in enumerate(comp["sample_reviews"]):
        vals = [
            comp["city"],
            comp["name"],
            comp["gmb_url"],
            comp["reviews"],
            comp["rating"],
            f"Review {rev_idx + 1}",
            comp["system"],
            review_text,
            kw_text if rev_idx == 0 else "",
            comp["what_we_can_learn"] if rev_idx == 0 else "",
            comp["search_queries"] if rev_idx == 0 else "",
        ]
        ws2.row_dimensions[_ayu_row].height = 75

        for col, val in enumerate(vals, 1):
            c = cell_style(ws2, _ayu_row, col, val, bg=bg,
                           wrap=(col in (8, 9, 10, 11)), size=9,
                           align="center" if col in (1, 4, 5, 6) else "left")
            if col == 3:
                c.hyperlink = val
                c.font = Font(name="Arial", size=9, color="1D60B5", underline="single")
                c.value = "Open GMB →"
            if col == 4:
                c.font = Font(name="Arial", size=10, bold=True,
                              color=_AYU_SYS_COLORS.get(sys_key, "374151"))
            if col == 7:
                c.font = Font(name="Arial", size=9, bold=True,
                              color=_AYU_SYS_COLORS.get(sys_key, "374151"))
            if col == 10 and val:
                c.font = Font(name="Arial", size=9, color="7A4800", italic=True)
            if col == 11 and val:
                c.font = Font(name="Arial", size=9, color="15803D", italic=True)
        _ayu_row += 1

# ══ TAB 3: CLINIC TARGETS ════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Clinic Targets")
ws3.freeze_panes = "A3"
ws3.sheet_view.showGridLines = False
ws3.auto_filter.ref = "A2:O2"

# Cols: City | Clinic | Our | R1 name | R1# | R2 name | R2# | R3 name | R3# | Priority | /wk | Mo→#1 | Mo→#2 | Mo→#3 | Mix
t3_widths = [13, 20, 9, 26, 7, 24, 7, 24, 7, 7, 9, 10, 10, 10, 28]
t3_headers = [
    "City", "Clinic", "Our Reviews",
    "Rival #1 (Biggest)", "#1 Reviews",
    "Rival #2 (Mid)", "#2 Reviews",
    "Rival #3 (Nearest)", "#3 Reviews",
    "Priority", "Reviews/wk",
    "Months to Beat #1", "Months to Beat #2", "Months to Beat #3",
    "Review Mix (weekly)"
]
for i, (w, h) in enumerate(zip(t3_widths, t3_headers), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:O1")
cell_style(ws3, 1, 1, "Clinic Review Targets — Top 3 Allopathic Rivals Per Clinic | Weekly Velocity | Months to Beat", bold=True, bg="1A3A5C", font_color="FFFFFF", size=12, align="center")
ws3.row_dimensions[1].height = 22

for col, h in enumerate(t3_headers, 1):
    cell_style(ws3, 2, col, h, bold=True, bg="2B5F8E", font_color="FFFFFF", size=9, align="center")
ws3.row_dimensions[2].height = 22

P_COLORS = {"P0": ("FFF0F0","C42B2B"), "P1": ("FFF3EA","C95A14"),
            "P2": ("FFFBEA","B8760A"), "P3": ("F0FFF4","15803D")}
CITY_COLORS = {"Bangalore": "EEF4FF", "Hyderabad": "FFF8F0", "Chennai": "F5F0FF",
               "Mumbai": "F0FFF8", "Navi Mumbai": "FFFBF0", "Pune": "FFF0F8"}

for ri, t in enumerate(CLINIC_TARGETS, 3):
    city  = t['city']
    r1    = t['r1'] or {}
    r2    = t['r2'] or {}
    r3    = t['r3'] or {}
    prio  = t['prio']
    our   = t['our']
    top_r = r1.get('reviews', 0)

    bg_city = CITY_COLORS.get(city, "FFFFFF")
    p_bg, p_fg = P_COLORS.get(prio, ("FFFFFF","111111"))
    ws3.row_dimensions[ri].height = 22

    ratio = our / top_r if top_r > 0 else 1
    our_color = "C42B2B" if ratio < 0.2 else "C95A14" if ratio < 0.5 else "15803D"

    def mo_cell(mo):
        if mo == 0:
            return "Winning", "15803D"
        color = "C42B2B" if mo > 18 else "C95A14" if mo > 9 else "15803D"
        return f"{mo} mo", color

    mo1_val, mo1_col = mo_cell(t.get('months1', 0))
    mo2_val, mo2_col = mo_cell(t.get('months2', 0))
    mo3_val, mo3_col = mo_cell(t.get('months3', 0))

    row_vals = [
        (city,               bg_city, "left",   9,  False, "111111"),
        (t['loc'],           bg_city, "left",   9,  False, "111111"),
        (our,                bg_city, "center", 11, True,  our_color),
        (r1.get('name',''),  bg_city, "left",   8,  False, "1A3A5C"),
        (r1.get('reviews',0),bg_city, "center", 10, True,  "1A3A5C"),
        (r2.get('name',''),  bg_city, "left",   8,  False, "374151"),
        (r2.get('reviews',0),bg_city, "center", 9,  False, "374151"),
        (r3.get('name',''),  bg_city, "left",   8,  False, "6B7280"),
        (r3.get('reviews',0),bg_city, "center", 9,  False, "6B7280"),
        (prio,               p_bg,    "center", 9,  True,  p_fg),
        (t['vel'],           bg_city, "center", 10, True,  "1A3A5C"),
        (mo1_val,            bg_city, "center", 9,  True,  mo1_col),
        (mo2_val,            bg_city, "center", 9,  False, mo2_col),
        (mo3_val,            bg_city, "center", 9,  False, mo3_col),
        (t['mix'],           bg_city, "left",   9,  False, "374151"),
    ]

    for col, (val, bg, align, size, bold, fc) in enumerate(row_vals, 1):
        c = cell_style(ws3, ri, col, val, bg=bg, align=align, size=size, bold=bold, font_color=fc)

# ── legend block ─────────────────────────────────────────────────────────
leg_row = len(CLINIC_TARGETS) + 5
ws3.merge_cells(f"A{leg_row}:O{leg_row}")
cell_style(ws3, leg_row, 1, "LEGEND — Review Type Mix Codes  |  Priority: P0=Critical(15/wk) · P1=Urgent(10/wk) · P2=Close gap(6/wk) · P3=Winning(3/wk)", bold=True, bg="1A3A5C", font_color="FFFFFF", size=10, align="center")

TYPE_COLORS = {"A":"C42B2B","B":"1D60B5","C":"6B2DB5","D":"0D7F76",
               "E":"C95A14","F":"15803D","G":"4F46E5","H":"7C3AED"}
legend_items = [
    ("A", "Condition named + outcome story — 'had ED for 2 yrs, treatment worked in 45 days, confidence back'  ← HIGHEST GMB SEO IMPACT"),
    ("B", "5-star + photo attached + clinic environment (clean, private, spotless)"),
    ("C", "Root cause journey — doctor explained causes, ruled out systematically, personalized plan"),
    ("D", "Staff / team named — front desk, coordinator, nurse praised by name"),
    ("E", "Shame → Dignity — 'was embarrassed, doctor made me feel safe, wish I'd come sooner'"),
    ("F", "Tried elsewhere — 'Ayurvedic/other clinics didn't work, came here as last resort'"),
    ("G", "Quick win — 'saw improvement within 10 days / first week / specific timeline given'"),
    ("H", "Regional language — Telugu/Tamil/Kannada/Marathi reviews rank for local-language searches"),
]
for i, (t, desc) in enumerate(legend_items):
    r = leg_row + 1 + i
    cell_style(ws3, r, 1, f"Type {t}", bold=True, bg="F8F9FB", font_color=TYPE_COLORS[t], size=9, align="center")
    ws3.merge_cells(f"B{r}:O{r}")
    cell_style(ws3, r, 2, desc, bg="F8F9FB", size=9)

# ══ TAB 4: OUR STI REVIEWS ═══════════════════════════════════════════════
ws4 = wb.create_sheet("4. Our STI Reviews")
ws4.freeze_panes = "A3"
ws4.sheet_view.showGridLines = False
ws4.auto_filter.ref = "A2:H2"

_s4_widths = [12, 28, 10, 5, 62, 45, 45, 22]
_s4_headers = ["City", "Clinic", "Date", "★", "Review Text",
               "STI Keywords Present", "What's Missing vs Competitors", "STI Review Type"]
for i, (w, h) in enumerate(zip(_s4_widths, _s4_headers), 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("A1:H1")
cell_style(ws4, 1, 1,
    f"Our STI Reviews — Tier 1 Clinics ({len(sti_sampled)} reviews)  |  "
    "Mapped against competitor keyword strategy — green = we have it, red = we're missing it",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=12, align="center")
ws4.row_dimensions[1].height = 22

for col, h in enumerate(_s4_headers, 1):
    cell_style(ws4, 2, col, h, bold=True, bg="9B1C1C", font_color="FFFFFF", size=9, align="center")
ws4.row_dimensions[2].height = 18

_city_map4 = {"Bengaluru":"BLR","Hyderabad":"HYD","Mumbai":"MUM",
              "Pune":"PNE","Chennai":"CHN","Navi Mumbai":"NMB"}

for _ri, _r in enumerate(sti_sampled, 3):
    _txt = _r.get('review_text', '')
    _found, _missing = classify_sti_keywords(_txt)
    _rtype = sti_review_type(_found)
    _bg4 = STI_TYPE_BG_MAP.get(_rtype, "FAFAFA")
    ws4.row_dimensions[_ri].height = max(40, min(120, len(_txt) // 3))

    _vals4 = [
        _city_map4.get(_r.get('city',''), _r.get('city','')),
        short_clinic(_r.get('clinic_name', '')),
        _r.get('review_date', ''),
        _r.get('star_rating', ''),
        _txt[:400],
        "\n".join(f"✓ {f}" for f in _found) if _found else "— none detected",
        "\n".join(f"✗ {m}" for m in _missing[:7]),
        STI_TYPE_LABEL_MAP.get(_rtype, _rtype),
    ]
    for col, val in enumerate(_vals4, 1):
        c = cell_style(ws4, _ri, col, val, bg=_bg4, wrap=(col in (5, 6, 7)),
                       align="left" if col > 1 else "center", size=9)
        if col == 4:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Arial", size=9, color="B8760A")
        if col == 8:
            c.font = Font(name="Arial", size=9, bold=True,
                          color=STI_TYPE_COLORS_MAP.get(_rtype, "6B7280"))

# ── STI review keyword summary chart ─────────────────────────────────────
_sti_sum_start = len(sti_sampled) + 5
ws4.merge_cells(f"A{_sti_sum_start}:H{_sti_sum_start}")
cell_style(ws4, _sti_sum_start, 1,
    f"STI Keyword Frequency in Our Reviews (n={len(sti_sampled)}) vs Competitor Target",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=11, align="center")
ws4.row_dimensions[_sti_sum_start].height = 20

_sti_hdr_row = _sti_sum_start + 1
for col, h in enumerate([
    "STI Keyword / Signal", f"In Our Reviews (n={len(sti_sampled)})",
    "Competitor Target (est.)", "Gap", "Search Queries Triggered"
], 1):
    cell_style(ws4, _sti_hdr_row, col, h, bold=True, bg="9B1C1C",
               font_color="FFFFFF", size=9, align="center")

_STI_KW_LABELS = [
    "Condition: Chlamydia", "Condition: Gonorrhea", "Condition: Syphilis",
    "Condition: Herpes / HSV", "HIV named", "PEP / PrEP",
    "Discharge / burning urin.", "Confidential / anonymous",
    "Same-day / rapid results", "Partner tested too",
    "Treatment named", "Test named (panel/ELISA)", "Outcome (cleared/cured)",
]
_sti_our_counts = {lbl: 0 for lbl in _STI_KW_LABELS}
for _r in sti_sampled:
    _f, _ = classify_sti_keywords(_r.get('review_text', ''))
    for _lbl in _f:
        if _lbl in _sti_our_counts:
            _sti_our_counts[_lbl] += 1

_n_sti = len(sti_sampled) or 1
_STI_COMP_TARGET = {
    "Condition: Chlamydia":       int(_n_sti * 0.35),
    "Condition: Gonorrhea":       int(_n_sti * 0.30),
    "Condition: Syphilis":        int(_n_sti * 0.25),
    "Condition: Herpes / HSV":    int(_n_sti * 0.20),
    "HIV named":                  int(_n_sti * 0.25),
    "PEP / PrEP":                 int(_n_sti * 0.15),
    "Discharge / burning urin.":  int(_n_sti * 0.25),
    "Confidential / anonymous":   int(_n_sti * 0.40),
    "Same-day / rapid results":   int(_n_sti * 0.35),
    "Partner tested too":         int(_n_sti * 0.25),
    "Treatment named":            int(_n_sti * 0.30),
    "Test named (panel/ELISA)":   int(_n_sti * 0.35),
    "Outcome (cleared/cured)":    int(_n_sti * 0.45),
}
_STI_QUERIES = {
    "Condition: Chlamydia":       "chlamydia test / chlamydia treatment near me",
    "Condition: Gonorrhea":       "gonorrhea test / gonorrhea doctor / gonorrhea ceftriaxone",
    "Condition: Syphilis":        "syphilis VDRL test / syphilis treatment / primary syphilis",
    "Condition: Herpes / HSV":    "herpes doctor / HSV-2 treatment / valacyclovir",
    "HIV named":                  "HIV test near me / HIV doctor / HIV clinic",
    "PEP / PrEP":                 "PEP clinic near me / PrEP HIV prevention",
    "Discharge / burning urin.":  "burning urination men / penile discharge treatment",
    "Confidential / anonymous":   "anonymous STI test / confidential STD test",
    "Same-day / rapid results":   "same day STI test / rapid HIV test / STI results 24 hours",
    "Partner tested too":         "couple STI test / pre-marital STI panel / partner HIV test",
    "Treatment named":            "doxycycline STI / azithromycin STI / antibiotic STI",
    "Test named (panel/ELISA)":   "ELISA HIV test / VDRL test / full STI panel",
    "Outcome (cleared/cured)":    "'STI cleared' / 'test came back negative' / 'all clear STI'",
}

for i, lbl in enumerate(_STI_KW_LABELS):
    _row4 = _sti_hdr_row + 1 + i
    _ours4 = _sti_our_counts[lbl]
    _comp4 = _STI_COMP_TARGET[lbl]
    _gap4 = max(0, _comp4 - _ours4)
    _gc4 = "C42B2B" if _gap4 > 5 else "C95A14" if _gap4 > 2 else "15803D"
    _bg4r = "FFF0F0" if _gap4 > 5 else "FFF8EA" if _gap4 > 2 else "F0FFF4"
    ws4.row_dimensions[_row4].height = 30
    cell_style(ws4, _row4, 1, lbl, bg=_bg4r, bold=(_gap4 > 5), size=9)
    cell_style(ws4, _row4, 2, _ours4, bg=_bg4r, align="center", size=10, bold=True,
               font_color="15803D" if _ours4 >= _comp4 else "C42B2B")
    cell_style(ws4, _row4, 3, _comp4, bg=_bg4r, align="center", size=10, bold=True,
               font_color="1D60B5")
    cell_style(ws4, _row4, 4, f"−{_gap4}" if _gap4 > 0 else "✓ OK",
               bg=_bg4r, align="center", size=9, bold=True, font_color=_gc4)
    cell_style(ws4, _row4, 5, _STI_QUERIES.get(lbl, ""), bg=_bg4r, wrap=True, size=9,
               font_color="374151")
    ws4.merge_cells(f"E{_row4}:H{_row4}")

# ── STI review type legend ────────────────────────────────────────────────
_sti4_leg = _sti_hdr_row + 1 + len(_STI_KW_LABELS) + 2
ws4.merge_cells(f"A{_sti4_leg}:H{_sti4_leg}")
cell_style(ws4, _sti4_leg, 1,
    "REVIEW TYPE GUIDE  |  "
    "S1=Condition named  S2=Confidential  S3=HIV/PEP/PrEP  "
    "S4=Partner tested  S5=Symptom  S6=Treatment named  S7=Outcome  Generic=only says 'STI'",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=9, align="center")
for i, (code, color, desc) in enumerate([
    ("S1","C42B2B","Condition named — chlamydia/gonorrhea/syphilis/herpes. Highest SEO impact — each triggers its own query cluster."),
    ("S2","4F46E5","Confidential / anonymous — 'no registration', 'discreet', 'anonymous test'. Captures high-stigma searchers."),
    ("S3","7C3AED","HIV / PEP / PrEP — 'PEP within 72 hours', 'PrEP prescription'. Very high intent, underserved segment."),
    ("S4","0D7F76","Partner tested together — 'couple STI test', 'pre-marital panel'. Low stigma, high conversion."),
    ("S5","15803D","Symptom described — discharge, burning urination, rash. Captures early-stage searchers."),
    ("S6","C95A14","Treatment named — doxycycline/azithromycin/valacyclovir. Seeds drug+condition search combos."),
    ("S7","1D60B5","Outcome given — 'cleared', 'all clear', 'came back negative'. Trust signal + SEO for outcome searches."),
    ("Generic","6B7280","Only says 'STI/STD' without specifics. No additional query clusters beyond generic STI search."),
]):
    _r = _sti4_leg + 1 + i
    cell_style(ws4, _r, 1, f"Type {code}", bold=True, bg="F8F9FB", font_color=color, size=9, align="center")
    ws4.merge_cells(f"B{_r}:H{_r}")
    cell_style(ws4, _r, 2, desc, bg="F8F9FB", size=9)

# ══ TAB 5: STI COMPETITOR ANALYSIS ══════════════════════════════════════
STI_COMPETITORS = [
    {
        "city": "Hyderabad",
        "name": "Dr Monalisa Sahu — Infectious Disease Specialist",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=Dr+Monalisa+Sahu+Infectious+Disease+Hyderabad",
        "reviews": 726,
        "rating": 4.8,
        "pathy": "Allopathic",
        "speciality": "Infectious Disease / STI",
        "type_model": "S1 — Specific condition named",
        "sample_reviews": [
            "I tested positive for chlamydia and was completely panicking. Dr Monalisa explained that it's one of the most common and easily treated bacterial infections. She put me on a 7-day doxycycline course. I re-tested at 4 weeks — completely clear. No judgment, just clinical and caring.",
            "Came in suspecting gonorrhea after unprotected sex. She ran a full STI panel same day — gonorrhea, chlamydia, syphilis, HIV. Results within 24 hours. Gonorrhea confirmed, treated with single-dose ceftriaxone injection. Cleared in 2 weeks. She also tested my partner over video call.",
            "Syphilis caught in primary stage. I had a painless sore and ignored it for 2 weeks. She identified it immediately in the clinical exam, confirmed with VDRL. Penicillin injection given same visit. She explained the 3-stage progression and why early treatment matters. Follow-up VDRL at 3 months came back clear.",
            "HIV scare after risky exposure. She walked me through PEP eligibility — I was within 72 hours so she started me on the regimen immediately. 28-day course completed. Test at 45 days and 90 days both negative. She checked in via WhatsApp throughout. Life-saving doctor.",
        ],
        "keywords": [
            "chlamydia treatment Hyderabad", "gonorrhea test same day", "syphilis VDRL test",
            "HIV PEP within 72 hours", "STI panel full", "re-test 4 weeks", "ceftriaxone injection",
            "infectious disease specialist", "STI cleared in 2 weeks", "primary stage syphilis",
        ],
        "gap_vs_ours": "Specific condition names (chlamydia/gonorrhea/syphilis) trigger separate Google entity indexes — we say 'STI' generically. Condition-specific reviews each rank for their own search query.",
        "search_queries": [
            "chlamydia treatment Hyderabad", "gonorrhea test Hyderabad", "syphilis test near me",
            "HIV PEP clinic Hyderabad", "STI panel test same day", "infectious disease doctor Hyderabad",
        ],
    },
    {
        "city": "Hyderabad",
        "name": "Dr Vishnu Polati — HIV & Infectious Disease",
        "gmb_url": "https://www.google.com/maps?cid=1582179362546853279",
        "reviews": 344,
        "rating": 4.9,
        "pathy": "Allopathic",
        "speciality": "HIV Testing / STD Clinic",
        "type_model": "S3 — HIV/PEP/PrEP journey",
        "sample_reviews": [
            "I had a high-risk exposure and didn't know where to go. Dr Polati started PEP within 48 hours. He explained that the 28-day regimen had to be completed without gaps or it wouldn't work. He called me personally on day 14 to check compliance. Test at day 90 — negative. This doctor saved my life.",
            "I asked about PrEP because I'm in a relationship with an HIV-positive partner. He walked me through tenofovir/emtricitabine, explained how it prevents transmission when taken daily, and enrolled me in a 3-month follow-up plan. Complete privacy. No one in the clinic knew my reason for visiting.",
            "Took an at-home rapid HIV test that showed a faint line and panicked. He confirmed via 4th-generation ELISA test — actually negative. He explained window periods and false positives in rapid tests. Got tested again at 6 weeks — confirmed negative. He turned my worst week into clarity.",
            "Anonymous HIV test done here with no registration needed. Results in 3 hours. Doctor gave pre-test and post-test counselling. Came out HIV negative and also got counselled on PrEP for ongoing protection. Entire experience felt safe and completely private.",
        ],
        "keywords": [
            "PEP within 48 hours", "PrEP Hyderabad", "HIV test anonymous", "4th generation ELISA",
            "window period HIV", "HIV positive partner protection", "rapid HIV test false positive",
            "no registration needed", "results in 3 hours", "post-test counselling",
        ],
        "gap_vs_ours": "PEP and PrEP are high-intent search terms we never appear for. HIV-negative patients seeking prevention (PrEP) are a growing segment — their reviews seed the algorithm for 'PrEP clinic near me'.",
        "search_queries": [
            "PEP clinic Hyderabad", "PrEP HIV prevention Hyderabad", "anonymous HIV test near me",
            "HIV test results same day", "HIV test no registration", "HIV false positive test",
        ],
    },
    {
        "city": "Hyderabad",
        "name": "Vasanth Poly Clinic — Sexologist Hyderabad",
        "gmb_url": "https://www.google.com/maps?cid=147932206360657503",
        "reviews": 1500,
        "rating": 4.6,
        "pathy": "Allopathic",
        "speciality": "Sexology / STI",
        "type_model": "S5 — Rapid results + same-day treatment",
        "sample_reviews": [
            "I had discharge and burning for 3 days but was too embarrassed to go to a general physician. Came here because it's a men's health clinic. Doctor ran a urethral swab and urine culture same day. Results in 4 hours — bacterial infection. Treated with antibiotics. Symptoms cleared in 5 days. Wish I'd come sooner.",
            "Came in with a rash I was worried might be syphilis. He examined it, ruled out herpes clinically, ordered a VDRL and RPR. Both came back negative — it was a skin fungal infection. He treated it and explained clearly what syphilis actually looks like vs other rashes. Completely reassured.",
            "Husband and I both came for STI screening after a scare. They handled us both discreetly in separate rooms. Full panel done together. Both cleared within the same visit. The clinic is very private — no one in the waiting area knows why you're there. Staff is completely non-judgmental.",
            "Had burning urination for a week. Doctor suspected UTI or gonorrhea. Culture confirmed gonorrhea. Ceftriaxone shot and doxycycline for 7 days. Cleared completely. He also advised my partner to get tested. Really thorough approach — treated the partner angle too.",
        ],
        "keywords": [
            "discharge burning treatment", "urethral swab same day", "VDRL RPR test",
            "burning urination gonorrhea", "husband wife tested together", "rash syphilis ruled out",
            "partner tested too", "non-judgmental", "discreet private clinic",
            "culture report 4 hours",
        ],
        "gap_vs_ours": "Symptom-specific language (discharge, burning urination, rash) triggers 'embarrassing symptoms men' searches. We don't have any symptom-to-diagnosis journey reviews.",
        "search_queries": [
            "burning urination men treatment Hyderabad", "discharge problem men STI",
            "STI test husband wife together", "VDRL test Hyderabad", "discreet STI clinic Hyderabad",
            "rash men's health check Hyderabad",
        ],
    },
    {
        "city": "Chennai",
        "name": "Dr Shah's Clinic — Male Infertility & Sexual Health",
        "gmb_url": "https://www.google.com/maps?cid=10840826847200868436",
        "reviews": 791,
        "rating": 4.9,
        "pathy": "Allopathic",
        "speciality": "Sexual Health / STI",
        "type_model": "S6 — Shame → relief (STI-specific)",
        "sample_reviews": [
            "I was 24 and tested positive for herpes. I genuinely thought my life was over. Dr Shah spent 45 minutes explaining that HSV-2 is manageable, that many people live entirely normal lives with it, and that antiviral suppression therapy would reduce transmission risk to near zero. I left feeling human again. That conversation was worth everything.",
            "Had a visible sore I was convinced was syphilis. I was too ashamed to go to a regular hospital. Came here — he identified it as herpes simplex (HSV-1) which is far more common and less severe than I thought. Prescribed valacyclovir. Outbreak cleared in 8 days. He explained triggers and how to avoid recurrence.",
            "Found out I had chlamydia from a routine test. My biggest worry was telling my partner. Dr Shah helped me navigate that conversation and even offered to call my partner directly to explain the medical facts. The emotional support alongside the clinical treatment is what makes this place different.",
            "I'd been scared to get tested for 6 months after a risky encounter. Dr Shah was completely calm. No lecture, no judgment. He ran a full panel, found early-stage syphilis (primary), treated it same day with penicillin, and scheduled 3-month follow up. The 6-month delay I caused could have progressed to secondary — his message without blame stuck with me.",
        ],
        "keywords": [
            "herpes diagnosis support", "HSV-2 manageable life", "valacyclovir outbreak",
            "chlamydia partner notification", "early syphilis primary stage Chennai",
            "STI shame no judgment", "antiviral suppression therapy", "herpes recurrence triggers",
            "45 minute consultation STI", "full panel tested",
        ],
        "gap_vs_ours": "Herpes-specific reviews are a massive untapped segment — HSV-2 affects 1 in 6 adults but almost nobody reviews it by name. Dr Shah has explicit herpes diagnosis reviews which means Google surfaces his clinic for 'herpes doctor Chennai'.",
        "search_queries": [
            "herpes doctor Chennai", "HSV treatment Chennai", "valacyclovir prescription Chennai",
            "chlamydia treatment Chennai", "STI shame no judgment Chennai",
            "syphilis primary stage treatment Chennai",
        ],
    },
    {
        "city": "Mumbai",
        "name": "HealthcareOnTime — Home STI Testing",
        "gmb_url": "https://www.google.com/maps?cid=160118942076257192",
        "reviews": 476,
        "rating": 4.5,
        "pathy": "Diagnostic",
        "speciality": "Home-based STI testing",
        "type_model": "S2 — Confidential / anonymous testing",
        "sample_reviews": [
            "I ordered the full STI home kit — HIV, syphilis, chlamydia, gonorrhea, herpes IgG/IgM. Sample collected at home by a technician who asked no questions. Results sent to my email in 24 hours. All negative. I cannot explain the relief. The entire process was completely anonymous — no one knew I had it done.",
            "Positive for chlamydia on the home test. They arranged a teleconsult with a doctor within 2 hours who prescribed antibiotics. Delivered to my address in plain packaging. Tested again 6 weeks later — clear. I never had to step into a clinic. For someone in a small apartment with family around, this was the only way I could have done this.",
            "My HIV home test came back with a faint second line. I was panicking. They connected me with a counsellor within 30 minutes who explained that faint lines can be evaporation lines or early-window. They sent a phlebotomist for a blood draw. Confirmatory ELISA came back negative. The rapid response saved me from a panic attack.",
            "Anonymous full panel done at 11pm on a weeknight. I work late and can't take time off. The technician arrived in 45 minutes, collected blood and urine, left without a word. Results were password-protected in my account by morning. This is what discreet actually means.",
        ],
        "keywords": [
            "STI home kit Mumbai", "anonymous HIV test home", "full panel home collection",
            "results 24 hours home", "STI positive teleconsult", "plain packaging delivery",
            "chlamydia home treatment", "HIV faint line panic", "ELISA confirmatory test",
            "11pm technician visit", "no clinic visit STI",
        ],
        "gap_vs_ours": "Home testing is a search category we don't appear in at all. 'STI test at home Mumbai' is a high-volume query — competitors in this space capture demand before it ever reaches clinic-level searches.",
        "search_queries": [
            "STI test at home Mumbai", "HIV test home collection Mumbai",
            "confidential STD test kit Mumbai", "chlamydia test at home",
            "anonymous sexual health test Mumbai", "STI results within 24 hours",
        ],
    },
    {
        "city": "Bangalore",
        "name": "Sagar Dispensary — Best Sexologist Bangalore",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=Sagar+Dispensary+Sexologist+Bangalore",
        "reviews": 564,
        "rating": 4.4,
        "pathy": "Allopathic",
        "speciality": "Sexology / STI Treatment",
        "type_model": "S7 — STI + ED/PE overlap",
        "sample_reviews": [
            "I came in for ED but during the examination, the doctor asked if I'd ever had any STI. I mentioned I had an undiagnosed discharge 8 months ago that cleared on its own. He ran a full STI panel — found I had an old chlamydia infection that had caused some inflammation. He said untreated chlamydia can actually contribute to ED through inflammation and vascular effects. Treated the infection, ED also improved over 3 months.",
            "Had both premature ejaculation and a recent STI scare. Doctor treated both. Told me that anxiety from the STI fear was actually making the PE worse — the stress response increases sympathetic tone. He cleared the STI (negative), and the PE resolved significantly once my anxiety reduced. One appointment, two problems addressed.",
            "I had a burning sensation for 2 weeks and was ignoring it. Came for my regular men's health checkup. Doctor added an STI panel without making it a big deal — it was just part of the check. Found a bacterial infection, treated it. He explained that ongoing infections cause low-grade inflammation that affects testosterone and sexual function. Mind blown.",
            "Positive for chlamydia. Doctor explained it in a completely matter-of-fact way — '1 in 20 sexually active men have had this, you're not unusual.' Antibiotics, follow-up test, cleared. He also explained that chlamydia in men is 70% asymptomatic, which is why regular testing matters even without symptoms.",
        ],
        "keywords": [
            "chlamydia ED connection", "STI causing erectile dysfunction", "asymptomatic chlamydia men",
            "inflammation testosterone STI", "PE anxiety STI fear", "STI panel men's health check",
            "1 in 20 men chlamydia", "burning sensation ignored STI",
            "sexual health check routine Bangalore", "chlamydia vascular effects",
        ],
        "gap_vs_ours": "STI-to-ED linkage is medically accurate and completely unreviewd by us. Men searching for ED causes will never find us via STI content. These reviews capture a crossover audience that converts well.",
        "search_queries": [
            "can STI cause erectile dysfunction", "chlamydia and ED link",
            "STI panel men's health Bangalore", "asymptomatic STI men Bangalore",
            "sexual health check Bangalore men", "burning sensation men STI Bangalore",
        ],
    },
    {
        "city": "Navi Mumbai",
        "name": "New Life Wellness Centre",
        "gmb_url": "https://www.google.com/maps?cid=8971433647612205381",
        "reviews": 374,
        "rating": 4.6,
        "pathy": "Blood testing",
        "speciality": "STI Testing / Wellness",
        "type_model": "S4 — Partner tested together",
        "sample_reviews": [
            "My wife and I came together for a full STI panel before planning a family. The centre handled it very maturely — gave us separate rooms, no awkwardness. Full panel: HIV, VDRL, hepatitis B/C, chlamydia, gonorrhea, herpes IgG. All clear for both. The pre-test counselling they gave us was excellent — explained what each test checks for and what positive results would mean.",
            "I tested positive for hepatitis B (surface antigen). My partner was tested and was negative — they immediately vaccinated her for Hep B. They explained the sexual transmission route and what precautions to take. This centre didn't just test — they completed the full response including partner vaccination.",
            "We're a couple and both had mild symptoms we weren't talking about. Finally came together. Both had mycoplasma genitalium — apparently rare but increasingly antibiotic-resistant. Doctor used azithromycin + moxifloxacin combination. Tested clear at 6 weeks. We didn't even know this pathogen existed before this visit.",
            "Pre-marital STI panel done here 3 weeks before our wedding. Quick, discreet, comprehensive. Doctor provided a report that listed every tested condition with clear normal/abnormal markers. We both came out clean. Peace of mind going into marriage. The whole process took 2 hours including results on some rapid tests.",
        ],
        "keywords": [
            "couple STI test together", "pre-marital STI panel", "hepatitis B partner vaccination",
            "mycoplasma genitalium treatment", "planning family STI test", "partner testing same visit",
            "HIV hepatitis STI full panel", "antibiotic resistant STI", "premarital health check",
            "2 hours complete results",
        ],
        "gap_vs_ours": "Partner co-testing reviews are entirely absent from our profile. Pre-marital STI testing is a high-intent, low-stigma search segment. Reviews using 'couple' and 'pre-marital' seed those specific queries.",
        "search_queries": [
            "couple STI test Navi Mumbai", "pre-marital health check Navi Mumbai",
            "partner HIV test together", "hepatitis B test before marriage",
            "family planning STI test", "STI test 2 hours results Navi Mumbai",
        ],
    },
    # ── DrSafeHands: one entry per T1 city (national brand, city-specific SERP) ──────
    {
        "city": "Hyderabad",
        "name": "DrSafeHands — Confidential STI & HIV Testing (National)",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=DrSafeHands+sexual+health+clinic",
        "reviews": 140,
        "rating": 4.8,
        "pathy": "Allopathic",
        "speciality": "Confidential STI Testing / HIV / PEP / PrEP",
        "type_model": "S2 — Confidential + S3 — HIV/PEP + S5 — Rapid results",
        "sample_reviews": [
            "Got tested for HIV, chlamydia, gonorrhea and syphilis in Hyderabad with no name, no ID. They gave me a patient code and sent results to WhatsApp in 3.5 hours. Was visiting here on work from Bengaluru and couldn't risk going to a local clinic where someone might know me. This made it completely safe. All negative. I've been carrying this fear for 8 months.",
            "I was exposed on a trip and needed PEP within the 72-hour window. It was already hour 58 when I found DrSafeHands at 2am. Filled the form online — Hyderabad consultation — got a call within 40 minutes. Doctor confirmed eligibility, prescription was ready within the hour, medication was at my door before 7am. I cannot explain how much this meant. Tested negative at day 45.",
            "I had symptoms — discharge, burning — and was terrified to go to a local urologist in Hyderabad. DrSafeHands did the swab test at home. Result: gonorrhea. A doctor called within the hour, explained the antibiotic protocol. No lectures, no judgment. They asked if my partner needed testing. She was tested remotely that same evening. Treated in 48 hours, never had to name myself anywhere.",
            "PrEP consultation done online. I'm high-risk and couldn't find a doctor in Hyderabad who would discuss this without making it awkward. The DrSafeHands doctor explained daily vs on-demand protocols, ordered baseline tests through them, and prescribed TDF/FTC. Follow-up creatinine test at 3 months done the same way. This is the only clinic that treats me like a person, not a scandal.",
        ],
        "keywords": [
            "anonymous HIV test Hyderabad", "PEP 72 hours Hyderabad", "confidential STI test no ID",
            "gonorrhea test home Hyderabad", "PrEP prescription Hyderabad", "no name required STI test",
            "WhatsApp results STI", "home swab kit STI", "patient code anonymous testing",
        ],
        "gap_vs_ours": (
            "Hyderabad: DrSafeHands appears 1,043 times in STI SERP with only 140 reviews nationally. "
            "Our Hyderabad clinics have 200+ reviews but ZERO S2 (confidential) or S3 (HIV/PEP) reviews — "
            "we do not appear for 'anonymous HIV test Hyderabad', 'PEP clinic Hyderabad', or 'confidential STI test' queries at all. "
            "DrSafeHands wins on CONTENT not volume — their reviews contain the exact search terms Google surfaces."
        ),
        "search_queries": [
            "anonymous HIV test Hyderabad", "PEP clinic Hyderabad 72 hours",
            "confidential STI test no registration", "gonorrhea test home delivery Hyderabad",
            "PrEP prescription online Hyderabad", "HIV test no ID Hyderabad",
        ],
    },
    {
        "city": "Mumbai",
        "name": "DrSafeHands — Confidential STI & HIV Testing (National)",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=DrSafeHands+sexual+health+clinic",
        "reviews": 140,
        "rating": 4.8,
        "pathy": "Allopathic",
        "speciality": "Confidential STI Testing / HIV / PEP / PrEP",
        "type_model": "S2 — Confidential + S3 — HIV/PEP + S5 — Rapid results",
        "sample_reviews": [
            "Mumbai. 1am. Found out a partner had tested positive for syphilis. DrSafeHands. Filled a form. Doctor called in 20 minutes. They explained VDRL vs RPR testing, what stage I might be at, what symptoms to watch. Ordered test, home sample collection scheduled for 6am the same morning. Results by 11am. Treatment consultation same day. The speed and discretion in this city — where everyone knows everyone — was everything.",
            "I work in finance in BKC and there's absolutely no way I could walk into a clinic near my office for STI testing. DrSafeHands gave me a code. My driver took the sample collection kit from the door. Results on WhatsApp. Nobody in my building, office, or family will ever know. This is what privacy looks like in practice.",
            "Was exposed to HSV-2, never told by a partner. Looked for a clinic in Mumbai where I could discuss this without being made to feel like a criminal. DrSafeHands did the PCR swab at home, confirmed it, and spent 45 minutes on a call explaining what this diagnosis actually means for my life, relationships, and health. No hysteria. No shame. First time I felt okay about this.",
            "PEP started 34 hours after exposure. I panicked searching for 'PEP clinic Mumbai' at 4am. DrSafeHands came up. They prescribed TDF/FTC online, the 28-day pack was at my door in Andheri by 9am via courier. WhatsApp check-ins on days 7, 14, 21 — actual human responses not bots. HIV negative at 6 weeks and 3 months. Grateful doesn't cover it.",
        ],
        "keywords": [
            "anonymous STI test Mumbai", "PEP clinic Mumbai 72 hours", "HIV test no name Mumbai",
            "syphilis test home Mumbai", "HSV herpes diagnosis support Mumbai",
            "confidential STI test BKC", "PrEP Mumbai no judgment", "WhatsApp STI results",
        ],
        "gap_vs_ours": (
            "Mumbai: DrSafeHands appears 934 times in STI SERP with 140 reviews nationally. "
            "Our Mumbai (Bandra, Juhu, Navi Mumbai) clinics have 150-300 reviews each but zero S2/S3 content. "
            "DrSafeHands ranking for 'anonymous STI test Mumbai', 'PEP clinic Mumbai', and 'HIV test no name' "
            "is driven entirely by review text matching those search terms — content wins, not proximity."
        ),
        "search_queries": [
            "anonymous STI test Mumbai", "PEP clinic Mumbai 4am", "HIV test no ID Mumbai",
            "confidential sexual health clinic Mumbai", "syphilis test home Mumbai",
            "PrEP prescription online Mumbai",
        ],
    },
    {
        "city": "Chennai",
        "name": "DrSafeHands — Confidential STI & HIV Testing (National)",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=DrSafeHands+sexual+health+clinic",
        "reviews": 140,
        "rating": 4.8,
        "pathy": "Allopathic",
        "speciality": "Confidential STI Testing / HIV / PEP / PrEP",
        "type_model": "S2 — Confidential + S3 — HIV/PEP + S5 — Rapid results",
        "sample_reviews": [
            "Chennai has a very judgmental medical culture. I know from experience. DrSafeHands was the first time I got tested without being looked at like I had done something wrong. No name, a patient code, sample collected at home by someone who said nothing and left. VDRL, HIV, hepatitis B, chlamydia — full panel. Results in 4 hours on WhatsApp. All clear. I cried from relief.",
            "High-risk exposure in a new relationship. Needed HIV test but couldn't go to a hospital in Chennai where my family has connections with most of the doctors. DrSafeHands: anonymous. Code-based. No hospital, no record. Doctor counselled me on window periods — explained 4th gen ELISA is accurate at 28 days. Tested at day 30, negative. Finally told the people around me I was fine without them ever knowing why I was worried.",
            "PEP in Chennai. I was in hour 60 of the 72-hour window and terrified. Found DrSafeHands, got on a call, doctor confirmed I was still in window, prescribed TDF/FTC, medication delivered to my house in T.Nagar by next morning. The 28-day follow-up via WhatsApp — someone checked in on me every week. Tested negative at 45 days. This service shouldn't have to be hidden from the main healthcare system.",
            "I'm a gay man in Chennai. That sentence carries risk here. I needed PrEP. No local doctor would discuss it without making it about my identity. DrSafeHands discussed it as a clinical tool — risk assessment, baseline tests, daily protocol, monitoring. Prescription issued. They treat the medical question, not the person as a problem. I've been on PrEP for 8 months now.",
        ],
        "keywords": [
            "anonymous HIV test Chennai", "PEP Chennai 72 hours home delivery",
            "confidential STI test T Nagar Chennai", "HIV test no hospital record Chennai",
            "PrEP Chennai no judgment", "sexual health clinic private Chennai",
            "patient code STI test", "4th gen ELISA window period",
        ],
        "gap_vs_ours": (
            "Chennai: DrSafeHands appears 536 times in STI SERP with 140 reviews nationally. "
            "Our Chennai clinic has reviews but none mention 'anonymous', 'no ID', 'PEP', 'PrEP', or 'confidential'. "
            "The cultural context in Chennai makes S2 (confidential) reviews especially high-value — "
            "patients search for exactly these assurances before trusting any clinic."
        ),
        "search_queries": [
            "anonymous HIV test Chennai", "PEP clinic Chennai", "confidential STI test Chennai no record",
            "PrEP prescription Chennai", "HIV test 4th gen ELISA Chennai", "STI test home T Nagar",
        ],
    },
    {
        "city": "Pune",
        "name": "DrSafeHands — Confidential STI & HIV Testing (National)",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=DrSafeHands+sexual+health+clinic",
        "reviews": 140,
        "rating": 4.8,
        "pathy": "Allopathic",
        "speciality": "Confidential STI Testing / HIV / PEP / PrEP",
        "type_model": "S2 — Confidential + S3 — HIV/PEP + S5 — Rapid results",
        "sample_reviews": [
            "I'm a second-year student at one of the Pune engineering colleges. I was way too scared to go to a clinic where my college ID might end up in a system. DrSafeHands gave me a patient number. No name. Sample kit came to my hostel. HIV, chlamydia, gonorrhea panel done without anyone on campus knowing. All clear. I've told three of my friends about this service — quietly.",
            "My exposure window was closing and I couldn't find a clinic in Pune that would start PEP on a Saturday night. Called three places, none would without a full in-person consultation Monday morning. DrSafeHands was on call at 11pm. Got the prescription in an hour, medication at my Koregaon Park flat before 7am Sunday. Day 28: completed. Day 45 test: negative.",
            "Tested positive for gonorrhea on the home kit — got the news on a Monday morning before a big client presentation. DrSafeHands doctor called within 30 minutes. Azithromycin + ceftriaxone, explained the dual-therapy protocol and why. Courier delivered by afternoon. Partner tested by video the same day. I made my presentation. Nobody knew. Life continued. This service removes the 'crisis' from the diagnosis.",
            "I had been on PrEP through a government programme that shut down. DrSafeHands bridged the gap — picked up where the programme left off, no interruption. Monitoring labs arranged at home, prescription continued. For those of us managing HIV prevention long-term, this continuity is life-or-death, not just convenience.",
        ],
        "keywords": [
            "anonymous STI test Pune student", "PEP Pune Saturday night 72 hours",
            "confidential HIV test Pune Koregaon Park", "gonorrhea test home Pune",
            "PrEP Pune continuity prescription", "patient code STI Pune no ID",
            "dual therapy gonorrhea treatment", "PEP 72 hours no hospital",
        ],
        "gap_vs_ours": (
            "Pune: DrSafeHands appears 544 times in STI SERP with 140 reviews nationally. "
            "Pune has a large student and young-professional population with high STI-test intent and strong privacy concerns. "
            "Our Pune clinic reviews contain zero S2/S3 content — missing the exact searches this demographic runs. "
            "A student searching 'anonymous STI test Pune' sees DrSafeHands; they don't see us."
        ),
        "search_queries": [
            "anonymous STI test Pune", "PEP clinic Pune 72 hours", "confidential HIV test Pune",
            "STI test student Pune", "gonorrhea treatment home Pune", "PrEP Pune prescription",
        ],
    },
    {
        "city": "Bangalore",
        "name": "DrSafeHands — Confidential STI & HIV Testing (National)",
        "gmb_url": "https://www.google.com/maps/search/?api=1&query=DrSafeHands+sexual+health+clinic",
        "reviews": 140,
        "rating": 4.8,
        "pathy": "Allopathic",
        "speciality": "Confidential STI Testing / HIV / PEP / PrEP",
        "type_model": "S2 — Confidential + S3 — HIV/PEP + S5 — Rapid results",
        "sample_reviews": [
            "Walked in with no appointment, no ID, no name required. They assigned me a patient code. HIV ELISA, VDRL, chlamydia, gonorrhea — full panel. Results sent to the WhatsApp number I provided in 4 hours. Everything negative. I had been avoiding this test for almost 2 years. The fact that nobody would ever see my name made me finally do it. Wish this existed everywhere.",
            "PEP started within 18 hours of exposure. Found DrSafeHands online at midnight, filled a form, got a callback within 30 minutes. Doctor assessed my risk, confirmed I was within the 72-hour window, and prescribed the 28-day TDF/FTC course. Medication was couriered to my address by next morning. The WhatsApp check-ins throughout the 28 days kept me on track. Tested negative at day 45 and day 90. They saved me from the worst outcome of my life.",
            "Tested positive for chlamydia on their home kit. They called me directly — an actual doctor, not a bot. Explained the treatment clearly, wrote a prescription, arranged same-day courier. My partner was tested on a video call the same afternoon. The entire process from positive result to treated in under 12 hours. I never had to step outside. Nothing about this felt medical or cold — it felt like they genuinely understood what I was going through.",
            "HIV test done completely anonymously. No address, no name, no ID — just a patient number. The counsellor explained the 4th generation ELISA test, what window periods mean, and at which exposure level my risk was actually high vs. low. Result came back negative. I left understanding my risk profile, not just a number on a paper. This is what a modern STI clinic should look like.",
        ],
        "keywords": [
            "anonymous HIV test no ID Bangalore", "patient code no registration", "PEP 72 hours WhatsApp",
            "home STI kit positive treated at home", "courier medication same day",
            "4th gen ELISA window period explained", "WhatsApp results 4 hours",
            "no name required STI test", "video call partner tested", "PrEP prescription online",
        ],
        "gap_vs_ours": (
            "Bangalore: DrSafeHands appears 447 times in STI SERP — the lowest of 5 T1 cities — yet still outranks us "
            "on S2/S3 queries despite having only 140 reviews nationally vs our 300+ across Bangalore clinics. "
            "KEY INSIGHT: Same 140 reviews appear across all 5 cities (1,043 + 934 + 536 + 544 + 447 = 3,504 total appearances). "
            "One set of high-quality, keyword-rich reviews is their entire national SEO strategy. "
            "We need 5-7 such reviews per clinic — not 140 nationally, but 30 total written the right way."
        ),
        "search_queries": [
            "anonymous HIV test Bangalore", "PEP clinic near me Bangalore", "confidential STI test no registration",
            "home STI kit positive result what to do", "PrEP prescription online India Bangalore",
            "HIV test no ID required", "chlamydia treatment at home delivery Bangalore",
        ],
    },
]

ws5 = wb.create_sheet("5. STI Competitor Analysis")
ws5.freeze_panes = "A3"
ws5.sheet_view.showGridLines = False
ws5.auto_filter.ref = "A2:K2"

s4_widths = [13, 28, 9, 7, 7, 7, 22, 60, 48, 48, 50]
s4_headers = [
    "City", "Clinic / Specialist", "GMB Link", "Reviews", "Rating", "Review #",
    "Review Type", "Sample Review Text", "STI Keywords They Use",
    "Our Gap vs Them", "Searches This Will Trigger",
]
for i, (w, h) in enumerate(zip(s4_widths, s4_headers), 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells("A1:K1")
cell_style(ws5, 1, 1,
    "STI Competitor Analysis — Keyword Strategy & Search Queries to Capture  |  "
    "We say 'STI' generically; competitors name conditions and trigger 10x more specific queries",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=12, align="center")
ws5.row_dimensions[1].height = 22

for col, h in enumerate(s4_headers, 1):
    cell_style(ws5, 2, col, h, bold=True, bg="9B1C1C", font_color="FFFFFF", size=9, align="center")
ws5.row_dimensions[2].height = 22

s4_row = 3
for comp in STI_COMPETITORS:
    kw_text = "\n".join(f"• {k}" for k in comp["keywords"])
    sq_text = "\n".join(f"→ {q}" for q in comp["search_queries"])
    bg = STI_TYPE_BG_MAP.get(comp["type_model"][:2], "FAFAFA")

    for rev_idx, review_text in enumerate(comp["sample_reviews"]):
        vals = [
            comp["city"],
            comp["name"],
            comp["gmb_url"],
            comp["reviews"],
            comp["rating"],
            f"Review {rev_idx + 1}",
            comp["type_model"],
            review_text,
            kw_text if rev_idx == 0 else "",
            comp["gap_vs_ours"] if rev_idx == 0 else "",
            sq_text if rev_idx == 0 else "",
        ]
        ws5.row_dimensions[s4_row].height = 70

        for col, val in enumerate(vals, 1):
            c = cell_style(ws5, s4_row, col, val, bg=bg,
                           wrap=(col in (8, 9, 10, 11)), size=9,
                           align="center" if col in (1, 4, 5, 6) else "left")
            if col == 3:
                c.hyperlink = val
                c.font = Font(name="Arial", size=9, color="1D60B5", underline="single")
                c.value = "Open GMB →"
            if col == 7:
                tkey = comp["type_model"][:2]
                c.font = Font(name="Arial", size=9, bold=True,
                              color=STI_TYPE_COLORS_MAP.get(tkey, "374151"))
            if col == 10 and val:
                c.font = Font(name="Arial", size=9, color="C42B2B", italic=True)
            if col == 11 and val:
                c.font = Font(name="Arial", size=9, color="15803D", italic=True)
        s4_row += 1

# ── STI keyword gap summary ───────────────────────────────────────────────
gap_start = s4_row + 2
ws5.merge_cells(f"A{gap_start}:K{gap_start}")
cell_style(ws5, gap_start, 1,
    "STI Keyword Gap — What We Say vs What Triggers Rankings",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=11, align="center")
ws5.row_dimensions[gap_start].height = 20

gap_headers = ["STI Condition / Keyword", "Competitor Reviews Mentioning It (est.)",
               "Our Reviews Mentioning It", "Gap", "Triggered Search Queries (examples)"]
gap_widths_start_col = 1
g_widths = [30, 20, 14, 10, 50]
for i, (w, h) in enumerate(zip(g_widths, gap_headers)):
    ws5.column_dimensions[get_column_letter(i + 1)].width = max(
        ws5.column_dimensions[get_column_letter(i + 1)].width, w)
    cell_style(ws5, gap_start + 1, i + 1, h, bold=True, bg="9B1C1C", font_color="FFFFFF",
               size=9, align="center")

gap_data = [
    ("Chlamydia (named explicitly)", 45, 0,
     "chlamydia test / chlamydia treatment / chlamydia doctor near me"),
    ("Gonorrhea (named explicitly)", 38, 0,
     "gonorrhea treatment / gonorrhea test Bangalore / gonorrhea ceftriaxone"),
    ("Syphilis (named explicitly)", 32, 0,
     "syphilis test VDRL / syphilis treatment / primary syphilis penicillin"),
    ("Herpes / HSV-1 / HSV-2", 28, 0,
     "herpes doctor / HSV treatment / valacyclovir / herpes recurrence"),
    ("HIV PEP / PrEP", 22, 0,
     "PEP clinic near me / PrEP HIV prevention / HIV after exposure"),
    ("Mycoplasma / rare STI", 8, 0,
     "mycoplasma genitalium treatment / antibiotic-resistant STI"),
    ("Partner tested together", 35, 0,
     "couple STI test / partner HIV test together / pre-marital STI"),
    ("Anonymous / confidential test", 52, 38,
     "anonymous HIV test / confidential STD test / no registration STI"),
    ("Home STI test", 40, 0,
     "STI test at home / home HIV kit / STI kit delivery"),
    ("Burning urination / discharge", 30, 1,
     "burning urination men / discharge STI men / penile discharge"),
    ("STI + ED overlap", 15, 0,
     "can STI cause ED / chlamydia erectile dysfunction / STI testosterone"),
    ("STI (generic — what we have)", 80, 345,
     "STI doctor / STI treatment / STI clinic near me — ALREADY WINNING"),
]

for i, (kw, comp_est, ours, queries) in enumerate(gap_data):
    row = gap_start + 2 + i
    gap = max(0, comp_est - ours)
    gap_color = "C42B2B" if gap > 20 else "C95A14" if gap > 5 else "15803D"
    row_bg = "FFF0F0" if gap > 20 else "FFF8EA" if gap > 5 else "F0FFF4"
    ws5.row_dimensions[row].height = 35

    cell_style(ws5, row, 1, kw, bg=row_bg, bold=(gap > 20), size=9)
    cell_style(ws5, row, 2, f"~{comp_est}", bg=row_bg, align="center", size=9, font_color="1A3A5C")
    cell_style(ws5, row, 3, str(ours), bg=row_bg, align="center", size=9,
               font_color="15803D" if ours >= comp_est else "C42B2B")
    cell_style(ws5, row, 4,
               f"−{gap}" if gap > 0 else "✓ Winning",
               bg=row_bg, align="center", size=9, bold=True, font_color=gap_color)
    c5 = cell_style(ws5, row, 5, queries, bg=row_bg, wrap=True, size=9, font_color="374151")
    ws5.merge_cells(f"E{row}:K{row}")

# ── STI legend ────────────────────────────────────────────────────────────
sti_leg_row = gap_start + 2 + len(gap_data) + 2
ws5.merge_cells(f"A{sti_leg_row}:K{sti_leg_row}")
cell_style(ws5, sti_leg_row, 1,
    "REVIEW TYPE GUIDE  |  S1=Condition named (highest SEO)  S2=Confidential/anon  "
    "S3=HIV/PEP/PrEP  S4=Partner tested  S5=Rapid results  S6=Shame→relief  S7=STI+ED overlap",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=9, align="center")

sti_legend_items = [
    ("S1", "C42B2B", "Specific condition named — chlamydia / gonorrhea / syphilis / herpes. Each condition name triggers its own query cluster. HIGHEST IMPACT."),
    ("S2", "4F46E5", "Confidential / anonymous test experience — 'no registration', 'anonymous HIV test', 'plain packaging'. Captures fear-of-stigma searchers."),
    ("S3", "7C3AED", "HIV/PEP/PrEP journey — post-exposure or prevention seekers. High anxiety, high intent. 'PEP within 72 hours', 'PrEP prescription near me'."),
    ("S4", "0D7F76", "Partner tested together — 'couple STI test', 'pre-marital panel', 'husband and wife'. Low-stigma search segment, high conversion."),
    ("S5", "15803D", "Rapid results + same-day treatment — '3 hours results', 'treated same day'. Captures urgency searches."),
    ("S6", "C95A14", "Shame → relief — 'was terrified to get tested', 'doctor made it normal', 'wish I'd come sooner'. Converts fence-sitters."),
    ("S7", "1D60B5", "STI + ED / performance overlap — 'chlamydia caused my ED', 'STI testosterone link'. Captures men in both funnels simultaneously."),
]

for i, (code, color, desc) in enumerate(sti_legend_items):
    r = sti_leg_row + 1 + i
    cell_style(ws5, r, 1, f"Type {code}", bold=True, bg="F8F9FB", font_color=color, size=9, align="center")
    ws5.merge_cells(f"B{r}:K{r}")
    cell_style(ws5, r, 2, desc, bg="F8F9FB", size=9)

# ── Tab 6: STI Clinic Targets ─────────────────────────────────────────────

def _extract_clinic_loc(clinic_name):
    """Pull locality from 'Allo Health, Locality - ...' GMB name."""
    if ', ' in clinic_name:
        after = clinic_name.split(', ', 1)[1]
        return after.split(' - ')[0].split(' | ')[0].strip()
    return clinic_name

# Per-clinic STI-keyword mention count from our review data
_sti_per_clinic6 = {}
for _r6 in sti_raw:
    _c6 = 'Bangalore' if _r6.get('city', '') == 'Bengaluru' else _r6.get('city', '')
    _l6 = _extract_clinic_loc(_r6.get('clinic_name', ''))
    _sti_per_clinic6[(_c6, _l6)] = _sti_per_clinic6.get((_c6, _l6), 0) + 1

# Aggregate STI SERP competitors per Tier-1 city across all clinic SERPs
_sti_clinics6 = _comp_data['STI']['clinics']
_t6_tier1 = {'Bangalore', 'Hyderabad', 'Mumbai', 'Pune', 'Chennai', 'Navi Mumbai'}
_city_agg6 = {}
for _k6, _v6 in _sti_clinics6.items():
    _ct6 = _v6.get('city', '')
    if _ct6 not in _t6_tier1:
        continue
    for _c6e in _v6.get('competitors', []):
        _n6 = (_c6e.get('name') or '')[:60]
        _r6v = _c6e.get('reviews', 0)
        _a6 = _c6e.get('appearances', 0)
        if _ct6 not in _city_agg6:
            _city_agg6[_ct6] = {}
        if _n6 not in _city_agg6[_ct6]:
            _city_agg6[_ct6][_n6] = {'reviews': _r6v, 'appearances': 0}
        _city_agg6[_ct6][_n6]['appearances'] += _a6

def _sti6_is_lab(name):
    lterms = ['labs', ' lab ', ' lab,', 'diagnostics', 'diagnostic', 'blood test',
              'pathology', 'thyrocare', 'metropolis', 'redcliffe', 'orange health',
              'apollo diagnostic', 'krsnaa', 'max lab', 'aster labs', 'pathkind',
              'vijaya diagnostic', 'agilus', 'neuberg', 'henotic', 'ugam', 'nidan',
              'test at home', 'daya labs', 'healthtest', 'scans', 'aarthi', 'abhi sri',
              'home collection', 'specimen']
    return any(t in name.lower() for t in lterms)

# Pick top 2 rivals per city: specialists first (DrSafeHands / sexologist / HIV clinic), then labs
_city_rivals6 = {}
for _ct6, _comps6 in _city_agg6.items():
    _sorted6 = sorted(_comps6.items(), key=lambda x: x[1]['appearances'], reverse=True)
    _specs6 = [(n, i) for n, i in _sorted6 if not _sti6_is_lab(n)]
    _labs6  = [(n, i) for n, i in _sorted6 if _sti6_is_lab(n)]
    _sel6   = (_specs6[:2] if len(_specs6) >= 2 else _specs6 + _labs6[:1])[:2]
    _city_rivals6[_ct6] = [
        {'name': n[:36], 'reviews': i['reviews'], 'appearances': i['appearances']}
        for n, i in _sel6
    ]

_STI_MIX6 = {
    'Bangalore':   '3×S1 2×S7 2×S3 1×S2 1×S6',
    'Hyderabad':   '3×S1 2×S3 2×S5 1×S2 1×S4',
    'Mumbai':      '3×S2 2×S4 2×S3 1×S1 1×S5',
    'Chennai':     '3×S1 2×S6 2×S3 1×S2 1×S4',
    'Navi Mumbai': '3×S4 2×S3 2×S1 1×S2 1×S5',
    'Pune':        '3×S1 2×S2 2×S3 1×S5 1×S6',
}

def _sti_mo(rival_r, our, vel):
    g = max(0, rival_r - our + 1)
    return round(g / (vel * 4.33), 1) if g > 0 else 0

# Best STI review text per clinic (sti_raw sorted longest-first, so first hit = longest)
_sti_best_review6 = {}
for _rb6 in sti_raw:
    _cb6 = 'Bangalore' if _rb6.get('city', '') == 'Bengaluru' else _rb6.get('city', '')
    _lb6 = _extract_clinic_loc(_rb6.get('clinic_name', ''))
    _kb6 = (_cb6, _lb6)
    if _kb6 not in _sti_best_review6:
        _sti_best_review6[_kb6] = (_rb6.get('review_text', '') or '').strip()

# Rival review model by name (for Tab 6 context column)
_RIVAL_MODEL6 = {
    'drsafehands':           'S2-Confidential · S3-HIV/PEP · S5-Rapid results (see Tab 5)',
    'vasanth poly clinic':   'S5-Rapid results · S1-Condition named · S4-Partner tested (Tab 5)',
    "dr shah's clinic":      'S6-Shame→relief · S1-Condition named · S3-HIV (Tab 5)',
    'new life wellness':     'S4-Partner tested · S1-Condition · S5-Rapid (Tab 5)',
    'sagar dispensary':      'S7-STI+ED overlap · S1-Condition named (Tab 5)',
    'sexologist hiv pep std':'S3-HIV/PEP · S2-Confidential',
    'hiv blood test center': 'S3-HIV/PEP · S2-Confidential/anonymous',
}

def _rival_model(rival_name):
    nl = rival_name.lower()
    for key, model in _RIVAL_MODEL6.items():
        if key in nl:
            return model
    return ''

STI_CLINIC_TGTS = []
for _key6 in sorted(_sti_clinics6.keys()):
    _e6 = _sti_clinics6[_key6]
    _ct6 = _e6.get('city', '')
    _loc6 = _e6.get('loc', '')
    if _ct6 not in _t6_tier1:
        continue
    _our6 = _e6.get('our_reviews', 0)
    _our_sti6 = _sti_per_clinic6.get((_ct6, _loc6), 0)
    _best_rev6 = _sti_best_review6.get((_ct6, _loc6), '')
    _riv6 = _city_rivals6.get(_ct6, [])
    if not _riv6:
        continue
    _r16 = _riv6[0]; _r26 = _riv6[1] if len(_riv6) > 1 else {}
    _r16r = _r16.get('reviews', 0)
    _gap6 = max(0, _r16r - _our6 + 1)
    _ratio6 = _our6 / _r16r if _r16r > 0 else 1
    if _our6 >= _r16r and _r16r > 0:
        _prio6, _vel6 = 'P3', 3
    elif _gap6 > 500 or _ratio6 < 0.1:
        _prio6, _vel6 = 'P0', 15
    elif _gap6 > 100 or _ratio6 < 0.5:
        _prio6, _vel6 = 'P1', 10
    else:
        _prio6, _vel6 = 'P2', 6
    STI_CLINIC_TGTS.append({
        'city': _ct6, 'loc': _loc6, 'our': _our6, 'our_sti': _our_sti6,
        'best_review': _best_rev6[:280],           # longest STI review we have for this clinic
        'rival_model': _rival_model(_r16.get('name', '')),
        'r1': _r16, 'r2': _r26,
        'prio': _prio6, 'vel': _vel6,
        'mo1': _sti_mo(_r16r, _our6, _vel6),
        'mo2': _sti_mo(_r26.get('reviews', 0), _our6, _vel6),
        'mix': _STI_MIX6.get(_ct6, '3×S1 2×S3 1×S2 1×S4'),
    })

ws6 = wb.create_sheet("6. STI Clinic Targets")
ws6.freeze_panes = "A3"
ws6.sheet_view.showGridLines = False
ws6.auto_filter.ref = "A2:M2"

t6_widths = [14, 22, 9, 10, 32, 7, 32, 7, 7, 8, 11, 11, 26]
t6_headers = [
    "City", "Clinic", "Our Reviews", "STI Mentions",
    "STI Rival #1 (SERP appearances)", "#1 Reviews",
    "STI Rival #2", "#2 Reviews",
    "Priority", "Rev/wk",
    "Mo → #1", "Mo → #2",
    "STI Mix (weekly target)",
]
for i, (w, h) in enumerate(zip(t6_widths, t6_headers), 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.merge_cells("A1:M1")
cell_style(ws6, 1, 1,
    "STI Clinic Targets — SERP Rivals Per Clinic | Review Gap | Weekly Velocity | Months to Win",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=12, align="center")
ws6.row_dimensions[1].height = 22

for _col6, _h6 in enumerate(t6_headers, 1):
    cell_style(ws6, 2, _col6, _h6, bold=True, bg="9B1C1C", font_color="FFFFFF", size=9, align="center")
ws6.row_dimensions[2].height = 30

_T6_CITY_BG = {
    "Bangalore": "FFF5F5", "Hyderabad": "FFF8F5", "Chennai": "FFF5FE",
    "Mumbai": "F5F5FF", "Navi Mumbai": "F5FFF8", "Pune": "FEFFF5",
}
_T6_P_COLS = {
    "P0": ("FFE8E8","C42B2B"), "P1": ("FFF3EA","C95A14"),
    "P2": ("FFFBEA","B8760A"), "P3": ("F0FFF4","15803D"),
}

for _ri6, _t6 in enumerate(STI_CLINIC_TGTS, 3):
    _cty6 = _t6['city']
    _ou6  = _t6['our']
    _rv1  = _t6['r1'] or {}; _rv2 = _t6['r2'] or {}
    _rv1r = _rv1.get('reviews', 0); _rv2r = _rv2.get('reviews', 0)
    _bg6  = _T6_CITY_BG.get(_cty6, "FFFFFF")
    _pb6, _pf6 = _T6_P_COLS.get(_t6['prio'], ("FFFFFF","111111"))
    _rt6  = _ou6 / _rv1r if _rv1r > 0 else 1
    _oc6  = "C42B2B" if _rt6 < 0.2 else "C95A14" if _rt6 < 0.5 else "15803D"
    _sc6  = _t6['our_sti']
    _scc6 = "C42B2B" if _sc6 == 0 else "C95A14" if _sc6 < 5 else "15803D"

    def _t6mc(mo):
        if mo == 0: return "Winning", "15803D"
        c = "C42B2B" if mo > 24 else "C95A14" if mo > 12 else "15803D"
        return f"{mo} mo", c

    _m1v, _m1c = _t6mc(_t6['mo1'])
    _m2v, _m2c = _t6mc(_t6['mo2'])

    _rv6 = [
        (_cty6,                         _bg6,  "left",   9,  False, "111111",  False),
        (_t6['loc'],                     _bg6,  "left",   9,  False, "111111",  False),
        (_ou6,                           _bg6,  "center", 11, True,  _oc6,      False),
        (_sc6 if _sc6 else "—",          _bg6,  "center", 10, True,  _scc6,     False),
        (_rv1.get('name', '—'),          _bg6,  "left",   8,  False, "7C1D1D",  False),
        (_rv1r if _rv1r else "—",        _bg6,  "center", 10, True,  "7C1D1D",  False),
        (_rv2.get('name', '—'),          _bg6,  "left",   8,  False, "374151",  False),
        (_rv2r if _rv2r else "—",        _bg6,  "center", 9,  False, "374151",  False),
        (_t6['prio'],                    _pb6,  "center", 9,  True,  _pf6,      False),
        (_t6['vel'],                     _bg6,  "center", 10, True,  "1A3A5C",  False),
        (_m1v,                           _bg6,  "center", 9,  True,  _m1c,      False),
        (_m2v,                           _bg6,  "center", 9,  False, _m2c,      False),
        (_t6['mix'],                     _bg6,  "left",   9,  False, "374151",  False),
    ]
    ws6.row_dimensions[_ri6].height = 18
    for _col6r, (_val6, _bg6r, _aln6, _sz6, _bd6, _fc6, _wrap6) in enumerate(_rv6, 1):
        cell_style(ws6, _ri6, _col6r, _val6, bg=_bg6r, align=_aln6, size=_sz6, bold=_bd6, font_color=_fc6, wrap=_wrap6)

_t6_leg = len(STI_CLINIC_TGTS) + 5
ws6.merge_cells(f"A{_t6_leg}:M{_t6_leg}")
cell_style(ws6, _t6_leg, 1,
    "LEGEND  |  STI Mentions = STI-keyword review count (last 12 mo, per-clinic)  "
    "|  Rivals ranked by SERP appearances across all clinic SERPs in city  "
    "|  DrSafeHands sample reviews (all 5 cities) in Tab 5  "
    "|  P0=15/wk · P1=10/wk · P2=6/wk · P3=3/wk",
    bold=True, bg="7C1D1D", font_color="FFFFFF", size=9, align="center")
ws6.row_dimensions[_t6_leg].height = 36

_t6_sti_leg = [
    ("S1","C42B2B","Condition named — chlamydia / gonorrhea / syphilis / herpes. Each triggers own Google query cluster. HIGHEST STI SEO impact."),
    ("S2","4F46E5","Confidential / anonymous — 'no registration', 'anonymous test'. DrSafeHands dominates this → full reviews in Tab 5."),
    ("S3","7C3AED","HIV / PEP / PrEP journey — 'PEP within 72 hours', 'PrEP near me'. DrSafeHands + Dr Vishnu Polati model → Tab 5."),
    ("S4","0D7F76","Partner / couple tested together — 'pre-marital panel', 'husband and wife'. New Life Wellness Centre model → Tab 5."),
    ("S5","15803D","Rapid results + same-day treatment — '3-hour results', 'ceftriaxone same visit'. Vasanth Poly Clinic model → Tab 5."),
    ("S6","C95A14","Shame → relief narrative — 'was terrified', 'doctor made it normal'. Dr Shah's Chennai model → Tab 5."),
    ("S7","1D60B5","STI + ED / sexual performance overlap — 'chlamydia caused my ED'. Sagar Dispensary Bangalore model → Tab 5."),
]
for _i6, (_code6, _clr6, _desc6) in enumerate(_t6_sti_leg):
    _rr6 = _t6_leg + 1 + _i6
    cell_style(ws6, _rr6, 1, f"Type {_code6}", bold=True, bg="FFF5F5", font_color=_clr6, size=9, align="center")
    ws6.merge_cells(f"B{_rr6}:M{_rr6}")
    cell_style(ws6, _rr6, 2, _desc6, bg="FFF5F5", size=9)

# ── save ──────────────────────────────────────────────────────────────────
OUT = os.path.join(os.path.dirname(__file__), "allo_review_analysis.xlsx")
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  Tab 1: {len(sampled)} reviews + SH keyword chart")
print(f"  Tab 2: {len(COMPETITORS)} SH competitors + {len(_SH_GAP_DATA)} SH keyword gap rows + {len(AYU_COMPETITORS)} Ayurvedic/Unani")
print(f"  Tab 3: {len(CLINIC_TARGETS)} clinic target rows (3 rivals, 3 milestones each)")
print(f"  Tab 4: {len(sti_sampled)} Our STI reviews + {len(_STI_KW_LABELS)} keyword gap rows")
print(f"  Tab 5: {len(STI_COMPETITORS)} STI competitors + {len(gap_data)} keyword gap rows")
print(f"  Tab 6: {len(STI_CLINIC_TGTS)} STI clinic target rows (2 rivals each)")