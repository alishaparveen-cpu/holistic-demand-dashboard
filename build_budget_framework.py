#!/usr/bin/env python3
"""
build_budget_framework.py
Google Ads budget & bid reallocation doc:
  Sheet 1 — Framework (decision tree, explained)
  Sheet 2 — Applied Decisions (per campaign, framework mapped to specific actions)
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(DIR, "allo_budget_framework.xlsx")

with open(os.path.join(DIR, "data_ga_campaigns.json")) as f:
    GA = json.load(f)
with open(os.path.join(DIR, "data_auction_insights.json")) as f:
    AI = json.load(f)

# ── Helpers ─────────────────────────────────────────────────────────────────
wb = Workbook()

def cs(ws, row, col, val, bg="FFFFFF", fc="111111", bold=False, size=9,
       align="left", wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.fill      = PatternFill("solid", fgColor=bg)

def merge(ws, r, c1, c2, val, **kw):
    ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
    cs(ws, r, c1, val, **kw)

def rh(ws, row, h): ws.row_dimensions[row].height = h
def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def v0(x):
    if isinstance(x, list): return x[0] if x else 0
    return x or 0

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Framework
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "1. Framework"
ws1.sheet_view.showGridLines = False
NC = 10
for i, w in enumerate([6,20,18,18,18,18,18,18,18,6], 1):
    cw(ws1, i, w)

r = 1

def title(ws, r, text, sub=""):
    merge(ws, r, 1, NC, text, bg="0F172A", fc="FFFFFF", bold=True, size=14, align="center")
    rh(ws, r, 28); r += 1
    if sub:
        merge(ws, r, 1, NC, sub, bg="1E293B", fc="94A3B8", size=9, italic=True, align="center")
        rh(ws, r, 14); r += 1
    return r

def section(ws, r, text, bg="1E3A5F"):
    merge(ws, r, 1, NC, text, bg=bg, fc="FFFFFF", bold=True, size=11)
    rh(ws, r, 20); r += 1
    return r

def node(ws, r, col_s, col_e, text, bg, fc="FFFFFF", sz=9, bold=True, wrap=True, ht=40):
    merge(ws, r, col_s, col_e, text, bg=bg, fc=fc, bold=bold, size=sz, align="left", wrap=wrap)
    rh(ws, r, ht); r += 1
    return r

def gap(ws, r, ht=8, bg="F8FAFC"):
    merge(ws, r, 1, NC, "", bg=bg); rh(ws, r, ht); r += 1
    return r

r = title(ws1, r,
    "Google Ads Budget & Bid Reallocation Framework",
    "Goals: IS% ≥ 70%  ·  loc% ≥ 70%  ·  QS ≥ 7  ·  util% 70–90%  ·  LP below-avg < 15%")

# ── Step 1 ───────────────────────────────────────────────────────────────────
r = section(ws1, r, "STEP 1 — Check Impression Share (IS%)  |  Goal: ≥ 70%")

merge(ws1, r, 1, 2, "IS% ≥ 70%?", bg="166534", fc="FFFFFF", bold=True, size=10, align="center")
merge(ws1, r, 3, 5, "✓  IS AT GOAL", bg="BBF7D0", fc="166534", bold=True, size=10, align="center")
merge(ws1, r, 6, NC, "→  Focus on loc% quality + QS maintenance. No budget action needed.", bg="F0FDF4", fc="166534", size=9, align="left", wrap=True)
rh(ws1, r, 32); r += 1

r = gap(ws1, r, 5)

merge(ws1, r, 1, 2, "IS% < 70%?", bg="C42B2B", fc="FFFFFF", bold=True, size=10, align="center")
merge(ws1, r, 3, NC, "  Diagnose the cause of IS loss: is it Budget Lost (BL%) or Rank Lost (RL%)?",
      bg="FEF2F2", fc="C42B2B", bold=True, size=9, align="left", wrap=True)
rh(ws1, r, 28); r += 1

r = gap(ws1, r, 5)

# BL branch
merge(ws1, r, 2, 3, "BL% > RL%\n(Budget-driven loss)", bg="1E3A5F", fc="FFFFFF", bold=True, size=9, align="center", wrap=True)
merge(ws1, r, 4, 6, "→  INCREASE BUDGET\nBudget is the binding constraint. More spend = more IS directly.", bg="EFF6FF", fc="1E3A5F", bold=False, size=9, align="left", wrap=True)
merge(ws1, r, 7, NC, "Check: is util% < 90%? If so, increase daily budget by the shortfall amount. "
      "If already at ceiling, raise the ceiling. BL > RL means bids are fine — money is the issue, not quality.",
      bg="F0F9FF", fc="374151", size=8, align="left", wrap=True)
rh(ws1, r, 44); r += 1

r = gap(ws1, r, 5)

# RL branch
merge(ws1, r, 2, 3, "RL% > BL%\n(Rank-driven loss)", bg="7C2D12", fc="FFFFFF", bold=True, size=9, align="center", wrap=True)
merge(ws1, r, 4, 6, "→  IMPROVE QUALITY SCORE\nRank loss = Google is choosing competitors over us in the auction.", bg="FFF7ED", fc="7C2D12", bold=False, size=9, align="left", wrap=True)
merge(ws1, r, 7, NC, "Check QS components:\n"
      "LP below avg (LP%) → Fix landing page experience (load speed, mobile, relevance, CTA).\n"
      "Ad Relevance below avg (AR%) → Rewrite RSA headlines to match keyword intent.\n"
      "Expected CTR below avg → Improve ad copy, test stronger CTAs.",
      bg="FFFBEB", fc="374151", size=8, align="left", wrap=True)
rh(ws1, r, 56); r += 1

r = gap(ws1, r, 5)

# QS too low to fix with bids
merge(ws1, r, 3, 4, "QS < 5 (critical)", bg="C42B2B", fc="FFFFFF", bold=True, size=9, align="center", wrap=True)
merge(ws1, r, 5, NC, "⛔  FREEZE BIDS — Do NOT increase bids until QS improves.\n"
      "Raising bids at QS < 5 is throwing money at a broken funnel. "
      "Google will still rank you below competitors because your ad/LP quality is poor. "
      "Fix LP first (LP% and AR% are your signals). Then reassess.",
      bg="FEF2F2", fc="C42B2B", bold=False, size=9, align="left", wrap=True)
rh(ws1, r, 54); r += 1

r = gap(ws1, r, 5)

merge(ws1, r, 3, 4, "QS 5-7 (moderate)\nLP/AR fixable", bg="A16207", fc="FFFFFF", bold=True, size=9, align="center", wrap=True)
merge(ws1, r, 5, NC, "⚠  FIX FIRST, then bid.\n"
      "Fix the identified QS weakness (LP or AR), wait 2 weeks for Google to reindex quality, "
      "then re-evaluate IS. If IS still < 70% after fix, consider bid increase. "
      "Do not increase bids while LP% or AR% are 'below avg' — wasted spend.",
      bg="FFFBEB", fc="A16207", bold=False, size=9, align="left", wrap=True)
rh(ws1, r, 54); r += 1

r = gap(ws1, r, 5)

merge(ws1, r, 3, 4, "QS ≥ 7 (good)\nLP/AR already OK", bg="166534", fc="FFFFFF", bold=True, size=9, align="center", wrap=True)
merge(ws1, r, 5, NC, "✓  QS is not the constraint. Consider BID ADJUSTMENT.\n"
      "Contingent on: (a) loc% ≥ 70% — is traffic quality good enough to warrant more bids? "
      "(b) loc → lead CVR — are location clicks converting to leads? "
      "If both good: raise bids. If loc% low: fix geography first. If CVR low: fix GMB/LP trust signals first.",
      bg="F0FDF4", fc="166534", bold=False, size=9, align="left", wrap=True)
rh(ws1, r, 54); r += 1

r = gap(ws1, r, 10)

# ── Step 2 ───────────────────────────────────────────────────────────────────
r = section(ws1, r, "STEP 2 — Check Location Click Rate (loc%)  |  Goal: ≥ 70%")

merge(ws1, r, 1, 2, "loc% ≥ 70%?", bg="166534", fc="FFFFFF", bold=True, size=10, align="center")
merge(ws1, r, 3, 5, "✓  LOCATION QUALITY OK", bg="BBF7D0", fc="166534", bold=True, size=10, align="center")
merge(ws1, r, 6, NC, "→  Traffic is qualifying as local intent. Proceed to competitor check (Step 3).", bg="F0FDF4", fc="166534", size=9, align="left", wrap=True)
rh(ws1, r, 28); r += 1

r = gap(ws1, r, 5)

merge(ws1, r, 1, 2, "loc% < 70%?", bg="A16207", fc="FFFFFF", bold=True, size=10, align="center")
merge(ws1, r, 3, NC, "  Traffic is diluted by non-local intent. Three fixes required:", bg="FFFBEB", fc="A16207", bold=True, size=9, align="left", wrap=True)
rh(ws1, r, 24); r += 1

fixes = [
    ("Keywords & geo cleanup",
     "Add negative keywords for generic/national searches. Add geo exclusions for areas you don't serve. "
     "Check for search terms that are pulling in non-local traffic (e.g. 'online consultation', generic health queries)."),
    ("Remove irrelevant sitelinks / headlines",
     "Sitelinks pointing to national content or brand pages dilute local intent. "
     "Remove any sitelink or headline that could attract non-local clicks. "
     "Every ad element should reinforce 'physical clinic, this city'."),
    ("+100% bid modifier for direct calls",
     "Call extensions with +100% bid adjustment ensures that any user about to call (highest local intent) "
     "sees us prominently. These are the highest-converting loc clicks — overpay for them."),
]
for title_f, body_f in fixes:
    merge(ws1, r, 2, 4, f"✦  {title_f}", bg="FED7AA", fc="A16207", bold=True, size=9, align="left")
    merge(ws1, r, 5, NC, body_f, bg="FFFBEB", fc="374151", size=8, align="left", wrap=True)
    rh(ws1, r, 40); r += 1
    r = gap(ws1, r, 4, "FFFBEB")

r = gap(ws1, r, 10)

# ── Step 3 ───────────────────────────────────────────────────────────────────
r = section(ws1, r, "STEP 3 — Check Paid Competitor Landscape (from Auction Insights)")

comp_branches = [
    ("Has local paid competitors\nOur organic is STRONG\n(GMB reviews dominant, map-pack #1)",
     "1E3A5F", "STANDARD BID",
     "Paid fills the gap organic can't reach. No need to overbid — our map-pack position already captures the majority of local intent. "
     "Bid enough to maintain IS% near goal. Overbidding here wastes money on traffic organic already gets you.",
     "BBF7D0", "166534"),
    ("Has local paid competitors\nOur organic is WEAK\n(low reviews, not in map-pack top 3)",
     "C42B2B", "INCREASE BID to maintain paid visibility",
     "Organic is not compensating. If we lose the paid auction too, we disappear entirely. "
     "Increase bids to stay in the auction. Track IS% vs specific competitor IS% — "
     "aim to match the top competitor's IS within 10%.",
     "FEF2F2", "C42B2B"),
    ("No local paid competitors\n(we are the only paid player)",
     "166534", "STANDARD BID (no need to overbid)",
     "No competitor to outbid. Standard bid is sufficient to win the auction. "
     "Focus budget on volume (more ad impressions) not on bid premium. "
     "Consider slightly lower bids to improve efficiency.",
     "F0FDF4", "166534"),
]

for cond, cb, dec, body, row_bg, row_fc in comp_branches:
    merge(ws1, r, 2, 4, cond, bg=cb, fc="FFFFFF", bold=True, size=9, align="center", wrap=True)
    merge(ws1, r, 5, 6, f"→  {dec}", bg=row_bg, fc=row_fc, bold=True, size=9, align="center", wrap=True)
    merge(ws1, r, 7, NC, body, bg="F8FAFC", fc="374151", size=8, align="left", wrap=True)
    rh(ws1, r, 52); r += 1
    r = gap(ws1, r, 5)

r = gap(ws1, r, 10)

# ── Step 4 ───────────────────────────────────────────────────────────────────
r = section(ws1, r, "STEP 4 — loc click → Lead CVR  |  Are location clicks converting?")

merge(ws1, r, 1, NC,
      "A location click = patient opened our GMB profile or clicked directions. "
      "If this is not converting to a lead (call / form), the issue is TRUST & CONTENT on the profile, not the ad.",
      bg="EFF6FF", fc="1E3A5F", size=9, bold=False, align="left", wrap=True)
rh(ws1, r, 28); r += 1

cvr_fixes = [
    ("GMB Reviews — condition-specific, outcome-driven",
     "Reviews that name conditions (PE, ED, STI, HIV) rank for those specific searches and convert "
     "searchers who find us via map-pack. Generic reviews ('doctor is nice') don't convert condition-seekers. "
     "See Chennai review strategy for template approach — apply same to all cities."),
    ("GMB Services — list every condition treated",
     "Add specific services: 'Premature Ejaculation Treatment', 'STI Testing & Treatment', 'HIV PEP Consultation', "
     "'Erectile Dysfunction Treatment'. Each listed service becomes a searchable entity on GMB."),
    ("Landing Page trust signals",
     "For paid loc clicks that land on our LP: add reviews widget, condition-specific content, "
     "visible phone number with click-to-call, clear pricing/consultation info. "
     "Every missing trust signal = a patient who bounces to a competitor."),
]
for title_c, body_c in cvr_fixes:
    merge(ws1, r, 2, 4, f"✦  {title_c}", bg="DDD6FE", fc="3B0764", bold=True, size=9, align="left")
    merge(ws1, r, 5, NC, body_c, bg="FAF5FF", fc="374151", size=8, align="left", wrap=True)
    rh(ws1, r, 40); r += 1
    r = gap(ws1, r, 4, "FAF5FF")

r = gap(ws1, r, 10)

# ── Summary ──────────────────────────────────────────────────────────────────
r = section(ws1, r, "DECISION SUMMARY — What to do in what order", "0F172A")

summary_rows = [
    ("1", "IS < 40%", "BL-driven", "—", "Increase budget immediately",            "INCREASE BUDGET", "FECACA","C42B2B"),
    ("2", "IS < 40%", "RL-driven", "QS < 5", "FREEZE. Fix LP, fix AR. No bid change.", "FIX FIRST (no spend)",  "FECACA","C42B2B"),
    ("3", "IS 40-70%", "RL-driven", "QS 5-7", "Fix LP/AR, wait 2wks, then raise bid",  "FIX THEN BID",  "FED7AA","A16207"),
    ("4", "IS 40-70%", "RL-driven", "QS ≥ 7", "Raise bids (contingent on loc% + CVR)",  "BID UP",        "FEF9C3","A16207"),
    ("5", "IS ≥ 70%",  "—",         "—",       "Maintain. Shift focus to loc% quality",  "MAINTAIN",      "BBF7D0","166534"),
    ("6", "loc% < 70%","—",         "—",       "Geo/keyword cleanup. +100% call bids",   "GEO CLEANUP",   "FED7AA","A16207"),
    ("7", "loc% ≥ 70%","No paid comps","—",    "Standard bid. Don't overbid.",           "STANDARD BID",  "BBF7D0","166534"),
    ("8", "loc% ≥ 70%","Comps, organic strong","—","Standard bid. Organic covers rest.", "STANDARD BID",  "BBF7D0","166534"),
    ("9", "loc% ≥ 70%","Comps, organic weak","—","Increase bid to maintain visibility.", "BID UP",        "FEF9C3","A16207"),
]

merge(ws1, r, 1, 1, "#",           bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center")
merge(ws1, r, 2, 3, "IS Situation",bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center")
merge(ws1, r, 4, 4, "Loss Type",   bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center")
merge(ws1, r, 5, 5, "QS Signal",   bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center")
merge(ws1, r, 6, 8, "Rule",        bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center")
merge(ws1, r, 9, NC,"Decision",    bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center")
rh(ws1, r, 20); r += 1

for num, is_sit, loss, qs, rule, dec, dec_bg, dec_fc in summary_rows:
    cs(ws1, r, 1, num,    bg="F1F5F9", fc="374151", bold=True, size=9, align="center")
    merge(ws1, r, 2, 3, is_sit,  bg="F8FAFC", fc="111111", bold=False, size=9, align="left")
    cs(ws1, r, 4, loss,   bg="F8FAFC", fc="374151", bold=False, size=8, align="left")
    cs(ws1, r, 5, qs,     bg="F8FAFC", fc="374151", bold=False, size=8, align="left")
    merge(ws1, r, 6, 8, rule, bg="F8FAFC", fc="111111", bold=False, size=8, align="left", wrap=True)
    merge(ws1, r, 9, NC, dec, bg=dec_bg, fc=dec_fc, bold=True, size=9, align="center")
    rh(ws1, r, 26); r += 1


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Applied Decisions (T1 + key T2 campaigns)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Applied Decisions")
ws2.sheet_view.showGridLines = False

# Column widths
col_widths = [28, 6, 6, 6, 6, 6, 6, 6, 6, 20, 14, 38, 36]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

NC2 = len(col_widths)

# Auction insights: per city/cat → top competitor IS
def top_comp_is(city, cat):
    city_data = AI.get("byCity", {}).get(city, {})
    cat_data  = city_data.get(cat, {})
    comps = cat_data.get("competitors", [])
    if not comps: return None, None, 0
    # exclude web-only platforms
    SKIP = {"practo.com","hexahealth.com","even.in","pristyncare.com","1mg.com","pharmeasy.in",
            "freehivtest.org.in","letsfighthpv.com","google.com"}
    clinic_comps = [c for c in comps if c.get("domain","") not in SKIP]
    if not clinic_comps: return None, None, 0
    def pct(s):
        if not s or s == "< 10%": return 5.0
        return float(str(s).replace("%","").strip())
    top = max(clinic_comps, key=lambda c: pct(c.get("is","0")))
    return top.get("domain",""), top.get("is","—"), pct(top.get("is","0"))

# Organic GMB strength per city (rough: our reviews vs top competitor)
ORGANIC_STRENGTH = {
    # city: (our_strongest_reviews, top_competitor_reviews, verdict)
    "Hyderabad": (320, 923, "WEAK"),   # our best clinic vs Kamaraj-equivalent
    "Bangalore": (388, 528, "MODERATE"),
    "Chennai":   (597, 923, "MODERATE"),  # Velachery 597 is decent
    "Mumbai":    (200, 338, "WEAK"),
    "Pune":      (200, 284, "WEAK"),
    "Navi Mumbai":(150, 300, "WEAK"),
    "Thane":     (100, 200, "WEAK"),
}

def organic_str(city):
    v = ORGANIC_STRENGTH.get(city, (100,500,"WEAK"))
    return v[2]

def campaign_city_cat(name):
    # parse city + cat from campaign name
    parts = name.split("_")
    cat_map = {"SH": "SH", "STD": "STI", "MH": "MH"}
    city = None; cat = None
    for p in parts:
        if p in cat_map: cat = cat_map[p]
        if p in ("Hyderabad","Bangalore","Chennai","Mumbai","Pune","Navi","Thane","Jaipur",
                 "Nagpur","Ahmedabad","Coimbatore","Surat","Ranchi","Vizag","Mangalore",
                 "Mysuru","Hubballi","Nashik","Bhopal","Amravati","Aurangabad","Vadodara",
                 "Gandhinagar","Jaipur","Bhopal","Nashik","Ranchi"):
            city = p
        if p == "Mumbai" and "Navi" in parts:
            city = "Navi Mumbai"
    return city, cat

# Decision logic
def classify(c, city, cat):
    is_   = float(v0(c.get("is", 0)) or 0)
    bl    = float(v0(c.get("bl", 0)) or 0)
    rl    = float(v0(c.get("rl", 0)) or 0)
    qs    = float(v0(c.get("qs", 0)) or 0)
    lp    = float(v0(c.get("lp", 0)) or 0)
    ar    = float(v0(c.get("ar", 0)) or 0)
    util  = float(c.get("util", 0) or 0)
    locpct= float(v0(c.get("locpct", 0)) or 0)
    bud   = float(c.get("bud", 0) or 0)

    top_dom, top_is_str, top_is = top_comp_is(city or "", cat or "SH")
    organic = organic_str(city or "")

    # Step 1: IS check
    if is_ >= 70:
        is_status = "✓ AT GOAL"
        is_bg = "BBF7D0"; is_fc = "166534"
    elif is_ >= 40:
        is_status = "⚠ BELOW GOAL"
        is_bg = "FEF9C3"; is_fc = "A16207"
    else:
        is_status = "✗ CRITICAL"
        is_bg = "FECACA"; is_fc = "C42B2B"

    # Loss type
    if bl > rl:
        loss_type = "Budget-lost"
        loss_bg = "BFDBFE"
    elif rl > 0:
        loss_type = "Rank-lost"
        loss_bg = "FED7AA"
    else:
        loss_type = "—"
        loss_bg = "F1F5F9"

    # QS signal
    if qs >= 7:
        qs_bg = "BBF7D0"; qs_fc = "166534"
    elif qs >= 5:
        qs_bg = "FEF9C3"; qs_fc = "A16207"
    else:
        qs_bg = "FECACA"; qs_fc = "C42B2B"

    # loc% signal
    if locpct >= 70:
        loc_bg = "BBF7D0"; loc_fc = "166534"
    elif locpct >= 50:
        loc_bg = "FEF9C3"; loc_fc = "A16207"
    else:
        loc_bg = "FECACA"; loc_fc = "C42B2B"

    # util signal
    if util > 95:
        util_bg = "FED7AA"; util_fc = "A16207"  # at ceiling
    elif util < 50:
        util_bg = "BFDBFE"; util_fc = "1E3A5F"  # underutilized
    else:
        util_bg = "F0FDF4"; util_fc = "166534"

    # ── Decision + Action ────────────────────────────────────────────────────
    decision = ""
    action   = ""
    dec_bg   = "F8FAFC"
    dec_fc   = "374151"

    fix_lp = lp >= 30
    fix_ar = ar >= 20
    fix_loc = locpct < 70

    quality_issues = []
    if fix_lp: quality_issues.append(f"LP: {lp:.0f}%↓")
    if fix_ar: quality_issues.append(f"AR: {ar:.0f}%↓")
    q_str = " + ".join(quality_issues) if quality_issues else "None"

    if is_ >= 70:
        decision = "MAINTAIN"
        action   = f"IS at goal. Monitor loc% ({locpct:.0f}% vs 70% goal). "
        if fix_loc: action += "Geography cleanup needed for loc%. "
        if fix_lp:  action += f"LP still below avg ({lp:.0f}%) — fix for quality retention. "
        dec_bg = "BBF7D0"; dec_fc = "166534"

    elif bl > rl:  # budget-driven
        decision = "INCREASE BUDGET"
        action   = (f"BL={bl:.0f}% > RL={rl:.0f}%: Budget is the binding constraint. "
                    f"Increase daily budget from ₹{bud:,.0f}. "
                    f"util={util:.0f}% — room to absorb more. "
                    f"QS={qs:.1f} is {'OK' if qs >= 6 else 'needs attention'}, "
                    f"so extra budget will convert to IS.")
        if fix_loc: action += f" Also fix loc% ({locpct:.0f}% < 70%): geo cleanup."
        dec_bg = "BFDBFE"; dec_fc = "1E3A5F"

    elif qs < 5:  # rank-lost, QS critical
        decision = "FREEZE — FIX QUALITY FIRST"
        action   = (f"QS={qs:.1f} is critical. RL={rl:.0f}% = rank loss is the issue. "
                    f"Raising bids at QS={qs:.1f} wastes money — Google ranks us below competitors regardless. "
                    f"Fix: {q_str}. No bid change until QS ≥ 6.")
        dec_bg = "FECACA"; dec_fc = "C42B2B"

    elif qs < 7 and (fix_lp or fix_ar):  # rank-lost, QS moderate, fixable
        decision = "FIX FIRST → THEN BID"
        action   = (f"RL={rl:.0f}% rank-driven. QS={qs:.1f} — fixable quality issues: {q_str}. "
                    f"Fix LP/AR first (2-week reindex), then re-evaluate IS. "
                    f"If IS still < 70% after fix: raise bid +10-15%.")
        if top_is > 0:
            action += f" Top paid competitor: {top_dom} at {top_is_str} IS."
        if fix_loc:
            action += f" Also: geo cleanup for loc% ({locpct:.0f}%)."
        dec_bg = "FED7AA"; dec_fc = "A16207"

    else:  # rank-lost, QS good — bid is the lever
        # Check competitor + organic
        if top_is > 0 and organic == "WEAK":
            decision = "INCREASE BID"
            action   = (f"QS={qs:.1f} is good (not the issue). RL={rl:.0f}% rank-driven. "
                        f"Paid competitor {top_dom} at {top_is_str} IS. Our organic is {organic}. "
                        f"Increase bids +15-25% to close IS gap. "
                        f"util={util:.0f}% — {'headroom available' if util < 80 else 'near ceiling, raise budget too'}.")
            dec_bg = "FED7AA"; dec_fc = "A16207"
        elif top_is > 0 and organic in ("MODERATE", "STRONG"):
            decision = "STANDARD BID"
            action   = (f"QS={qs:.1f} good. RL={rl:.0f}% present but organic ({organic}) compensates. "
                        f"Competitor {top_dom} at {top_is_str} IS. Standard bid — organic fills the gap paid leaves. "
                        f"Monitor: if organic deteriorates, raise bid.")
            dec_bg = "F0FDF4"; dec_fc = "166534"
        else:
            decision = "STANDARD BID"
            action   = (f"QS={qs:.1f} good, no strong local paid competitor. Standard bid. "
                        f"util={util:.0f}%.")
            dec_bg = "F0FDF4"; dec_fc = "166534"

        if fix_loc:
            action += f" Geo cleanup needed (loc%={locpct:.0f}% < 70%)."

        # Special: util very low → increase budget regardless
        if util < 50 and is_ < 70:
            decision = "INCREASE BUDGET (underutilised)"
            action = (f"util={util:.0f}% — we're not spending our own budget! "
                      f"IS={is_:.0f}% is low but we have budget headroom. "
                      f"QS={qs:.1f} OK. "
                      f"Raise daily budget from ₹{bud:,.0f} or check geo/keyword constraints limiting delivery.")
            dec_bg = "BFDBFE"; dec_fc = "1E3A5F"

    # Loc% fix note
    loc_action = ""
    if fix_loc:
        loc_action = f"Geo/keyword cleanup (loc%={locpct:.0f}%). +100% call bid modifier."

    return {
        "is_": is_, "is_status": is_status, "is_bg": is_bg, "is_fc": is_fc,
        "loss_type": loss_type, "loss_bg": loss_bg,
        "bl": bl, "rl": rl,
        "qs": qs, "qs_bg": qs_bg, "qs_fc": qs_fc,
        "lp": lp, "ar": ar, "util": util, "util_bg": util_bg, "util_fc": util_fc,
        "locpct": locpct, "loc_bg": loc_bg, "loc_fc": loc_fc,
        "top_dom": top_dom or "—", "top_is_str": top_is_str or "—",
        "organic": organic,
        "q_str": q_str,
        "decision": decision, "action": action,
        "dec_bg": dec_bg, "dec_fc": dec_fc,
        "loc_action": loc_action,
    }

# ── Headers ──────────────────────────────────────────────────────────────────
r2 = 1
merge(ws2, r2, 1, NC2,
      "Google Ads Budget & Bid Reallocation — Framework Applied Per Campaign",
      bg="0F172A", fc="FFFFFF", bold=True, size=13, align="center")
rh(ws2, r2, 24); r2 += 1

merge(ws2, r2, 1, NC2,
      "Source: Google Ads live data (data_ga_campaigns.json)  |  Auction Insights (data_auction_insights.json)  "
      "|  Decision = framework from Sheet 1 applied to each campaign's current metrics",
      bg="1E293B", fc="94A3B8", size=8, italic=True, align="center")
rh(ws2, r2, 13); r2 += 1

merge(ws2, r2, 1, NC2,
      "IS goal=70% · loc% goal=70% · QS goal≥7 · util target 70-90%  "
      "|  IS colour: 🟢≥70% · 🟡40-70% · 🔴<40%  "
      "|  QS colour: 🟢≥7 · 🟡5-7 · 🔴<5  "
      "|  loc%: 🟢≥70% · 🟡50-70% · 🔴<50%  "
      "|  Decision: BLUE=budget, ORANGE=fix first, RED=freeze, GREEN=maintain",
      bg="F8FAFC", fc="374151", size=8, align="center")
rh(ws2, r2, 13); r2 += 1

HDRS = [
    "Campaign", "IS%", "BL%", "RL%", "QS", "LP↓%", "AR↓%",
    "util%", "loc%", "Top Paid Rival", "Organic\nStrength",
    "Decision", "Specific Action"
]
for col, h in enumerate(HDRS, 1):
    cs(ws2, r2, col, h, bg="1E293B", fc="FFFFFF", bold=True, size=8, align="center", wrap=True)
rh(ws2, r2, 30); r2 += 1
ws2.freeze_panes = f"A{r2}"

# ── Campaign rows ─────────────────────────────────────────────────────────────
# Filter: T1 Exact-Local campaigns + notable others
FOCUS_CAMPAIGNS = [
    c for c in GA["campaigns"]
    if ("Exact_Local" in c["n"] or "Exact-Local" in c["n"])
    and any(city in c["n"] for city in
            ["Hyderabad","Bangalore","Chennai","Mumbai","Navi_Mumbai","Thane","Pune"])
]
# Also include key non-exact-local that matter
ALSO_INCLUDE = ["T1_Bangalore_SH_Exact", "T1_Hyderabad_SH_Exact", "T1_Pune_SH_Exact",
                "T1_Mumbai_SH_Exact", "T1_Chennai_SH_Exact",
                "T1_Bangalore_SH_Phrase_Local", "ONL_LT_SH_HighIntent_South",
                "ONL_LT_SH_HighIntent_Hindi"]
for c in GA["campaigns"]:
    if c["n"] in ALSO_INCLUDE:
        FOCUS_CAMPAIGNS.append(c)

# Sort: T1 Exact-Local first (by cat then city), then others
def sort_key(c):
    n = c["n"]
    tier = 0 if "T1" in n else 1
    cat_order = {"SH": 0, "STD": 1, "MH": 2}
    cat = next((v for k, v in cat_order.items() if f"_{k}_" in n), 9)
    city_order = {"Bangalore":0,"Chennai":1,"Hyderabad":2,"Mumbai":3,"Navi":4,"Thane":5,"Pune":6}
    city = next((v for k, v in city_order.items() if k in n), 9)
    return (tier, cat, city)

FOCUS_CAMPAIGNS.sort(key=sort_key)

# Category section breaks
prev_group = None

for c in FOCUS_CAMPAIGNS:
    name = c["n"]
    city, cat = campaign_city_cat(name)
    cl = classify(c, city, cat)

    # Group header
    group = "SH Exact-Local" if "SH_Exact_Local" in name else \
            "STI (STD) Exact-Local" if "STD_Exact_Local" in name else \
            "MH Exact-Local" if "MH_Exact_Local" in name else \
            "SH Exact (non-local)" if "SH_Exact" in name and "Local" not in name else \
            "SH Phrase-Local" if "SH_Phrase" in name else \
            "Online / National"

    if group != prev_group:
        merge(ws2, r2, 1, NC2, f"  {group}", bg={
            "SH Exact-Local": "1E3A5F",
            "STI (STD) Exact-Local": "7C2D12",
            "MH Exact-Local": "3B0764",
            "SH Exact (non-local)": "334155",
            "SH Phrase-Local": "374151",
            "Online / National": "18181B",
        }.get(group,"1E293B"), fc="FFFFFF", bold=True, size=10)
        rh(ws2, r2, 18); r2 += 1
        prev_group = group

    row_bg = {"SH Exact-Local":"EFF6FF","STI (STD) Exact-Local":"FFF7ED",
              "MH Exact-Local":"FAF5FF"}.get(group,"F8FAFC")

    # clean campaign name display
    display_name = (name.replace("T1_","").replace("T2_","").replace("_Exact_Local","")
                    .replace("_Exact","").replace("_Phrase_Local","").replace("_"," "))

    vals = [
        (display_name,         row_bg,        "left",   8,  True,  "111111",      False),
        (f"{cl['is_']:.0f}%",  cl["is_bg"],   "center", 10, True,  cl["is_fc"],   False),
        (f"{cl['bl']:.0f}%",   cl["loss_bg"] if cl["bl"]>cl["rl"] else row_bg, "center",8,False,"374151",False),
        (f"{cl['rl']:.0f}%",   cl["loss_bg"] if cl["rl"]>cl["bl"] else row_bg, "center",8,False,"374151",False),
        (f"{cl['qs']:.1f}",    cl["qs_bg"],   "center", 10, True,  cl["qs_fc"],   False),
        (f"{cl['lp']:.0f}%",   "FECACA" if cl["lp"]>=30 else "FEF9C3" if cl["lp"]>=15 else row_bg,
                                               "center",  8, False, "374151",      False),
        (f"{cl['ar']:.0f}%",   "FED7AA" if cl["ar"]>=20 else row_bg,
                                               "center",  8, False, "374151",      False),
        (f"{cl['util']:.0f}%", cl["util_bg"], "center",  9, False, cl["util_fc"], False),
        (f"{cl['locpct']:.0f}%",cl["loc_bg"], "center",  9, True,  cl["loc_fc"],  False),
        (f"{cl['top_dom']}\n{cl['top_is_str']} IS", row_bg,"left",7,False,"374151",True),
        (cl["organic"],        "F0FDF4" if cl["organic"]=="STRONG" else "FEF9C3" if cl["organic"]=="MODERATE" else "FFF7ED",
                                               "center",  8, True,
                                               "166534" if cl["organic"]=="STRONG" else "A16207" if cl["organic"]=="MODERATE" else "C42B2B",
                                               False),
        (cl["decision"],       cl["dec_bg"],  "left",    8, True,  cl["dec_fc"],  True),
        (cl["action"],         "FAFAFA",      "left",    8, False, "111111",      True),
    ]

    rh(ws2, r2, 64)
    for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(vals, 1):
        cs(ws2, r2, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)
    r2 += 1

# ── Budget reallocation summary table ────────────────────────────────────────
r2 += 1
merge(ws2, r2, 1, NC2, "REALLOCATION SUMMARY — Where Budget Moves", bg="0F172A", fc="FFFFFF", bold=True, size=11)
rh(ws2, r2, 20); r2 += 1

REALLOC = [
    ("INCREASE BUDGET", "BFDBFE","1E3A5F",
     "NMB SH Exact-Local (util=45%, QS=7.1) · Thane SH Exact-Local (util=43%, IS=69% — just below goal)"),
    ("INCREASE BID (+15-25%)", "FED7AA","A16207",
     "MUM STD Exact-Local (DrSafeHands IS=60%, posAbove=97%, our IS=29%) · "
     "HYD STD Exact-Local (IS=38%, DrSafeHands IS=53%, util=60% headroom) · "
     "BLR STD Exact-Local (DrSafeHands IS=64%, our IS=42%) — AFTER LP fix"),
    ("FIX LP/AR FIRST (no spend change)", "FEF9C3","A16207",
     "BLR SH Exact-Local (AR=26%, RL=49%) · CHN SH Exact-Local (LP=53%, RL=42%) · "
     "MUM SH Exact-Local (LP=42%, loc%=43%) · HYD SH Exact-Local (LP=50%, RL=41%)"),
    ("FREEZE — FIX QUALITY BEFORE TOUCHING", "FECACA","C42B2B",
     "BLR MH Exact-Local (QS=3.8, LP=62%↓ — DO NOT RAISE BIDS. Fix LP first.) · "
     "ONL SH Hindi (QS=3.8, LP=90%↓ — functionally broken) · "
     "CHN SH Exact non-local (util=6%, LP=94%↓ — dead campaign, LP not working at all)"),
    ("GEO CLEANUP + CALL BID", "EDE9FE","3B0764",
     "CHN STD (loc%=32%) · MUM SH (loc%=43%) · HYD STD (loc%=43%) · CHN SH (loc%=49%) · "
     "All campaigns with loc% < 50% → negative geo, remove non-clinic sitelinks, +100% call bid modifier"),
    ("MAINTAIN — STANDARD BID", "F0FDF4","166534",
     "PUNE SH Exact-Local (IS=66%, close to goal) · CHN SH (organic moderate, compensates) · "
     "NMB SH after budget increase · Thane SH after budget increase"),
]

for dec_label, bg, fc, camps_list in REALLOC:
    merge(ws2, r2, 1, 3, dec_label, bg=bg, fc=fc, bold=True, size=9, align="left")
    merge(ws2, r2, 4, NC2, camps_list, bg="F8FAFC", fc="111111", size=8, wrap=True)
    rh(ws2, r2, 48); r2 += 1
    merge(ws2, r2, 1, NC2, "", bg="F1F5F9"); rh(ws2, r2, 5); r2 += 1

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  Sheet 1: Framework (decision tree + summary)")
print(f"  Sheet 2: {len(FOCUS_CAMPAIGNS)} campaigns with applied decisions + reallocation summary")
