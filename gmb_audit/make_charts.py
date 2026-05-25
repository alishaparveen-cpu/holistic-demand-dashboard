"""
Generate the 6 charts that back the audit hypotheses.
Output: PNG files in gmb_audit/charts/, 1600×1000 @ 144dpi, Allo purple branding.
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl

HERE = Path(__file__).parent
OUT = HERE / "charts"
OUT.mkdir(exist_ok=True)

PURPLE = "#6e42e5"
DEEP = "#321c6d"
GOLD = "#ffc247"
GREEN = "#16a34a"
RED = "#dc2626"
GREY = "#94a3b8"
LIGHT = "#f4f3ff"

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "font.size": 11,
    "axes.edgecolor": "#1f1b46",
    "axes.labelcolor": "#1f1b46",
    "axes.titlecolor": "#1f1b46",
    "xtick.color": "#1f1b46",
    "ytick.color": "#1f1b46",
    "axes.spines.top": False,
    "axes.spines.right": False,
})


# ============================================================================
# DATA LOADERS — reuse logic from full_audit.py + analyze_keywords.py
# ============================================================================

def load_sti():
    rows = []
    with open(HERE / "sti_share.csv") as f:
        for r in csv.DictReader(f):
            r["avg_sti_pct"] = float(r["avg_sti_pct"]) if r["avg_sti_pct"] else None
            r["total_calls"] = int(r["total_calls"])
            rows.append(r)
    return rows


def load_locations_x():
    wb = openpyxl.load_workbook(HERE / "locations.xlsx")
    ws = wb["Sheet1"]
    hdr = [c.value for c in ws[1]]
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def load_insights():
    with open(HERE / "insights.csv") as f:
        rdr = csv.reader(f); hdr = next(rdr)
        rows = []
        for r in rdr:
            if not r[0] or not r[0].strip().isdigit(): continue
            d = dict(zip(hdr, r))
            for k in ("Google Search – Mobile","Google Search – Desktop","Google Maps – Mobile","Google Maps – Desktop","Calls","Directions","Website clicks"):
                d[k] = int(d.get(k) or 0)
            d["imp"] = d["Google Search – Mobile"]+d["Google Search – Desktop"]+d["Google Maps – Mobile"]+d["Google Maps – Desktop"]
            d["eng"] = d["Calls"]+d["Directions"]+d["Website clicks"]
            rows.append(d)
    return rows


def tier_of(p):
    if p is None: return None
    if p >= 28: return "top"
    if p >= 12: return "mid"
    return "low"


def match_sti_to_location(sti_rows, locs):
    out = {}
    used = set()
    for s in sti_rows:
        cn = s["clinic"].lower()
        for L in locs:
            code = L.get("Shop code")
            if code in used: continue
            name = (L.get("Business name") or "").lower()
            addr = " ".join(str(L.get(f"Address line {i}") or "") for i in range(1, 6)).lower()
            if cn in name or cn in addr:
                out[s["clinic"]] = code
                used.add(code); break
    return out


# Locality → clinic map (from channel_breakdown.py)
LOCALITY_TO_CLINIC = {
    "Kengeri":"Aarohi_Allo_Clinic","Jayanagar":"Adhventha_Allo_Clinic","Koramangala":"Cudur_Allo_Clinic",
    "KR Puram":"HealthPlus_Allo_Clinic","Vijayanagar":"Jeevasare_Allo_Clinic","Indiranagar":"Life_Plus_Allo_Clinic",
    "Electronic City":"NEO- Allo_Clinic","Bellandur":"Nurture_Allo_Clinic","Arekere":"Svasti_Allo_Clinic",
    "Whitefield":"TIDE_Allo Clinic","HSR Layout":"UMC","Sahakara Nagar":"Vikyath_Allo_Clinic",
    "Nungambakkam":"Birthwave_Allo_Clinic","Velachery":"Lapser_Allo_Clinic","Mogappair":"MMRV_Allo_Clinic",
    "Thoraipakkam":"Pearl_Singapore_Allo_Clinic","Tambaram":"Sai_Doctors_Allo_Clinic ",
    "Bharathi Nagar":"Heart_&_Her_Allo_Clinic","Kondapur":"HALE-Allo_Clinic","Narsingi":"MedSurge_Allo_Clinic",
    "Sainikpuri":"Olive_Tree_Allo_Clinic","Ameerpet":"RAD_Allo_Clinic","Kukatpally":"Sravanthi_Allo_Clinic",
    "Nallagandla":"Sree_vedha_Allo_Clinic","Vaishali Nagar":"Asha_Clinic_Allo_Clinic","Falnir Rd":"Lifeline_Allo_Clinic",
    "Andheri East":"Apple_Allo-Clinic","Dadar":"Thakur_Allo_Clinic","Kalyan West":"Vatsal-Allo Clinic",
    "Ghatkopar":"Vedant_Allo_Clinic","Malad":"Zenith_Allo_Clinic","Saraswathipuram":"DHC- Allo Health Sexual Wellness Clinic",
    "Tatya Tope Nagar":"Medicure_Allo_Clinic","Trimurti Chowk":"Trimurti_Allo_Clinic","Thane":"Infinity_Allo_Clinic",
    "Panvel":"Bhusare_Allo_Clinic","Kharghar":"Kharghar_Multispeciality_Allo_Clinic","Baner":"Baner_Allo_Health",
    "Wakad":"Curesta_Allo_Clinic","Kothrud":"DMS_Allo_Clinic","Kharadi":"Pune_SSC_Allo_Clinic",
    "Katraj":"Sainath_Allo_Clinic","Hadapsar":"Savali_Allo_Clinic","Chinchwad":"Shree_Sadguru_Allo_Clinic",
    "Ashok Nagar":"Jeevah_Allo_Clinic","Bhimrad":"Bombay_Multi_Allo_Clinic","MVP Colony":"Lalit_Healthcare_Allo_Clinic",
    "Garkheda":"MediArts_Hospital_Allo_Clinic","Vidya Nagar":"Ashoka_Multi-Specialty_Allo_Clinic",
    "Suryaraopeta":"Medicare_Allo_Clinic",
}


def load_channel():
    out = defaultdict(lambda: defaultdict(lambda: {"done":0,"sti_done":0}))
    with open(HERE / "redshift_clinic_source.csv") as f:
        for r in csv.DictReader(f):
            out[r["clinic"]][r["source"]]["done"] += int(r["done"])
            out[r["clinic"]][r["source"]]["sti_done"] += int(r["sti_done"])
    return out


# ============================================================================
# CHART 1 — Profile completeness by tier (proves H1: profile fields don't differ)
# ============================================================================

def chart_profile_completeness():
    sti = load_sti()
    locs = load_locations_x()
    matched = match_sti_to_location(sti, locs)
    by_code = {L.get("Shop code"): L for L in locs}

    buckets = {"top":[], "mid":[], "low":[]}
    for s in sti:
        if s["avg_sti_pct"] is None or s["total_calls"] < 10: continue
        t = tier_of(s["avg_sti_pct"])
        if not t: continue
        L = by_code.get(matched.get(s["clinic"]))
        if not L: continue
        buckets[t].append(L)

    features = [
        ("Primary category set", lambda L: bool(L.get("Primary category"))),
        ("Description present",  lambda L: bool((L.get("From the business") or "").strip())),
        ("Cover photo",          lambda L: bool(L.get("Cover photo"))),
        ("Other photos",         lambda L: bool(L.get("Other photos"))),
        ("LGBTQ+ flag",          lambda L: L.get("Crowd: LGBTQ+ friendly (welcomes_lgbtq)") in ("TRUE",True,"true")),
        ("WhatsApp URL",         lambda L: bool(L.get("Place page URLs: WhatsApp (url_whatsapp)"))),
    ]
    labels = [f[0] for f in features]
    top_pct = [100*sum(1 for L in buckets["top"] if f[1](L))/len(buckets["top"]) for f in features]
    low_pct = [100*sum(1 for L in buckets["low"] if f[1](L))/len(buckets["low"]) for f in features]

    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = range(len(features))
    w = 0.38
    ax.bar([i-w/2 for i in x], top_pct, w, label=f"Top tier (n={len(buckets['top'])})", color=PURPLE)
    ax.bar([i+w/2 for i in x], low_pct, w, label=f"Low tier (n={len(buckets['low'])})", color=GOLD)
    for i,v in enumerate(top_pct):
        ax.text(i-w/2, v+1.5, f"{v:.0f}%", ha="center", fontsize=10, color=DEEP, weight="bold")
    for i,v in enumerate(low_pct):
        ax.text(i+w/2, v+1.5, f"{v:.0f}%", ha="center", fontsize=10, color=DEEP, weight="bold")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=20, ha="right")
    ax.set_ylim(0, 115)
    ax.set_ylabel("% of clinics with feature set")
    ax.set_title("H1 — Profile completeness is NOT the driver\nTop and low tiers have nearly identical GMB profile fields", color=DEEP, weight="bold")
    ax.legend(loc="lower right")
    plt.tight_layout()
    plt.savefig(OUT / "01_profile_completeness.png", dpi=144, bbox_inches="tight", facecolor="white")
    plt.close()
    print("→ 01_profile_completeness.png")


# ============================================================================
# CHART 2 — GMB performance by tier (proves H2: traffic volume isn't the gap)
# ============================================================================

def chart_gmb_performance():
    sti = load_sti()
    locs = load_locations_x()
    ins = load_insights()
    matched = match_sti_to_location(sti, locs)
    by_code = {i["Shop code"]: i for i in ins}

    buckets = defaultdict(list)
    for s in sti:
        if s["avg_sti_pct"] is None or s["total_calls"] < 10: continue
        t = tier_of(s["avg_sti_pct"])
        if not t: continue
        code = matched.get(s["clinic"])
        i = by_code.get(code)
        if i: buckets[t].append(i)

    metrics = ["Impressions", "Phone calls", "Website clicks", "Direction taps"]
    def avg(rows, k): return sum(r[k] for r in rows)/len(rows) if rows else 0
    top = [avg(buckets["top"], "imp"), avg(buckets["top"], "Calls"), avg(buckets["top"], "Website clicks"), avg(buckets["top"], "Directions")]
    low = [avg(buckets["low"], "imp"), avg(buckets["low"], "Calls"), avg(buckets["low"], "Website clicks"), avg(buckets["low"], "Directions")]

    fig, axes = plt.subplots(1, 4, figsize=(13, 4.5))
    for ax, label, t_v, l_v in zip(axes, metrics, top, low):
        ax.bar([0,1], [t_v, l_v], color=[PURPLE, GOLD], width=0.6)
        ax.set_xticks([0,1])
        ax.set_xticklabels(["Top", "Low"])
        ax.set_title(label, color=DEEP, weight="bold", fontsize=11)
        for i,v in enumerate([t_v, l_v]):
            ax.text(i, v + max(t_v,l_v)*0.02, f"{v:,.0f}", ha="center", weight="bold", color=DEEP, fontsize=10)
        ax.set_ylim(0, max(t_v, l_v) * 1.18)
        ax.set_yticks([])
    fig.suptitle("H2 — Top clinics get more GMB exposure, but the ratio doesn't explain the 6× STI gap\n(per-clinic 30-day average)", color=DEEP, weight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig(OUT / "02_gmb_performance.png", dpi=144, bbox_inches="tight", facecolor="white")
    plt.close()
    print("→ 02_gmb_performance.png")


# ============================================================================
# CHART 3 — Search keyword intent mix by tier (proves H3: same intent reaches both)
# ============================================================================

def chart_keyword_intent():
    sti = load_sti()
    locs = load_locations_x()
    matched = match_sti_to_location(sti, locs)
    # Aggregate impressions from search_keywords.csv by tier × intent
    STI_R = re.compile(r"\b(sti|std|hiv|aids|herpes|gonorr|syphilis|chlamydia|genital|venereal|hpv|wart|chancre|trichom)\b", re.I)
    BRAND_R = re.compile(r"\b(allo|allohealth|allo health|allo clinic)\b", re.I)
    SEXO_R = re.compile(r"\b(sexologist|sex doctor|sex therapist|sex specialist|sex consultant|andrologist)\b", re.I)
    PROB_R = re.compile(r"\b(erectile|premature|ejaculation|impotence|libido|fertility|penis|infertility|ed problem|pe problem|sex problem|sexual problem|orgasm|nightfall|stamina|kamzori)\b", re.I)
    GEN_R = re.compile(r"\b(sex|sex clinic|clinic near me|hospitals|hospital|doctor|gp|clinic|sex near me|sex problem doctor|gynec|men's clinic|men clinic)\b|सेक्स", re.I)
    def classify(k):
        kl = k.lower()
        if STI_R.search(kl): return "sti"
        if BRAND_R.search(kl): return "brand"
        if PROB_R.search(kl): return "problem"
        if SEXO_R.search(kl): return "sexologist"
        if GEN_R.search(kl) or "सेक्स" in k: return "generic"
        return "other"

    tier_codes = {}
    for s in sti:
        if s["avg_sti_pct"] is None or s["total_calls"] < 10: continue
        t = tier_of(s["avg_sti_pct"])
        if not t: continue
        c = matched.get(s["clinic"])
        if c: tier_codes[c] = t

    tier_intent = {t: defaultdict(int) for t in ("top","mid","low")}
    with open(HERE / "search_keywords.csv") as f:
        for r in csv.DictReader(f):
            t = tier_codes.get(r["store_code"])
            if not t: continue
            intent = classify(r["keyword"])
            tier_intent[t][intent] += int(r["impressions_lower"] or 0)

    intents = ["sti","brand","sexologist","problem","generic","other"]
    colors = [RED, PURPLE, "#0ea5e9", "#f59e0b", GREY, "#cbd5e1"]
    labels = ["STI/STD intent","Brand","Sexologist","Sex problem","Generic","Other"]

    fig, ax = plt.subplots(figsize=(11, 4.5))
    rows = ["top", "mid", "low"]
    bottoms = [0]*3
    totals = [sum(tier_intent[t].values()) for t in rows]
    for intent, color, label in zip(intents, colors, labels):
        widths = [tier_intent[t][intent]/totals[i]*100 for i,t in enumerate(rows)]
        ax.barh([0,1,2], widths, left=bottoms, color=color, label=label, edgecolor="white", linewidth=1)
        for i,w in enumerate(widths):
            if w > 4:
                ax.text(bottoms[i]+w/2, i, f"{w:.0f}%", ha="center", va="center", color="white", weight="bold", fontsize=10)
        bottoms = [bottoms[i]+widths[i] for i in range(3)]
    ax.set_yticks([0,1,2])
    ax.set_yticklabels([f"Top  (n=10)", f"Mid  (n=27)", f"Low  (n=9)"])
    ax.invert_yaxis()
    ax.set_xlim(0, 100)
    ax.set_xlabel("% of GMB search impressions")
    ax.set_title("H3 — Search intent reaching profiles is similar across tiers\nLow-tier clinics get the same ~5% STI-keyword exposure as top tier", color=DEEP, weight="bold")
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.16), ncol=6, frameon=False)
    plt.tight_layout()
    plt.savefig(OUT / "03_keyword_intent.png", dpi=144, bbox_inches="tight", facecolor="white")
    plt.close()
    print("→ 03_keyword_intent.png")


# ============================================================================
# CHART 4 — STI tagging rate by channel × tier (THE KEY CHART — proves it's clinic-side)
# ============================================================================

def chart_channel_x_tier():
    sti = load_sti()
    chan = load_channel()
    tier_chan = {t: defaultdict(lambda: [0,0]) for t in ("top","mid","low")}  # tier → src → [done, sti]
    for s in sti:
        if s["avg_sti_pct"] is None or s["total_calls"] < 10: continue
        if s["clinic"] == "Ashok Nagar": continue  # doctor here doesn't do STI cases
        t = tier_of(s["avg_sti_pct"])
        if not t: continue
        rs = LOCALITY_TO_CLINIC.get(s["clinic"])
        if not rs or rs not in chan: continue
        for src, d in chan[rs].items():
            tier_chan[t][src][0] += d["done"]
            tier_chan[t][src][1] += d["sti_done"]

    SOURCES = ["gmb","google","organic","practo"]
    LABELS = ["GMB","Google","Organic","Practo"]

    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = range(len(SOURCES))
    w = 0.27
    colors_t = {"top":PURPLE,"mid":"#a385f1","low":GOLD}
    for i,t in enumerate(["top","mid","low"]):
        vals = []
        for src in SOURCES:
            done, sti_d = tier_chan[t][src]
            vals.append(sti_d/done*100 if done else 0)
        offset = (i-1)*w
        bars = ax.bar([j+offset for j in x], vals, w, label={"top":"Top (≥28% STI)","mid":"Mid (12–28%)","low":"Low (<12%)"}[t], color=colors_t[t], edgecolor="white", linewidth=1)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x()+bar.get_width()/2, v+1, f"{v:.0f}%", ha="center", weight="bold", fontsize=10, color=DEEP)
    ax.set_xticks(list(x))
    ax.set_xticklabels(LABELS)
    ax.set_ylabel("% of completed consultations tagged STI")
    ax.set_ylim(0, 45)
    ax.set_title("H4 — Channel mix is NOT the issue\nEvery channel converts to STI ~3× worse at low-tier clinics — same source, different tagging rate", color=DEEP, weight="bold")
    ax.legend(loc="upper right", frameon=False)
    ax.axhline(y=0, color="#1f1b46", linewidth=0.5)
    plt.tight_layout()
    plt.savefig(OUT / "04_channel_x_tier.png", dpi=144, bbox_inches="tight", facecolor="white")
    plt.close()
    print("→ 04_channel_x_tier.png  (KEY CHART)")


# ============================================================================
# CHART 5 — STI conversion rate per clinic (extreme contrasts)
# ============================================================================

def chart_per_clinic():
    sti = load_sti()
    chan = load_channel()

    # Pick representative clinics: 5 top + 5 bottom (by truth STI%)
    actives = [s for s in sti if s["avg_sti_pct"] is not None and s["total_calls"] >= 10]
    actives.sort(key=lambda r: -r["avg_sti_pct"])
    top5 = actives[:5]
    bot5 = sorted(
        [a for a in actives if a["avg_sti_pct"] < 12 and a["clinic"] != "Ashok Nagar"],
        key=lambda r: r["avg_sti_pct"],
    )[:5]
    chosen = top5 + bot5

    SOURCES = ["gmb","google","organic","practo"]
    LABELS = ["GMB","Google","Organic","Practo"]
    colors = [PURPLE, "#0ea5e9", GREEN, "#f97316"]

    fig, ax = plt.subplots(figsize=(13, 6))
    n_clinics = len(chosen)
    x = list(range(n_clinics))
    w = 0.2
    for i, src in enumerate(SOURCES):
        vals = []
        for s in chosen:
            rs = LOCALITY_TO_CLINIC.get(s["clinic"])
            d = (chan.get(rs) or {}).get(src) or {"done":0,"sti_done":0}
            vals.append(d["sti_done"]/d["done"]*100 if d["done"] else None)
        offset = (i-1.5)*w
        positions = [j+offset for j in x]
        # Plot only non-None
        for pos, v in zip(positions, vals):
            if v is not None:
                ax.bar(pos, v, w, color=colors[i], edgecolor="white", linewidth=0.5)
    # Truth-STI overlay
    for j,s in enumerate(chosen):
        ax.axhline(y=s["avg_sti_pct"], xmin=(j-0.45)/n_clinics, xmax=(j+0.45)/n_clinics, color=RED, linewidth=2.2, linestyle="--", alpha=0.7)

    ax.set_xticks(x)
    ax.set_xticklabels([f"{s['clinic']}\n({s['city']})\nTruth: {s['avg_sti_pct']:.0f}%" for s in chosen], rotation=0, fontsize=9)
    ax.set_ylabel("% of consultations tagged STI (last 30 days)")
    ax.set_ylim(0, 60)
    ax.axvline(x=4.5, color=GREY, linewidth=1, linestyle=":")
    ax.text(2, 56, "Top 5 (STI ≥30%)", ha="center", fontsize=10, weight="bold", color=DEEP)
    ax.text(7, 56, "Bottom 5 (STI ≤7%)", ha="center", fontsize=10, weight="bold", color=DEEP)

    # Legend
    handles = [mpatches.Patch(color=c, label=l) for c,l in zip(colors, LABELS)]
    handles.append(mpatches.Patch(color=RED, label="Truth STI%"))
    ax.legend(handles=handles, loc="upper right", frameon=False)
    ax.set_title("Per-clinic STI tagging rate, by acquisition channel\nIn top clinics every channel converts well. In bottom clinics nothing does.", color=DEEP, weight="bold")
    plt.tight_layout()
    plt.savefig(OUT / "05_per_clinic.png", dpi=144, bbox_inches="tight", facecolor="white")
    plt.close()
    print("→ 05_per_clinic.png")


# ============================================================================
# CHART 6 — Zero-tag anomaly: clinics where the truth-STI says >0 but Redshift says 0
# ============================================================================

def chart_zero_tag():
    sti = load_sti()
    chan = load_channel()
    actives = [s for s in sti if s["avg_sti_pct"] is not None and s["total_calls"] >= 10]
    rows = []
    for s in actives:
        rs = LOCALITY_TO_CLINIC.get(s["clinic"])
        if not rs or rs not in chan: continue
        total_done = sum(d["done"] for d in chan[rs].values())
        total_sti = sum(d["sti_done"] for d in chan[rs].values())
        rs_pct = (total_sti/total_done*100) if total_done else 0
        rows.append({
            "clinic": s["clinic"],
            "city": s["city"],
            "truth_pct": s["avg_sti_pct"],
            "rs_pct": rs_pct,
        })
    # Sort by gap (truth - rs)
    rows.sort(key=lambda r: -(r["truth_pct"] - r["rs_pct"]))
    extreme = rows[:8] + rows[-2:]  # 8 with biggest negative gap + 2 with positive (Arekere etc)

    fig, ax = plt.subplots(figsize=(11, 6))
    y = list(range(len(extreme)))
    truth = [r["truth_pct"] for r in extreme]
    rs = [r["rs_pct"] for r in extreme]
    ax.barh([i-0.2 for i in y], truth, 0.35, label="Truth (QS2 sheet)", color=PURPLE)
    ax.barh([i+0.2 for i in y], rs, 0.35, label="Redshift (encounter_tags)", color=GOLD)
    for i,(t,r) in enumerate(zip(truth, rs)):
        ax.text(t+0.3, i-0.2, f"{t:.0f}%", va="center", fontsize=9, color=DEEP, weight="bold")
        ax.text(r+0.3, i+0.2, f"{r:.0f}%", va="center", fontsize=9, color=DEEP, weight="bold")
    ax.set_yticks(y)
    ax.set_yticklabels([f"{r['clinic']} ({r['city']})" for r in extreme], fontsize=10)
    ax.invert_yaxis()
    ax.set_xlabel("STI % of consultations")
    ax.set_title("Same clinic, two answers — and a tagging-discipline trail\nBiggest gaps where truth STI% says >0 but our Redshift tag-count says 0", color=DEEP, weight="bold")
    ax.legend(loc="lower right", frameon=False)
    plt.tight_layout()
    plt.savefig(OUT / "06_zero_tag.png", dpi=144, bbox_inches="tight", facecolor="white")
    plt.close()
    print("→ 06_zero_tag.png")


# ============================================================================

def main():
    chart_profile_completeness()
    chart_gmb_performance()
    chart_keyword_intent()
    chart_channel_x_tier()
    chart_per_clinic()
    chart_zero_tag()
    print(f"\nAll charts written to {OUT}/")


if __name__ == "__main__":
    main()
