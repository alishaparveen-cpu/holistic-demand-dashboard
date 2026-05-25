"""
Analyze GBP search-keyword data by STI tier.

For each clinic in the STI list, joins to the matching GBP location, classifies
each keyword by intent, and computes per-tier intent share.

Intent buckets:
  sti          — explicit STI/STD/HIV/herpes/genital ulcer terms
  brand        — allo / allo health / allo clinic queries
  sexologist   — sexologist / sex doctor / sex specialist queries
  problem      — sex problem / erectile / premature / ED / PE / impotence
  generic      — generic "sex"/"clinic near me" with no condition signal
  other        — everything else
"""

from __future__ import annotations

import csv
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent

STI_TERMS = re.compile(r"\b(sti|std|hiv|aids|herpes|gonorr|syphilis|chlamydia|genital|venereal|hpv|wart|chancre|trichom)\b", re.I)
BRAND_TERMS = re.compile(r"\b(allo|allohealth|allo health|allo clinic)\b", re.I)
SEXOLOGIST_TERMS = re.compile(r"\b(sexologist|sex doctor|sex therapist|sex specialist|sex consultant|andrologist)\b", re.I)
PROBLEM_TERMS = re.compile(r"\b(erectile|premature|ejaculation|impotence|libido|fertility|penis|infertility|टाइम|टाईम|समस्या|ed problem|pe problem|sex problem|sexual problem|orgasm|nightfall|swapnadosh|शीघ्र|stamina|hard|kamzori|कमजोर)\b", re.I)
GENERIC_TERMS = re.compile(r"\b(sex|sex clinic|clinic near me|hospitals|hospital|doctor|gp|डॉक्टर|clinic|sex near me|sex problem doctor|gynec|men's clinic|men clinic)\b", re.I)


def classify(kw: str) -> str:
    k = kw.lower()
    if STI_TERMS.search(k): return "sti"
    if BRAND_TERMS.search(k): return "brand"
    if PROBLEM_TERMS.search(k): return "problem"
    if SEXOLOGIST_TERMS.search(k): return "sexologist"
    if GENERIC_TERMS.search(k): return "generic"
    return "other"


def load_sti():
    rows = []
    with open(HERE / "sti_share.csv") as f:
        for r in csv.DictReader(f):
            r["avg_sti_pct"] = float(r["avg_sti_pct"]) if r["avg_sti_pct"] else None
            r["total_calls"] = int(r["total_calls"])
            r["sti_calls"] = int(r["sti_calls"])
            rows.append(r)
    return rows


def load_keywords():
    rows = []
    with open(HERE / "search_keywords.csv") as f:
        for r in csv.DictReader(f):
            r["impressions_lower"] = int(r["impressions_lower"] or 0)
            r["intent"] = classify(r["keyword"])
            rows.append(r)
    return rows


def load_locations_x():
    wb = openpyxl.load_workbook(HERE / "locations.xlsx")
    ws = wb["Sheet1"]
    hdr = [c.value for c in ws[1]]
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def tier(p):
    if p is None: return None
    if p >= 28: return "top"
    if p >= 12: return "mid"
    return "low"


def main():
    sti = load_sti()
    kws = load_keywords()
    locs = load_locations_x()

    # Match STI clinic → location (by clinic-name substring in title/address)
    used = set()
    sti_to_loc = {}
    for s in sti:
        clinic = s["clinic"].lower()
        for L in locs:
            code = L.get("Shop code")
            if code in used: continue
            name = (L.get("Business name") or "").lower()
            addr = " ".join(str(L.get(f"Address line {i}") or "") for i in range(1, 6)).lower()
            if clinic in name or clinic in addr:
                sti_to_loc[s["clinic"]] = code
                used.add(code)
                break

    # Index keywords by store_code → list of (kw, intent, imp)
    by_store = defaultdict(list)
    for kw in kws:
        by_store[kw["store_code"]].append(kw)

    # Aggregate per tier
    tier_impressions = defaultdict(lambda: defaultdict(int))  # tier → intent → impressions
    tier_clinics = defaultdict(list)

    per_clinic_rows = []
    for s in sti:
        if s["avg_sti_pct"] is None or s["total_calls"] < 10: continue
        t = tier(s["avg_sti_pct"])
        if not t: continue
        store_code = sti_to_loc.get(s["clinic"])
        if not store_code: continue
        clinic_kws = by_store.get(store_code, [])
        if not clinic_kws: continue
        intent_imp = defaultdict(int)
        for k in clinic_kws:
            intent_imp[k["intent"]] += k["impressions_lower"]
        total_imp = sum(intent_imp.values())
        for intent, n in intent_imp.items():
            tier_impressions[t][intent] += n
        tier_clinics[t].append(s["clinic"])
        per_clinic_rows.append({
            "clinic": s["clinic"], "city": s["city"], "tier": t,
            "sti_pct": s["avg_sti_pct"],
            "total_imp": total_imp,
            **{f"imp_{intent}": intent_imp.get(intent, 0) for intent in ("sti","brand","sexologist","problem","generic","other")},
            **{f"share_{intent}": (intent_imp.get(intent, 0)/total_imp if total_imp else 0)*100 for intent in ("sti","brand","sexologist","problem","generic","other")},
        })

    # ==== REPORT ====
    print("=" * 80)
    print("SEARCH-KEYWORD INTENT MIX BY STI TIER")
    print("=" * 80)
    print(f"{'Tier':<8} {'n':>3} {'total imp':>12} {'sti%':>7} {'brand%':>8} {'sexol%':>8} {'problem%':>10} {'generic%':>10} {'other%':>7}")
    for t in ("top","mid","low"):
        ti = tier_impressions[t]
        tot = sum(ti.values())
        if not tot: continue
        share = {k: ti.get(k,0)/tot*100 for k in ("sti","brand","sexologist","problem","generic","other")}
        print(f"{t.upper():<8} {len(tier_clinics[t]):>3} {tot:>12,} {share['sti']:>6.1f}% {share['brand']:>7.1f}% {share['sexologist']:>7.1f}% {share['problem']:>9.1f}% {share['generic']:>9.1f}% {share['other']:>6.1f}%")
    print()

    # Per-clinic table sorted by STI share
    print("=" * 80)
    print("PER-CLINIC INTENT SHARE (sorted by STI share desc)")
    print("=" * 80)
    print(f"{'Clinic':<22} {'City':<14} {'STI%':>5} {'total imp':>10} {'sti%':>6} {'brand%':>7} {'sexol%':>7} {'prob%':>6} {'gen%':>6}")
    per_clinic_rows.sort(key=lambda r: -r["sti_pct"])
    for r in per_clinic_rows:
        print(f"{r['clinic']:<22} {r['city']:<14} {r['sti_pct']:>4.0f}% {r['total_imp']:>10,} {r['share_sti']:>5.1f}% {r['share_brand']:>6.1f}% {r['share_sexologist']:>6.1f}% {r['share_problem']:>5.1f}% {r['share_generic']:>5.1f}%")
    print()

    # Top STI-tagged keywords across all
    print("=" * 80)
    print("TOP STI-INTENT KEYWORDS (across all clinics, by impressions)")
    print("=" * 80)
    sti_kws = Counter()
    for k in kws:
        if k["intent"] == "sti":
            sti_kws[k["keyword"].lower()] += k["impressions_lower"]
    for kw, imp in sti_kws.most_common(20):
        print(f"  {imp:>6,}  {kw}")

    # Save per-clinic to CSV
    with open(HERE / "per_clinic_intent.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(per_clinic_rows[0].keys()))
        w.writeheader()
        w.writerows(per_clinic_rows)
    print(f"\nSaved per-clinic intent table → per_clinic_intent.csv ({len(per_clinic_rows)} rows)")


if __name__ == "__main__":
    main()
