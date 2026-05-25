"""
GMB full audit — joins three sources:
  1. sti_share.csv     (STI share % per clinic, internal data)
  2. locations.xlsx    (full GMB profile fields per location)
  3. insights.csv      (GMB performance metrics, last 30d)

For each location we compute:
  - STI tier (top/mid/low) from internal data
  - Profile completeness (categories, description, hours, photos, LGBTQ flag, links)
  - GMB performance (impressions, calls, directions, conversion rate)

Output:
  - Bucket comparison: what % of top-tier clinics have feature X vs low-tier
  - Per-clinic edit list for the low tier
"""

from __future__ import annotations

import csv
import re
import statistics
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent


# ----------- LOADERS -----------

def load_sti() -> list[dict]:
    rows = []
    with open(HERE / "sti_share.csv") as f:
        for r in csv.DictReader(f):
            r["avg_sti_pct"] = float(r["avg_sti_pct"]) if r["avg_sti_pct"] else None
            r["sti_calls"] = int(r["sti_calls"])
            r["total_calls"] = int(r["total_calls"])
            rows.append(r)
    return rows


def load_locations() -> list[dict]:
    wb = openpyxl.load_workbook(HERE / "locations.xlsx")
    ws = wb["Sheet1"]
    hdr = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(hdr, r)))
    return rows


def load_insights() -> list[dict]:
    """Insights CSV has a definition row at index 1 we need to skip."""
    with open(HERE / "insights.csv") as f:
        rdr = csv.reader(f)
        hdr = next(rdr)
        rows = []
        for r in rdr:
            if not r[0] or not r[0].strip().isdigit():
                continue  # skip definition rows
            d = dict(zip(hdr, r))
            for k in ("Google Search – Mobile","Google Search – Desktop","Google Maps – Mobile",
                      "Google Maps – Desktop","Calls","Messages","Bookings","Directions","Website clicks"):
                d[k] = int(d.get(k) or 0)
            d["impressions"] = d["Google Search – Mobile"] + d["Google Search – Desktop"] + d["Google Maps – Mobile"] + d["Google Maps – Desktop"]
            d["engagements"] = d["Calls"] + d["Directions"] + d["Website clicks"]
            d["engage_rate"] = d["engagements"] / d["impressions"] if d["impressions"] else 0
            rows.append(d)
    return rows


# ----------- MATCHING -----------

def match_sti_to_location(sti_rows, loc_rows):
    """Match by clinic-name substring against business name OR address."""
    out = []
    used = set()
    for s in sti_rows:
        clinic_lc = s["clinic"].lower()
        best = None
        for L in loc_rows:
            code = L.get("Shop code")
            if code in used:
                continue
            name = (L.get("Business name") or "").lower()
            addr = " ".join(str(L.get(f"Address line {i}") or "") for i in range(1, 6)).lower()
            if clinic_lc in name or clinic_lc in addr:
                best = L
                break
        if best:
            used.add(best.get("Shop code"))
        out.append({"sti": s, "loc": best})
    return out


def attach_insights(matched, insights):
    by_code = {i["Shop code"]: i for i in insights}
    for m in matched:
        loc = m["loc"]
        if loc:
            m["ins"] = by_code.get(loc.get("Shop code"))
        else:
            m["ins"] = None
    return matched


# ----------- PROFILE FEATURES -----------

def parse_addl_categories(s: str | None) -> list[str]:
    if not s:
        return []
    return [c.strip() for c in str(s).split(",") if c.strip()]


def profile_features(L: dict | None) -> dict:
    """Return a flat feature dict for one location row."""
    if not L:
        return {}
    addl = parse_addl_categories(L.get("Additional categories"))
    addl_updated = parse_addl_categories(L.get("[UPDATED] Additional categories"))
    addl_all = list({*addl, *addl_updated})  # union
    descr = L.get("From the business") or ""
    descr_lc = descr.lower()
    hours_cells = [L.get(f"{day} hours") for day in
                   ("Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday")]
    return {
        "status": L.get("Status"),
        "name": L.get("Business name"),
        "addr_l1": L.get("Address line 1"),
        "city": L.get("Locality"),
        "primary_category": L.get("Primary category"),
        "n_additional_categories": len(addl_all),
        "additional_categories": addl_all,
        "has_description": bool(descr.strip()),
        "description_len": len(descr),
        "description_has_sti": "sti" in descr_lc or "std" in descr_lc,
        "description_has_hiv": "hiv" in descr_lc,
        "description_has_discreet": "discreet" in descr_lc or "confidential" in descr_lc,
        "has_logo": bool(L.get("Logo photo")),
        "has_cover": bool(L.get("Cover photo")),
        "has_other_photos": bool(L.get("Other photos")),
        "hours_filled": sum(1 for h in hours_cells if h and str(h).strip()),
        "has_website": bool(L.get("Website")),
        "has_appointment_link": bool(L.get("Place page URLs: Appointment links (url_appointment)")),
        "has_instagram": bool(L.get("Place page URLs: Instagram (url_instagram)")),
        "has_facebook": bool(L.get("Place page URLs: Facebook (url_facebook)")),
        "has_whatsapp": bool(L.get("Place page URLs: WhatsApp (url_whatsapp)")),
        "lgbtq_friendly": L.get("Crowd: LGBTQ+ friendly (welcomes_lgbtq)") in ("TRUE", True, "true", "Yes"),
        "appt_required": L.get("Planning: Appointment required (requires_appointments)") in ("TRUE", True, "true", "Yes"),
        "phone_set": bool(L.get("Primary phone")),
        "wheelchair_entrance": L.get("Accessibility: Wheelchair-accessible entrance (has_wheelchair_accessible_entrance)") in ("TRUE", True, "true", "Yes"),
    }


# ----------- BUCKETING + AGGREGATION -----------

def tier_of(sti_pct):
    if sti_pct is None:
        return None
    if sti_pct >= 28:
        return "top"
    if sti_pct >= 12:
        return "mid"
    return "low"


def aggregate(matched: list[dict]):
    buckets = {"top": [], "mid": [], "low": []}
    for m in matched:
        s = m["sti"]
        if s["total_calls"] < 10 or s["avg_sti_pct"] is None:
            continue
        tier = tier_of(s["avg_sti_pct"])
        if tier:
            feats = profile_features(m["loc"])
            insights = m["ins"] or {}
            buckets[tier].append({"sti": s, "feat": feats, "ins": insights})
    return buckets


def pct(rows, key):
    if not rows:
        return 0.0
    return 100 * sum(1 for r in rows if r["feat"].get(key)) / len(rows)


def mean(rows, key, default=0.0):
    vals = [r["feat"].get(key) for r in rows if r["feat"].get(key) is not None]
    return statistics.mean(vals) if vals else default


def mean_ins(rows, key, default=0.0):
    vals = [r["ins"].get(key) for r in rows if r["ins"] and r["ins"].get(key) is not None]
    return statistics.mean(vals) if vals else default


# ----------- REPORTING -----------

def print_bucket_summary(b):
    print("=" * 80)
    print(f"{'BUCKET':<10} {'n':>4}  {'avg STI%':>10}  {'avg impressions':>18}  {'avg calls':>12}  {'engage %':>10}")
    print("-" * 80)
    for name in ("top", "mid", "low"):
        rows = b[name]
        if not rows:
            continue
        avg_sti = statistics.mean(r["sti"]["avg_sti_pct"] for r in rows)
        imp = statistics.mean(r["ins"].get("impressions", 0) for r in rows if r["ins"])
        cal = statistics.mean(r["ins"].get("Calls", 0) for r in rows if r["ins"])
        eng = statistics.mean(r["ins"].get("engage_rate", 0) for r in rows if r["ins"]) * 100
        print(f"{name.upper():<10} {len(rows):>4}  {avg_sti:>9.1f}%  {imp:>18,.0f}  {cal:>12,.0f}  {eng:>9.1f}%")
    print()


def print_feature_diff(b):
    feats = [
        ("Primary category set",                 "primary_category"),
        ("Description present",                  "has_description"),
        ("Description mentions STI/STD",         "description_has_sti"),
        ("Description mentions HIV",             "description_has_hiv"),
        ("Description says discreet/confidential", "description_has_discreet"),
        ("Logo photo uploaded",                  "has_logo"),
        ("Cover photo uploaded",                 "has_cover"),
        ("Other photos uploaded",                "has_other_photos"),
        ("LGBTQ+ friendly flag set",             "lgbtq_friendly"),
        ("Appointment required flag set",        "appt_required"),
        ("Appointment URL set",                  "has_appointment_link"),
        ("Instagram link set",                   "has_instagram"),
        ("Facebook link set",                    "has_facebook"),
        ("WhatsApp link set",                    "has_whatsapp"),
        ("Wheelchair-accessible entrance flag",  "wheelchair_entrance"),
    ]
    print(f"{'Feature':<42} {'Top':>10} {'Mid':>10} {'Low':>10}  {'Δ Top-Low':>10}")
    print("-" * 82)
    for label, key in feats:
        t, m, l = pct(b["top"], key), pct(b["mid"], key), pct(b["low"], key)
        print(f"{label:<42} {t:>9.0f}% {m:>9.0f}% {l:>9.0f}%  {t-l:>+9.0f}pp")
    print()
    print(f"{'Description length (chars, avg)':<42} {mean(b['top'],'description_len'):>9.0f}  {mean(b['mid'],'description_len'):>9.0f}  {mean(b['low'],'description_len'):>9.0f}")
    print(f"{'# Additional categories (avg)':<42} {mean(b['top'],'n_additional_categories'):>9.1f}  {mean(b['mid'],'n_additional_categories'):>9.1f}  {mean(b['low'],'n_additional_categories'):>9.1f}")
    print(f"{'# Hours filled (out of 7)':<42} {mean(b['top'],'hours_filled'):>9.1f}  {mean(b['mid'],'hours_filled'):>9.1f}  {mean(b['low'],'hours_filled'):>9.1f}")
    print()


def print_primary_categories(b):
    print("PRIMARY CATEGORIES BY TIER")
    print("-" * 80)
    for name in ("top", "mid", "low"):
        rows = b[name]
        c = defaultdict(int)
        for r in rows:
            cat = r["feat"].get("primary_category") or "(not set)"
            c[cat] += 1
        print(f"{name.upper()}:")
        for cat, n in sorted(c.items(), key=lambda x: -x[1]):
            print(f"  {n:>3}  {cat}")
    print()


def print_additional_categories_top(b, top_n=15):
    print(f"ADDITIONAL CATEGORIES — frequency by tier (top {top_n})")
    print("-" * 80)
    counts = {name: defaultdict(int) for name in ("top","mid","low")}
    for name in counts:
        for r in b[name]:
            for c in r["feat"].get("additional_categories", []):
                counts[name][c] += 1
    all_cats = sorted({c for d in counts.values() for c in d}, key=lambda c: -(counts['top'].get(c,0)+counts['low'].get(c,0)))
    print(f"{'Category':<45} {'Top':>6} {'Mid':>6} {'Low':>6}")
    for c in all_cats[:top_n]:
        n_top, n_mid, n_low = counts['top'].get(c,0), counts['mid'].get(c,0), counts['low'].get(c,0)
        print(f"{c:<45} {n_top:>6} {n_mid:>6} {n_low:>6}")
    print()


def print_low_tier_details(b):
    print("=" * 80)
    print(f"LOW-TIER CLINICS  n={len(b['low'])}")
    print("=" * 80)
    for r in sorted(b['low'], key=lambda r: r["sti"]["avg_sti_pct"]):
        s, f, i = r["sti"], r["feat"], r["ins"]
        print(f"\n  ── {s['clinic']} ({s['city']}) ────────────────────────────────")
        print(f"     STI %         : {s['avg_sti_pct']:.0f}%  ({s['sti_calls']}/{s['total_calls']} calls)")
        if i:
            print(f"     GMB impressions (30d): {i.get('impressions',0):,}    calls={i.get('Calls',0)}  dirs={i.get('Directions',0)}  web={i.get('Website clicks',0)}")
        if not f:
            print("     PROFILE        : NOT FOUND in GMB export")
            continue
        print(f"     STATUS         : {f.get('status')}")
        print(f"     Primary cat    : {f.get('primary_category')}")
        addl = f.get("additional_categories") or []
        print(f"     Additional cats: {', '.join(addl) if addl else '(none)'}")
        flags = []
        if not f.get("has_description"): flags.append("NO description")
        if not f.get("description_has_sti"): flags.append("description missing STI/STD keywords")
        if not f.get("description_has_discreet"): flags.append("description missing 'discreet/confidential'")
        if not f.get("has_logo"): flags.append("no logo")
        if not f.get("has_cover"): flags.append("no cover photo")
        if not f.get("has_other_photos"): flags.append("no other photos")
        if not f.get("lgbtq_friendly"): flags.append("LGBTQ+ flag off")
        if not f.get("has_appointment_link"): flags.append("no appointment URL")
        if not f.get("has_whatsapp"): flags.append("no WhatsApp URL")
        if (f.get("hours_filled") or 0) < 7: flags.append(f"hours only {f.get('hours_filled')}/7 days")
        print(f"     ISSUES         : {'; '.join(flags) if flags else '(none from this data)'}")


def main():
    sti = load_sti()
    locs = load_locations()
    ins = load_insights()
    print(f"Loaded: {len(sti)} STI rows, {len(locs)} locations, {len(ins)} insights rows\n")

    matched = match_sti_to_location(sti, locs)
    attach_insights(matched, ins)
    nm = sum(1 for m in matched if m["loc"])
    print(f"Matched STI→GMB: {nm}/{len(matched)}")
    nim = sum(1 for m in matched if m["ins"])
    print(f"With insights:   {nim}/{len(matched)}\n")

    b = aggregate(matched)
    print_bucket_summary(b)
    print_feature_diff(b)
    print_primary_categories(b)
    print_additional_categories_top(b)
    print_low_tier_details(b)


if __name__ == "__main__":
    main()
