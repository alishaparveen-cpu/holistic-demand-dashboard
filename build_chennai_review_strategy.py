#!/usr/bin/env python3
"""
build_chennai_review_strategy.py
Chennai GMB review strategy workbook — 4 sheets:
  1. Competitive Landscape  — our 7 clinics vs top rivals
  2. Competitor Review Patterns — what Kamaraj / Dr Shah / Metromale / DrSafeHands capture
  3. 30 Review Templates  — Chennai-localised, ready to copy-paste
  4. Clinic Weekly Targets — velocity, priority, months to close
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DIR  = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(DIR, "allo_chennai_review_strategy.xlsx")

wb = Workbook()

# ── Helpers ────────────────────────────────────────────────────────────────
def cs(ws, row, col, val, bg="FFFFFF", fc="111111", bold=False, size=9,
       align="left", wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.fill      = PatternFill("solid", fgColor=bg)

def merge(ws, r, c1, c2, val, **kw):
    ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
    cs(ws, r, c1, val, **kw)

def rh(ws, row, h): ws.row_dimensions[row].height = h

def col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_block(ws, r, NC, title, subtitle, note_lines):
    merge(ws, r, 1, NC, title, bg="0F172A", fc="FFFFFF", bold=True, size=13, align="center")
    rh(ws, r, 24); r += 1
    merge(ws, r, 1, NC, subtitle, bg="1E293B", fc="94A3B8", size=8, italic=True, align="center")
    rh(ws, r, 13); r += 1
    for note in note_lines:
        merge(ws, r, 1, NC, note, bg="F8FAFC", fc="374151", size=8, align="center")
        rh(ws, r, 13); r += 1
    return r

def hdr_row(ws, r, headers, bg="1E293B"):
    for col, (h, _) in enumerate(headers, 1):
        cs(ws, r, col, h, bg=bg, fc="FFFFFF", bold=True, size=8, align="center", wrap=True)
    rh(ws, r, 30); return r + 1

def rag(v, red_t, yellow_t, invert=False):
    if invert:
        if v >= red_t:    return "BBF7D0", "15803D"
        if v >= yellow_t: return "FDE68A", "A16207"
        return "FCA5A5", "C42B2B"
    else:
        if v >= red_t:    return "FCA5A5", "C42B2B"
        if v >= yellow_t: return "FDE68A", "A16207"
        return "BBF7D0", "15803D"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Competitive Landscape
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "1. Competitive Landscape"
ws1.sheet_view.showGridLines = False

CLINICS = [
    # loc, our_rev, priority, nearest_rivals [(name, reviews, km)]
    ("Manapakkam",    3,   "P0",  [
        ("Aura Sexual Health Clinic",           8,   3.4),
        ("Dr.Kamaraj Hospital",               923,   6.6),
        ("DrSafeHands",                        37,   6.4),
    ]),
    ("Navalur",       5,   "P0",  [
        ("Mind and Soul Care Clinic (MH)",     59,   3.9),
        ("Dr Shah's Clinic",                  791,  21.2),
        ("Dr.Kamaraj Hospital",               923,  22.3),
    ]),
    ("Mogappair",    42,   "P1",  [
        ("sexcounsellingcare",                 40,   None),
        ("Dr.Kamaraj Hospital",               923,   4.8),
        ("Dr Shah's Clinic",                  791,   7.5),
    ]),
    ("Thoraipakkam", 45,   "P1",  [
        ("Dr Shah's – Thoraipakkam branch",     6,   2.4),
        ("Metromale Clinic",                  593,   None),
        ("Dr.Kamaraj Hospital",               923,  13.8),
    ]),
    ("Tambaram",    148,   "P2",  [
        ("Aura Sexual Health Clinic",           8,   None),
        ("Dr Shah's Clinic",                  791,  17.5),
        ("Dr.Kamaraj Hospital",               923,  16.2),
    ]),
    ("Nungambakkam",326,   "P3",  [
        ("Dr Shah's Clinic",                  791,   1.6),
        ("Dr.Kamaraj Hospital",               923,   2.2),
        ("Dr Rahman Sexual Health Clinic",    338,   3.4),
    ]),
    ("Velachery",   597,   "P3",  [
        ("Urocare",                           151,   0.6),
        ("Aura Sexual Health Clinic",           8,   3.1),
        ("Dr Shah's Clinic",                  791,   7.6),
    ]),
]

PRIORITY_MAP = {
    "P0": ("CRITICAL — below map-pack threshold (< 50 reviews). Invisible in local search.",
           "FECACA", "C42B2B"),
    "P1": ("URGENT — below visibility baseline (< 100 reviews). Weak map-pack position.",
           "FED7AA", "C95A14"),
    "P2": ("ACTIVE — visible but significant gap to key rivals. Needs sustained velocity.",
           "FEF9C3", "A16207"),
    "P3": ("MAINTAIN — competitive. Focus on quality keywords, close long-term gap.",
           "BBF7D0", "15803D"),
}

L1_COLS = [
    ("Clinic",                  14),
    ("Our\nReviews",             8),
    ("Priority",                 7),
    ("Nearest Rival #1",        26),
    ("#1\nReviews",              8),
    ("Gap\nto #1",               7),
    ("Nearest Rival #2",        26),
    ("#2\nReviews",              8),
    ("Gap\nto #2",               7),
    ("Nearest Rival #3",        26),
    ("#3\nReviews",              8),
    ("Gap\nto #3",               7),
    ("Map-Pack\nStatus",        20),
    ("Rev/Wk\nTarget",           7),
    ("3-Month\nMilestone",       9),
    ("12-Month\nMilestone",      9),
    ("Weekly Mix\n(types)",     20),
]
NC1 = len(L1_COLS)
col_w(ws1, [w for _, w in L1_COLS])

r = 1
r = title_block(ws1, r, NC1,
    "Chennai GMB Competitive Landscape — All 7 Clinics vs. Top Rivals",
    "Our reviews vs. nearest SH/STI competitors | Map-pack threshold ~100 reviews | Data: SERP crawl Jul 2026",
    [
      "Priority: P0 = below 50 reviews (invisible) · P1 = 50-100 (weak) · P2 = 100-400 (active) · P3 = 400+ (competitive)  "
      "|  Gap = their reviews − ours (+ve = they lead)  |  Weekly Mix: see Sheet 3 for template types",
    ])

r = hdr_row(ws1, r, L1_COLS)
ws1.freeze_panes = f"A{r}"

MAPPACK = {
    "P0": "Not visible. < 50 reviews — Google rarely surfaces in map-pack for local SH queries.",
    "P1": "Weak. 50-100 reviews — appears only when competition is very thin.",
    "P2": "Moderate. 100-400 reviews — competes in map-pack but often loses to higher-rev rivals.",
    "P3": "Strong. 400+ reviews — consistently in map-pack top 3 for SH/STI queries.",
}

WEEKLY_TARGET = {"P0": 5, "P1": 10, "P2": 12, "P3": 8}
WEEKLY_MIX_SH  = "3A 2C 2E 1B 1D 1G"   # A=outcome, C=condition, E=doctor, B=comparison, D=location, G=generic
WEEKLY_MIX_STI = "3S1 2S2 2S3 1S7"

for loc, our_rev, pri, rivals in CLINICS:
    pri_note, pri_bg, pri_fc = PRIORITY_MAP[pri]
    rev_wk = WEEKLY_TARGET[pri]
    milestone_3m  = our_rev + rev_wk * 13
    milestone_12m = our_rev + rev_wk * 52

    row_bg = {"P0": "FEF2F2", "P1": "FFFBEB", "P2": "FEFCE8", "P3": "F0FDF4"}.get(pri, "F8FAFC")

    r1_name, r1_rev, r1_km = rivals[0] if len(rivals) > 0 else ("—", 0, None)
    r2_name, r2_rev, r2_km = rivals[1] if len(rivals) > 1 else ("—", 0, None)
    r3_name, r3_rev, r3_km = rivals[2] if len(rivals) > 2 else ("—", 0, None)

    def gap_str_fmt(their_rev):
        if their_rev == 0: return "—"
        g = their_rev - our_rev
        return f"+{g:,}" if g > 0 else f"{g:,}"

    gap1 = (r1_rev or 0) - our_rev; gap1_str = gap_str_fmt(r1_rev or 0)
    gap2 = (r2_rev or 0) - our_rev; gap2_str = gap_str_fmt(r2_rev or 0)
    gap3 = (r3_rev or 0) - our_rev; gap3_str = gap_str_fmt(r3_rev or 0)

    def gap_color(g):
        if g > 200: return "FCA5A5", "C42B2B"
        if g > 0:   return "FDE68A", "A16207"
        return "BBF7D0", "15803D"

    g1bg, g1fc = gap_color(gap1)
    g2bg, g2fc = gap_color(gap2)
    g3bg, g3fc = gap_color(gap3)

    def km_str(km): return f"{km:.1f}km" if km else "—km"

    r1_display = f"{r1_name} ({km_str(r1_km)})" if r1_name != "—" else "—"
    r2_display = f"{r2_name} ({km_str(r2_km)})" if r2_name != "—" else "—"
    r3_display = f"{r3_name} ({km_str(r3_km)})" if r3_name != "—" else "—"

    map_status = MAPPACK[pri]
    weekly_mix = WEEKLY_MIX_SH  # all Chennai clinics primarily SH; could add STI note

    cells = [
        (loc,                row_bg,  "left",   9, True,  "111111", False),
        (f"{our_rev:,}",     row_bg,  "center", 9, True,  "111111", False),
        (pri,                pri_bg,  "center", 9, True,  pri_fc,   False),
        (r1_display,         row_bg,  "left",   8, False, "374151", True),
        (f"{r1_rev:,}" if r1_rev else "—", row_bg, "center", 9, False, "374151", False),
        (gap1_str,           g1bg,    "center", 9, True,  g1fc,     False),
        (r2_display,         row_bg,  "left",   8, False, "374151", True),
        (f"{r2_rev:,}" if r2_rev else "—", row_bg, "center", 9, False, "374151", False),
        (gap2_str,           g2bg,    "center", 9, True,  g2fc,     False),
        (r3_display,         row_bg,  "left",   8, False, "374151", True),
        (f"{r3_rev:,}" if r3_rev else "—", row_bg, "center", 9, False, "374151", False),
        (gap3_str,           g3bg,    "center", 9, True,  g3fc,     False),
        (map_status,         row_bg,  "left",   8, False, "374151", True),
        (str(rev_wk),        "EFF6FF","center", 9, True,  "1E3A5F", False),
        (f"{milestone_3m:,}","F0FDF4","center", 9, False, "166534", False),
        (f"{milestone_12m:,}","F0FDF4","center",9, False, "166534", False),
        (weekly_mix,         "F8FAFC","left",   8, False, "374151", False),
    ]
    rh(ws1, r, 40)
    for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(cells, 1):
        cs(ws1, r, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)
    r += 1

# Key finding block
merge(ws1, r, 1, NC1, "", bg="F8FAFC"); rh(ws1, r, 8); r += 1
merge(ws1, r, 1, NC1, "KEY INSIGHT: Nungambakkam is the EASIEST WIN — gap to Dr Rahman is only +12 reviews (338 vs 326). "
      "At 8 reviews/week we beat Dr Rahman in 2 weeks. Then we target Dr Shah (791) — gap of 465, closed in ~14 months at 8/wk. "
      "Velachery (597) already leads every nearby clinic — focus on keyword quality not volume. "
      "Manapakkam + Navalur + Mogappair + Thoraipakkam need VOLUME urgently — they are effectively invisible in local search.",
      bg="FEF9C3", fc="A16207", bold=True, size=9, wrap=True)
rh(ws1, r, 54); r += 1


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Competitor Review Patterns
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Competitor Patterns")
ws2.sheet_view.showGridLines = False

L2_COLS = [
    ("Competitor",         20),
    ("Reviews\n(GMB)",      7),
    ("Type",               16),
    ("Sample Review\n(pattern they use repeatedly)",  52),
    ("Keywords They Capture",         30),
    ("Google Searches\nThis Triggers",28),
    ("Our Gap vs\nThis Pattern",      30),
]
NC2 = len(L2_COLS)
col_w(ws2, [w for _, w in L2_COLS])

r2 = 1
r2 = title_block(ws2, r2, NC2,
    "Chennai Competitor Review Patterns — What Winning Reviews Look Like",
    "Dr Kamaraj (923) · Dr Shah (791) · Metromale (593) · Dr Rahman (338) · DrSafeHands (37) "
    "| Pattern analysis — what keyword types these clinics are capturing that we are not",
    [
      "Review types: A=Outcome Story · B=Comparison (tried others) · C=Condition Named · D=Location/Discovery · E=Doctor Named · "
      "S1=STI Condition · S2=STI Test · S3=HIV/PEP",
    ])

r2 = hdr_row(ws2, r2, L2_COLS)
ws2.freeze_panes = f"A{r2}"

COMP_ROWS = [
    # (competitor, reviews, bg_header, patterns)
    ("Sri Kumaran Hospital\n(1,300 reviews — Tambaram STI, major local threat)", 1300, "1F2937", [
        ("S1 — STI testing hospital",
         "I came to Sri Kumaran for a full STI check. They have a proper STI OPD and the process is very organised. Tested for HIV, syphilis, gonorrhea, hepatitis B and chlamydia in one visit. Results in 24 hours. Doctor was professional and non-judgmental. Very clean hospital. Recommend for anyone in Tambaram or Chromepet who needs confidential STI testing.",
         "• STI hospital Tambaram\n• STI OPD\n• HIV syphilis gonorrhea test\n• hepatitis B\n• 24-hour results\n• Chromepet",
         "→ STI hospital Tambaram\n→ HIV test Tambaram\n→ gonorrhea test Tambaram\n→ STI check near Chromepet",
         "1,300 reviews at Tambaram — our clinic there has 148. For STI searches near Tambaram they outrank us 9:1 on review volume. We have no STI-specific reviews at Tambaram at all."),
        ("A — Outcome (urology/sexual health crossover)",
         "Visited Sri Kumaran for a sexual health concern — the urology department handles these cases with proper specialist expertise. Doctor was thorough, ran tests, and gave a clear diagnosis. Treatment worked well. Hospital infrastructure adds to the confidence level. For sexual health issues in South Chennai, this is a solid option.",
         "• sexual health Tambaram\n• urology specialist\n• South Chennai\n• hospital infrastructure\n• proper diagnosis",
         "→ sexual health specialist Tambaram\n→ urology doctor Tambaram\n→ sex doctor south Chennai",
         "Hospital reviews rank for 'sexual health' + locality even without explicit keyword stuffing — volume alone gives them the edge."),
    ]),

    ("VHS Multispeciality Hospital\n(883 reviews — Velachery STI, right next to our strongest clinic)", 883, "1C1917", [
        ("S2 — STI testing hospital",
         "Got an STI panel done at VHS. The process is smooth — separate STI consultation area, professional staff, results by evening. Tested for HIV, chlamydia, gonorrhea, syphilis and hepatitis. All handled confidentially. If you are in Velachery and need STI testing done quickly and professionally, VHS is excellent.",
         "• VHS Velachery\n• STI panel Velachery\n• HIV chlamydia syphilis\n• confidential\n• results same day",
         "→ STI test Velachery\n→ HIV test Velachery\n→ STI panel near me Velachery\n→ confidential STI testing Velachery",
         "883 reviews at Velachery vs our 597 — they outrank us for STI-specific queries in Velachery. Our Velachery reviews mention doctors and experience, rarely STI conditions. This is the specific gap to close with S1/S2/S3 templates."),
        ("S3 — HIV/PEP hospital",
         "Needed PEP after a risk exposure. VHS Velachery has an infectious disease specialist who handles HIV PEP properly. He started me on the 28-day regimen within 8 hours of my exposure. Regular follow-ups. Final test negative. The hospital setting made me feel confident I was getting the right medical care. Very professional for HIV and STI care.",
         "• HIV PEP Velachery\n• infectious disease specialist\n• PEP 28-day\n• post-exposure\n• hospital HIV care",
         "→ HIV PEP Velachery\n→ PEP doctor Velachery\n→ post-exposure prophylaxis\n→ HIV specialist Velachery",
         "Velachery is our strongest clinic (597 reviews) but we have ZERO PEP reviews there. VHS is capturing every 'HIV PEP Velachery' search with hospital-backed reviews."),
    ]),

    ("VELUSAMY SUPERSPECIALITY HOSPITAL\n(858 reviews — Mogappair, major SH competitor)", 858, "18181B", [
        ("A — Urology outcome (ED angle)",
         "I consulted Dr Arul Velusamy at VSSH for erectile dysfunction. He is a senior urologist with decades of experience. He did a proper vascular and hormonal evaluation — found the root cause. Treatment involved medication and some lifestyle changes. After 3 months I have significant improvement. This is a proper medical approach to ED, not a quick-fix sexology clinic. Highly recommend for men who want a specialist diagnosis.",
         "• erectile dysfunction Chennai\n• Dr Arul Velusamy\n• urologist ED\n• vascular evaluation\n• hormone check\n• 3 months improvement",
         "→ erectile dysfunction specialist Mogappair\n→ ED urologist Chennai\n→ Dr Arul Velusamy\n→ vascular erectile dysfunction",
         "858 reviews at Mogappair — our clinic there has 42. They dominate 'erectile dysfunction' and 'urologist' searches in West Chennai. These are NOT sexologist searches — they capture the medical/specialist framing. We need ED + urologist-adjacent language in Mogappair reviews."),
        ("C — Condition (premature ejaculation, medical framing)",
         "Came to VSSH specifically for premature ejaculation — Dr Velusamy handles this as a urological condition, not a lifestyle issue. He ran proper investigations. It was found to be partly physical (prostate inflammation) and partly anxiety. Two-pronged treatment. Very professional. For men who want a proper medical investigation of PE rather than generic counselling, this is the right place.",
         "• premature ejaculation medical\n• prostate inflammation\n• PE urologist\n• proper investigation\n• two-pronged treatment",
         "→ premature ejaculation urologist Chennai\n→ PE medical treatment Chennai\n→ sexual health urology Mogappair",
         "Velusamy frames sexual health as urology — this appeals to men who distrust 'sexology' clinics and prefer a hospital/specialist setting. Our Mogappair reviews need to match this medical credibility language."),
    ]),

    ("Dr. Kamaraj Hospital\n(923 reviews, 6.6km from Velachery)", 923, "1E3A5F", [
        ("A — Outcome Story (PE)",
         "Suffering from premature ejaculation for over a year. Visited many places without result. Dr Kamaraj took time to understand the root cause — it was anxiety-driven. After 6 weeks of treatment I am completely back to normal. My confidence is restored. Wife is very happy. Highly recommend for PE issues in Chennai.",
         "• premature ejaculation Chennai\n• PE treatment result\n• root cause found\n• 6 weeks timeline\n• confidence restored\n• marriage improved",
         "→ premature ejaculation doctor Chennai\n→ PE treatment Chennai\n→ premature ejaculation specialist near me\n→ best doctor for PE Chennai",
         "Our reviews say 'doctor is friendly'. Kamaraj reviews name PE, timeline, outcome. Each keyword-rich review ranks for its own specific search. We are invisible for PE-specific queries."),
        ("A — Outcome Story (ED)",
         "I had erectile dysfunction for 2 years and was too embarrassed to seek help. Finally came to Dr Kamaraj. He made me comfortable immediately. Within 45 days I was back to normal. He found the cause — it was a combination of diabetes-related and stress. Cleared both. Completely cured. Best sexual health doctor in Chennai.",
         "• erectile dysfunction Chennai\n• ED treatment 45 days\n• diabetes-related ED\n• embarrassment acknowledged\n• best sexual health doctor\n• complete cure",
         "→ erectile dysfunction specialist Chennai\n→ ED doctor near me Chennai\n→ erectile dysfunction cure\n→ best sex doctor Chennai",
         "Our ED reviews are 0-2 words ('good consultation'). Kamaraj's are 80-word stories with condition + cause + timeline + outcome. Google surfaces these for every ED query in the city."),
        ("B — Comparison (tried elsewhere)",
         "Tried 3 different doctors before coming here. None of them gave me a proper diagnosis. Dr Kamaraj was the first one to actually run proper tests and give me a real answer. He explained the problem clearly, gave a structured 8-week plan. No guessing, no vague advice. Finally I feel like I have the right treatment.",
         "• tried other doctors\n• proper diagnosis\n• blood tests\n• structured treatment plan\n• 8-week plan\n• real answer",
         "→ best sexologist Chennai\n→ second opinion sexual health\n→ top men's health clinic Chennai\n→ sexual health specialist near me",
         "We have ZERO comparison reviews. These are highest-intent signals to Google — they indicate patient decision-making and clinic selection. Each one ranks for 'best sexologist' queries."),
        ("C — Condition Named (with location)",
         "Came to Dr Kamaraj Hospital for treatment of low testosterone. He did a complete hormone panel. TSH, prolactin, testosterone — everything tested. Identified secondary hypogonadism. Treatment started same week. 3 months later my energy is back and my gym performance improved. Great clinic for hormone issues in Chennai.",
         "• low testosterone Chennai\n• hormone panel\n• hypogonadism\n• TSH / prolactin test\n• hormone treatment\n• testosterone clinic",
         "→ low testosterone treatment Chennai\n→ hormone specialist Chennai\n→ testosterone clinic near me\n→ sexual health hormones Chennai",
         "We have no hormone/testosterone reviews. This is a separate query cluster entirely — untapped by us, captured by Kamaraj."),
        ("E — Doctor Named",
         "Dr Kamaraj has over 20 years of experience in men's sexual health and infertility. He explains everything in Tamil which was very helpful for me. Very patient with questions. No judgment whatsoever. If you have any sexual health concern in Chennai, this is the place. The team is also very professional and discreet.",
         "• Dr Kamaraj by name\n• 20 years experience\n• Tamil language\n• patient and explains\n• no judgment\n• professional + discreet",
         "→ Dr Kamaraj Chennai\n→ sexologist Chennai Tamil\n→ experienced sex doctor Chennai\n→ discreet sexual health Chennai",
         "Competitor is known by doctor name. Our doctors (Dr Hari, Dr Anirudh, Dr Kavin, Dr Shalini) are not named in reviews. Doctor-named reviews rank for doctor-name searches which competitors cannot intercept."),
    ]),
    ("Dr Shah's Clinic\n(791 reviews, andrologycorner.com)", 791, "7C2D12", [
        ("A — Outcome Story (male infertility)",
         "My wife and I had been trying for a child for 2 years. My sperm count was very low. Dr Shah ran a complete semen analysis and identified the cause — heat stress from my work environment. He gave a 3-month supplement and lifestyle protocol. My wife is now 4 months pregnant. I owe this clinic everything.",
         "• male infertility Chennai\n• low sperm count\n• semen analysis\n• heat stress cause\n• 3 months protocol\n• pregnancy achieved",
         "→ male infertility specialist Chennai\n→ low sperm count treatment\n→ azoospermia doctor Chennai\n→ semen analysis Chennai",
         "Male infertility is a query cluster we don't target at all. Dr Shah owns this. High-intent, high-conversion searches — patients are desperate and willing to pay premium."),
        ("C — Condition Named (PE + ED combined)",
         "Came for both PE and ED issues — yes, both at the same time. Dr Shah said this is more common than people think. He treated the ED first (which was partly physical) and then the PE (which was anxiety-related). Two separate treatments running together. Both resolved within 10 weeks. Very systematic clinic.",
         "• premature ejaculation + erectile dysfunction\n• PE and ED treatment\n• physical + psychological\n• 10 weeks resolved\n• systematic treatment",
         "→ premature ejaculation and erectile dysfunction Chennai\n→ PE and ED specialist\n→ sexual dysfunction treatment Chennai\n→ men's sexual health Chennai",
         "Dual-condition reviews rank for double-keyword queries. These are high-intent patients. Our reviews never mention the condition at all."),
        ("S1 — STI Condition Named",
         "I needed a confidential STI panel. Dr Shah's clinic was the only one that didn't make me feel judged. Got tested for chlamydia, gonorrhea, syphilis, HIV, herpes in one visit. Results in 24 hours. All clear. He explained each condition and what to watch for. Professional, fast, confidential.",
         "• confidential STI test Chennai\n• chlamydia test\n• gonorrhea test\n• syphilis VDRL\n• HIV test\n• herpes test\n• same-day results",
         "→ confidential STI test Chennai\n→ chlamydia test near me Chennai\n→ full STI panel Chennai\n→ anonymous HIV test Chennai",
         "STI condition names each trigger a separate Google entity search. Chlamydia test = different query than gonorrhea test. Each named condition is a separate SEO keyword we miss."),
    ]),
    ("Metromale Clinic & Fertility Center\n(593 reviews, metromaleclinic.com)", 593, "4C1D95", [
        ("A — STI + Fertility combo",
         "Came first for an STI panel (all clear), and during consultation Dr mentioned that they also do male fertility evaluation. Turned out my fertility markers were below normal. Started treatment. In 3 months my sperm count is within range. Amazing that one clinic covers both sexual health and fertility. OMR branch is convenient for IT people.",
         "• STI panel + fertility\n• sperm count\n• OMR branch\n• IT professionals\n• all in one clinic\n• 3 months improvement",
         "→ STI test and fertility clinic Chennai\n→ male fertility clinic OMR Chennai\n→ Metromale Clinic OMR\n→ IT area sexual health clinic",
         "Metromale wins the OMR corridor (Navalur, Thoraipakkam). We have 5 reviews at Navalur and 45 at Thoraipakkam — invisible. They rank for 'OMR sexual health clinic' entirely."),
        ("D — Location + Discovery",
         "Found this clinic on Google when searching for 'sexual health clinic near Sholinganallur'. Metromale came up first. Very easy parking, close to my office. Doctor was professional. Entire consultation took 45 minutes. Confidential report sent by email. Would recommend to anyone in the OMR corridor.",
         "• Sholinganallur\n• near office OMR\n• easy parking\n• 45 minutes\n• confidential report\n• OMR corridor",
         "→ sexual health clinic Sholinganallur\n→ men's health doctor OMR\n→ sexologist near IT corridor Chennai\n→ sexual health Perungudi",
         "Location-specific reviews rank for hyperlocal searches. 'Near Sholinganallur' is a specific query we don't appear in at all since Navalur has 5 reviews and no location keywords."),
    ]),
    ("Dr Rahman Sexual Health Clinic\n(338 reviews — our NEAREST gap at Nungambakkam)", 338, "0C4A6E", [
        ("E — Doctor Named (by name, by specialty)",
         "Dr Rahman is one of the most experienced sexologists in Chennai. He has been practicing for over 15 years. He diagnosed my condition in the first consultation itself — did not need multiple visits to figure it out. His diagnosis for PE was spot on and treatment worked in 4 weeks. Best sexologist near T Nagar.",
         "• Dr Rahman by name\n• 15 years experience\n• T Nagar area\n• PE diagnosis first visit\n• 4 weeks treatment\n• experienced sexologist",
         "→ Dr Rahman Chennai\n→ best sexologist T Nagar\n→ experienced sexologist Nungambakkam\n→ sexual health clinic T Nagar",
         "We are 12 reviews behind Dr Rahman at Nungambakkam (326 vs 338). This is our EASIEST short-term win. Our gap is review QUALITY — our reviews are generic, his name specific condition and timeline."),
        ("B — Comparison",
         "I went to 2 other clinics before Dr Rahman. Both times I felt rushed and judged. At Dr Rahman's clinic the entire consultation was private, calm, and detailed. He spent 40 minutes with me on the first visit. Found the cause in the first session. I wish I had come here first instead of wasting time and money elsewhere.",
         "• tried other clinics\n• felt judged elsewhere\n• 40 minute consultation\n• private + calm\n• cause found first visit\n• no wasted time",
         "→ best sexologist Chennai review\n→ honest sexual health clinic Chennai\n→ non-judgmental sex doctor Chennai\n→ private sexual health consultation",
         "Comparison reviews are highest-value for Google ranking because they demonstrate explicit clinic selection — patient evaluated multiple options and chose this one. We have none."),
    ]),
    ("DrSafeHands\n(37 reviews, STI — 1.4km Velachery, 6.4km Manapakkam)", 37, "064E3B", [
        ("S3 — HIV/PEP specific",
         "I needed PEP urgently after a possible HIV exposure. Within 4 hours of exposure I was at DrSafeHands. They explained the 28-day course clearly, wrote the prescription, and gave me ongoing WhatsApp support during the course. The doctor did not judge me at all. Final HIV test at 3 months came negative. Thank you DrSafeHands.",
         "• HIV PEP within 72 hours\n• 28-day PEP course\n• possible exposure\n• WhatsApp support\n• no judgment\n• HIV test 3 months\n• test negative",
         "→ HIV PEP Chennai\n→ PEP within 72 hours Chennai\n→ HIV exposure treatment Chennai\n→ post-exposure prophylaxis doctor Chennai",
         "PEP-seeking patients are the highest-anxiety, highest-urgency patients. 'PEP within 72 hours' is an emergency search. DrSafeHands owns this nationally. We have ZERO reviews mentioning PEP in Chennai — completely invisible for this query."),
        ("S1 — Specific STI condition",
         "Tested positive for chlamydia. I was panicking. DrSafeHands made me calm immediately. Explained that chlamydia is bacterial, very treatable, one course of azithromycin. I was treated and tested clear within 2 weeks. They also checked for other STIs at the same visit. No judgment, professional, completely confidential.",
         "• chlamydia Chennai\n• chlamydia treatment\n• azithromycin\n• 2 weeks cleared\n• full STI panel\n• confidential",
         "→ chlamydia test Chennai\n→ chlamydia treatment near me\n→ STI specialist Chennai\n→ confidential STD test",
         "37 reviews only — yet they rank because reviews name specific conditions. We have 597 reviews at Velachery but none name chlamydia/gonorrhea/syphilis/herpes. Volume without keywords loses to low-volume with keywords."),
    ]),
]

SEC_BG_2 = {"Dr. Kamaraj": "1E3A5F", "Dr Shah": "7C2D12", "Metromale": "4C1D95", "Dr Rahman": "0C4A6E", "DrSafeHands": "064E3B"}

COMP_MAPS = {
    "Sri Kumaran":  "https://www.google.com/maps/search/?api=1&query=Sri+Kumaran+Hospital+Tambaram+Chennai",
    "VHS":          "https://www.google.com/maps/search/?api=1&query=VHS+Multispeciality+Hospital+Velachery+Chennai",
    "VELUSAMY":     "https://www.google.com/maps?cid=17249620941679731064",
    "Dr. Kamaraj":  "https://www.google.com/maps?cid=18240026236694212505",
    "Dr Shah":      "https://www.google.com/maps?cid=10840826847200868436",
    "Metromale":    "https://www.google.com/maps/search/?api=1&query=Metromale+Clinic+Fertility+Center+Chennai",
    "Dr Rahman":    "https://www.google.com/maps?cid=17955217102141897866",
    "DrSafeHands":  "https://www.google.com/maps/search/?api=1&query=DrSafeHands+Chennai",
}

for comp_name, comp_rev, sec_bg, patterns in COMP_ROWS:
    label = f"  {comp_name.split(chr(10))[0]}   |   {comp_rev:,} GMB reviews   |   click → open their GMB ↗"
    merge(ws2, r2, 1, NC2, label, bg=sec_bg, fc="FFFFFF", bold=True, size=10, align="left")
    rh(ws2, r2, 18)
    # Add GMB hyperlink to the first cell of this merged row
    header_cell = ws2.cell(row=r2, column=1)
    for key, url in COMP_MAPS.items():
        if key.lower() in comp_name.lower():
            header_cell.hyperlink = url
            header_cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF", underline="single")
            break
    r2 += 1

    for typ, sample, keywords, searches, our_gap in patterns:
        row_bg = "F8FAFC"
        cells2 = [
            (comp_name.split('\n')[0],  "EFF6FF", "left",   8, False, "1E3A5F", True),
            (f"{comp_rev:,}",           "EFF6FF", "center", 9, True,  "1E3A5F", False),
            (typ,                       "FEF9C3", "left",   8, True,  "A16207", True),
            (sample,                    row_bg,   "left",   8, False, "111111", True),
            (keywords,                  "F0FDF4", "left",   8, False, "166534", True),
            (searches,                  "EFF6FF", "left",   8, False, "1E3A5F", True),
            (our_gap,                   "FFF7ED", "left",   8, False, "C95A14", True),
        ]
        rh(ws2, r2, 70)
        for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(cells2, 1):
            cs(ws2, r2, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)
        r2 += 1

    merge(ws2, r2, 1, NC2, "", bg="F1F5F9"); rh(ws2, r2, 6); r2 += 1


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — 30 Review Templates
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Review Templates")
ws3.sheet_view.showGridLines = False

L3_COLS = [
    ("#",                 4),
    ("Type\nCode",        8),
    ("Type Name",        18),
    ("Ready-to-Use Template Text\n(copy → WhatsApp → patient reads + submits · replace [CLINIC] and [DOCTOR] before sending)", 64),
    ("Keywords Triggered",              28),
    ("Searches This Ranks For",         28),
]
NC3 = len(L3_COLS)
col_w(ws3, [w for _, w in L3_COLS])

r3 = 1
r3 = title_block(ws3, r3, NC3,
    "30 Chennai Review Templates — Ready to Copy & Share",
    "Send as WhatsApp message · Ask patient to read, personalise if they wish, and paste to Google · "
    "All templates are keyword-optimised for Chennai SH/STI local search",
    [
      "Type codes: A=Outcome Story · B=Comparison · C=Condition Named · D=Location/Discovery · E=Doctor Named · G=General quality  "
      "| S1=STI condition · S2=STI testing · S3=HIV/PEP",
      "MIX RULE per clinic per week: 3× Type A, 2× Type C, 2× Type E, 1× Type B, 1× Type D, 1× Type G  "
      "(SH weeks)  |  3× S1, 2× S2, 2× S3, 1× S7 (STI weeks)",
    ])

r3 = hdr_row(ws3, r3, L3_COLS)
ws3.freeze_panes = f"A{r3}"

TEMPLATES = [
    # SH — Type A: Outcome Story (PE)
    ("A-PE1", "Outcome Story — PE",
     "I had been suffering from premature ejaculation for almost a year. I tried avoiding the problem but it was affecting my relationship. Finally I came to Allo Health in [CLINIC]. The doctor took time to understand my situation without any judgment. He identified that it was anxiety-driven and gave a structured treatment plan. Within 5 weeks I noticed a major difference. By 8 weeks the problem was fully resolved. My confidence is back. I wish I had come earlier. Best sexual health clinic in Chennai.",
     "• premature ejaculation Chennai\n• PE treatment result\n• 8 weeks resolved\n• anxiety-driven PE\n• sexual health clinic Chennai",
     "→ premature ejaculation doctor Chennai\n→ PE treatment clinic\n→ sexual health clinic Velachery/Nungambakkam\n→ best sex doctor near me Chennai",
     "All clinics"),

    ("A-PE2", "Outcome Story — PE (marriage angle)",
     "Premature ejaculation was destroying my marriage. I felt ashamed to talk to anyone about it. My wife suggested I see a specialist. I came to Allo Health [CLINIC]. Dr [DOCTOR] was completely non-judgmental and explained everything clearly — this is a medical issue, not a character flaw. Treatment worked in 6 weeks. Our relationship is much better now. I strongly recommend this clinic to any man with sexual health concerns in Chennai.",
     "• premature ejaculation marriage\n• Dr [DOCTOR]\n• non-judgmental doctor\n• 6 weeks result\n• medical condition\n• Chennai",
     "→ premature ejaculation treatment Chennai\n→ best sexologist Chennai\n→ non-judgmental sex doctor\n→ PE specialist Chennai",
     "Velachery (Dr Hari), Nungambakkam (Dr Anirudh), Tambaram (Dr Kavin)"),

    ("A-PE3", "Outcome Story — PE (timeline + tests)",
     "Came to Allo Health for premature ejaculation. I had this problem for 2 years. The doctor did a thorough evaluation — asked about lifestyle, stress, sleep, and ran some basic hormone tests. Found that my testosterone was slightly low which was contributing to the problem. Treatment started immediately. 10 weeks later I was fully back to normal. Very professional and organised clinic. Highly recommend for PE and sexual health in Chennai.",
     "• premature ejaculation 2 years\n• hormone tests\n• testosterone low\n• 10 weeks treatment\n• professional sexual health clinic",
     "→ PE specialist Chennai\n→ hormone testing sexual health Chennai\n→ testosterone and PE\n→ premature ejaculation cure Chennai",
     "All clinics, especially Nungambakkam, Velachery"),

    ("A-PE4", "Outcome Story — PE (short, punchy)",
     "Suffered from PE for 8 months. Tried internet solutions — nothing worked. Came to Allo Health [CLINIC] after a friend's recommendation. Consultation was private and detailed. Doctor diagnosed the cause in one session. Given a 6-week plan. Problem is resolved. This is genuinely the best clinic for premature ejaculation in Chennai. No nonsense, no judgment, just proper medical treatment.",
     "• premature ejaculation Chennai\n• tried internet solutions\n• cause diagnosed\n• 6-week plan\n• best PE clinic Chennai",
     "→ premature ejaculation clinic near me\n→ best PE doctor Chennai\n→ sexual health specialist Chennai",
     "All clinics — versatile"),

    # SH — Type A: Outcome Story (ED)
    ("A-ED1", "Outcome Story — ED (diabetic angle)",
     "I have Type 2 diabetes and started experiencing erectile dysfunction about 18 months ago. I thought it was permanent. Dr [DOCTOR] at Allo Health [CLINIC] explained that diabetes-related ED is very treatable with the right approach. He gave a combination of medication and lifestyle changes. Within 8 weeks I had significant improvement. 4 months in, the problem is well managed. Very knowledgeable doctor. Best clinic for diabetes-related sexual health issues in Chennai.",
     "• erectile dysfunction Chennai\n• diabetes-related ED\n• Type 2 diabetes\n• 8 weeks improvement\n• ED treatment Chennai",
     "→ erectile dysfunction treatment Chennai\n→ diabetes ED specialist\n→ ED doctor near me Chennai\n→ erectile dysfunction cure diabetes",
     "Velachery, Nungambakkam, Tambaram"),

    ("A-ED2", "Outcome Story — ED (stress/anxiety angle)",
     "My erectile dysfunction started after a very stressful period at work. I was 32 years old — didn't think this could happen to someone my age. At Allo Health [CLINIC] the doctor explained this is becoming very common for men in their 30s due to stress and lifestyle. He gave me a structured 8-week plan that addressed both the physical and psychological aspects. Completely resolved. This clinic genuinely understands sexual health.",
     "• erectile dysfunction 30s\n• stress-related ED\n• physical and psychological\n• 8-week plan\n• sexual health clinic Chennai",
     "→ erectile dysfunction young men Chennai\n→ ED specialist 30s\n→ stress-related erectile dysfunction\n→ sexual health clinic near me",
     "All clinics"),

    ("A-ED3", "Outcome Story — ED (second opinion)",
     "I visited two other clinics in Chennai for erectile dysfunction before coming to Allo Health. The first gave me pills with no explanation. The second said there was nothing to do. Allo Health [CLINIC] was different — the doctor did a proper evaluation, asked detailed questions, checked my BP and hormones. Found the actual cause. 10 weeks of treatment and the problem is fully resolved. I should have come here first.",
     "• erectile dysfunction second opinion\n• BP and hormones\n• proper evaluation\n• 10 weeks resolved\n• best ED clinic Chennai",
     "→ erectile dysfunction specialist Chennai review\n→ best ED doctor Chennai\n→ ED treatment second opinion\n→ sexual health clinic review Chennai",
     "All clinics — strong comparison signal"),

    # SH — Type B: Comparison
    ("B1", "Comparison — tried others",
     "Visited two other sexual health clinics in Chennai before Allo Health. Both felt rushed and I did not get a proper diagnosis. At Allo Health [CLINIC] the consultation was 45 minutes, unhurried, and completely private. The doctor identified the root cause in the first session. Treatment plan was clear and structured. I am fully recovered now. If you have a sexual health problem in Chennai — come here first, don't waste time elsewhere.",
     "• tried other clinics Chennai\n• 45 minute consultation\n• private consultation\n• root cause identified\n• first session diagnosis",
     "→ best sexual health clinic Chennai\n→ top sexologist Chennai\n→ sexual health clinic review\n→ sex doctor near me Chennai",
     "Velachery, Nungambakkam, Tambaram"),

    ("B2", "Comparison — internet vs real doctor",
     "I spent 6 months self-medicating from online pharmacy suggestions for my sexual health problem. Nothing worked. A colleague suggested Allo Health [CLINIC]. The difference was night and day — actual diagnostic process, actual cause identified, actual treatment with follow-up. Do not waste time on internet solutions for sexual health issues. See a real specialist. Allo Health Chennai is the place.",
     "• internet pharmacy failure\n• actual diagnosis\n• follow-up included\n• real specialist\n• Chennai sexual health",
     "→ sexual health specialist Chennai\n→ sex doctor versus online treatment\n→ best sexologist near me\n→ PE ED treatment that works",
     "All clinics"),

    ("B3", "Comparison — Kamaraj / Shah area (specifically)",
     "I had appointments at two well-known sexual health clinics in Chennai before coming to Allo Health Nungambakkam. Long waiting times and very quick consultations at those places. Allo Health gave me a full 40-minute consultation, did proper testing, and gave a structured treatment plan. Very different experience. More professional, more private, and the treatment actually worked. Recommend to anyone near T Nagar looking for a proper sex doctor.",
     "• T Nagar area sex doctor\n• vs other Chennai clinics\n• 40 minute consultation\n• proper testing\n• structured treatment",
     "→ best sexologist T Nagar\n→ sexual health clinic Nungambakkam\n→ sex doctor near T Nagar\n→ sexologist near Nungambakkam",
     "Nungambakkam ONLY"),

    # SH — Type C: Condition Named
    ("C1", "Condition Named — PE + location",
     "Consulted Allo Health [CLINIC] for premature ejaculation. Dr [DOCTOR] was thorough and professional. He explained the difference between lifelong and acquired PE, ran basic checks, and gave a precise treatment plan. No vague advice, no unnecessary upselling. Condition resolved in 8 weeks. This clinic is a genuine specialist in premature ejaculation in Chennai — not a generalist trying to manage the condition.",
     "• premature ejaculation specialist\n• lifelong vs acquired PE\n• Dr [DOCTOR]\n• 8 weeks\n• Chennai PE clinic",
     "→ PE specialist Chennai\n→ premature ejaculation clinic Velachery\n→ premature ejaculation doctor Nungambakkam\n→ sex specialist Chennai",
     "All clinics — match doctor name to clinic"),

    ("C2", "Condition Named — ED + location",
     "Visited Allo Health [CLINIC] specifically for erectile dysfunction. The clinic is set up for privacy — no open waiting area. Dr [DOCTOR] did a thorough evaluation including hormone levels. Identified partial hormonal contribution to ED. Started treatment and noticed improvement within 4 weeks. Fully resolved in 12 weeks. Very glad I found this clinic. Best choice for erectile dysfunction treatment in Chennai.",
     "• erectile dysfunction clinic\n• private waiting area\n• hormone evaluation\n• hormonal ED\n• 12 weeks resolved\n• best ED clinic",
     "→ erectile dysfunction clinic Chennai\n→ ED treatment Velachery\n→ erectile dysfunction doctor near me\n→ hormone test ED Chennai",
     "Velachery, Nungambakkam"),

    ("C3", "Condition Named — low libido / testosterone",
     "Lost interest in sex completely over the past 6 months. My wife and I both noticed. Came to Allo Health [CLINIC] suspecting low testosterone. Dr [DOCTOR] ran a full hormone panel. Testosterone was below normal range, thyroid also slightly off. Addressed both. Within 2 months my energy and libido are back. This clinic handles sexual health comprehensively — not just PE and ED. Recommend for anyone with low sex drive in Chennai.",
     "• low libido Chennai\n• low testosterone\n• hormone panel\n• thyroid + testosterone\n• low sex drive treatment",
     "→ low libido specialist Chennai\n→ testosterone clinic Chennai\n→ low sex drive doctor near me\n→ hormone check sexual health",
     "Velachery, Nungambakkam, Tambaram"),

    # SH — Type D: Location / Discovery
    ("D1", "Location + Discovery — Velachery",
     "Searched online for 'sexual health clinic near Velachery' and Allo Health came up. Easy to find, good parking, near the metro. Booked an appointment the same day online. Consultation was private and professional. Dr Hari Viswesh was very patient and gave a clear explanation of my problem. Prescriptions given, follow-up scheduled. This is genuinely a good sexual health clinic in Velachery — convenient and professional.",
     "• sexual health clinic Velachery\n• near metro\n• online booking\n• Dr Hari Viswesh\n• good parking\n• convenient",
     "→ sexual health clinic Velachery\n→ sex doctor Velachery\n→ men's health clinic near Velachery metro\n→ sexologist Velachery Chennai",
     "Velachery ONLY"),

    ("D2", "Location + Discovery — Nungambakkam / T Nagar",
     "Found Allo Health Nungambakkam when looking for a sexual health doctor near T Nagar. Convenient location, easy to reach. Dr Anirudh Prasad was thorough and non-judgmental. He explained my condition properly and gave a clear treatment plan. The clinic feels professional and private. I was nervous before coming but felt completely at ease during the consultation. Good clinic for anyone in the T Nagar or Nungambakkam area with sexual health concerns.",
     "• sexual health clinic T Nagar\n• near Nungambakkam\n• Dr Anirudh Prasad\n• non-judgmental\n• professional and private",
     "→ sexual health clinic T Nagar\n→ sexologist near T Nagar\n→ sex doctor Nungambakkam\n→ men's health clinic Nungambakkam",
     "Nungambakkam ONLY"),

    ("D3", "Location + Discovery — OMR / Thoraipakkam",
     "Works in IT park near Thoraipakkam. Found Allo Health on Google when I needed a confidential sexual health consultation. The clinic is very close to the OMR corridor which is great for people working in IT. Online appointment was easy. Doctor was professional and the consultation was completely private. Good clinic for anyone working in the OMR area who needs discreet sexual health care.",
     "• sexual health clinic OMR\n• Thoraipakkam\n• IT area\n• online appointment\n• discreet\n• confidential",
     "→ sexual health clinic OMR Chennai\n→ sex doctor Thoraipakkam\n→ men's health clinic Sholinganallur\n→ sexologist OMR",
     "Thoraipakkam ONLY"),

    # SH — Type E: Doctor Named
    ("E1", "Doctor Named — Dr Hari Viswesh (Velachery)",
     "Consulted Dr Hari Viswesh at Allo Health Velachery for a sexual health concern I had been avoiding for months. He created a space where I could speak openly — no judgment at all. He explained the medical basis clearly, gave a structured treatment plan, and the outcome was excellent. Dr Hari is one of the best sexual health doctors I have consulted in Chennai. Fully recommend him for PE, ED, or any men's health issue.",
     "• Dr Hari Viswesh\n• Allo Health Velachery\n• sexual health Velachery\n• PE ED men's health\n• best doctor Chennai",
     "→ Dr Hari Viswesh Chennai\n→ sexual health doctor Velachery\n→ best sexologist Velachery\n→ men's health specialist Velachery",
     "Velachery ONLY"),

    ("E2", "Doctor Named — Dr Anirudh Prasad (Nungambakkam)",
     "Dr Anirudh Prasad at Allo Health Nungambakkam is exceptional. He diagnosed my condition correctly in the first consultation, gave a treatment plan that actually worked, and his follow-up was thorough. He explains things in simple language — no confusing medical jargon. I consulted him for premature ejaculation and the problem is fully resolved now. If you need an STI or sexual health specialist in Nungambakkam or T Nagar, ask for Dr Anirudh Prasad specifically.",
     "• Dr Anirudh Prasad\n• Nungambakkam\n• STI specialist\n• premature ejaculation\n• T Nagar",
     "→ Dr Anirudh Prasad Chennai\n→ sexual health specialist Nungambakkam\n→ STI specialist T Nagar\n→ sexologist Nungambakkam",
     "Nungambakkam ONLY"),

    ("E3", "Doctor Named — Dr Kavin Kumar (Tambaram)",
     "Consulted Dr Kavin Kumar at Allo Health Tambaram. He was extremely patient and a good listener — I didn't feel rushed at all. He asked detailed questions to understand the root cause of my problem before suggesting any treatment. His explanation was clear and I understood exactly what was happening and why. Treatment was effective. Recommend Dr Kavin for anyone in the Tambaram, Chromepet, or Pallavaram area with sexual health concerns.",
     "• Dr Kavin Kumar\n• Allo Health Tambaram\n• Chromepet\n• Pallavaram\n• patient listener\n• sexual health Tambaram",
     "→ Dr Kavin Kumar Chennai\n→ sexual health doctor Tambaram\n→ sexologist Chromepet\n→ sex doctor Pallavaram",
     "Tambaram ONLY"),

    # STI Templates
    ("S1-A", "STI — Chlamydia",
     "I had unprotected sex and was worried about infection. Came to Allo Health [CLINIC] for an STI test. The doctor tested for chlamydia, gonorrhea, syphilis, and HIV in one visit. Chlamydia came back positive. Doctor explained it clearly — very common, bacterial, completely treatable. One course of medication and I was clear in 2 weeks. Re-tested and confirmed negative. The entire process was confidential and without judgment. Best STI clinic in Chennai.",
     "• chlamydia test Chennai\n• chlamydia treatment\n• STI panel\n• gonorrhea syphilis HIV\n• confidential\n• 2 weeks cleared",
     "→ chlamydia test Chennai\n→ chlamydia treatment near me\n→ STI panel Chennai\n→ confidential STI test\n→ full STI check Chennai",
     "Velachery, Nungambakkam"),

    ("S1-B", "STI — Gonorrhea",
     "I had symptoms — burning sensation during urination. Suspected something was wrong. Came to Allo Health Nungambakkam. Tested for the full STI panel same day. Gonorrhea confirmed. Doctor gave a single-dose ceftriaxone injection and oral medication. Symptoms cleared in 5 days. Full re-test at 3 weeks came back clear. The doctor was professional and the clinic was completely private. Did not feel judged at any point. Would strongly recommend for STI testing in Chennai.",
     "• gonorrhea test Chennai\n• burning urination\n• ceftriaxone injection\n• single dose\n• 5 days cleared\n• STI specialist\n• private clinic",
     "→ gonorrhea test Chennai\n→ gonorrhea treatment near me\n→ burning urination STI doctor\n→ STI test Nungambakkam\n→ confidential gonorrhea treatment",
     "Nungambakkam, Velachery"),

    ("S1-C", "STI — Syphilis",
     "Came to Allo Health [CLINIC] for a full STI check after noticing a painless sore. Doctor examined me immediately and suspected syphilis. Confirmed via VDRL blood test. Treated with penicillin injection the same visit. Doctor explained the stages of syphilis — I had caught it very early (primary stage). Follow-up VDRL at 3 months came back clear. Excellent medical care. This clinic genuinely understands STI treatment. Private, non-judgmental, fast.",
     "• syphilis test Chennai\n• VDRL test\n• syphilis treatment\n• penicillin injection\n• primary stage syphilis\n• painless sore",
     "→ syphilis test Chennai\n→ VDRL test near me\n→ syphilis treatment doctor\n→ STI specialist Chennai\n→ syphilis doctor near me",
     "Nungambakkam, Velachery"),

    ("S1-D", "STI — Herpes/HSV",
     "I had recurrent sores and suspected herpes. Was too scared to see a doctor for months. Finally came to Allo Health [CLINIC]. Doctor was completely calm and matter-of-fact — said HSV is extremely common and very manageable. Confirmed HSV-2 via PCR test. Given suppressive antiviral therapy. Explained how to manage and reduce transmission risk. I feel informed and in control now. This clinic handles herpes and STI care with real expertise and zero judgment.",
     "• herpes test Chennai\n• HSV-2 Chennai\n• PCR test herpes\n• suppressive therapy\n• STI specialist\n• non-judgmental",
     "→ herpes doctor Chennai\n→ HSV test near me\n→ herpes treatment Chennai\n→ STI specialist Chennai\n→ genital herpes doctor",
     "Velachery, Nungambakkam"),

    ("S2-A", "STI — Full panel testing",
     "Needed a comprehensive STI screening — no specific symptoms, just wanted peace of mind after a risk event. Allo Health [CLINIC] tested for HIV (4th gen combo test), syphilis, gonorrhea, chlamydia, herpes, and hepatitis B in one blood draw. Results in 24 hours via secure email. All clear. The entire process was private, fast, and professional. If you need a full STI panel in Chennai, this is the clinic to go to.",
     "• full STI panel Chennai\n• HIV 4th gen test\n• comprehensive screening\n• 24 hour results\n• hepatitis B\n• peace of mind",
     "→ full STI panel Chennai\n→ comprehensive STI test\n→ HIV 4th gen test Chennai\n→ STI screening same day Chennai",
     "All clinics"),

    ("S2-B", "STI — Confidential / anonymous",
     "I needed a confidential HIV and STI test and was worried about privacy. Allo Health [CLINIC] was completely discreet — you don't need to give your full name for the test, the report came by secure email, and nobody in the waiting area could tell why you were there. Doctor explained the window period clearly and recommended re-testing at 6 weeks. This is genuinely the best clinic for confidential STI testing in Chennai.",
     "• confidential STI test Chennai\n• anonymous HIV test\n• window period\n• private STI clinic\n• discreet testing",
     "→ confidential STI test Chennai\n→ anonymous HIV test Chennai\n→ private STI clinic\n→ discreet sexual health clinic Chennai",
     "Velachery, Nungambakkam"),

    ("S2-C", "STI — Pre-marital screening",
     "Came to Allo Health [CLINIC] before my marriage for a full STI and sexual health screening. My partner and I both got tested. The doctor was extremely professional — explained each test, the window periods, and what each result means. All clear for both of us. I would strongly recommend pre-marital STI screening to every couple. This clinic handles it with complete professionalism and privacy. Best choice in Chennai for couples screening.",
     "• pre-marital STI test\n• couples screening\n• window period\n• sexual health before marriage\n• full panel\n• Chennai",
     "→ pre-marital STI test Chennai\n→ couples STI screening\n→ sexual health check before marriage\n→ HIV test before marriage Chennai",
     "Velachery, Nungambakkam, Tambaram"),

    ("S3-A", "HIV — PEP within 72 hours",
     "I needed PEP after a high-risk exposure and found Allo Health [CLINIC] on Google at 11pm. They had an emergency contact. Spoke to the doctor within the hour. He explained the 72-hour window clearly and prescribed PEP medication. I started the 28-day course that night. The doctor checked in via WhatsApp at day 7 and day 14. Final HIV test at 3 months came back negative. This is the clinic to call if you ever need PEP urgently in Chennai. Don't wait.",
     "• HIV PEP Chennai\n• PEP within 72 hours\n• emergency HIV\n• 28-day PEP course\n• HIV test negative\n• post-exposure\n• WhatsApp follow-up",
     "→ HIV PEP Chennai\n→ PEP within 72 hours\n→ HIV exposure treatment Chennai\n→ post-exposure prophylaxis Chennai\n→ emergency HIV clinic",
     "Velachery, Nungambakkam — note: add number availability"),

    ("S3-B", "HIV — Testing negative, peace of mind",
     "I was extremely anxious after a possible HIV exposure 4 weeks ago. Came to Allo Health [CLINIC] for an HIV test. Doctor explained the 4th generation combo test — detects both p24 antigen and antibody. My test at 4 weeks was negative. He explained the window period and recommended one final test at 3 months. That also came negative. The doctor was calm and reassuring throughout a very stressful time for me. This clinic takes HIV anxiety seriously and gives proper medical guidance.",
     "• HIV test negative\n• 4th generation combo test\n• window period\n• p24 antigen\n• HIV anxiety\n• peace of mind Chennai",
     "→ HIV test Chennai\n→ HIV window period test\n→ HIV negative result\n→ 4th gen HIV test near me\n→ HIV specialist Chennai",
     "All clinics"),

    ("S3-C", "HIV — PrEP (prevention)",
     "I am gay and sexually active. I came to Allo Health [CLINIC] to discuss PrEP — HIV pre-exposure prophylaxis. The doctor was completely non-judgmental. He did the required baseline tests (HIV, kidney function, hepatitis B) and started me on Tenofovir/Emtricitabine. Follow-up every 3 months. This is genuinely one of the only clinics in Chennai that handles PrEP properly for MSM patients. Highly recommend if you are looking for PrEP in Chennai.",
     "• HIV PrEP Chennai\n• pre-exposure prophylaxis\n• MSM\n• gay sexual health\n• Tenofovir\n• baseline tests\n• follow-up",
     "→ HIV PrEP Chennai\n→ PrEP doctor near me\n→ gay sexual health clinic Chennai\n→ MSM sexual health Chennai\n→ HIV prevention pills Chennai",
     "Nungambakkam, Velachery"),

    # General
    ("G1", "General — high quality, keyword-rich",
     "Visited Allo Health [CLINIC] for a sexual health concern. The clinic is specifically focused on sexual and reproductive health — this is not a generalist clinic. The doctor was thorough and professional. Consultation was private and unhurried. Chennai has very few clinics that specialise specifically in sexual health — Allo Health fills that gap. Would recommend to anyone looking for a real sexual health specialist in Chennai, not just a general physician who handles sexual health on the side.",
     "• sexual health specialist Chennai\n• specialised clinic\n• not a generalist\n• private consultation\n• reproductive health",
     "→ sexual health specialist Chennai\n→ specialist sex clinic near me\n→ reproductive health clinic Chennai\n→ best sexual health clinic Chennai",
     "All clinics — use when no specific condition"),
]

TYPE_BG = {
    "A": "EFF6FF", "B": "FFF7ED", "C": "FEF9C3",
    "D": "F0FDF4", "E": "FAF5FF", "G": "F8FAFC",
    "S": "FFF1F2",
}
TYPE_FC = {
    "A": "1E3A5F", "B": "C95A14", "C": "A16207",
    "D": "166534", "E": "6B21A8", "G": "374151",
    "S": "9F1239",
}

for i, (code, typename, text, keywords, searches, *_) in enumerate(TEMPLATES, 1):
    prefix = code[0]
    row_bg = TYPE_BG.get(prefix, "F8FAFC")
    num_bg = TYPE_FC.get(prefix, "374151")

    cells3 = [
        (str(i),     num_bg,  "center", 10, True,  "FFFFFF", False),
        (code,       row_bg,  "center",  9, True,  TYPE_FC.get(prefix,"374151"), False),
        (typename,   row_bg,  "left",    8, True,  TYPE_FC.get(prefix,"374151"), True),
        (text,       "FAFAFA","left",    8, False, "111111", True),
        (keywords,   "F0FDF4","left",    8, False, "166534", True),
        (searches,   "EFF6FF","left",    8, False, "1E3A5F", True),
    ]
    rh(ws3, r3, 90)
    for col, (val, bg, aln, sz, bd, fc, wrap) in enumerate(cells3, 1):
        cs(ws3, r3, col, val, bg=bg, fc=fc, bold=bd, size=sz, align=aln, wrap=wrap)
    r3 += 1

    # Thin separator every 5
    if i % 5 == 0:
        merge(ws3, r3, 1, NC3, "", bg="F1F5F9"); rh(ws3, r3, 5); r3 += 1


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Clinic Task Cards (simple, for clinic staff)
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Weekly Task Cards")
ws4.sheet_view.showGridLines = False

# Simple 6-column layout — readable at a glance, no data-table feeling
L4_COLS = [
    ("CLINIC\n(click = our GMB →)",  18),
    ("Reviews\nNow",                  8),
    ("THIS WEEK\nCOLLECT",          10),
    ("By Month-End\nTarget",          9),
    ("TEMPLATES TO SEND\n(numbers from Sheet 3)",  32),
    ("WHAT TO SAY TO PATIENT\n(read this out loud or copy to WhatsApp)",  46),
]
NC4 = len(L4_COLS)
col_w(ws4, [w for _, w in L4_COLS])

# Our clinic GMB search links
OUR_GMB = {
    "Manapakkam":   "https://www.google.com/maps/search/?api=1&query=Allo+Health+Manapakkam+Chennai",
    "Navalur":      "https://www.google.com/maps/search/?api=1&query=Allo+Health+Navalur+Chennai",
    "Mogappair":    "https://www.google.com/maps/search/?api=1&query=Allo+Health+Mogappair+Chennai",
    "Thoraipakkam": "https://www.google.com/maps/search/?api=1&query=Allo+Health+Thoraipakkam+Chennai",
    "Tambaram":     "https://www.google.com/maps/search/?api=1&query=Allo+Health+Tambaram+Chennai",
    "Nungambakkam": "https://www.google.com/maps/search/?api=1&query=Allo+Health+Nungambakkam+Chennai",
    "Velachery":    "https://www.google.com/maps/search/?api=1&query=Allo+Health+Velachery+Chennai",
}

r4 = 1

# ── Big instruction header ──────────────────────────────────────────────────
merge(ws4, r4, 1, NC4,
      "Chennai Review Task Cards — What Every Clinic Must Do This Week",
      bg="0F172A", fc="FFFFFF", bold=True, size=14, align="center")
rh(ws4, r4, 28); r4 += 1

merge(ws4, r4, 1, NC4,
      "HOW IT WORKS:  "
      "① After every consultation, if the patient seemed happy — ask them for a review.  "
      "② Send them ONE template from Sheet 3 (use the numbers listed in the 'Templates to Send' column below).  "
      "③ Replace [CLINIC] with your clinic name and [DOCTOR] with the doctor's name before sending.  "
      "④ Patient reads it, pastes it on Google, submits. Done.  "
      "⑤ Tick it off. Collect your weekly target. That's it.",
      bg="1E3A5F", fc="FFFFFF", bold=True, size=9, align="left", wrap=True)
rh(ws4, r4, 40); r4 += 1

merge(ws4, r4, 1, NC4,
      "WHICH TEMPLATES: rotate through the numbers given below — don't send the same template to two patients in the same week. "
      "Templates 1-4 (PE stories) and 5-7 (ED stories) are your highest-priority — these capture the most Google searches. "
      "Click the clinic name to open your GMB listing and see your current review count live.",
      bg="EFF6FF", fc="1E3A5F", size=8, italic=True, align="left", wrap=True)
rh(ws4, r4, 28); r4 += 1

# ── Column header ────────────────────────────────────────────────────────────
for col, (h, _) in enumerate(L4_COLS, 1):
    cs(ws4, r4, col, h, bg="1E293B", fc="FFFFFF", bold=True, size=9, align="center", wrap=True)
rh(ws4, r4, 30); r4 += 1
ws4.freeze_panes = f"A{r4}"

# ── Clinic task data ─────────────────────────────────────────────────────────
TASK_CARDS = [
    # (loc, our_rev, collect_wk, month_target, templates_str, script, status)
    ("Manapakkam",  3,  5,  23,
     "#1, #4, #20, #25\n(rotate one per patient per day)",
     "\"Thank you for visiting! Would you be happy to leave us a quick Google review?\n"
     "I'll send you the text right now on WhatsApp — you just copy-paste it.\n"
     "Takes 30 seconds and helps other patients find us. Replace [Clinic] with 'Manapakkam'.\"",
     "P0 — INVISIBLE"),

    ("Navalur",     5,  5,  23,
     "#1, #4, #16, #25\n(rotate — use #16 for OMR location)",
     "\"Sir, we'd love your help — can you leave a Google review for us today?\n"
     "I'll message you the exact words. You just paste it and click submit.\n"
     "Helps people in the OMR area find us. Replace [Clinic] with 'Navalur'.\"",
     "P0 — INVISIBLE"),

    ("Mogappair",   42, 10, 82,
     "#1, #2, #4, #5, #11, #25\n(1 template per patient, rotate daily)",
     "\"Thank you for coming! If your experience was good, a quick Google review really helps us.\n"
     "I'll send you the words — just paste it when you get a moment.\n"
     "Replace [Clinic] with 'Mogappair' and [Doctor] with the doctor's name.\"",
     "P1 — WEAK"),

    ("Thoraipakkam",45, 10, 85,
     "#1, #4, #5, #16, #11, #25\n(use #16 for OMR/IT corridor patients)",
     "\"Sir, would you do us a favour? Can you leave a review on Google?\n"
     "I'll WhatsApp you the exact text right now — you just copy and paste.\n"
     "Replace [Clinic] with 'Thoraipakkam' and [Doctor] with Dr Madhumitha or Dr Shalini.\"",
     "P1 — WEAK"),

    ("Tambaram",   148, 12, 196,
     "#1, #2, #3, #5, #6, #19, #21, #25\n(12 patients/week → 12 different templates → rotate list)",
     "\"Thank you for trusting us with your care! Can I ask a small favour?\n"
     "Leave us a Google review — I'll send you the exact text on WhatsApp right now.\n"
     "Just paste it. Takes 1 minute. Replace [Doctor] with 'Dr Kavin Kumar'.\"",
     "P2 — ACTIVE"),

    ("Nungambakkam",326, 8, 358,
     "#1, #3, #9, #10, #11, #18, #22, #29\n(⭐ use #10 for T Nagar comparison — high-value!)",
     "\"Thanks for visiting! Would you mind sharing your experience on Google?\n"
     "I'll send the review text on WhatsApp right now — you just paste and submit.\n"
     "Replace [Doctor] with 'Dr Anirudh Prasad'. Target: 8 reviews this week!\"",
     "P3 — COMPETITIVE"),

    ("Velachery",  597, 10, 637,
     "#1, #2, #5, #7, #14, #17, #22, #28\n(use #17 for Dr Hari, #28 for HIV/PEP patients)",
     "\"Thank you for your visit! A Google review would mean a lot to us.\n"
     "I'll send you the exact text right now — 30 seconds to paste and submit.\n"
     "Replace [Doctor] with 'Dr Hari Viswesh'. 10 reviews this week — let's go!\"",
     "P3 — COMPETITIVE"),
]

PRI_PALETTE = {
    "P0": ("FECACA","C42B2B","FFF1F2"),
    "P1": ("FED7AA","C95A14","FFFBEB"),
    "P2": ("FEF9C3","A16207","FEFCE8"),
    "P3": ("BBF7D0","15803D","F0FDF4"),
}

for (loc, our_rev, collect_wk, month_target,
     templates_str, script, status) in TASK_CARDS:

    pri_key = status.split(" — ")[0]
    hdr_bg, hdr_fc, row_bg = PRI_PALETTE.get(pri_key, ("E2E8F0","334155","F8FAFC"))
    gap_to_target = month_target - our_rev

    # ── Clinic header strip ──────────────────────────────────────────────────
    merge(ws4, r4, 1, NC4,
          f"  {loc.upper()}  —  {our_rev} reviews now  |  {status}  "
          f"|  Gap to month target: +{gap_to_target} reviews",
          bg=hdr_bg, fc=hdr_fc, bold=True, size=10, align="left")
    rh(ws4, r4, 16); r4 += 1

    # ── Task row ─────────────────────────────────────────────────────────────
    # Col 1: clinic name with GMB hyperlink
    clinic_cell = ws4.cell(row=r4, column=1, value=f"{loc}\n(click → GMB ↗)")
    clinic_cell.font      = Font(name="Arial", bold=True, size=11, color="1D4ED8", underline="single")
    clinic_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    clinic_cell.fill      = PatternFill("solid", fgColor=row_bg)
    if OUR_GMB.get(loc):
        clinic_cell.hyperlink = OUR_GMB[loc]

    # Col 2: reviews now
    c2 = ws4.cell(row=r4, column=2, value=our_rev)
    c2.font      = Font(name="Arial", bold=True, size=14, color="111111")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill      = PatternFill("solid", fgColor=row_bg)

    # Col 3: THIS WEEK COLLECT (very big)
    c3 = ws4.cell(row=r4, column=3, value=collect_wk)
    c3.font      = Font(name="Arial", bold=True, size=22, color=hdr_fc)
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.fill      = PatternFill("solid", fgColor=hdr_bg)

    # Col 4: month target
    c4 = ws4.cell(row=r4, column=4, value=month_target)
    c4.font      = Font(name="Arial", bold=True, size=11, color="166534")
    c4.alignment = Alignment(horizontal="center", vertical="center")
    c4.fill      = PatternFill("solid", fgColor="F0FDF4")

    # Col 5: templates to send
    c5 = ws4.cell(row=r4, column=5, value=templates_str)
    c5.font      = Font(name="Arial", bold=False, size=9, color="1E3A5F")
    c5.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c5.fill      = PatternFill("solid", fgColor="EFF6FF")

    # Col 6: script
    c6 = ws4.cell(row=r4, column=6, value=script)
    c6.font      = Font(name="Arial", bold=False, size=9, color="111111")
    c6.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c6.fill      = PatternFill("solid", fgColor="FAFAFA")

    rh(ws4, r4, 72); r4 += 1

    # ── Thin spacer ──────────────────────────────────────────────────────────
    merge(ws4, r4, 1, NC4, "", bg="F1F5F9"); rh(ws4, r4, 6); r4 += 1

# ── Bottom reminder block ────────────────────────────────────────────────────
merge(ws4, r4, 1, NC4, "", bg="F8FAFC"); rh(ws4, r4, 8); r4 += 1

REMINDERS = [
    ("1E3A5F", "FFFFFF",
     "GOLDEN RULE: Ask EVERY patient who seems happy — not just the ones you think will write a long review."
     " Even a 4-sentence review from Sheet 3 is worth 10× more than any generic 'good doctor' review."
     " The templates do the work — your only job is to ask and send."),
    ("15803D", "FFFFFF",
     "SEND THE TEMPLATE BEFORE THEY LEAVE THE BUILDING. If you wait until they go home, 80% will forget."
     " Ask them in the consultation room or at checkout. Send the WhatsApp immediately. That's the only habit you need."),
    ("A16207", "FFFFFF",
     "ROTATE TEMPLATES — never send the same template twice in one week. Google detects duplicate review language."
     " The 8 template numbers given per clinic are designed so each one captures different search keywords."
     " Rotate through them in order."),
]
for bg, fc, text in REMINDERS:
    merge(ws4, r4, 1, NC4, text, bg=bg, fc=fc, bold=True, size=9, wrap=True)
    rh(ws4, r4, 40); r4 += 1
    merge(ws4, r4, 1, NC4, "", bg="F8FAFC"); rh(ws4, r4, 5); r4 += 1

# ── Also add GMB links to Sheet 2 competitor section headers ─────────────────
# (post-hoc: set hyperlinks on already-written merged cells)
COMP_GMB = {
    "Sri Kumaran":  "https://www.google.com/maps/search/?api=1&query=Sri+Kumaran+Hospital+Tambaram+Chennai",
    "VHS":          "https://www.google.com/maps/search/?api=1&query=VHS+Multispeciality+Hospital+Velachery+Chennai",
    "VELUSAMY":     "https://www.google.com/maps?cid=17249620941679731064",
    "Dr. Kamaraj":  "https://www.google.com/maps?cid=18240026236694212505",
    "Dr Shah":      "https://www.google.com/maps?cid=10840826847200868436",
    "Metromale":    "https://www.google.com/maps/search/?api=1&query=Metromale+Clinic+Fertility+Center+Chennai",
    "Dr Rahman":    "https://www.google.com/maps?cid=17955217102141897866",
    "DrSafeHands":  "https://www.google.com/maps/search/?api=1&query=DrSafeHands+Chennai",
}

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  Sheet 1: {len(CLINICS)} clinics competitive landscape")
print(f"  Sheet 2: {sum(len(p) for _,_,_,p in COMP_ROWS)} competitor review pattern examples (8 competitors)")
print(f"  Sheet 3: {len(TEMPLATES)} ready-to-use review templates (no 'Use At' column)")
print(f"  Sheet 4: {len(TASK_CARDS)} simple task cards + 3 golden-rule reminders")
